#!/usr/bin/env python3
"""
Script independiente para unificar archivos Excel de pedidos.
Extrae NOMBRE, TELEFONO, CORREO ELECTRONICO de múltiples .xls y genera
un único Excel organizado alfabéticamente por nombre.

Archivos de entrada:
  - pedidos san vicente 38 actualizado.xls
  - pedidos por alberto aguilera 12.xls
  - pedidos martinez izquierdo.xls
  - pedidos ibiza.xls
  - pedidos doctor castelo 13.xls
  - pedidos chamberi.xls
  - pedido general diaz porlier 18.xls

Requisito: pip install xlrd openpyxl
"""

import os
import re
import unicodedata
from pathlib import Path

try:
    import xlrd
except ImportError:
    print("ERROR: Necesitas instalar xlrd para leer archivos .xls")
    print("  pip install xlrd")
    raise SystemExit(1)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ARCHIVOS = [
    "pedidos san vicente 38 actualizado.xls",
    "pedidos por alberto aguilera 12.xls",
    "pedidos martinez izquierdo.xls",
    "pedidos ibiza.xls",
    "pedidos doctor castelo 13.xls",
    "pedidos chamberi.xls",
    "pedido general diaz porlier 18.xls",
]

# Mapeo columnas origen -> destino
# Header: ORIGEN, ESTADO, DESCRIPCIÓN, REF. CLIENTE, USUARIO, EMAIL, TELÉFONO, MENSAJE, FECHA
COL_ORIGEN = 0
COL_ESTADO = 1
COL_DESCRIPCION = 2
COL_REF_CLIENTE = 3
COL_USUARIO = 4   # -> NOMBRE
COL_EMAIL = 5     # -> CORREO ELECTRONICO
COL_TELEFONO = 6  # -> TELEFONO
COL_MENSAJE = 7
COL_FECHA = 8


def clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return s


def clean_phone(val):
    s = clean(val)
    s = re.sub(r"[^\d+]", "", s)
    return s


def sort_key(name):
    s = (name or "").upper().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]", "", s)
    return s


def leer_archivo(ruta: str):
    """Lee un .xls y devuelve lista de dicts con NOMBRE, TELEFONO, CORREO ELECTRONICO, ORIGEN_ARCHIVO."""
    if not os.path.exists(ruta):
        return []
    registros = []
    wb = xlrd.open_workbook(ruta)
    for sh in wb.sheets():
        for row_idx in range(1, sh.nrows):
            usuario = clean(sh.cell_value(row_idx, COL_USUARIO))
            email = clean(sh.cell_value(row_idx, COL_EMAIL))
            telefono = clean(sh.cell_value(row_idx, COL_TELEFONO))
            if not usuario and not email and not telefono:
                continue
            nombre = usuario if usuario else ""
            telefono_limpio = telefono
            if telefono and not re.search(r"\d", telefono):
                telefono_limpio = ""
            registros.append({
                "NOMBRE": nombre,
                "TELEFONO": telefono_limpio,
                "CORREO ELECTRONICO": email,
                "ORIGEN_ARCHIVO": os.path.basename(ruta),
            })
    return registros


def deduplicar(registros: list[dict]) -> list[dict]:
    """Elimina duplicados por (nombre, telefono, email). Mantiene el primero."""
    vistos = set()
    unicos = []
    for r in registros:
        clave = (
            (r["NOMBRE"] or "").strip().lower(),
            (r["TELEFONO"] or "").replace(" ", ""),
            (r["CORREO ELECTRONICO"] or "").strip().lower(),
        )
        if clave in vistos:
            continue
        if clave == ("", "", ""):
            continue
        vistos.add(clave)
        unicos.append(r)
    return unicos


def main():
    base_dir = Path(__file__).parent
    os.chdir(base_dir)

    todos = []
    for f in ARCHIVOS:
        ruta = base_dir / f
        if not ruta.exists():
            print(f"  [AVISO] No existe: {f}")
            continue
        regs = leer_archivo(str(ruta))
        print(f"  {f}: {len(regs)} registros")
        todos.extend(regs)

    print(f"\nTotal antes de deduplicar: {len(todos)}")
    todos = deduplicar(todos)
    print(f"Total después de deduplicar: {len(todos)}")

    # Ordenar: primero por nombre (A-Z), luego los sin nombre por teléfono, luego por email
    def orden(r):
        nombre = (r["NOMBRE"] or "").strip()
        tel = (r["TELEFONO"] or "").replace(" ", "")
        email = (r["CORREO ELECTRONICO"] or "").strip().lower()
        if nombre:
            return (0, sort_key(nombre), tel, email)
        if tel:
            return (1, tel, email, "")
        return (2, email, tel, "")

    todos.sort(key=orden)

    # Separar: con nombre (por letra A-Z) vs solo números (sin nombre)
    con_nombre = [r for r in todos if (r.get("NOMBRE") or "").strip()]
    solo_numeros = [r for r in todos if not (r.get("NOMBRE") or "").strip()]

    def primera_letra(nombre):
        key = sort_key(nombre)
        if key and key[0].isalpha():
            return key[0]
        return "#"

    from collections import OrderedDict
    grupos = OrderedDict()
    for r in con_nombre:
        letra = primera_letra(r.get("NOMBRE", ""))
        if letra not in grupos:
            grupos[letra] = []
        grupos[letra].append(r)

    # Crear Excel estilo LLAMADAS TORREJON: hojas por letra + hoja "Solo números"
    wb = Workbook()
    wb.remove(wb.active)

    headers = ["NOMBRE", "TELEFONO", "CORREO ELECTRONICO"]
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="C0C0C0"),
        right=Side(style="thin", color="C0C0C0"),
        top=Side(style="thin", color="C0C0C0"),
        bottom=Side(style="thin", color="C0C0C0"),
    )
    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    data_align = Alignment(vertical="center")

    def escribir_hoja(ws, registros):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for r, reg in enumerate(registros, 2):
            for c, key in enumerate(headers, 1):
                cell = ws.cell(row=r, column=c, value=reg.get(key, ""))
                cell.border = thin_border
                cell.alignment = data_align
                if r % 2 == 0:
                    cell.fill = even_fill
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 40
        if registros:
            ws.auto_filter.ref = f"A1:C{len(registros) + 1}"
        ws.freeze_panes = "A2"

    # Hoja RESUMEN
    ws_res = wb.create_sheet("RESUMEN", 0)
    ws_res.cell(row=1, column=1, value="RESUMEN - PEDIDOS UNIFICADOS (por letra A-Z)").font = Font(bold=True, size=14, color="1B5E20")
    ws_res.merge_cells("A1:C1")
    ws_res.cell(row=3, column=1, value="LETRA").font = header_font
    ws_res.cell(row=3, column=1).fill = header_fill
    ws_res.cell(row=3, column=2, value="CANTIDAD").font = header_font
    ws_res.cell(row=3, column=2).fill = header_fill
    ws_res.cell(row=3, column=3, value="HOJA").font = header_font
    ws_res.cell(row=3, column=3).fill = header_fill
    row_res = 4
    for letra, regs in grupos.items():
        ws_res.cell(row=row_res, column=1, value=letra)
        ws_res.cell(row=row_res, column=2, value=len(regs))
        ws_res.cell(row=row_res, column=3, value=f"Ver hoja '{letra}'")
        row_res += 1
    if solo_numeros:
        ws_res.cell(row=row_res, column=1, value="Solo números")
        ws_res.cell(row=row_res, column=2, value=len(solo_numeros))
        ws_res.cell(row=row_res, column=3, value="Ver hoja 'Solo números'")
        row_res += 1
    ws_res.cell(row=row_res + 1, column=1, value="TOTAL").font = Font(bold=True)
    ws_res.cell(row=row_res + 1, column=2, value=len(todos)).font = Font(bold=True)
    ws_res.column_dimensions["A"].width = 15
    ws_res.column_dimensions["B"].width = 12
    ws_res.column_dimensions["C"].width = 25

    # Hoja "Solo números" (solo teléfono/email, sin nombre)
    if solo_numeros:
        ws_num = wb.create_sheet("Solo números")
        escribir_hoja(ws_num, solo_numeros)
        print(f"  Hoja 'Solo números': {len(solo_numeros)} registros")

    # Hojas por letra A-Z
    for letra in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        if letra in grupos:
            ws = wb.create_sheet(letra)
            escribir_hoja(ws, grupos[letra])
            print(f"  Hoja '{letra}': {len(grupos[letra])} registros")

    output_path = base_dir / "PEDIDOS_UNIFICADOS.xlsx"
    wb.save(str(output_path))
    print(f"\n*** Archivo generado: {output_path} ***")
    print(f"*** {len(todos)} registros | {len(grupos)} letras + 1 hoja 'Solo números' ***")


if __name__ == "__main__":
    print("Unificando pedidos...")
    main()
