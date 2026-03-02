#!/usr/bin/env python3
"""
🏭 GENERADOR DE CLIENTES FICTICIOS v2
Escribe TODOS los datos del cliente en TODAS las hojas:
- inf.gral 1: Titular, empresa, NIF, dirección, teléfono, email, año
- 2.1. DATOS PARCELAS: Parcelas con municipios y cultivos del cliente
- inf.trat 1: Tratamientos
"""
import sys
import json
import random
import shutil
from pathlib import Path
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

sys.path.insert(0, str(Path(__file__).parent))


# ============================================
# Clientes Ficticios con TODOS los datos
# ============================================

CLIENTES = {
    "juan_garcia_guisantes": {
        "id": "juan_garcia_guisantes",
        # DATOS EMPRESA (inf.gral 1)
        "razon_social": "GRANJA DE GUISANTES GARCIA S.L.",
        "titular": "JUAN GARCIA HERNANDEZ",
        "nif": "B37123456",
        "nif_titular": "12345678A",
        "reg_nacional": "ES370001234",
        "reg_autonomico": "371234567",
        "direccion": "CTRA. SALAMANCA KM 23",
        "localidad": "ALBA DE TORMES",
        "cp": "37800", 
        "provincia": "SALAMANCA",
        "telefono_fijo": "923456789",
        "telefono_movil": "687123456",
        "email": "juan@granjagarcia.es",
        "year": 2026,
        # DATOS PARCELAS
        "municipios": [
            "108-ALBA DE TORMES",
            "152-TERRADILLOS", 
            "089-CALVARRASA DE ARRIBA",
            "178-MOZARBEZ"
        ],
        "cultivos": [
            ("GUISANTE", 0.40),
            ("LENTEJA", 0.25),
            ("GARBANZO", 0.20),
            ("VEZA", 0.15)
        ],
        "productos": [
            {"nombre": "ROUNDUP ENERGY PRO", "registro": "21001", "dosis": 3.0, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "CLORPIRIFOS 48%", "registro": "19887", "dosis": 1.5, "unidad": "l/ha", "plaga": "PULGON"},
            {"nombre": "AZUFRE MOJABLE 80%", "registro": "15234", "dosis": 2.5, "unidad": "kg/ha", "plaga": "OIDIO"},
            {"nombre": "DECIS PROTECH", "registro": "22456", "dosis": 0.5, "unidad": "l/ha", "plaga": "GORGOJO"}
        ],
        "parcelas": 85,
        "aplicaciones_por_parcela": 2
    },
    
    "pedro_perez_maiz": {
        "id": "pedro_perez_maiz",
        # DATOS EMPRESA
        "razon_social": "SIEMBRA DE MAIZ PEREZ E HIJOS S.A.",
        "titular": "PEDRO PEREZ MARTINEZ",
        "nif": "A47654321",
        "nif_titular": "87654321B",
        "reg_nacional": "ES470005678",
        "reg_autonomico": "475678901",
        "direccion": "AVDA. DE LA AGRICULTURA 45",
        "localidad": "MEDINA DEL CAMPO",
        "cp": "47400",
        "provincia": "VALLADOLID",
        "telefono_fijo": "983654321",
        "telefono_movil": "612987654",
        "email": "pedro@siembraperez.com",
        "year": 2026,
        # DATOS PARCELAS
        "municipios": [
            "201-MEDINA DEL CAMPO",
            "234-RUEDA",
            "189-LA SECA",
            "215-POZALDEZ",
            "223-RODILANA"
        ],
        "cultivos": [
            ("MAIZ", 0.50),
            ("GIRASOL", 0.25),
            ("REMOLACHA AZUCARERA", 0.15),
            ("ALFALFA", 0.10)
        ],
        "productos": [
            {"nombre": "LAUDIS OD", "registro": "23567", "dosis": 1.2, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "CALLISTO", "registro": "24789", "dosis": 1.0, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "KARATE ZEON", "registro": "18234", "dosis": 0.4, "unidad": "l/ha", "plaga": "TALADRO DEL MAIZ"},
            {"nombre": "FOLICUR 25 WG", "registro": "20567", "dosis": 1.0, "unidad": "l/ha", "plaga": "ROYA"}
        ],
        "parcelas": 120,
        "aplicaciones_por_parcela": 3
    },
    
    "enrique_iglesias_agricola": {
        "id": "enrique_iglesias_agricola",
        # DATOS EMPRESA
        "razon_social": "GRANJA AGRICOLA IGLESIAS HNOS. S.L.",
        "titular": "ENRIQUE IGLESIAS LOPEZ",
        "nif": "B49112233",
        "nif_titular": "11223344C",
        "reg_nacional": "ES490009012",
        "reg_autonomico": "499012345",
        "direccion": "CAMINO DE LA VEGA S/N",
        "localidad": "TORO",
        "cp": "49800",
        "provincia": "ZAMORA",
        "telefono_fijo": "980123456",
        "telefono_movil": "678456123",
        "email": "enrique@granjaiglesias.es",
        "year": 2026,
        # DATOS PARCELAS
        "municipios": [
            "301-TORO",
            "298-MORALES DE TORO",
            "312-VENIALBO",
            "325-VILLABUENA DEL PUENTE",
            "287-PELEAGONZALO",
            "334-SANZOLES"
        ],
        "cultivos": [
            ("VID", 0.35),
            ("CEBADA", 0.25),
            ("TRIGO BLANDO", 0.20),
            ("ALMENDRO", 0.12),
            ("OLIVO", 0.08)
        ],
        "productos": [
            {"nombre": "TOUCHDOWN PREMIUM", "registro": "21001", "dosis": 3.0, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "CUPROSAN 50 PM", "registro": "16890", "dosis": 2.0, "unidad": "kg/ha", "plaga": "MILDIU"},
            {"nombre": "KUMULUS DF", "registro": "15234", "dosis": 3.0, "unidad": "kg/ha", "plaga": "OIDIO"},
            {"nombre": "MOVENTO 150 O-TEQ", "registro": "25123", "dosis": 0.6, "unidad": "l/ha", "plaga": "ARAÑA ROJA"},
            {"nombre": "JUVINAL 10 EC", "registro": "23890", "dosis": 0.5, "unidad": "l/ha", "plaga": "COCHINILLA"}
        ],
        "parcelas": 150,
        "aplicaciones_por_parcela": 3
    }
}


# ============================================
# Generadores
# ============================================

def write_safe(ws, row, col, value):
    """Escribe en celda solo si no es merged"""
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def generar_parcelas(config: dict, seed: int = None) -> list:
    """Genera parcelas realistas"""
    if seed:
        random.seed(seed)
    
    parcelas = []
    num_parcelas = config["parcelas"]
    municipios = config["municipios"]
    cultivos = config["cultivos"]
    
    poligono_actual = 1
    parcela_actual = 1
    
    for i in range(num_parcelas):
        municipio = municipios[i % len(municipios)]
        
        # Seleccionar cultivo según peso
        rand = random.random()
        cumulative = 0
        cultivo = cultivos[0][0]
        for c, prob in cultivos:
            cumulative += prob
            if rand <= cumulative:
                cultivo = c
                break
        
        # Recintos
        num_recintos = random.choices([1, 2], weights=[0.8, 0.2])[0]
        
        for recinto in range(1, num_recintos + 1):
            superficie = round(random.uniform(1.0, 20.0), 2)
            
            parcelas.append({
                "nro_orden": len(parcelas) + 1,
                "municipio": municipio,
                "polygon": poligono_actual,
                "parcel": parcela_actual,
                "recinto": recinto,
                "superficie_sigpac": round(superficie * 1.02, 2),
                "superficie_cultivada": superficie,
                "crop": cultivo
            })
        
        parcela_actual += 1
        if parcela_actual > 200:
            parcela_actual = 1
            poligono_actual += 1
    
    return parcelas


def generar_tratamientos(parcelas: list, config: dict, seed: int = None) -> list:
    """Genera tratamientos únicos"""
    if seed:
        random.seed(seed + 1000)
    
    tratamientos = []
    productos = config["productos"]
    aplicaciones = config["aplicaciones_por_parcela"]
    
    fecha_base = datetime.now()
    
    for parcela in parcelas:
        num_aplicaciones = random.randint(1, aplicaciones)
        productos_seleccionados = random.sample(productos, min(num_aplicaciones, len(productos)))
        
        for i, producto in enumerate(productos_seleccionados):
            dias_atras = random.randint(0, 120)
            fecha = (fecha_base - timedelta(days=dias_atras)).strftime("%d/%m/%Y")
            
            tratamientos.append({
                "id_parcelas": parcela["nro_orden"],
                "polygon": parcela["polygon"],
                "parcel": parcela["parcel"],
                "recinto": parcela["recinto"],
                "crop": parcela["crop"],
                "surface": parcela["superficie_cultivada"],
                "product": producto["nombre"],
                "registry_number": producto["registro"],
                "dose": {"value": producto["dosis"], "unit": producto["unidad"]},
                "fecha": fecha,
                "pest": producto["plaga"],
                "notes": ""
            })
    
    return tratamientos


def escribir_datos_cliente(wb, config: dict):
    """Escribe los datos del cliente en inf.gral 1"""
    if "inf.gral 1" not in wb.sheetnames:
        return
    
    ws = wb["inf.gral 1"]
    
    # Según la estructura del template:
    # (6,13): Año
    write_safe(ws, 6, 13, config["year"])
    
    # (9,4): Razón social
    write_safe(ws, 9, 4, config["razon_social"])
    
    # (9,12): NIF
    write_safe(ws, 9, 12, config["nif"])
    
    # (10,4): Reg Nacional (aproximado)
    write_safe(ws, 10, 4, config["reg_nacional"])
    
    # (10,12): Reg Autonómico
    write_safe(ws, 10, 12, config["reg_autonomico"])
    
    # (11,2): Dirección
    write_safe(ws, 11, 2, config["direccion"])
    
    # (11,7): Localidad
    write_safe(ws, 11, 7, config["localidad"])
    
    # (11,10): CP
    write_safe(ws, 11, 10, config["cp"])
    
    # (11,13): Provincia
    write_safe(ws, 11, 13, config["provincia"])
    
    # (12,2): Teléfono fijo
    write_safe(ws, 12, 2, config["telefono_fijo"])
    
    # (12,6): Teléfono móvil
    write_safe(ws, 12, 6, config["telefono_movil"])
    
    # (12,9): Email
    write_safe(ws, 12, 9, config["email"])
    
    # (14,3): Nombre titular (es fórmula =+D9, pero la sobreescribimos)
    write_safe(ws, 14, 3, config["titular"])
    
    # (14,10): NIF titular
    write_safe(ws, 14, 10, config["nif_titular"])


def escribir_parcelas(wb, parcelas: list):
    """Escribe parcelas en 2.1. DATOS PARCELAS"""
    if "2.1. DATOS PARCELAS" not in wb.sheetnames:
        return
    
    ws = wb["2.1. DATOS PARCELAS"]
    
    # Limpiar datos existentes
    for row in range(14, min(ws.max_row + 1, 500)):
        for col in range(1, 15):
            try:
                write_safe(ws, row, col, None)
            except:
                pass
    
    # Escribir parcelas
    for i, p in enumerate(parcelas):
        row = 14 + i
        write_safe(ws, row, 2, p["nro_orden"])
        write_safe(ws, row, 4, p["municipio"])
        write_safe(ws, row, 7, p["polygon"])
        write_safe(ws, row, 8, p["parcel"])
        write_safe(ws, row, 9, p["recinto"])
        write_safe(ws, row, 11, p["superficie_sigpac"])
        write_safe(ws, row, 12, p["superficie_cultivada"])
        write_safe(ws, row, 13, p["crop"])


def limpiar_tratamientos(wb):
    """Limpia inf.trat 1 para escribir nuevos"""
    if "inf.trat 1" not in wb.sheetnames:
        return
    
    ws = wb["inf.trat 1"]
    for row in range(11, 200):
        for col in range(1, 14):
            try:
                write_safe(ws, row, col, None)
            except:
                pass


def crear_cuaderno_cliente(config: dict, parcelas: list, 
                           output_path: str, golden_path: str):
    """Crea cuaderno completo del cliente"""
    shutil.copy(golden_path, output_path)
    
    wb = load_workbook(output_path)
    
    # 1. Escribir datos del cliente
    escribir_datos_cliente(wb, config)
    
    # 2. Escribir parcelas
    escribir_parcelas(wb, parcelas)
    
    # 3. Limpiar tratamientos (se escriben después con el resolver)
    limpiar_tratamientos(wb)
    
    wb.save(output_path)
    wb.close()


def generar_cliente_completo(cliente_id: str, base_dir: Path, golden_path: str) -> dict:
    """Genera TODOS los archivos para un cliente"""
    config = CLIENTES[cliente_id]
    
    print(f"\n{'='*60}")
    print(f"  🏭 {config['razon_social']}")
    print(f"     Titular: {config['titular']}")
    print(f"     NIF: {config['nif']}")
    print(f"     {config['localidad']}, {config['provincia']}")
    print(f"{'='*60}")
    
    seed = hash(cliente_id) % 10000
    
    # Generar datos
    parcelas = generar_parcelas(config, seed)
    tratamientos = generar_tratamientos(parcelas, config, seed)
    
    print(f"📊 Parcelas: {len(parcelas)}")
    print(f"📊 Tratamientos: {len(tratamientos)}")
    print(f"📊 Municipios: {config['municipios']}")
    print(f"📊 Cultivos: {[c[0] for c in config['cultivos']]}")
    
    # Crear directorio
    client_dir = base_dir / "clientes_reales" / cliente_id
    client_dir.mkdir(parents=True, exist_ok=True)
    
    # Crear cuaderno del cliente (con TODOS sus datos)
    input_path = client_dir / f"{config['titular'].replace(' ', '_')}_2026.xlsx"
    crear_cuaderno_cliente(config, parcelas, str(input_path), golden_path)
    
    # También crear golden (para el resolver)
    golden_dest = client_dir / "golden.xlsx"
    shutil.copy(golden_path, golden_dest)
    
    # Guardar tratamientos
    trat_path = client_dir / "tratamientos.json"
    with open(trat_path, "w", encoding="utf-8") as f:
        json.dump(tratamientos, f, indent=2, ensure_ascii=False)
    
    # Guardar config completa
    config_path = client_dir / "config.json"
    config_serializable = {k: v for k, v in config.items() if not callable(v)}
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config_serializable, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"✅ Cuaderno: {input_path.name}")
    print(f"✅ Directorio: {client_dir}")
    
    return {
        "cliente_id": cliente_id,
        "empresa": config["razon_social"],
        "titular": config["titular"],
        "cuaderno_path": str(input_path),
        "parcelas": len(parcelas),
        "tratamientos": len(tratamientos)
    }


def main():
    print(f"\n{'🏭'*20}")
    print(f"  GENERADOR DE CLIENTES v2")
    print(f"  Cuadernos COMPLETOS con datos reales")
    print(f"{'🏭'*20}")
    
    BASE_DIR = Path(__file__).parent
    GOLDEN_PATH = "/Volumes/Uniclick4TB/organizadorhndezbueno/PABLO PEREZ RUBIO 2025 RESUELTO.XLSX"
    
    if not Path(GOLDEN_PATH).exists():
        print(f"❌ Golden no encontrado: {GOLDEN_PATH}")
        return 1
    
    resultados = []
    
    for cliente_id in CLIENTES.keys():
        resultado = generar_cliente_completo(cliente_id, BASE_DIR, GOLDEN_PATH)
        resultados.append(resultado)
    
    # Resumen
    print(f"\n{'='*60}")
    print(f"  📊 CUADERNOS GENERADOS")
    print(f"{'='*60}")
    
    for r in resultados:
        print(f"\n  📁 {r['titular']}")
        print(f"     Empresa: {r['empresa']}")
        print(f"     Parcelas: {r['parcelas']}")
        print(f"     Tratamientos: {r['tratamientos']}")
        print(f"     Archivo: {Path(r['cuaderno_path']).name}")
    
    print(f"\n✅ 3 cuadernos con datos REALES de clientes ficticios.\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
