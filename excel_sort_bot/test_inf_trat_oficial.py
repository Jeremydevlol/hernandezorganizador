#!/usr/bin/env python3
"""
TEST: Escritura en inf.trat 1 (formato RESUELTO oficial)
Valida que los tratamientos se escriben en la hoja oficial del cuaderno
"""
import sys
import shutil
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from src.ai_processor import AIProcessor, resolve_date
from src.parcel_manager import ParcelManager
from openpyxl import load_workbook


# Configuración
BASE_DIR = Path("/Volumes/Uniclick4TB/organizadorhndezbueno")
INPUT_FILE = BASE_DIR / "PABLO PEREZ RUBIO 2025.xlsx"
OUTPUT_FILE = Path("/tmp/OUT_inf_trat_oficial.xlsx")
REFERENCE_FILE = BASE_DIR / "PABLO PEREZ RUBIO 2025 RESUELTO.XLSX"


def print_section(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def count_rows_in_sheet(filepath, sheet_name, start_row=11):
    """Cuenta filas con datos en una hoja"""
    wb = load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[sheet_name]
    count = 0
    for row in range(start_row, ws.max_row + 1):
        if ws.cell(row, 1).value is not None:
            count += 1
    wb.close()
    return count


def main():
    print_section("TEST: Escritura en inf.trat 1")
    
    # 1. Setup
    print("\n📁 Setup...")
    if not INPUT_FILE.exists():
        print(f"❌ Archivo no encontrado: {INPUT_FILE}")
        return False
    
    shutil.copy(INPUT_FILE, OUTPUT_FILE)
    print(f"✅ Copiado a: {OUTPUT_FILE}")
    
    # 2. Verificar estado inicial
    print_section("Estado Inicial")
    
    initial_rows = count_rows_in_sheet(OUTPUT_FILE, "inf.trat 1")
    print(f"📊 Filas en inf.trat 1 (inicial): {initial_rows}")
    
    # 3. Cargar parcelas
    pm = ParcelManager(str(OUTPUT_FILE))
    parcels = pm.get_parcels()
    print(f"📊 Parcelas cargadas: {len(parcels)}")
    
    # Mostrar algunos nro_orden
    print("\n📋 Algunos Nº Orden:")
    for p in parcels[:5]:
        print(f"   Orden {p.nro_orden}: Pol {p.polygon}, Parc {p.parcel}, Rec {p.recinto}, {p.crop}")
    
    # 4. Interpretar comando
    print_section("Interpretación con IA")
    
    context = pm.get_context_for_ai()
    message = "Aplica Glifosato 3 l/ha hoy en las 3 primeras cebadas. Notas: prueba oficial."
    
    # Simplificar: tomar las primeras 3 cebadas manualmente
    cebadas = [p for p in parcels if "CEBADA" in p.crop][:3]
    print(f"💬 Seleccionando primeras 3 CEBADAS: {len(cebadas)} parcelas")
    
    for p in cebadas:
        print(f"   → Orden {p.nro_orden}: Pol {p.polygon}, {p.crop}, {p.surface_ha} ha")
    
    # 5. Escribir en inf.trat 1
    print_section("Escribir en inf.trat 1")
    
    treatment_data = {
        "product": "Glifosato",
        "dose": {"value": 3, "unit": "l/ha"},
        "pest": "MALAS HIERBAS",
        "notes": "Prueba oficial"
    }
    resolved_date = "05/02/2026"
    
    output_path, rows_written = pm.write_treatments_official(
        cebadas, 
        treatment_data, 
        resolved_date,
        output_path=str(OUTPUT_FILE)
    )
    
    print(f"✅ Escritas {rows_written} filas en inf.trat 1")
    
    # 6. Verificar
    print_section("Verificación")
    
    final_rows = count_rows_in_sheet(OUTPUT_FILE, "inf.trat 1")
    print(f"📊 Filas en inf.trat 1 (final): {final_rows}")
    
    # Leer las filas escritas
    wb = load_workbook(OUTPUT_FILE, data_only=True)
    ws = wb["inf.trat 1"]
    
    print("\n📋 Últimas filas escritas:")
    start_row = final_rows - rows_written + 11  # Offset de header
    for row in range(start_row, start_row + min(3, rows_written)):
        id_parc = ws.cell(row, 1).value
        especie = ws.cell(row, 2).value
        superficie = ws.cell(row, 4).value
        fecha = ws.cell(row, 5).value
        producto = ws.cell(row, 9).value
        dosis = ws.cell(row, 11).value
        print(f"   Row {row}: Id={id_parc}, {especie}, {superficie}ha, {producto}, {dosis}")
    
    wb.close()
    
    # 7. Comparar con RESUELTO
    print_section("Comparación con RESUELTO")
    
    if REFERENCE_FILE.exists():
        ref_rows = count_rows_in_sheet(REFERENCE_FILE, "inf.trat 1")
        print(f"📊 Filas en RESUELTO: {ref_rows}")
        print(f"📊 Filas en nuestro output: {final_rows}")
    
    # Checks
    checks = {
        "rows_written": rows_written == len(cebadas),
        "final_has_more": final_rows > initial_rows,
        "no_errors": True  # Si llegamos aquí, no hubo errores
    }
    
    print_section("Checklist")
    for k, v in checks.items():
        print(f"  {'✅' if v else '❌'} {k}")
    
    all_passed = all(checks.values())
    print(f"\n{'🎉 TEST ESCRITURA OFICIAL: PASSED!' if all_passed else '❌ FAILED'}")
    print(f"\n📁 Output generado: {OUTPUT_FILE}")
    print("   Abre el archivo y verifica la hoja 'inf.trat 1'")
    
    return all_passed


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
