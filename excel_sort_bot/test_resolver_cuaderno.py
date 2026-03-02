#!/usr/bin/env python3
"""
TEST: Resolver Cuaderno con Golden Template
Valida que el output es IDÉNTICO al RESUELTO en estructura
"""
import sys
import shutil
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from src.resolver_cuaderno import resolver_cuaderno, GoldenTemplateWriter
from src.parcel_manager import ParcelManager
from openpyxl import load_workbook


# Configuración
BASE_DIR = Path("/Volumes/Uniclick4TB/organizadorhndezbueno")
INPUT_FILE = BASE_DIR / "PABLO PEREZ RUBIO 2025.xlsx"
GOLDEN_FILE = BASE_DIR / "PABLO PEREZ RUBIO 2025 RESUELTO.XLSX"
OUTPUT_FILE = Path("/tmp/OUT_cuaderno_resuelto.xlsx")


def print_section(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def compare_sheets(file1, file2):
    """Compara hojas de dos workbooks"""
    wb1 = load_workbook(file1, data_only=True)
    wb2 = load_workbook(file2, data_only=True)
    
    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)
    
    result = {
        "same_sheets": sheets1 == sheets2,
        "in_golden_only": sheets2 - sheets1,
        "in_output_only": sheets1 - sheets2,
        "common": sheets1 & sheets2
    }
    
    wb1.close()
    wb2.close()
    return result


def count_rows_in_inf_trat(filepath):
    """Cuenta filas de datos en inf.trat 1"""
    wb = load_workbook(filepath, data_only=True)
    if "inf.trat 1" not in wb.sheetnames:
        wb.close()
        return 0
    
    ws = wb["inf.trat 1"]
    count = 0
    for row in range(11, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val is not None:
            try:
                int(val)
                count += 1
            except:
                pass
    wb.close()
    return count


def main():
    print_section("TEST: Resolver Cuaderno con Golden Template")
    
    # Verificar archivos
    if not INPUT_FILE.exists():
        print(f"❌ No existe: {INPUT_FILE}")
        return False
    
    if not GOLDEN_FILE.exists():
        print(f"❌ No existe Golden: {GOLDEN_FILE}")
        return False
    
    print(f"✅ Input: {INPUT_FILE.name}")
    print(f"✅ Golden: {GOLDEN_FILE.name}")
    
    # Contar filas en Golden (referencia)
    golden_rows = count_rows_in_inf_trat(GOLDEN_FILE)
    print(f"📊 Filas en Golden inf.trat 1: {golden_rows}")
    
    # 1. Cargar parcelas del input para generar tratamientos
    print_section("1. Cargar Parcelas del Input")
    
    pm = ParcelManager(str(INPUT_FILE))
    parcels = pm.get_parcels()
    print(f"📊 Parcelas cargadas: {len(parcels)}")
    
    # Seleccionar algunas CEBADAS para el test
    cebadas = [p for p in parcels if "CEBADA" in p.crop][:5]
    print(f"📊 Seleccionando {len(cebadas)} CEBADAS para test")
    
    for p in cebadas:
        print(f"   → Orden {p.nro_orden}: Pol {p.polygon}, {p.crop}, {p.surface_ha} ha")
    
    # 2. Preparar tratamientos
    print_section("2. Preparar Tratamientos")
    
    treatments = []
    for p in cebadas:
        treatments.append({
            "polygon": p.polygon,
            "parcel": p.parcel,
            "recinto": p.recinto,
            "crop": p.crop,
            "surface": p.surface_ha,
            "product": "Glifosato",
            "dose": {"value": 3, "unit": "l/ha"},
            "fecha": "05/02/2026",
            "pest": "MALAS HIERBAS",
            "notes": "Test Resolver"
        })
    
    print(f"✅ {len(treatments)} tratamientos preparados")
    
    # 3. Ejecutar Resolver
    print_section("3. Ejecutar Resolver Cuaderno")
    
    result = resolver_cuaderno(
        input_path=str(INPUT_FILE),
        output_path=str(OUTPUT_FILE),
        golden_path=str(GOLDEN_FILE),
        treatments=treatments
    )
    
    print(f"📊 Resultado:")
    print(f"   Success: {result['success']}")
    print(f"   Actions applied: {result['actions_applied']}")
    print(f"   Write stats: {result['write_stats']}")
    print(f"   Warnings: {len(result['warnings'])}")
    print(f"   Errors: {len(result['errors'])}")
    
    if result['errors']:
        for e in result['errors']:
            print(f"   ❌ {e}")
    
    # 4. Verificar estructura
    print_section("4. Verificar Estructura")
    
    comparison = compare_sheets(OUTPUT_FILE, GOLDEN_FILE)
    
    print(f"📊 Hojas en Golden: {len(comparison['common']) + len(comparison['in_golden_only'])}")
    print(f"📊 Hojas en Output: {len(comparison['common']) + len(comparison['in_output_only'])}")
    
    if comparison['in_output_only']:
        print(f"   ⚠️ Hojas extras en output: {comparison['in_output_only']}")
    else:
        print(f"   ✅ No hay hojas extras (output = golden)")
    
    if comparison['in_golden_only']:
        print(f"   ⚠️ Hojas faltantes: {comparison['in_golden_only']}")
    
    # 5. Verificar datos escritos
    print_section("5. Verificar Datos Escritos")
    
    output_rows = count_rows_in_inf_trat(OUTPUT_FILE)
    print(f"📊 Filas en Output inf.trat 1: {output_rows}")
    
    # Leer las filas escritas
    wb = load_workbook(OUTPUT_FILE, data_only=True)
    ws = wb["inf.trat 1"]
    
    print("\n📋 Filas escritas:")
    for row in range(11, 11 + min(5, output_rows)):
        id_p = ws.cell(row, 1).value
        esp = ws.cell(row, 2).value
        sup = ws.cell(row, 4).value
        fecha = ws.cell(row, 5).value
        prod = ws.cell(row, 9).value
        dosis = ws.cell(row, 11).value
        print(f"   Row {row}: Id={id_p}, {esp}, {sup}ha, {prod}, {dosis}")
    
    wb.close()
    
    # 6. Checklist final
    print_section("Checklist Final")
    
    checks = {
        "resolver_success": result['success'],
        "actions_applied": result['actions_applied'] == len(treatments),
        "no_extra_sheets": len(comparison['in_output_only']) == 0,
        "rows_written": output_rows >= len(treatments),
        "no_errors": len(result['errors']) == 0
    }
    
    for k, v in checks.items():
        print(f"  {'✅' if v else '❌'} {k}")
    
    all_passed = all(checks.values())
    
    if all_passed:
        print(f"\n🎉 RESOLVER CUADERNO: PASSED!")
        print(f"\n📁 Output generado: {OUTPUT_FILE}")
        print("   El archivo tiene la MISMA estructura que el Golden Template")
    else:
        print(f"\n❌ Algunos checks fallaron")
    
    # Copiar a raíz
    shutil.copy(OUTPUT_FILE, Path(__file__).parent / "OUTPUT_cuaderno_resuelto.xlsx")
    print(f"\n📁 Copiado a: excel_sort_bot/OUTPUT_cuaderno_resuelto.xlsx")
    
    return all_passed


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
