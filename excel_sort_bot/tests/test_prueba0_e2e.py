"""
PRUEBA 0 — Test E2E Enterprise
Valida: Fingerprint Gate + Preview 300 parcelas + Commit + Idempotencia + Exclusión
"""
import sys
import json
import uuid
import shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.ai_processor import AIProcessor, resolve_date
from src.parcel_manager import ParcelManager
from src.template_fingerprint import fingerprint_workbook, load_registry, match_template

# ============================================================
# CONFIGURACIÓN - RUTAS
# ============================================================
BASE_DIR = Path("/Volumes/Uniclick4TB/organizadorhndezbueno")

FILE_INPUT_300 = BASE_DIR / "PRUEBA0_input_300parcelas_clean.xlsx"
FILE_RESUELTO_300 = BASE_DIR / "PRUEBA0_resuelto_300parcelas.xlsx"
FILE_BASE = BASE_DIR / "PABLO PEREZ RUBIO 2025.xlsx"

OUTPUT_DIR = Path("/tmp")
OUTPUT_COMMIT = OUTPUT_DIR / "OUT_prueba0_commit.xlsx"
OUTPUT_EXCLUSION = OUTPUT_DIR / "OUT_prueba0_exclusion.xlsx"

# ============================================================
# UTILIDADES
# ============================================================
def print_section(title, level=1):
    if level == 1:
        print(f"\n{'='*70}")
        print(f"  {title}")
        print(f"{'='*70}")
    else:
        print(f"\n{'─'*50}")
        print(f"  {title}")
        print(f"{'─'*50}")


def count_treatment_rows(filepath):
    """Cuenta filas en Tratamientos_IA (excluyendo header)"""
    wb = load_workbook(filepath, data_only=True)
    if "Tratamientos_IA" not in wb.sheetnames:
        wb.close()
        return 0
    sheet = wb["Tratamientos_IA"]
    count = sheet.max_row - 1
    wb.close()
    return max(0, count)


def validate_treatment_rows(filepath, expected_count=None, exclude_recintos=None):
    """Valida filas de tratamiento con checks enterprise"""
    exclude_recintos = exclude_recintos or []
    results = {
        "rows_found": 0,
        "duplicates": 0,
        "invalid_dates": 0,
        "zero_surfaces": 0,
        "excluded_recintos_present": 0,
        "errors": []
    }
    
    wb = load_workbook(filepath, data_only=True)
    
    if "Tratamientos_IA" not in wb.sheetnames:
        results["errors"].append("No existe hoja Tratamientos_IA")
        wb.close()
        return results
    
    sheet = wb["Tratamientos_IA"]
    seen_keys = set()
    
    for row in range(2, sheet.max_row + 1):
        results["rows_found"] += 1
        
        # Extraer datos (ajustar columnas según tu layout)
        fecha = sheet.cell(row, 2).value
        producto = sheet.cell(row, 4).value
        dosis = sheet.cell(row, 6).value
        unidad = sheet.cell(row, 7).value
        superficie = sheet.cell(row, 9).value
        poligono = sheet.cell(row, 11).value
        parcela = sheet.cell(row, 12).value
        recinto = sheet.cell(row, 13).value
        
        # Validar fecha
        if fecha:
            try:
                if isinstance(fecha, str) and "/" in fecha:
                    datetime.strptime(fecha, "%d/%m/%Y")
            except:
                results["invalid_dates"] += 1
        
        # Validar superficie
        try:
            if superficie is None or float(superficie) <= 0:
                results["zero_surfaces"] += 1
        except:
            results["zero_surfaces"] += 1
        
        # Validar recinto excluido
        try:
            rec_int = int(recinto) if recinto else 0
            if rec_int in exclude_recintos:
                results["excluded_recintos_present"] += 1
        except:
            pass
        
        # Detectar duplicados (clave de negocio + superficie para detectar bugs del sistema)
        key = (str(fecha), str(producto), str(poligono), str(parcela), str(recinto), str(superficie))
        if key in seen_keys:
            results["duplicates"] += 1
        seen_keys.add(key)
    
    wb.close()
    return results


# ============================================================
# MOCK PROPOSAL STORE (simula servidor)
# ============================================================
class MockProposalStore:
    def __init__(self):
        self.proposals = {}
        self.committed = {}
    
    def create(self, proposal_id, data):
        self.proposals[proposal_id] = {"created_at": datetime.now(), **data}
    
    def commit(self, proposal_id):
        if proposal_id in self.committed:
            return {"success": False, "status_code": 409, "error": "ALREADY_COMMITTED",
                    "message": f"Ya ejecutada el {self.committed[proposal_id].strftime('%d/%m/%Y %H:%M:%S')}"}
        
        proposal = self.proposals.get(proposal_id)
        if not proposal:
            return {"success": False, "status_code": 404, "error": "NOT_FOUND"}
        
        try:
            pm = ParcelManager(proposal["file_path"])
            rows = proposal["rows"]
            output_path, rows_written = pm.write_treatments(rows, output_path=proposal["output_path"])
            
            self.committed[proposal_id] = datetime.now()
            del self.proposals[proposal_id]
            
            return {"success": True, "status_code": 200, "rows_written": rows_written, "output_path": output_path}
        except Exception as e:
            return {"success": False, "status_code": 500, "error": str(e)}


# ============================================================
# PASO 1: FINGERPRINT GATE
# ============================================================
def test_fingerprint_gate():
    print_section("PASO 1: FINGERPRINT GATE")
    
    files_to_test = [
        ("FILE_INPUT_300", FILE_INPUT_300),
        ("FILE_RESUELTO_300", FILE_RESUELTO_300),
        ("FILE_BASE", FILE_BASE),
    ]
    
    results = {}
    registry = load_registry()
    
    for name, path in files_to_test:
        if not path.exists():
            print(f"⚠️ {name}: No existe ({path})")
            results[name] = {"exists": False}
            continue
        
        fp = fingerprint_workbook(str(path), parcels_sheet_hint_contains=["2.1.", "PARCELAS"])
        decision = match_template(fp, registry)
        
        print(f"\n📋 {name}:")
        print(f"   • Sheets: {len(fp.sheetnames)}")
        print(f"   • Parcels sheet: {fp.parcels_sheet}")
        print(f"   • Header row: {fp.header_row}")
        print(f"   • Template ID: {decision['template_id']}")
        print(f"   • Score: {decision['score']}")
        print(f"   • Status: {decision['status']}")
        
        results[name] = {
            "exists": True,
            "template_id": decision["template_id"],
            "score": decision["score"],
            "status": decision["status"]
        }
    
    # Validaciones
    all_valid = all(r.get("status") in ["ACCEPT", "WARN"] for r in results.values() if r.get("exists"))
    print(f"\n{'✅' if all_valid else '❌'} Fingerprint gate: {'PASS' if all_valid else 'FAIL'}")
    
    return results, all_valid


# ============================================================
# PASO 2: PREVIEW + COMMIT 300 PARCELAS
# ============================================================
def test_preview_commit_300():
    print_section("PASO 2: PREVIEW + COMMIT (300 parcelas)")
    
    # Usar el archivo que tenga más parcelas
    test_file = FILE_INPUT_300 if FILE_INPUT_300.exists() else FILE_BASE
    
    if not test_file.exists():
        print(f"❌ No existe archivo de prueba: {test_file}")
        return None, False
    
    # Copiar a output
    shutil.copy(test_file, OUTPUT_COMMIT)
    
    # Cargar parcelas
    pm = ParcelManager(str(OUTPUT_COMMIT))
    parcels = pm.get_parcels()
    context = pm.get_context_for_ai()
    
    print(f"📊 Parcelas cargadas: {len(parcels)}")
    print(f"   Cultivos: {context['crops'][:5]}...")
    print(f"   Municipios: {context['municipios'][:3]}...")
    
    # Preview con IA
    print_section("Preview con IA", level=2)
    
    message = "Aplica Glifosato 3 l/ha hoy en todas las parcelas. Notas: preventivo."
    print(f"💬 Comando: \"{message}\"")
    
    ai = AIProcessor()
    intent = ai.interpret_command(message, context)
    
    if "error" in intent:
        print(f"❌ Error IA: {intent['error']}")
        return None, False
    
    print(f"✅ Intent recibido")
    print(f"   Scope: {intent.get('filters', {}).get('targets', {}).get('scope')}")
    
    # Filtrar
    filters = intent.get("filters", {})
    targets = filters.get("targets", {})
    logic = filters.get("logic", "AND")
    scope = targets.get("scope", "ALL")
    treatment_data = intent.get("data", {})
    
    matched = pm.filter_parcels(scope, targets, logic)
    print(f"📊 Matched: {len(matched)} parcelas")
    
    requires_confirmation = len(matched) > 20
    print(f"   requires_confirmation: {requires_confirmation}")
    
    # Generar filas
    date_obj = treatment_data.get("date", {})
    resolved_date = resolve_date(date_obj)
    rows = pm.generate_treatment_rows(matched, treatment_data, resolved_date, start_order=1)
    
    # Commit
    print_section("Commit", level=2)
    
    proposal_id = str(uuid.uuid4())[:8].upper()
    store = MockProposalStore()
    store.create(proposal_id, {
        "file_path": str(OUTPUT_COMMIT),
        "output_path": str(OUTPUT_COMMIT),
        "rows": rows,
        "intent": intent
    })
    
    result = store.commit(proposal_id)
    
    if result["success"]:
        print(f"✅ Commit OK: {result['rows_written']} filas escritas")
    else:
        print(f"❌ Commit FAIL: {result}")
        return None, False
    
    # Validar output
    print_section("Validaciones Post-Commit", level=2)
    
    validation = validate_treatment_rows(str(OUTPUT_COMMIT))
    print(f"📊 Validación:")
    print(f"   • Filas encontradas: {validation['rows_found']}")
    print(f"   • Duplicados: {validation['duplicates']}")
    print(f"   • Fechas inválidas: {validation['invalid_dates']}")
    print(f"   • Superficies 0: {validation['zero_surfaces']}")
    
    checks = {
        "rows_match": validation["rows_found"] == len(matched),
        "no_duplicates": validation["duplicates"] == 0,
        "no_invalid_dates": validation["invalid_dates"] == 0,
    }
    
    all_passed = all(checks.values())
    for k, v in checks.items():
        print(f"   {'✅' if v else '❌'} {k}")
    
    print(f"\n{'✅' if all_passed else '❌'} Preview+Commit 300: {'PASS' if all_passed else 'FAIL'}")
    
    return {"proposal_id": proposal_id, "store": store, "rows_written": result["rows_written"]}, all_passed


# ============================================================
# PASO 3: IDEMPOTENCIA (DOUBLE COMMIT ATTACK)
# ============================================================
def test_idempotency(prev_result):
    print_section("PASO 3: IDEMPOTENCIA (Double Commit Attack)")
    
    if not prev_result:
        print("⚠️ Saltando: no hay resultado previo")
        return False
    
    proposal_id = prev_result["proposal_id"]
    store = prev_result["store"]
    rows_before = count_treatment_rows(str(OUTPUT_COMMIT))
    
    print(f"🔑 Proposal ID: {proposal_id}")
    print(f"📊 Filas antes del ataque: {rows_before}")
    
    # Segundo commit (ataque)
    result2 = store.commit(proposal_id)
    
    print(f"\n📊 Resultado del 2do commit (ataque):")
    print(f"   Status: {result2['status_code']}")
    print(f"   Success: {result2['success']}")
    
    blocked = result2["status_code"] == 409 and result2.get("error") == "ALREADY_COMMITTED"
    
    if blocked:
        print(f"   ✅ BLOQUEADO: {result2['message']}")
    else:
        print(f"   ❌ NO BLOQUEADO: {result2}")
    
    # Verificar que no se duplicaron filas
    rows_after = count_treatment_rows(str(OUTPUT_COMMIT))
    no_duplicates = rows_after == rows_before
    
    print(f"\n📊 Filas después del ataque: {rows_after}")
    print(f"   {'✅' if no_duplicates else '❌'} Sin duplicados en Excel")
    
    passed = blocked and no_duplicates
    print(f"\n{'✅' if passed else '❌'} Idempotencia: {'PASS' if passed else 'FAIL'}")
    
    return passed


# ============================================================
# PASO 4: EXCLUSIÓN DE RECINTO
# ============================================================
def test_recinto_exclusion():
    print_section("PASO 4: EXCLUSIÓN DE RECINTO")
    
    test_file = FILE_BASE
    
    if not test_file.exists():
        print(f"❌ No existe archivo: {test_file}")
        return False
    
    # Copiar
    shutil.copy(test_file, OUTPUT_EXCLUSION)
    
    pm = ParcelManager(str(OUTPUT_EXCLUSION))
    context = pm.get_context_for_ai()
    
    # Mensaje con exclusión
    message = "Aplica Glifosato 3 l/ha hoy en polígono 8 parcela 148, excepto recinto 2. Notas: test exclusión."
    print(f"💬 Comando: \"{message}\"")
    
    ai = AIProcessor()
    intent = ai.interpret_command(message, context)
    
    if "error" in intent:
        print(f"❌ Error IA: {intent['error']}")
        return False
    
    filters = intent.get("filters", {})
    targets = filters.get("targets", {})
    logic = filters.get("logic", "AND")
    scope = targets.get("scope", "")
    treatment_data = intent.get("data", {})
    exclude_recinto = targets.get("exclude_recinto", [])
    
    print(f"📊 Intent:")
    print(f"   Scope: {scope}")
    print(f"   exclude_recinto: {exclude_recinto}")
    
    # Filtrar
    matched = pm.filter_parcels(scope, targets, logic)
    print(f"📊 Matched: {len(matched)} parcelas")
    
    # Verificar que recinto 2 no está
    recintos_matched = [p.recinto for p in matched]
    recinto_2_present = 2 in recintos_matched
    print(f"   Recintos matched: {recintos_matched}")
    print(f"   {'❌' if recinto_2_present else '✅'} Recinto 2 {'PRESENTE (mal)' if recinto_2_present else 'EXCLUIDO (bien)'}")
    
    if recinto_2_present:
        return False
    
    # Commit
    date_obj = treatment_data.get("date", {})
    resolved_date = resolve_date(date_obj)
    rows = pm.generate_treatment_rows(matched, treatment_data, resolved_date, start_order=1)
    
    proposal_id = str(uuid.uuid4())[:8].upper()
    store = MockProposalStore()
    store.create(proposal_id, {
        "file_path": str(OUTPUT_EXCLUSION),
        "output_path": str(OUTPUT_EXCLUSION),
        "rows": rows
    })
    
    result = store.commit(proposal_id)
    
    if not result["success"]:
        print(f"❌ Commit FAIL: {result}")
        return False
    
    print(f"✅ Commit OK: {result['rows_written']} filas")
    
    # Validar que recinto 2 no aparece en Excel
    validation = validate_treatment_rows(str(OUTPUT_EXCLUSION), exclude_recintos=[2])
    
    no_excluded = validation["excluded_recintos_present"] == 0
    print(f"   {'✅' if no_excluded else '❌'} Recinto 2 no aparece en Excel")
    
    passed = no_excluded
    print(f"\n{'✅' if passed else '❌'} Exclusión recinto: {'PASS' if passed else 'FAIL'}")
    
    return passed


# ============================================================
# MAIN
# ============================================================
def main():
    print("\n" + "🔥"*35)
    print("  PRUEBA 0 — TEST E2E ENTERPRISE")
    print("🔥"*35)
    
    checklist = {}
    
    # Paso 1: Fingerprint
    _, checklist["fingerprint_gate"] = test_fingerprint_gate()
    
    # Paso 2: Preview + Commit 300
    prev_result, checklist["preview_commit_300"] = test_preview_commit_300()
    
    # Paso 3: Idempotencia
    checklist["idempotency"] = test_idempotency(prev_result)
    
    # Paso 4: Exclusión recinto
    checklist["recinto_exclusion"] = test_recinto_exclusion()
    
    # Resumen
    print_section("RESUMEN FINAL")
    
    for key, passed in checklist.items():
        status = "✅" if passed else "❌"
        print(f"  {status} {key}")
    
    all_passed = all(checklist.values())
    
    print(f"\n{'🎉 PRUEBA 0 E2E ENTERPRISE: ALL PASSED!' if all_passed else '⚠️ Algunos tests fallaron'}")
    print(f"\n📁 Outputs generados:")
    print(f"   • {OUTPUT_COMMIT}")
    print(f"   • {OUTPUT_EXCLUSION}")
    
    return all_passed


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
