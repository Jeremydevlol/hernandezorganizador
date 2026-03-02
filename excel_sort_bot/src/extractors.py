"""
Extractores robustos para encontrar datos en Excels desordenados
"""
import re
import unicodedata
from typing import Optional, Any
from openpyxl.utils import get_column_letter

from .types import (
    SourceConfig, SourceType, TakeDirection, ExtractedValue, PatternType
)
from .io_excel import ExcelReader, get_cell_reference


# ============================================================
# PATRONES REGEX
# ============================================================

PATTERNS = {
    # NIF/NIE español: 8 dígitos + letra o X/Y/Z + 7 dígitos + letra
    "nif_nie": r"(?:[0-9]{8}[A-Z]|[XYZ][0-9]{7}[A-Z]|[0-9]-?[0-9]{7,8}-?[A-Z])",
    
    # Dinero: acepta €, EUR, separadores de miles, decimales
    "money": r"(?:€\s*)?(?:\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?|\d+(?:[.,]\d{1,2})?)(?:\s*€|\s*EUR)?",
    
    # Fechas: dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd
    "date": r"(?:\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{4}[-/]\d{1,2}[-/]\d{1,2})",
    
    # Email
    "email": r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}",
    
    # Teléfono español
    "phone": r"(?:\+34\s?)?(?:[69]\d{2}[\s.-]?\d{3}[\s.-]?\d{3}|\d{3}[\s.-]?\d{2}[\s.-]?\d{2}[\s.-]?\d{2})",
    
    # Código postal español
    "postal_code": r"\b(?:0[1-9]|[1-4][0-9]|5[0-2])\d{3}\b",
}


def normalize_text(text: str) -> str:
    """
    Normaliza texto para comparación (quita tildes, mayúsculas, espacios extra)
    """
    if not text:
        return ""
    # Convertir a string
    text = str(text)
    # Quitar tildes
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    # Minúsculas y limpiar espacios
    text = text.lower().strip()
    text = re.sub(r'\s+', ' ', text)
    return text


def labels_match(cell_value: str, labels: list[str]) -> bool:
    """
    Comprueba si el valor de celda coincide con alguna etiqueta
    """
    if not cell_value:
        return False
    
    normalized_cell = normalize_text(cell_value)
    
    for label in labels:
        normalized_label = normalize_text(label)
        # Coincidencia exacta o la celda contiene la etiqueta
        if normalized_label == normalized_cell or normalized_label in normalized_cell:
            return True
        # También si la etiqueta está al inicio
        if normalized_cell.startswith(normalized_label):
            return True
    
    return False


def extract_by_pattern(text: str, pattern_name: str) -> Optional[str]:
    """
    Extrae un valor usando un patrón regex predefinido
    
    Returns:
        El primer match encontrado o None
    """
    if not text or pattern_name not in PATTERNS:
        return None
    
    text_str = str(text)
    pattern = PATTERNS[pattern_name]
    
    match = re.search(pattern, text_str, re.IGNORECASE)
    if match:
        return match.group(0)
    
    return None


def find_value_by_pattern(reader: ExcelReader, pattern_name: str,
                          sheet_hint: Optional[str] = None) -> list[tuple[str, int, int, str]]:
    """
    Busca valores que coincidan con un patrón en todo el workbook
    
    Returns:
        Lista de tuplas (sheet_name, row, col, matched_value)
    """
    results = []
    sheets = [sheet_hint] if sheet_hint else reader.sheet_names
    
    for sheet_name in sheets:
        if sheet_name not in reader.sheet_names:
            continue
        for row, col, value in reader.iter_cells_with_values(sheet_name):
            if value is None:
                continue
            matched = extract_by_pattern(str(value), pattern_name)
            if matched:
                results.append((sheet_name, row, col, matched))
    
    return results


class RobustExtractor:
    """
    Extractor robusto que implementa múltiples estrategias de búsqueda
    """
    
    def __init__(self, reader: ExcelReader, debug: bool = False):
        self.reader = reader
        self.debug = debug
        self._cache: dict[str, list] = {}
    
    def _log(self, message: str):
        """Log de debug"""
        if self.debug:
            print(f"[Extractor] {message}")
    
    def extract_by_label(self, source: SourceConfig) -> Optional[ExtractedValue]:
        """
        Extrae un valor buscando por etiquetas
        
        Strategy:
        1. Buscar celdas que contengan alguna de las etiquetas
        2. Para cada match, buscar el valor en la dirección indicada
        3. Si está vacío, intentar hasta max_distance celdas
        4. Devolver el primer valor encontrado
        """
        sheets_to_search = [source.sheet_hint] if source.sheet_hint else self.reader.sheet_names
        
        for sheet_name in sheets_to_search:
            if sheet_name not in self.reader.sheet_names:
                continue
            
            for row, col, cell_value in self.reader.iter_cells_with_values(sheet_name):
                if not labels_match(str(cell_value), source.labels):
                    continue
                
                self._log(f"Label encontrado: '{cell_value}' en {sheet_name}!{get_cell_reference(row, col)}")
                
                # Buscar valor adyacente
                direction = source.take.value
                
                for distance in range(1, source.max_distance + 1):
                    new_row, new_col, value = self.reader.get_adjacent_cell(
                        sheet_name, row, col, direction, distance
                    )
                    
                    if value is not None and str(value).strip():
                        # Verificar que no sea otra etiqueta
                        if not any(labels_match(str(value), source.labels[:3]) for _ in [1]):
                            self._log(f"  -> Valor encontrado: '{value}' en {get_cell_reference(new_row, new_col)}")
                            return ExtractedValue(
                                key="",  # Se asigna después
                                raw_value=value,
                                transformed_value=value,
                                source_sheet=sheet_name,
                                source_cell=get_cell_reference(new_row, new_col),
                                extraction_method="label",
                                confidence=1.0 - (distance - 1) * 0.1
                            )
                
                # Si no encontró a la derecha, probar abajo
                if source.take == TakeDirection.RIGHT:
                    for distance in range(1, source.max_distance + 1):
                        new_row, new_col, value = self.reader.get_adjacent_cell(
                            sheet_name, row, col, "down", distance
                        )
                        if value is not None and str(value).strip():
                            self._log(f"  -> Valor encontrado (down): '{value}' en {get_cell_reference(new_row, new_col)}")
                            return ExtractedValue(
                                key="",
                                raw_value=value,
                                transformed_value=value,
                                source_sheet=sheet_name,
                                source_cell=get_cell_reference(new_row, new_col),
                                extraction_method="label_down",
                                confidence=0.8 - (distance - 1) * 0.1
                            )
        
        return None
    
    def extract_by_pattern(self, source: SourceConfig) -> Optional[ExtractedValue]:
        """
        Extrae un valor buscando por patrones regex
        """
        for pattern_name in source.patterns:
            results = find_value_by_pattern(self.reader, pattern_name, source.sheet_hint)
            
            if results:
                # Tomar el primero (podría mejorarse con heurísticas)
                sheet_name, row, col, matched_value = results[0]
                self._log(f"Patrón '{pattern_name}' encontrado: '{matched_value}' en {sheet_name}!{get_cell_reference(row, col)}")
                
                return ExtractedValue(
                    key="",
                    raw_value=matched_value,
                    transformed_value=matched_value,
                    source_sheet=sheet_name,
                    source_cell=get_cell_reference(row, col),
                    extraction_method=f"pattern:{pattern_name}",
                    confidence=0.9
                )
        
        return None
    
    def extract_by_label_or_pattern(self, source: SourceConfig) -> Optional[ExtractedValue]:
        """
        Intenta primero por label, luego por pattern
        """
        # Primero intentar por label
        result = self.extract_by_label(source)
        if result:
            return result
        
        # Luego por pattern
        self._log("Label no encontrado, intentando por patrón...")
        result = self.extract_by_pattern(source)
        if result:
            return result
        
        return None
    
    def extract(self, field_key: str, source: SourceConfig) -> Optional[ExtractedValue]:
        """
        Método principal de extracción que selecciona la estrategia adecuada
        """
        self._log(f"\n--- Extrayendo campo: {field_key} ---")
        
        result: Optional[ExtractedValue] = None
        
        if source.type == SourceType.LABEL:
            result = self.extract_by_label(source)
        elif source.type == SourceType.PATTERN:
            result = self.extract_by_pattern(source)
        elif source.type == SourceType.LABEL_OR_PATTERN:
            result = self.extract_by_label_or_pattern(source)
        
        if result:
            result.key = field_key
        
        return result
    
    def extract_table(self, sheet_name: str, header_row: int,
                      column_mapping: dict[str, str]) -> list[dict[str, Any]]:
        """
        Extrae datos de una tabla estructurada
        
        Args:
            sheet_name: Nombre de la hoja
            header_row: Fila donde están los encabezados
            column_mapping: Dict de {nombre_columna_excel: nombre_campo}
            
        Returns:
            Lista de diccionarios con los datos de cada fila
        """
        sheet = self.reader.get_sheet(sheet_name)
        if not sheet:
            return []
        
        # Encontrar columnas por nombre
        col_indices = {}
        for col in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=header_row, column=col).value
            if header_value:
                normalized_header = normalize_text(str(header_value))
                for excel_name, field_name in column_mapping.items():
                    if normalize_text(excel_name) in normalized_header:
                        col_indices[field_name] = col
                        break
        
        self._log(f"Columnas encontradas: {col_indices}")
        
        # Extraer datos
        results = []
        row = header_row + 1
        
        while row <= sheet.max_row:
            # Verificar si la fila tiene datos
            has_data = False
            row_data = {}
            
            for field_name, col_idx in col_indices.items():
                value = sheet.cell(row=row, column=col_idx).value
                if value is not None:
                    has_data = True
                row_data[field_name] = value
            
            if not has_data:
                # Línea vacía, terminar
                break
            
            results.append(row_data)
            row += 1
        
        return results


def choose_best_match(matches: list[ExtractedValue], 
                      validation_func: Optional[callable] = None) -> Optional[ExtractedValue]:
    """
    Elige el mejor match entre varios candidatos
    
    Criterios:
    1. Mayor confianza
    2. Pasa validación (si se proporciona)
    """
    if not matches:
        return None
    
    if len(matches) == 1:
        return matches[0]
    
    # Ordenar por confianza
    sorted_matches = sorted(matches, key=lambda x: x.confidence, reverse=True)
    
    # Si hay función de validación, preferir los que la pasen
    if validation_func:
        for match in sorted_matches:
            if validation_func(match.transformed_value):
                return match
    
    return sorted_matches[0]
