"""
Template Fingerprint: Detecta y valida plantillas de cuadernos
Multi-template con scoring enterprise
"""
from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, Any, List, Optional
import json
import re
import unicodedata
from pathlib import Path
from openpyxl import load_workbook


def _norm(s: str) -> str:
    """Normaliza texto para comparación (quita tildes, mayúsculas, espacios)"""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("\n", " ").strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


def _tokenize_headers(row_values: List[Any]) -> List[str]:
    """Extrae tokens normalizados de una fila"""
    out = []
    for v in row_values:
        t = _norm(v)
        if t:
            out.append(t)
    return out


@dataclass
class Fingerprint:
    """Huella digital de un workbook"""
    sheetnames: List[str]
    parcels_sheet: Optional[str]
    header_row: Optional[int]
    header_tokens: List[str]
    columns_map: Dict[str, int]  # logical_name -> 1-based col index
    evidence: Dict[str, Any]


def fingerprint_workbook(
    file_path: str,
    parcels_sheet_hint_contains: List[str] = None,
    header_scan_max_rows: int = 60
) -> Fingerprint:
    """
    Genera fingerprint de un workbook Excel.
    
    Args:
        file_path: Ruta al archivo Excel
        parcels_sheet_hint_contains: Tokens que debe contener el nombre de la hoja de parcelas
        header_scan_max_rows: Máximo de filas a escanear buscando headers
    
    Returns:
        Fingerprint con toda la información detectada
    """
    parcels_sheet_hint_contains = parcels_sheet_hint_contains or ["PARCELAS"]

    wb = load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames[:]

    # 1) Detect parcels sheet by name contains
    parcels_sheet = None
    for sn in sheetnames:
        n = _norm(sn)
        if all(_norm(x) in n for x in parcels_sheet_hint_contains):
            parcels_sheet = sn
            break

    header_row = None
    header_tokens = []
    columns_map: Dict[str, int] = {}

    evidence: Dict[str, Any] = {
        "parcels_sheet_candidates": [],
        "header_row_candidates": []
    }

    if parcels_sheet:
        sh = wb[parcels_sheet]
        
        # 2) Find header row by scanning for tokens indicating SIGPAC table
        for r in range(1, min(header_scan_max_rows, sh.max_row) + 1):
            row_vals = [sh.cell(r, c).value for c in range(1, min(sh.max_column, 40) + 1)]
            toks = _tokenize_headers(row_vals)

            # Heuristic: needs at least 2 of these key tokens
            hit = 0
            key_tokens = [
                "Nº DE POLIGONO", "POLIGONO",
                "Nº DE PARCELA", "PARCELA",
                "Nº DE RECINTO", "RECINTO",
                "ESPECIE", "ESPECIE/ VARIEDAD", "ESPECIE/VARIEDAD",
                "TERMINO MUNICIPAL", "TÉRMINO MUNICIPAL"
            ]
            for key in key_tokens:
                if any(key in t for t in toks):
                    hit += 1

            if hit >= 2:
                evidence["header_row_candidates"].append({
                    "row": r,
                    "hits": hit,
                    "tokens": toks[:10]
                })
                # Choose first strong match
                header_row = r
                header_tokens = toks
                break

        # 3) Build columns_map from header_tokens positions
        if header_row:
            row_vals = [sh.cell(header_row, c).value for c in range(1, min(sh.max_column, 60) + 1)]
            for idx, v in enumerate(row_vals, start=1):
                t = _norm(v)
                if not t:
                    continue
                
                # Municipio
                if "TERMINO MUNICIPAL" in t or "TÉRMINO MUNICIPAL" in t:
                    columns_map.setdefault("municipio", idx)
                
                # Polígono
                if "Nº DE POLIGONO" in t or ("POLIGONO" in t and "polygon" not in columns_map):
                    columns_map.setdefault("polygon", idx)
                
                # Parcela
                if "Nº DE PARCELA" in t or ("PARCELA" in t and "parcel" not in columns_map):
                    columns_map.setdefault("parcel", idx)
                
                # Recinto
                if "Nº DE RECINTO" in t or ("RECINTO" in t and "recinto" not in columns_map):
                    columns_map.setdefault("recinto", idx)
                
                # Superficie
                if "SUPERFICIE CULTIVADA" in t or "SUP. CULTIVADA" in t:
                    columns_map.setdefault("surface", idx)
                
                # Cultivo
                if "ESPECIE" in t or "VARIEDAD" in t:
                    columns_map.setdefault("crop", idx)

    wb.close()

    return Fingerprint(
        sheetnames=sheetnames,
        parcels_sheet=parcels_sheet,
        header_row=header_row,
        header_tokens=header_tokens,
        columns_map=columns_map,
        evidence=evidence
    )


def load_registry(registry_path: str = None) -> Dict[str, Any]:
    """Carga el registro de templates"""
    if registry_path is None:
        registry_path = Path(__file__).parent / "templates_registry.json"
    
    with open(registry_path, "r", encoding="utf-8") as f:
        return json.load(f)


def match_template(fp: Fingerprint, registry: Dict[str, Any]) -> Dict[str, Any]:
    """
    Matchea un fingerprint contra los templates registrados.
    
    Returns:
        Dict con template_id, score, status (ACCEPT/WARN/REJECT), reasons
    """
    templates = registry.get("templates", [])
    policy = registry.get("policy", {"accept_threshold": 0.85, "warn_threshold": 0.65})

    def has_required_sheets_any_of(rule_sets: List[List[str]]) -> bool:
        sheetset = set(_norm(s) for s in fp.sheetnames)
        for rs in rule_sets:
            if all(_norm(x) in sheetset for x in rs):
                return True
        return False

    best = {
        "template_id": None,
        "score": 0.0,
        "status": "REJECT",
        "reasons": [],
        "policy": policy
    }

    for t in templates:
        score = 0.0
        reasons = []

        # A) Sheet match
        any_of = t.get("required_sheets_any_of", [])
        if any_of and has_required_sheets_any_of(any_of):
            score += t["scoring"]["sheet_match"]
        else:
            reasons.append("missing_required_sheets")

        # B) Parcels sheet name contains
        contains = t.get("parcels_sheet_name_contains", [])
        if fp.parcels_sheet:
            ps = _norm(fp.parcels_sheet)
            if all(_norm(x) in ps for x in contains):
                score += 0.05  # Minor boost
        else:
            reasons.append("parcels_sheet_not_found")

        # C) Header tokens required
        required_tokens = t.get("header_tokens_required", [])
        hit = 0
        for rt in required_tokens:
            rt_norm = _norm(rt)
            if any(rt_norm in ht for ht in fp.header_tokens):
                hit += 1
        
        if required_tokens:
            ratio = hit / max(1, len(required_tokens))
            score += t["scoring"]["header_tokens"] * ratio
            if ratio < 1.0:
                reasons.append(f"header_tokens_missing_{len(required_tokens)-hit}")
        else:
            score += t["scoring"]["header_tokens"]

        # D) Columns map required
        required_cols = t.get("columns_required", [])
        col_hit = sum(1 for c in required_cols if c in fp.columns_map)
        
        if required_cols:
            ratio2 = col_hit / max(1, len(required_cols))
            score += t["scoring"]["columns_map"] * ratio2
            if ratio2 < 1.0:
                reasons.append(f"columns_missing_{len(required_cols)-col_hit}")
        else:
            score += t["scoring"]["columns_map"]

        if score > best["score"]:
            best = {
                "template_id": t["template_id"],
                "score": round(score, 3),
                "status": "REJECT",
                "reasons": reasons,
                "policy": policy
            }

    # Decide status based on thresholds
    if best["score"] >= policy["accept_threshold"]:
        best["status"] = "ACCEPT"
    elif best["score"] >= policy["warn_threshold"]:
        best["status"] = "WARN"
    else:
        best["status"] = "REJECT"

    return best


def validate_template(file_path: str) -> Dict[str, Any]:
    """
    Validación completa de template (convenience function).
    
    Returns:
        Dict con fingerprint, decision, y resumen
    """
    fp = fingerprint_workbook(file_path, parcels_sheet_hint_contains=["2.1.", "PARCELAS"])
    registry = load_registry()
    decision = match_template(fp, registry)
    
    return {
        "fingerprint": {
            "sheetnames": fp.sheetnames,
            "parcels_sheet": fp.parcels_sheet,
            "header_row": fp.header_row,
            "columns_map": fp.columns_map,
            "header_tokens_count": len(fp.header_tokens)
        },
        "decision": decision,
        "summary": {
            "template_id": decision["template_id"],
            "score": decision["score"],
            "status": decision["status"],
            "is_valid": decision["status"] in ["ACCEPT", "WARN"]
        }
    }
