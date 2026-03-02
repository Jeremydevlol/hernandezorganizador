"""
Módulo de entrada/salida para archivos Excel
"""
import os
from pathlib import Path
from typing import Optional, Iterator
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelReader:
    """
    Lector de archivos Excel con caché y métodos de utilidad
    """
    
    def __init__(self, filepath: str, data_only: bool = True):
        """
        Inicializa el lector de Excel
        
        Args:
            filepath: Ruta al archivo Excel
            data_only: Si True, solo lee valores (no fórmulas)
        """
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {filepath}")
        
        self.workbook = load_workbook(str(self.filepath), data_only=data_only)
        self._cell_cache: dict[str, dict] = {}
    
    @property
    def sheet_names(self) -> list[str]:
        """Devuelve los nombres de las hojas"""
        return self.workbook.sheetnames
    
    def get_sheet(self, name: str) -> Optional[Worksheet]:
        """Obtiene una hoja por nombre"""
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        return None
    
    def get_cell_value(self, sheet_name: str, cell_ref: str) -> any:
        """
        Obtiene el valor de una celda específica
        
        Args:
            sheet_name: Nombre de la hoja
            cell_ref: Referencia de celda (ej: "A1", "B5")
        """
        sheet = self.get_sheet(sheet_name)
        if sheet is None:
            return None
        return sheet[cell_ref].value
    
    def iter_cells_with_values(self, sheet_name: str) -> Iterator[tuple[int, int, any]]:
        """
        Itera sobre todas las celdas con valores en una hoja
        
        Yields:
            Tuplas de (fila, columna, valor)
        """
        sheet = self.get_sheet(sheet_name)
        if sheet is None:
            return
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=col).value
                if value is not None:
                    yield (row, col, value)
    
    def find_cell_by_value(self, value: str, sheet_name: Optional[str] = None,
                           case_sensitive: bool = False) -> list[tuple[str, int, int]]:
        """
        Busca celdas que contengan un valor específico
        
        Args:
            value: Valor a buscar
            sheet_name: Hoja específica o None para buscar en todas
            case_sensitive: Si la búsqueda es sensible a mayúsculas
            
        Returns:
            Lista de tuplas (nombre_hoja, fila, columna)
        """
        results = []
        search_value = value if case_sensitive else value.lower()
        
        sheets_to_search = [sheet_name] if sheet_name else self.sheet_names
        
        for sn in sheets_to_search:
            for row, col, cell_value in self.iter_cells_with_values(sn):
                if cell_value is None:
                    continue
                cell_str = str(cell_value)
                compare_value = cell_str if case_sensitive else cell_str.lower()
                if search_value in compare_value:
                    results.append((sn, row, col))
        
        return results
    
    def get_adjacent_cell(self, sheet_name: str, row: int, col: int,
                          direction: str, distance: int = 1) -> tuple[int, int, any]:
        """
        Obtiene el valor de una celda adyacente
        
        Args:
            sheet_name: Nombre de la hoja
            row: Fila actual
            col: Columna actual
            direction: "right", "left", "up", "down"
            distance: Distancia en celdas
            
        Returns:
            Tupla (nueva_fila, nueva_col, valor)
        """
        sheet = self.get_sheet(sheet_name)
        if sheet is None:
            return (row, col, None)
        
        new_row, new_col = row, col
        
        if direction == "right":
            new_col = col + distance
        elif direction == "left":
            new_col = col - distance
        elif direction == "up":
            new_row = row - distance
        elif direction == "down":
            new_row = row + distance
        
        if new_row < 1 or new_col < 1:
            return (new_row, new_col, None)
        
        try:
            value = sheet.cell(row=new_row, column=new_col).value
            return (new_row, new_col, value)
        except Exception:
            return (new_row, new_col, None)
    
    def close(self):
        """Cierra el workbook"""
        self.workbook.close()


class TemplateWriter:
    """
    Escritor de plantillas Excel que preserva estilos y fórmulas
    """
    
    def __init__(self, template_path: str):
        """
        Inicializa el escritor a partir de una plantilla
        
        Args:
            template_path: Ruta a la plantilla Excel
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Plantilla no encontrada: {template_path}")
        
        # Cargar sin data_only para preservar fórmulas
        self.workbook = load_workbook(str(self.template_path), data_only=False)
        self._written_cells: list[tuple[str, str, any]] = []
    
    @property
    def sheet_names(self) -> list[str]:
        """Devuelve los nombres de las hojas"""
        return self.workbook.sheetnames
    
    def get_sheet(self, name: str) -> Optional[Worksheet]:
        """Obtiene una hoja por nombre"""
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        return None
    
    def write_value(self, sheet_name: str, cell_ref: str, value: any) -> bool:
        """
        Escribe un valor en una celda específica, preservando estilos
        
        Args:
            sheet_name: Nombre de la hoja
            cell_ref: Referencia de celda (ej: "A1", "B5")
            value: Valor a escribir
            
        Returns:
            True si se escribió correctamente
        """
        sheet = self.get_sheet(sheet_name)
        if sheet is None:
            return False
        
        try:
            cell = sheet[cell_ref]
            
            # Si la celda es parte de un rango fusionado, encontrar la celda principal
            if hasattr(cell, 'coordinate'):
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Obtener la celda superior izquierda del rango fusionado
                        main_cell = sheet.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col
                        )
                        main_cell.value = value
                        self._written_cells.append((sheet_name, cell_ref, value))
                        return True
            
            # Si no está fusionada, escribir directamente
            cell.value = value
            self._written_cells.append((sheet_name, cell_ref, value))
            return True
        except AttributeError:
            # Es una MergedCell, intentar encontrar la celda principal
            try:
                row, col = parse_cell_reference(cell_ref)
                for merged_range in sheet.merged_cells.ranges:
                    if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                        main_cell = sheet.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col
                        )
                        main_cell.value = value
                        self._written_cells.append((sheet_name, cell_ref, value))
                        return True
            except Exception as inner_e:
                print(f"Error escribiendo en celda fusionada {sheet_name}!{cell_ref}: {inner_e}")
                return False
        except Exception as e:
            print(f"Error escribiendo en {sheet_name}!{cell_ref}: {e}")
            return False
        
        return False
    
    def add_log_sheet(self, sheet_name: str = "LOG") -> Worksheet:
        """
        Añade una hoja de log al workbook
        
        Args:
            sheet_name: Nombre de la hoja de log
            
        Returns:
            La hoja creada
        """
        if sheet_name in self.workbook.sheetnames:
            # Eliminar si ya existe
            del self.workbook[sheet_name]
        
        ws = self.workbook.create_sheet(sheet_name)
        
        # Encabezados del log
        headers = [
            "Timestamp", "Campo", "Estado", "Método", 
            "Valor Original", "Valor Transformado",
            "Hoja Origen", "Celda Origen", 
            "Hoja Destino", "Celda Destino", "Mensaje"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        return ws
    
    def write_log_entries(self, sheet: Worksheet, entries: list, start_row: int = 2):
        """
        Escribe las entradas de log en la hoja
        
        Args:
            sheet: Hoja de log
            entries: Lista de LogEntry
            start_row: Fila inicial
        """
        from .types import LogEntry
        
        for idx, entry in enumerate(entries):
            row = start_row + idx
            sheet.cell(row=row, column=1, value=entry.timestamp.isoformat())
            sheet.cell(row=row, column=2, value=entry.field_key)
            sheet.cell(row=row, column=3, value=entry.status.value)
            sheet.cell(row=row, column=4, value=entry.extraction_method)
            sheet.cell(row=row, column=5, value=str(entry.raw_value) if entry.raw_value else "")
            sheet.cell(row=row, column=6, value=str(entry.transformed_value) if entry.transformed_value else "")
            sheet.cell(row=row, column=7, value=entry.source_sheet or "")
            sheet.cell(row=row, column=8, value=entry.source_cell or "")
            sheet.cell(row=row, column=9, value=entry.target_sheet or "")
            sheet.cell(row=row, column=10, value=entry.target_cell or "")
            sheet.cell(row=row, column=11, value=entry.message)
    
    def save(self, output_path: str):
        """
        Guarda el workbook en la ruta especificada
        
        Args:
            output_path: Ruta del archivo de salida
        """
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(str(output))
    
    def close(self):
        """Cierra el workbook"""
        self.workbook.close()


def get_cell_reference(row: int, col: int) -> str:
    """Convierte fila y columna a referencia de celda (ej: A1)"""
    return f"{get_column_letter(col)}{row}"


def parse_cell_reference(cell_ref: str) -> tuple[int, int]:
    """Convierte referencia de celda a fila y columna"""
    # Separar letras de números
    col_letters = ""
    row_num = ""
    
    for char in cell_ref:
        if char.isalpha():
            col_letters += char
        elif char.isdigit():
            row_num += char
    
    col = column_index_from_string(col_letters)
    row = int(row_num)
    
    return (row, col)
