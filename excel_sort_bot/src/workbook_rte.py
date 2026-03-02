"""
WORKBOOK REAL-TIME EDITOR (RTE) v2.0
Motor de edición determinístico con guardrails.

v2.0 Features:
- INSERT_ROWS / DELETE_ROWS
- RENAME_SHEET / ADD_SHEET / DELETE_SHEET
- SET_FORMULA (controlado)
- Validación de fórmulas

Regla de Oro:
- IA propone ops
- Backend valida y ejecuta
- Auditoría completa
- Rollback siempre posible
"""
import json
import shutil
import hashlib
import uuid
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple, Set
from dataclasses import dataclass, field, asdict
from enum import Enum
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


# ============================================
# ENUMS Y TIPOS
# ============================================

class OpType(str, Enum):
    # v1 ops
    SET_CELL = "SET_CELL"
    SET_RANGE = "SET_RANGE"
    CLEAR_RANGE = "CLEAR_RANGE"
    FIND_REPLACE = "FIND_REPLACE"
    # v2 ops - Rows
    INSERT_ROWS = "INSERT_ROWS"
    DELETE_ROWS = "DELETE_ROWS"
    COPY_ROW = "COPY_ROW"
    # v2 ops - Sheets
    RENAME_SHEET = "RENAME_SHEET"
    ADD_SHEET = "ADD_SHEET"
    DELETE_SHEET = "DELETE_SHEET"
    COPY_SHEET = "COPY_SHEET"
    # v2 ops - Formulas
    SET_FORMULA = "SET_FORMULA"
    # v2 ops - Columns
    INSERT_COLS = "INSERT_COLS"
    DELETE_COLS = "DELETE_COLS"


class ValueType(str, Enum):
    STRING = "string"
    NUMBER = "number"
    DATE = "date"
    FORMULA = "formula"
    BOOLEAN = "boolean"


class PermissionMode(str, Enum):
    STRICT = "STRICT"    # Solo SET_CELL/SET_RANGE en hojas permitidas
    POWER = "POWER"      # + insert/delete rows, find/replace
    ADMIN = "ADMIN"      # + rename/delete sheet, todo


# ============================================
# DATA CLASSES
# ============================================

@dataclass
class Op:
    """Operación individual sobre el workbook"""
    op: OpType
    sheet: str
    cell: Optional[str] = None
    range: Optional[str] = None
    value: Any = None
    value_type: ValueType = ValueType.STRING
    # Para FIND_REPLACE
    find: Optional[str] = None
    replace: Optional[str] = None
    # Para INSERT/DELETE ROWS
    row_start: Optional[int] = None
    row_count: Optional[int] = None
    # Para INSERT/DELETE COLS
    col_start: Optional[int] = None
    col_count: Optional[int] = None
    # Para RENAME/COPY SHEET
    new_name: Optional[str] = None
    source_sheet: Optional[str] = None
    # Para formulas
    formula: Optional[str] = None
    
    def to_dict(self) -> Dict:
        return {k: v.value if isinstance(v, Enum) else v 
                for k, v in asdict(self).items() if v is not None}


@dataclass
class OpResult:
    """Resultado de ejecutar una op"""
    success: bool
    op: Op
    cells_affected: int = 0
    rows_affected: int = 0
    old_values: Dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None


@dataclass
class Proposal:
    """Propuesta de cambios (preview)"""
    proposal_id: str
    session_id: str
    instruction: str
    ops: List[Op]
    diff_preview: Dict[str, Any]
    requires_confirmation: bool = False
    warnings: List[str] = field(default_factory=list)
    questions: List[str] = field(default_factory=list)
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass 
class AuditEntry:
    """Registro de auditoría"""
    audit_id: str
    session_id: str
    proposal_id: str
    ops_applied: List[Dict]
    cells_changed: int
    rows_changed: int
    input_checksum: str
    output_checksum: str
    user_id: Optional[str] = None
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())


# ============================================
# FORMULA VALIDATOR
# ============================================

class FormulaValidator:
    """Valida fórmulas permitidas"""
    
    # Fórmulas permitidas (whitelist)
    ALLOWED_FUNCTIONS = {
        # Matemáticas
        "SUM", "AVERAGE", "COUNT", "COUNTA", "COUNTIF", "COUNTIFS",
        "MAX", "MIN", "ROUND", "ROUNDUP", "ROUNDDOWN",
        # Lógicas
        "IF", "AND", "OR", "NOT", "IFERROR", "IFNA",
        # Texto
        "CONCATENATE", "CONCAT", "LEFT", "RIGHT", "MID", "LEN", "UPPER", "LOWER",
        # Fechas
        "TODAY", "NOW", "DATE", "YEAR", "MONTH", "DAY",
        # Búsqueda
        "VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "SUMIF", "SUMIFS",
    }
    
    # Funciones bloqueadas (peligrosas)
    BLOCKED_FUNCTIONS = {
        "INDIRECT", "OFFSET", "HYPERLINK", "WEBSERVICE", "CALL",
    }
    
    @classmethod
    def validate(cls, formula: str) -> Tuple[bool, Optional[str]]:
        """Valida si una fórmula es permitida"""
        if not formula.startswith("="):
            return False, "Fórmula debe empezar con ="
        
        # Extraer funciones usadas
        functions_used = set(re.findall(r'([A-Z]+)\s*\(', formula.upper()))
        
        # Verificar bloqueadas
        blocked = functions_used & cls.BLOCKED_FUNCTIONS
        if blocked:
            return False, f"Funciones bloqueadas: {blocked}"
        
        # Verificar que solo use permitidas
        unknown = functions_used - cls.ALLOWED_FUNCTIONS
        if unknown:
            return False, f"Funciones no permitidas: {unknown}"
        
        return True, None


# ============================================
# WORKBOOK CONTRACT (Reglas por hoja)
# ============================================

class WorkbookContract:
    """Define qué se puede editar y cómo en cada hoja"""
    
    def __init__(self, dict_path: str = None):
        if dict_path is None:
            dict_path = Path(__file__).parent / "workbook_dictionary.json"
        
        with open(dict_path, "r", encoding="utf-8") as f:
            self.data = json.load(f)
        
        self.sheets = self.data.get("sheets", {})
        
        # Reglas duras
        self.protected_rows = {
            "inf.trat 1": list(range(1, 11)),  # Headers fijos
            "inf.trat 2": list(range(1, 11)),
            "inf.trat 3": list(range(1, 11)),
            "inf.trat 4": list(range(1, 11)),
            "2.1. DATOS PARCELAS": list(range(1, 14)),
            "inf.gral 1": list(range(1, 100)),  # Toda la hoja protegida
            "inf.gral 2": list(range(1, 100)),
        }
        
        # Hojas que no se pueden eliminar/renombrar
        self.protected_sheets = {
            "inf.gral 1", "inf.gral 2", "2.1. DATOS PARCELAS",
            "inf.trat 1", "inf.trat 2", "inf.trat 3", "inf.trat 4",
            "reg.prod", "reg.fert.", "reg. cosecha"
        }
        
        self.editable_columns = {
            "inf.trat 1": {
                "A": "id_parcelas",
                "B": "especie", 
                "C": "variedad",
                "D": "superficie_tratada",
                "E": "fecha",
                "F": "problema_fito",
                "G": "nro_orden_tratamiento",
                "H": "nro_aplicacion",
                "I": "producto",
                "J": "nro_registro",
                "K": "dosis",
                "L": "eficacia",
                "M": "observaciones"
            }
        }
        
        self.column_validators = {
            "E": self._validate_date,
            "D": self._validate_number,
            "K": self._validate_dose,
        }
    
    def is_row_protected(self, sheet: str, row: int) -> bool:
        """Verifica si una fila está protegida"""
        protected = self.protected_rows.get(sheet, [])
        return row in protected
    
    def is_sheet_protected(self, sheet: str) -> bool:
        """Verifica si una hoja está protegida"""
        return sheet in self.protected_sheets
    
    def is_column_editable(self, sheet: str, col: str) -> bool:
        """Verifica si una columna es editable"""
        editable = self.editable_columns.get(sheet)
        if editable is None:
            return True  # Si no hay reglas, todo es editable
        return col in editable
    
    def validate_value(self, col: str, value: Any) -> Tuple[bool, Optional[str]]:
        """Valida un valor según la columna"""
        validator = self.column_validators.get(col)
        if validator:
            return validator(value)
        return True, None
    
    def _validate_date(self, value) -> Tuple[bool, Optional[str]]:
        if value is None:
            return True, None
        if isinstance(value, datetime):
            return True, None
        if isinstance(value, str):
            try:
                datetime.strptime(value, "%Y-%m-%d")
                return True, None
            except:
                try:
                    datetime.strptime(value, "%d/%m/%Y")
                    return True, None
                except:
                    return False, "Fecha inválida"
        return False, "Tipo de fecha no reconocido"
    
    def _validate_number(self, value) -> Tuple[bool, Optional[str]]:
        if value is None:
            return True, None
        try:
            float(str(value).replace(",", "."))
            return True, None
        except:
            return False, "Número inválido"
    
    def _validate_dose(self, value) -> Tuple[bool, Optional[str]]:
        if value is None:
            return True, None
        val_str = str(value).strip()
        if val_str.endswith(" l") or val_str.endswith(" kg"):
            return True, None
        try:
            float(val_str.replace(",", "."))
            return True, None
        except:
            return False, "Dosis inválida (formato: X,XX l o X,XX kg)"


# ============================================
# WORKBOOK REAL-TIME EDITOR v2.0
# ============================================

class WorkbookRTE:
    """
    Motor de edición en tiempo real v2.0
    
    Flujo:
    1. start_session() -> carga workbook
    2. preview(instruction) -> genera ops sin ejecutar
    3. commit(proposal_id) -> ejecuta ops
    4. undo(audit_id) -> rollback
    
    v2.0 Features:
    - INSERT_ROWS / DELETE_ROWS
    - RENAME_SHEET / ADD_SHEET / DELETE_SHEET
    - SET_FORMULA (con validación)
    """
    
    VERSION = "2.0"
    
    def __init__(self, 
                 file_path: str,
                 contract: WorkbookContract = None,
                 mode: PermissionMode = PermissionMode.STRICT,
                 checkpoints_dir: str = None):
        self.file_path = Path(file_path)
        self.contract = contract or WorkbookContract()
        self.mode = mode
        
        # Directorio para checkpoints/rollback
        if checkpoints_dir:
            self.checkpoints_dir = Path(checkpoints_dir)
        else:
            self.checkpoints_dir = self.file_path.parent / ".rte_checkpoints"
        self.checkpoints_dir.mkdir(exist_ok=True)
        
        # Estado
        self.session_id = str(uuid.uuid4())[:8]
        self.wb = None
        self.proposals: Dict[str, Proposal] = {}
        self.audit_log: List[AuditEntry] = []
        self.committed_keys: Set[str] = set()  # Idempotencia
    
    def open(self):
        """Abre el workbook para edición"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self.file_path}")
        self.wb = load_workbook(self.file_path)
        self._create_checkpoint("initial")
    
    def close(self):
        """Guarda y cierra"""
        if self.wb:
            self.wb.save(self.file_path)
            self.wb.close()
            self.wb = None
    
    def _create_checkpoint(self, name: str) -> str:
        """Crea checkpoint para rollback"""
        checkpoint_id = f"{self.session_id}_{name}_{datetime.now().strftime('%H%M%S')}"
        checkpoint_path = self.checkpoints_dir / f"{checkpoint_id}.xlsx"
        shutil.copy(self.file_path, checkpoint_path)
        return checkpoint_id
    
    def _file_checksum(self) -> str:
        """Calcula checksum del archivo"""
        with open(self.file_path, "rb") as f:
            return hashlib.md5(f.read()).hexdigest()[:12]
    
    # ============================================
    # PREVIEW (Dry-run)
    # ============================================
    
    def preview(self, 
                instruction: str,
                ops: List[Op] = None,
                selection: Dict = None) -> Proposal:
        """Preview de cambios sin ejecutar"""
        proposal_id = str(uuid.uuid4())[:8]
        warnings = []
        requires_confirmation = False
        
        if ops is None:
            ops = []
        
        validated_ops = []
        diff_preview = {"cells": [], "rows": [], "sheets": [], "summary": ""}
        
        for op in ops:
            # Verificar permisos
            if not self._check_permission(op):
                warnings.append(f"Operación no permitida en modo {self.mode.value}: {op.op.value}")
                requires_confirmation = True
                continue
            
            # Validaciones específicas por tipo de op
            if op.op in [OpType.SET_CELL, OpType.SET_RANGE]:
                if op.cell:
                    col = ''.join(filter(str.isalpha, op.cell))
                    row = int(''.join(filter(str.isdigit, op.cell)))
                    
                    if self.contract.is_row_protected(op.sheet, row):
                        warnings.append(f"Fila {row} protegida en {op.sheet}")
                        requires_confirmation = True
                        continue
                    
                    if not self.contract.is_column_editable(op.sheet, col):
                        warnings.append(f"Columna {col} no editable en {op.sheet}")
                        requires_confirmation = True
                        continue
                    
                    valid, error = self.contract.validate_value(col, op.value)
                    if not valid:
                        warnings.append(f"Valor inválido para {op.cell}: {error}")
                        requires_confirmation = True
            
            elif op.op in [OpType.DELETE_ROWS, OpType.INSERT_ROWS]:
                if op.row_start:
                    for r in range(op.row_start, op.row_start + (op.row_count or 1)):
                        if self.contract.is_row_protected(op.sheet, r):
                            warnings.append(f"No se puede modificar fila protegida {r}")
                            requires_confirmation = True
                            break
                
                # GUARDRAIL: Operaciones destructivas SIEMPRE requieren confirmación
                if op.op == OpType.DELETE_ROWS:
                    warnings.append(f"⚠️ DELETE_ROWS: Se eliminarán {op.row_count or 1} fila(s) desde la {op.row_start}")
                    requires_confirmation = True
                
                diff_preview["rows"].append({
                    "op": op.op.value,
                    "sheet": op.sheet,
                    "row_start": op.row_start,
                    "row_count": op.row_count
                })
            
            elif op.op in [OpType.RENAME_SHEET, OpType.DELETE_SHEET]:
                if self.contract.is_sheet_protected(op.sheet):
                    warnings.append(f"Hoja protegida: {op.sheet}")
                    requires_confirmation = True
                    continue
                
                # GUARDRAIL: DELETE_SHEET siempre requiere confirmación
                if op.op == OpType.DELETE_SHEET:
                    warnings.append(f"⚠️ DELETE_SHEET: Se eliminará la hoja '{op.sheet}' permanentemente")
                    requires_confirmation = True
                
                diff_preview["sheets"].append({
                    "op": op.op.value,
                    "sheet": op.sheet,
                    "new_name": op.new_name
                })
            
            elif op.op == OpType.SET_FORMULA:
                valid, error = FormulaValidator.validate(op.formula or op.value)
                if not valid:
                    warnings.append(f"Fórmula no permitida: {error}")
                    requires_confirmation = True
                    continue
            
            validated_ops.append(op)
            
            # Diff preview para SET_CELL
            if op.op == OpType.SET_CELL:
                current = self._get_cell_value(op.sheet, op.cell)
                diff_preview["cells"].append({
                    "cell": f"{op.sheet}!{op.cell}",
                    "old": current,
                    "new": op.value
                })
        
        diff_preview["summary"] = (
            f"{len(validated_ops)} ops, "
            f"{len(diff_preview['cells'])} celdas, "
            f"{len(diff_preview['rows'])} rows, "
            f"{len(diff_preview['sheets'])} sheets"
        )
        
        proposal = Proposal(
            proposal_id=proposal_id,
            session_id=self.session_id,
            instruction=instruction,
            ops=validated_ops,
            diff_preview=diff_preview,
            requires_confirmation=requires_confirmation,
            warnings=warnings
        )
        
        self.proposals[proposal_id] = proposal
        return proposal
    
    # ============================================
    # COMMIT
    # ============================================
    
    def commit(self, 
               proposal_id: str,
               idempotency_key: str = None) -> Dict[str, Any]:
        """Ejecuta las operaciones de una propuesta"""
        # Idempotencia
        if idempotency_key:
            if idempotency_key in self.committed_keys:
                return {
                    "success": False,
                    "error": "DUPLICATE_COMMIT",
                    "message": f"Key {idempotency_key} ya fue ejecutada"
                }
            self.committed_keys.add(idempotency_key)
        
        proposal = self.proposals.get(proposal_id)
        if not proposal:
            return {"success": False, "error": "PROPOSAL_NOT_FOUND"}
        
        if proposal.requires_confirmation:
            return {
                "success": False, 
                "error": "REQUIRES_CONFIRMATION",
                "warnings": proposal.warnings
            }
        
        # Checkpoint
        input_checksum = self._file_checksum()
        checkpoint_id = self._create_checkpoint(f"pre_{proposal_id}")
        
        # Ejecutar ops
        results = []
        cells_changed = 0
        rows_changed = 0
        
        for op in proposal.ops:
            result = self._execute_op(op)
            results.append(result)
            if result.success:
                cells_changed += result.cells_affected
                rows_changed += result.rows_affected
        
        # Guardar
        self.wb.save(self.file_path)
        output_checksum = self._file_checksum()
        
        # Auditoría
        audit_entry = AuditEntry(
            audit_id=str(uuid.uuid4())[:8],
            session_id=self.session_id,
            proposal_id=proposal_id,
            ops_applied=[op.to_dict() for op in proposal.ops],
            cells_changed=cells_changed,
            rows_changed=rows_changed,
            input_checksum=input_checksum,
            output_checksum=output_checksum
        )
        self.audit_log.append(audit_entry)
        
        return {
            "success": True,
            "audit_id": audit_entry.audit_id,
            "cells_changed": cells_changed,
            "rows_changed": rows_changed,
            "checksum": output_checksum,
            "checkpoint_id": checkpoint_id
        }
    
    # ============================================
    # UNDO
    # ============================================
    
    def undo(self, checkpoint_id: str = None) -> Dict[str, Any]:
        """Restaura desde un checkpoint"""
        if checkpoint_id is None:
            checkpoints = sorted(self.checkpoints_dir.glob(f"{self.session_id}_*.xlsx"))
            if not checkpoints:
                return {"success": False, "error": "NO_CHECKPOINTS"}
            checkpoint_path = checkpoints[-1]
        else:
            checkpoint_path = self.checkpoints_dir / f"{checkpoint_id}.xlsx"
        
        if not checkpoint_path.exists():
            return {"success": False, "error": "CHECKPOINT_NOT_FOUND"}
        
        if self.wb:
            self.wb.close()
        
        shutil.copy(checkpoint_path, self.file_path)
        self.wb = load_workbook(self.file_path)
        
        return {
            "success": True,
            "restored_from": checkpoint_path.name,
            "checksum": self._file_checksum()
        }
    
    # ============================================
    # EJECUTORES DE OPS
    # ============================================
    
    def _execute_op(self, op: Op) -> OpResult:
        """Ejecuta una operación"""
        try:
            # v1 ops
            if op.op == OpType.SET_CELL:
                return self._exec_set_cell(op)
            elif op.op == OpType.SET_RANGE:
                return self._exec_set_range(op)
            elif op.op == OpType.CLEAR_RANGE:
                return self._exec_clear_range(op)
            elif op.op == OpType.FIND_REPLACE:
                return self._exec_find_replace(op)
            # v2 ops - Rows
            elif op.op == OpType.INSERT_ROWS:
                return self._exec_insert_rows(op)
            elif op.op == OpType.DELETE_ROWS:
                return self._exec_delete_rows(op)
            elif op.op == OpType.COPY_ROW:
                return self._exec_copy_row(op)
            # v2 ops - Sheets
            elif op.op == OpType.RENAME_SHEET:
                return self._exec_rename_sheet(op)
            elif op.op == OpType.ADD_SHEET:
                return self._exec_add_sheet(op)
            elif op.op == OpType.DELETE_SHEET:
                return self._exec_delete_sheet(op)
            elif op.op == OpType.COPY_SHEET:
                return self._exec_copy_sheet(op)
            # v2 ops - Formulas
            elif op.op == OpType.SET_FORMULA:
                return self._exec_set_formula(op)
            # v2 ops - Columns
            elif op.op == OpType.INSERT_COLS:
                return self._exec_insert_cols(op)
            elif op.op == OpType.DELETE_COLS:
                return self._exec_delete_cols(op)
            else:
                return OpResult(success=False, op=op, error=f"Op no soportada: {op.op}")
        except Exception as e:
            return OpResult(success=False, op=op, error=str(e))
    
    # --- v1 ops ---
    
    def _exec_set_cell(self, op: Op) -> OpResult:
        """SET_CELL: poner valor en una celda"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        cell = ws[op.cell]
        
        if isinstance(cell, MergedCell):
            return OpResult(success=False, op=op, error=f"Celda mergeada: {op.cell}")
        
        old_value = cell.value
        value = self._convert_value(op.value, op.value_type)
        cell.value = value
        
        return OpResult(
            success=True,
            op=op,
            cells_affected=1,
            old_values={op.cell: old_value}
        )
    
    def _exec_set_range(self, op: Op) -> OpResult:
        """SET_RANGE: poner valores en un rango"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        old_values = {}
        cells_affected = 0
        
        if not isinstance(op.value, list):
            return OpResult(success=False, op=op, error="SET_RANGE requiere lista de valores")
        
        start, end = op.range.split(":")
        start_col = column_index_from_string(''.join(filter(str.isalpha, start)))
        start_row = int(''.join(filter(str.isdigit, start)))
        
        for i, row_values in enumerate(op.value):
            row = start_row + i
            for j, val in enumerate(row_values):
                col = start_col + j
                cell = ws.cell(row, col)
                if not isinstance(cell, MergedCell):
                    old_values[f"{get_column_letter(col)}{row}"] = cell.value
                    cell.value = val
                    cells_affected += 1
        
        return OpResult(success=True, op=op, cells_affected=cells_affected, old_values=old_values)
    
    def _exec_clear_range(self, op: Op) -> OpResult:
        """CLEAR_RANGE: limpiar un rango"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        old_values = {}
        cells_affected = 0
        
        for row in ws[op.range]:
            for cell in row:
                if not isinstance(cell, MergedCell):
                    coord = f"{cell.column_letter}{cell.row}"
                    old_values[coord] = cell.value
                    cell.value = None
                    cells_affected += 1
        
        return OpResult(success=True, op=op, cells_affected=cells_affected, old_values=old_values)
    
    def _exec_find_replace(self, op: Op) -> OpResult:
        """FIND_REPLACE: buscar y reemplazar"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        old_values = {}
        cells_affected = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if op.find in cell.value:
                        coord = f"{cell.column_letter}{cell.row}"
                        old_values[coord] = cell.value
                        cell.value = cell.value.replace(op.find, op.replace)
                        cells_affected += 1
        
        return OpResult(success=True, op=op, cells_affected=cells_affected, old_values=old_values)
    
    # --- v2 ops - Rows ---
    
    def _exec_insert_rows(self, op: Op) -> OpResult:
        """INSERT_ROWS: insertar filas"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        row_start = op.row_start or 1
        row_count = op.row_count or 1
        
        ws.insert_rows(row_start, row_count)
        
        return OpResult(
            success=True,
            op=op,
            rows_affected=row_count
        )
    
    def _exec_delete_rows(self, op: Op) -> OpResult:
        """DELETE_ROWS: eliminar filas"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        row_start = op.row_start or 1
        row_count = op.row_count or 1
        
        # Guardar valores antes de borrar
        old_values = {}
        for r in range(row_start, row_start + row_count):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.value:
                    old_values[f"{get_column_letter(c)}{r}"] = cell.value
        
        ws.delete_rows(row_start, row_count)
        
        return OpResult(
            success=True,
            op=op,
            rows_affected=row_count,
            old_values=old_values
        )
    
    def _exec_copy_row(self, op: Op) -> OpResult:
        """COPY_ROW: copiar una fila a otra posición"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        source_row = op.row_start
        target_row = op.row_count  # Usamos row_count como destino
        
        if not source_row or not target_row:
            return OpResult(success=False, op=op, error="COPY_ROW requiere row_start y row_count (destino)")
        
        cells_affected = 0
        for c in range(1, ws.max_column + 1):
            source_cell = ws.cell(source_row, c)
            target_cell = ws.cell(target_row, c)
            if not isinstance(target_cell, MergedCell):
                target_cell.value = source_cell.value
                cells_affected += 1
        
        return OpResult(success=True, op=op, cells_affected=cells_affected, rows_affected=1)
    
    # --- v2 ops - Sheets ---
    
    def _exec_rename_sheet(self, op: Op) -> OpResult:
        """RENAME_SHEET: renombrar hoja"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        if not op.new_name:
            return OpResult(success=False, op=op, error="RENAME_SHEET requiere new_name")
        
        if op.new_name in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Ya existe hoja: {op.new_name}")
        
        ws = self.wb[op.sheet]
        old_name = ws.title
        ws.title = op.new_name
        
        return OpResult(
            success=True,
            op=op,
            old_values={"old_name": old_name}
        )
    
    def _exec_add_sheet(self, op: Op) -> OpResult:
        """ADD_SHEET: añadir nueva hoja"""
        new_name = op.new_name or op.sheet
        
        if new_name in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Ya existe hoja: {new_name}")
        
        self.wb.create_sheet(title=new_name)
        
        return OpResult(success=True, op=op)
    
    def _exec_delete_sheet(self, op: Op) -> OpResult:
        """DELETE_SHEET: eliminar hoja"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        del self.wb[op.sheet]
        
        return OpResult(success=True, op=op)
    
    def _exec_copy_sheet(self, op: Op) -> OpResult:
        """COPY_SHEET: copiar hoja"""
        source = op.source_sheet or op.sheet
        
        if source not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja origen no existe: {source}")
        
        new_name = op.new_name or f"{source}_copy"
        
        if new_name in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Ya existe hoja: {new_name}")
        
        source_ws = self.wb[source]
        new_ws = self.wb.copy_worksheet(source_ws)
        new_ws.title = new_name
        
        return OpResult(success=True, op=op)
    
    # --- v2 ops - Formulas ---
    
    def _exec_set_formula(self, op: Op) -> OpResult:
        """SET_FORMULA: poner fórmula en celda"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        formula = op.formula or op.value
        
        # Validar fórmula
        valid, error = FormulaValidator.validate(formula)
        if not valid:
            return OpResult(success=False, op=op, error=error)
        
        ws = self.wb[op.sheet]
        cell = ws[op.cell]
        
        if isinstance(cell, MergedCell):
            return OpResult(success=False, op=op, error=f"Celda mergeada: {op.cell}")
        
        old_value = cell.value
        cell.value = formula
        
        return OpResult(
            success=True,
            op=op,
            cells_affected=1,
            old_values={op.cell: old_value}
        )
    
    # --- v2 ops - Columns ---
    
    def _exec_insert_cols(self, op: Op) -> OpResult:
        """INSERT_COLS: insertar columnas"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        col_start = op.col_start or 1
        col_count = op.col_count or 1
        
        ws.insert_cols(col_start, col_count)
        
        return OpResult(success=True, op=op, cells_affected=col_count * ws.max_row)
    
    def _exec_delete_cols(self, op: Op) -> OpResult:
        """DELETE_COLS: eliminar columnas"""
        if op.sheet not in self.wb.sheetnames:
            return OpResult(success=False, op=op, error=f"Hoja no existe: {op.sheet}")
        
        ws = self.wb[op.sheet]
        col_start = op.col_start or 1
        col_count = op.col_count or 1
        
        ws.delete_cols(col_start, col_count)
        
        return OpResult(success=True, op=op, cells_affected=col_count * ws.max_row)
    
    # ============================================
    # HELPERS
    # ============================================
    
    def _check_permission(self, op: Op) -> bool:
        """Verifica permisos según modo"""
        if self.mode == PermissionMode.ADMIN:
            return True
        
        if self.mode == PermissionMode.STRICT:
            return op.op in [OpType.SET_CELL, OpType.SET_RANGE]
        
        if self.mode == PermissionMode.POWER:
            return op.op in [
                OpType.SET_CELL, OpType.SET_RANGE, 
                OpType.CLEAR_RANGE, OpType.FIND_REPLACE,
                OpType.INSERT_ROWS, OpType.DELETE_ROWS,
                OpType.SET_FORMULA
            ]
        
        return False
    
    def _get_cell_value(self, sheet: str, cell: str) -> Any:
        """Obtiene valor actual de una celda"""
        if sheet not in self.wb.sheetnames:
            return None
        ws = self.wb[sheet]
        return ws[cell].value
    
    def _convert_value(self, value: Any, value_type: ValueType) -> Any:
        """Convierte valor al tipo especificado"""
        if value is None:
            return None
        
        if value_type == ValueType.DATE:
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                for fmt in ["%Y-%m-%d", "%d/%m/%Y"]:
                    try:
                        return datetime.strptime(value, fmt)
                    except:
                        pass
            return value
        
        if value_type == ValueType.NUMBER:
            try:
                return float(str(value).replace(",", "."))
            except:
                return value
        
        if value_type == ValueType.BOOLEAN:
            return bool(value)
        
        if value_type == ValueType.FORMULA:
            return value  # Fórmulas se pasan tal cual
        
        return value
    
    def get_session_info(self) -> Dict:
        """Info de la sesión actual"""
        return {
            "session_id": self.session_id,
            "version": self.VERSION,
            "file_path": str(self.file_path),
            "mode": self.mode.value,
            "sheets": self.wb.sheetnames if self.wb else [],
            "proposals_count": len(self.proposals),
            "commits_count": len(self.audit_log),
            "checksum": self._file_checksum() if self.file_path.exists() else None
        }
    
    def get_capabilities(self) -> List[str]:
        """Devuelve capabilities según modo"""
        if self.mode == PermissionMode.ADMIN:
            return [op.value for op in OpType]
        
        if self.mode == PermissionMode.POWER:
            return [
                "SET_CELL", "SET_RANGE", "CLEAR_RANGE", "FIND_REPLACE",
                "INSERT_ROWS", "DELETE_ROWS", "SET_FORMULA"
            ]
        
        # STRICT
        return ["SET_CELL", "SET_RANGE"]
