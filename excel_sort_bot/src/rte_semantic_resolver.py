"""
RTE SEMANTIC RESOLVER v1.0
Traduce referencias semánticas a coordenadas Excel.

Ejemplos:
- "orden 39" → buscar en inf.trat 1 col A y mapear a fila real
- "parcela 15-234-1" → buscar pol/parc/rec en 2.1 DATOS PARCELAS
- "último tratamiento de Glifosato" → buscar por producto/fecha
- "todas las cebadas de Rasueros" → filtrar por cultivo+municipio
"""
import re
from typing import Dict, Any, List, Optional, Tuple, Set
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ============================================
# DATA CLASSES
# ============================================

@dataclass
class SemanticTarget:
    """Resultado de resolver una referencia semántica"""
    sheet: str
    cells: List[str]  # ["E15", "E16", ...] o ["A15:M15"]
    rows: List[int]   # [15, 16, ...]
    description: str  # "Orden 39 (fila 15)"
    confidence: float = 1.0  # 0-1
    ambiguous: bool = False
    alternatives: List[str] = field(default_factory=list)


@dataclass
class TreatmentRow:
    """Representa una fila de tratamiento"""
    row: int
    orden: int
    parcela_id: Any
    cultivo: str
    fecha: Any
    producto: str
    dosis: str
    municipio: Optional[str] = None


@dataclass
class ParcelRow:
    """Representa una fila de parcela"""
    row: int
    nro_orden: int
    municipio: str
    polygon: int
    parcel: int
    recinto: int
    cultivo: str
    superficie: float


# ============================================
# SEMANTIC RESOLVER
# ============================================

class SemanticResolver:
    """
    Resuelve referencias semánticas a coordenadas Excel.
    
    Uso:
        resolver = SemanticResolver(workbook_path)
        target = resolver.resolve("orden 39", "inf.trat 1")
        # target.cells = ["A15:M15"], target.rows = [15]
    """
    
    def __init__(self, workbook_path: str = None, workbook = None):
        self.wb_path = workbook_path
        self.wb = workbook
        self._owns_workbook = False
        
        if workbook is None and workbook_path:
            self.wb = load_workbook(workbook_path, data_only=True)
            self._owns_workbook = True
        
        # Cache de índices
        self._tratamientos_index: Dict[int, TreatmentRow] = {}
        self._parcelas_index: Dict[str, ParcelRow] = {}  # "pol-parc-rec" -> row
        self._productos_index: Dict[str, List[int]] = {}  # producto -> [rows]
        self._cultivos_index: Dict[str, List[int]] = {}   # cultivo -> [rows]
        
        # Construir índices
        if self.wb:
            self._build_indices()
    
    def close(self):
        if self._owns_workbook and self.wb:
            self.wb.close()
    
    def _build_indices(self):
        """Construye índices de búsqueda rápida"""
        self._index_tratamientos()
        self._index_parcelas()
    
    def _index_tratamientos(self):
        """Indexa tratamientos de inf.trat 1-4"""
        for sheet_name in ["inf.trat 1", "inf.trat 2", "inf.trat 3", "inf.trat 4"]:
            if sheet_name not in self.wb.sheetnames:
                continue
            
            ws = self.wb[sheet_name]
            
            for row in range(11, ws.max_row + 1):
                orden = ws.cell(row, 1).value  # Col A
                if orden is None:
                    continue
                
                try:
                    orden_int = int(orden)
                except (ValueError, TypeError):
                    continue
                
                tratamiento = TreatmentRow(
                    row=row,
                    orden=orden_int,
                    parcela_id=ws.cell(row, 1).value,
                    cultivo=str(ws.cell(row, 2).value or "").upper(),
                    fecha=ws.cell(row, 5).value,
                    producto=str(ws.cell(row, 9).value or "").upper(),
                    dosis=str(ws.cell(row, 11).value or "")
                )
                
                # Indexar por orden
                self._tratamientos_index[orden_int] = tratamiento
                
                # Indexar por producto
                if tratamiento.producto:
                    prod_key = tratamiento.producto.strip().upper()
                    if prod_key not in self._productos_index:
                        self._productos_index[prod_key] = []
                    self._productos_index[prod_key].append(row)
                
                # Indexar por cultivo
                if tratamiento.cultivo:
                    cult_key = tratamiento.cultivo.strip().upper()
                    if cult_key not in self._cultivos_index:
                        self._cultivos_index[cult_key] = []
                    self._cultivos_index[cult_key].append(row)
    
    def _index_parcelas(self):
        """Indexa parcelas de 2.1. DATOS PARCELAS"""
        if "2.1. DATOS PARCELAS" not in self.wb.sheetnames:
            return
        
        ws = self.wb["2.1. DATOS PARCELAS"]
        
        for row in range(14, ws.max_row + 1):
            nro_orden = ws.cell(row, 2).value  # Col B
            if nro_orden is None:
                continue
            
            try:
                polygon = int(ws.cell(row, 7).value or 0)
                parcel = int(ws.cell(row, 8).value or 0)
                recinto = int(ws.cell(row, 9).value or 0)
            except (ValueError, TypeError):
                continue
            
            key = f"{polygon}-{parcel}-{recinto}"
            
            self._parcelas_index[key] = ParcelRow(
                row=row,
                nro_orden=int(nro_orden) if nro_orden else 0,
                municipio=str(ws.cell(row, 4).value or ""),
                polygon=polygon,
                parcel=parcel,
                recinto=recinto,
                cultivo=str(ws.cell(row, 13).value or "").upper(),
                superficie=float(ws.cell(row, 12).value or 0)
            )
    
    # ============================================
    # PUBLIC API
    # ============================================
    
    def resolve(self, reference: str, default_sheet: str = "inf.trat 1") -> SemanticTarget:
        """
        Resuelve una referencia semántica.
        
        Args:
            reference: "orden 39", "parcela 15-234-1", "todas las cebadas", etc.
            default_sheet: Hoja por defecto si no se especifica
        
        Returns:
            SemanticTarget con las celdas/filas resueltas
        """
        ref_lower = reference.lower().strip()
        
        # 1. Referencia directa a celda (A1, E15, etc.)
        match = re.match(r'^([A-Z]+)(\d+)$', reference.upper())
        if match:
            return SemanticTarget(
                sheet=default_sheet,
                cells=[reference.upper()],
                rows=[int(match.group(2))],
                description=f"Celda {reference.upper()}"
            )
        
        # 2. Referencia a fila directa ("fila 15", "row 15")
        match = re.search(r'fila\s*(\d+)|row\s*(\d+)', ref_lower)
        if match:
            row = int(match.group(1) or match.group(2))
            return SemanticTarget(
                sheet=default_sheet,
                cells=[f"A{row}:M{row}"],
                rows=[row],
                description=f"Fila {row}"
            )
        
        # 3. Referencia por ORDEN ("orden 39", "tratamiento 39")
        match = re.search(r'(?:orden|tratamiento|nro|#)\s*(\d+)', ref_lower)
        if match:
            orden = int(match.group(1))
            return self._resolve_by_orden(orden, default_sheet)
        
        # 4. Referencia por parcela SIGPAC ("parcela 15-234-1", "pol 15 parc 234 rec 1")
        match = re.search(r'(?:parcela\s+)?(\d+)[/-](\d+)[/-](\d+)', ref_lower)
        if not match:
            match = re.search(r'pol(?:igono)?\s*(\d+).*parc(?:ela)?\s*(\d+).*rec(?:into)?\s*(\d+)', ref_lower)
        if match:
            pol, parc, rec = int(match.group(1)), int(match.group(2)), int(match.group(3))
            return self._resolve_by_sigpac(pol, parc, rec, default_sheet)
        
        # 5. Referencia por producto ("glifosato", "tratamientos con roundup")
        for producto, rows in self._productos_index.items():
            if producto.lower() in ref_lower or ref_lower in producto.lower():
                return SemanticTarget(
                    sheet=default_sheet,
                    cells=[f"A{r}:M{r}" for r in rows],
                    rows=rows,
                    description=f"Tratamientos con {producto} ({len(rows)} filas)"
                )
        
        # 6. Referencia por cultivo ("todas las cebadas", "tratamientos de vid")
        match = re.search(r'(?:todas?\s+(?:las?\s+)?|tratamientos?\s+(?:de\s+)?|cultivo\s+)(\w+)', ref_lower)
        if match:
            cultivo_ref = match.group(1).upper()
            matching_rows = []
            for cultivo, rows in self._cultivos_index.items():
                if cultivo_ref in cultivo or cultivo in cultivo_ref:
                    matching_rows.extend(rows)
            if matching_rows:
                return SemanticTarget(
                    sheet=default_sheet,
                    cells=[f"A{r}:M{r}" for r in matching_rows],
                    rows=matching_rows,
                    description=f"Tratamientos de {cultivo_ref} ({len(matching_rows)} filas)"
                )
        
        # 7. Referencia temporal ("último tratamiento", "tratamientos de ayer")
        if "último" in ref_lower or "ultima" in ref_lower:
            return self._resolve_last_treatment(default_sheet)
        
        # 8. Rango de órdenes ("orden 10 a 20", "tratamientos 10-20")
        match = re.search(r'(?:orden(?:es)?|tratamientos?)\s*(\d+)\s*(?:a|hasta|-)\s*(\d+)', ref_lower)
        if match:
            start, end = int(match.group(1)), int(match.group(2))
            return self._resolve_orden_range(start, end, default_sheet)
        
        # No se pudo resolver
        return SemanticTarget(
            sheet=default_sheet,
            cells=[],
            rows=[],
            description=f"No se pudo resolver: {reference}",
            confidence=0.0,
            ambiguous=True,
            alternatives=["Usa 'orden X', 'fila X', 'parcela X-Y-Z', o nombre de producto/cultivo"]
        )
    
    def _resolve_by_orden(self, orden: int, sheet: str) -> SemanticTarget:
        """Resuelve por número de orden"""
        if orden in self._tratamientos_index:
            trat = self._tratamientos_index[orden]
            return SemanticTarget(
                sheet=sheet,
                cells=[f"A{trat.row}:M{trat.row}"],
                rows=[trat.row],
                description=f"Orden {orden} → fila {trat.row} ({trat.producto})"
            )
        
        return SemanticTarget(
            sheet=sheet,
            cells=[],
            rows=[],
            description=f"Orden {orden} no encontrado",
            confidence=0.0,
            ambiguous=True
        )
    
    def _resolve_by_sigpac(self, pol: int, parc: int, rec: int, sheet: str) -> SemanticTarget:
        """Resuelve por coordenadas SIGPAC"""
        key = f"{pol}-{parc}-{rec}"
        
        if key in self._parcelas_index:
            parcela = self._parcelas_index[key]
            # Buscar tratamientos de esta parcela
            matching_rows = [
                row for orden, trat in self._tratamientos_index.items()
                if trat.orden == parcela.nro_orden
            ]
            
            if matching_rows:
                return SemanticTarget(
                    sheet=sheet,
                    cells=[f"A{r}:M{r}" for r in matching_rows],
                    rows=matching_rows,
                    description=f"Parcela {key} → {len(matching_rows)} tratamientos"
                )
            else:
                return SemanticTarget(
                    sheet="2.1. DATOS PARCELAS",
                    cells=[f"A{parcela.row}:N{parcela.row}"],
                    rows=[parcela.row],
                    description=f"Parcela {key} (sin tratamientos)"
                )
        
        return SemanticTarget(
            sheet=sheet,
            cells=[],
            rows=[],
            description=f"Parcela {key} no encontrada",
            confidence=0.0,
            ambiguous=True
        )
    
    def _resolve_last_treatment(self, sheet: str) -> SemanticTarget:
        """Resuelve el último tratamiento por fecha"""
        if not self._tratamientos_index:
            return SemanticTarget(
                sheet=sheet, cells=[], rows=[],
                description="No hay tratamientos", confidence=0.0
            )
        
        # Buscar el de mayor fecha
        latest = None
        latest_date = None
        
        for trat in self._tratamientos_index.values():
            if trat.fecha:
                fecha = trat.fecha
                if isinstance(fecha, datetime):
                    if latest_date is None or fecha > latest_date:
                        latest_date = fecha
                        latest = trat
        
        if latest:
            return SemanticTarget(
                sheet=sheet,
                cells=[f"A{latest.row}:M{latest.row}"],
                rows=[latest.row],
                description=f"Último tratamiento: orden {latest.orden}, fila {latest.row}"
            )
        
        # Fallback: último por número de orden
        max_orden = max(self._tratamientos_index.keys())
        trat = self._tratamientos_index[max_orden]
        return SemanticTarget(
            sheet=sheet,
            cells=[f"A{trat.row}:M{trat.row}"],
            rows=[trat.row],
            description=f"Último tratamiento: orden {trat.orden}, fila {trat.row}"
        )
    
    def _resolve_orden_range(self, start: int, end: int, sheet: str) -> SemanticTarget:
        """Resuelve un rango de órdenes"""
        rows = []
        for orden in range(start, end + 1):
            if orden in self._tratamientos_index:
                rows.append(self._tratamientos_index[orden].row)
        
        if rows:
            return SemanticTarget(
                sheet=sheet,
                cells=[f"A{r}:M{r}" for r in rows],
                rows=rows,
                description=f"Órdenes {start}-{end} → {len(rows)} tratamientos"
            )
        
        return SemanticTarget(
            sheet=sheet, cells=[], rows=[],
            description=f"Órdenes {start}-{end} no encontradas",
            confidence=0.0
        )
    
    # ============================================
    # COLUMN RESOLVER
    # ============================================
    
    def resolve_column(self, field_name: str, sheet: str = "inf.trat 1") -> Optional[str]:
        """
        Resuelve un nombre de campo a columna.
        
        Ejemplos:
            "fecha" → "E"
            "producto" → "I"
            "dosis" → "K"
        """
        # Mapeo semántico → columna
        COLUMN_MAP = {
            "inf.trat 1": {
                "orden": "A", "id": "A", "parcela": "A",
                "cultivo": "B", "especie": "B", "crop": "B",
                "variedad": "C",
                "superficie": "D", "surface": "D", "ha": "D",
                "fecha": "E", "date": "E",
                "plaga": "F", "problema": "F", "pest": "F",
                "nro_tratamiento": "G",
                "aplicacion": "H", "nro_aplicacion": "H",
                "producto": "I", "product": "I", "fitosanitario": "I",
                "registro": "J", "nro_registro": "J",
                "dosis": "K", "dose": "K",
                "eficacia": "L",
                "observaciones": "M", "notas": "M", "notes": "M"
            },
            "2.1. DATOS PARCELAS": {
                "nro_orden": "B", "orden": "B",
                "municipio": "D",
                "poligono": "G", "polygon": "G",
                "parcela": "H", "parcel": "H",
                "recinto": "I",
                "superficie_sigpac": "K",
                "superficie": "L", "surface": "L",
                "cultivo": "M", "crop": "M"
            }
        }
        
        sheet_map = COLUMN_MAP.get(sheet, {})
        field_lower = field_name.lower().strip()
        
        return sheet_map.get(field_lower)
    
    def get_cell_for_field(self, reference: str, field_name: str, 
                           default_sheet: str = "inf.trat 1") -> Optional[str]:
        """
        Combina resolver de referencia + columna.
        
        Ejemplo:
            get_cell_for_field("orden 39", "fecha") → "E15"
        """
        target = self.resolve(reference, default_sheet)
        if not target.rows:
            return None
        
        col = self.resolve_column(field_name, target.sheet)
        if not col:
            return None
        
        # Devuelve la primera celda
        return f"{col}{target.rows[0]}"
    
    def get_range_for_field(self, reference: str, field_name: str,
                            default_sheet: str = "inf.trat 1") -> List[str]:
        """
        Devuelve todas las celdas para un campo en múltiples filas.
        
        Ejemplo:
            get_range_for_field("todas las cebadas", "dosis") → ["K15", "K20", "K25"]
        """
        target = self.resolve(reference, default_sheet)
        if not target.rows:
            return []
        
        col = self.resolve_column(field_name, target.sheet)
        if not col:
            return []
        
        return [f"{col}{row}" for row in target.rows]
