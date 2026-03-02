"""
Módulo para gestionar Parcelas y Tratamientos
VERSION 2.0 - Con soporte completo de recinto y separación cultivo/municipio
"""
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell


class Parcel:
    """Estructura de datos de una parcela/recinto"""
    def __init__(self, polygon: int, parcel: int, recinto: int = 1, 
                 crop: str = "", surface_ha: float = 0.0, 
                 municipio: str = "", nro_orden: int = 0, raw_data: dict = None):
        self.polygon = polygon
        self.parcel = parcel
        self.recinto = recinto
        self.crop = crop.upper().strip() if crop else ""
        self.crop_norm = self._normalize_crop(self.crop)
        self.surface_ha = surface_ha
        self.municipio = municipio.upper().strip() if municipio else ""
        self.nro_orden = nro_orden  # Id. Parcelas para inf.trat 1
        self.raw_data = raw_data or {}
    
    def _normalize_crop(self, crop: str) -> str:
        """Normaliza nombres de cultivo comunes para búsquedas flexibles"""
        crop_upper = crop.upper()
        norm_map = {
            "VID": "VIÑEDO", "VIÑA": "VIÑEDO", "VINA": "VIÑEDO", "UVA": "VIÑEDO",
            "OLIVO": "OLIVAR", "OLIVA": "OLIVAR", "ACEITE": "OLIVAR", "ACEITUNA": "OLIVAR",
            "ALMENDRO": "ALMENDRO", "ALMENDRA": "ALMENDRO",
            "TRIGO": "CEREALES", "CEBADA": "CEREALES", "AVENA": "CEREALES",
            "GIRASOL": "GIRASOL", "MAIZ": "MAÍZ", "MAÍZ": "MAÍZ",
        }
        for key, val in norm_map.items():
            if key in crop_upper:
                return val
        return crop_upper
    
    def matches_filter(self, scope: str, targets: Dict[str, Any], logic: str = "AND") -> bool:
        """
        Evalúa si esta parcela coincide con el filtro dado
        
        Args:
            scope: "ALL", "BY_CROP", "BY_MUNICIPIO", "BY_IDS", etc.
            targets: dict con crop, municipio, polygon, parcel, recinto, exclude_*
            logic: "AND" o "OR" para combinar criterios
        """
        # Exclusiones primero (siempre aplican)
        if targets.get("exclude_polygon") and self.polygon in targets["exclude_polygon"]:
            return False
        if targets.get("exclude_parcel") and self.parcel in targets["exclude_parcel"]:
            return False
        if targets.get("exclude_recinto") and self.recinto in targets["exclude_recinto"]:
            return False
        
        # Filtro de recinto (si se especifica, debe coincidir)
        recinto_filter = targets.get("recinto")
        if recinto_filter and self.recinto not in recinto_filter:
            return False
        
        # Evaluar scope principal
        scope_match = self._match_scope(scope, targets)
        
        # Si logic=AND, también verificar criterios adicionales
        if logic == "AND":
            # Si hay municipio especificado además del scope principal
            if targets.get("municipio") and scope != "BY_MUNICIPIO":
                target_municipio = targets["municipio"].upper()
                if target_municipio not in self.municipio and self.municipio not in target_municipio:
                    return False
            
            # Si hay crop especificado además del scope principal (para BY_MUNICIPIO)
            if targets.get("crop") and scope == "BY_MUNICIPIO":
                target_crop = targets["crop"].upper()
                if not (target_crop in self.crop or target_crop in self.crop_norm or
                        self.crop in target_crop or self.crop_norm in target_crop):
                    return False
        
        return scope_match
    
    def _match_scope(self, scope: str, targets: Dict[str, Any]) -> bool:
        """Evalúa si coincide con el scope principal"""
        if scope == "ALL":
            return True
        
        if scope == "BY_CROP":
            target_crop = (targets.get("crop") or "").upper()
            return (target_crop in self.crop or 
                    target_crop in self.crop_norm or
                    self.crop in target_crop or
                    self.crop_norm in target_crop)
        
        if scope == "BY_MUNICIPIO":
            target_municipio = (targets.get("municipio") or "").upper()
            return target_municipio in self.municipio or self.municipio in target_municipio
        
        if scope == "BY_POLYGON":
            polygon_list = targets.get("polygon") or []
            return self.polygon in polygon_list
        
        if scope == "BY_IDS":
            polygon_list = targets.get("polygon") or []
            parcel_list = targets.get("parcel") or []
            in_polygon = self.polygon in polygon_list if polygon_list else True
            in_parcel = self.parcel in parcel_list if parcel_list else True
            return in_polygon or in_parcel
        
        if scope == "BY_POLYGON_AND_PARCEL":
            polygon_list = targets.get("polygon") or []
            parcel_list = targets.get("parcel") or []
            in_polygon = self.polygon in polygon_list if polygon_list else True
            in_parcel = self.parcel in parcel_list if parcel_list else True
            return in_polygon and in_parcel
        
        return False
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "polygon": self.polygon,
            "parcel": self.parcel,
            "recinto": self.recinto,
            "crop": self.crop,
            "crop_norm": self.crop_norm,
            "surface_ha": self.surface_ha,
            "municipio": self.municipio,
            "nro_orden": self.nro_orden
        }


class ParcelManager:
    """
    Gestor de parcelas y tratamientos profesional v2
    Con separación clara de cultivo y municipio
    """
    
    PARCEL_SHEET_PATTERNS = [
        r"datos.*parcela",
        r"id.*parc",
        r"parcela",
        r"recinto",
        r"sigpac",
    ]
    
    TREATMENT_SHEET_PATTERNS = [
        r"inf.*trat",
        r"tratamiento",
        r"aplicacion",
    ]
    
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {filepath}")
        self._parcels: List[Parcel] = []
        self._crops: List[str] = []
        self._municipios: List[str] = []
        
    def get_parcels(self) -> List[Parcel]:
        """Extrae la lista de parcelas del archivo"""
        if self._parcels:
            return self._parcels
            
        wb = load_workbook(self.filepath, data_only=True)
        
        sheet = self._find_sheet_by_pattern(wb, self.PARCEL_SHEET_PATTERNS)
        if not sheet:
            for name in ["2.1. DATOS PARCELAS", "id.parc 2"]:
                if name in wb.sheetnames:
                    sheet = wb[name]
                    break
        if not sheet:
            sheet = wb.active
            
        headers_row, col_map = self._find_headers_advanced(sheet)
        
        if not headers_row:
            wb.close()
            return []
        
        for row in range(headers_row + 1, sheet.max_row + 1):
            try:
                polygon = self._safe_int(self._get_cell_value(sheet, row, col_map.get('polygon', 7)))
                parcel_num = self._safe_int(self._get_cell_value(sheet, row, col_map.get('parcel', 8)))
                
                if polygon is None and parcel_num is None:
                    continue
                
                # Nº de Orden está en columna 2 (fija en este template)
                nro_orden = self._safe_int(self._get_cell_value(sheet, row, 2)) or 0
                    
                recinto = self._safe_int(self._get_cell_value(sheet, row, col_map.get('recinto', 9))) or 1
                municipio = str(self._get_cell_value(sheet, row, col_map.get('municipio', 4)) or "")
                crop = str(self._get_cell_value(sheet, row, col_map.get('crop', 13)) or "")
                surface = self._safe_float(self._get_cell_value(sheet, row, col_map.get('surface', 12)))
                
                p = Parcel(
                    polygon=polygon or 0,
                    parcel=parcel_num or 0,
                    recinto=recinto,
                    crop=crop,
                    surface_ha=surface,
                    municipio=municipio,
                    nro_orden=nro_orden
                )
                self._parcels.append(p)
                
            except Exception:
                continue
        
        wb.close()
        return self._parcels
    
    def get_unique_crops(self) -> List[str]:
        """Devuelve lista de CULTIVOS únicos (NO municipios)"""
        if self._crops:
            return self._crops
        parcels = self.get_parcels()
        crops_set = set()
        for p in parcels:
            if p.crop:
                crops_set.add(p.crop)
        self._crops = sorted(list(crops_set))
        return self._crops
    
    def get_unique_municipios(self) -> List[str]:
        """Devuelve lista de MUNICIPIOS únicos (NO cultivos)"""
        if self._municipios:
            return self._municipios
        parcels = self.get_parcels()
        muni_set = set()
        for p in parcels:
            if p.municipio:
                muni_set.add(p.municipio)
        self._municipios = sorted(list(muni_set))
        return self._municipios
    
    def get_context_for_ai(self) -> Dict[str, List[str]]:
        """Devuelve contexto estructurado para la IA"""
        return {
            "crops": self.get_unique_crops(),
            "municipios": self.get_unique_municipios()
        }
    
    def filter_parcels(self, scope: str, targets: Dict[str, Any], logic: str = "AND") -> List[Parcel]:
        """Filtra parcelas según el scope, targets y logic del intent"""
        parcels = self.get_parcels()
        return [p for p in parcels if p.matches_filter(scope, targets, logic)]
    
    def generate_treatment_rows(self, parcels: List[Parcel], 
                                 treatment_data: Dict[str, Any],
                                 resolved_date: str,
                                 start_order: int = 1) -> List[List[Any]]:
        """Genera las filas de tratamiento para insertar"""
        rows = []
        dose = treatment_data.get("dose", {})
        
        for i, p in enumerate(parcels):
            row = [
                start_order + i,                    # A: Nº Orden
                resolved_date,                      # B: Fecha Inicio
                resolved_date,                      # C: Fecha Fin
                treatment_data.get("product", ""),  # D: Producto
                treatment_data.get("registry_number") or "",  # E: Nº Registro
                dose.get("value", 0),               # F: Dosis valor
                dose.get("unit", ""),               # G: Unidad
                treatment_data.get("pest") or "",   # H: Plaga
                p.surface_ha,                       # I: Superficie
                p.crop or p.municipio,              # J: Cultivo
                p.polygon,                          # K: Polígono
                p.parcel,                           # L: Parcela
                p.recinto                           # M: Recinto
            ]
            rows.append(row)
        
        return rows
    
    def write_treatments(self, rows: List[List[Any]], 
                         output_path: Optional[str] = None) -> Tuple[str, int]:
        """Escribe las filas de tratamiento en una nueva hoja"""
        wb = load_workbook(self.filepath)
        
        new_sheet_name = "Tratamientos_IA"
        if new_sheet_name in wb.sheetnames:
            del wb[new_sheet_name]
        
        ws = wb.create_sheet(new_sheet_name)
        
        headers = ["Nº", "Fecha Inicio", "Fecha Fin", "Producto", "Nº Registro", 
                   "Dosis", "Unidad", "Plaga/Enfermedad", "Superficie (ha)", 
                   "Cultivo", "Polígono", "Parcela", "Recinto"]
        ws.append(headers)
        
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for row_data in rows:
            ws.append(row_data)
        
        column_widths = [6, 12, 12, 20, 12, 8, 10, 18, 12, 15, 10, 10, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        save_path = output_path or str(self.filepath).replace(".xlsx", "_TRATAMIENTOS.xlsx")
        wb.save(save_path)
        wb.close()
        
        return save_path, len(rows)
    
    def write_treatments_official(self, parcels: List[Parcel], 
                                   treatment_data: Dict[str, Any],
                                   resolved_date: str,
                                   output_path: Optional[str] = None) -> Tuple[str, int]:
        """
        Escribe tratamientos en la hoja oficial 'inf.trat 1' del cuaderno.
        Formato idéntico al archivo RESUELTO de referencia.
        
        Columnas en inf.trat 1:
        A: Id. Parcelas (Nº Orden de 2.1. DATOS PARCELAS)
        B: Especie (cultivo)
        C: Variedad (vacío normalmente)
        D: Superf. tratada (ha)
        E: Fecha (datetime o string)
        F: Problema fitosanitario
        G: Aplicador (default 1)
        H: Equipo (default 1)
        I: Nombre comercial (producto)
        J: Nº Registro
        K: Dosis (formato "X,XX l" o "X kg")
        L: Eficacia (default "BUENA")
        M: Observaciones/Notas
        """
        from datetime import datetime as dt
        from decimal import Decimal, ROUND_HALF_UP
        
        wb = load_workbook(self.filepath)
        
        # Buscar hoja inf.trat 1
        sheet_name = None
        for name in wb.sheetnames:
            if "inf.trat 1" in name.lower() or "inf.trat" in name.lower():
                sheet_name = name
                break
        
        if not sheet_name:
            # Fallback: crear hoja nueva
            wb.close()
            return self.write_treatments(
                self.generate_treatment_rows(parcels, treatment_data, resolved_date),
                output_path
            )
        
        ws = wb[sheet_name]
        
        # Row 8-9 son headers, Row 10 suele estar vacío, Row 11 empieza datos
        first_data_row = 11
        
        # Buscar el bloque de notas al pie (empieza con "[1]", "[2]", etc.)
        # Este bloque típicamente está mergeado y contiene texto explicativo
        footer_start_row = ws.max_row + 1
        for row in range(first_data_row, ws.max_row + 1):
            val_a = str(ws.cell(row, 1).value or "")
            # Las notas al pie empiezan con [1], [2], etc. o texto largo
            if val_a.startswith("[1]") or val_a.startswith("[2]") or len(val_a) > 50:
                footer_start_row = row
                break
        
        # Buscar la última fila con datos REALES (Id numérico) ANTES del footer
        last_data_row = first_data_row - 1
        for row in range(first_data_row, footer_start_row):
            val_a = ws.cell(row, 1).value
            # Solo contar filas con Id numérico (no texto de headers/footers)
            if val_a is not None:
                try:
                    int(val_a)  # Verificar que es un número
                    last_data_row = row
                except (ValueError, TypeError):
                    pass
        
        # Escribir a partir de la siguiente fila después de los datos
        # o desde first_data_row si no hay datos
        write_row = max(first_data_row, last_data_row + 1)
        
        # Detectar merged cells para evitar escribir en ellas
        merged_rows = set()
        for merged_range in ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                merged_rows.add(row)
        
        # Preparar datos del tratamiento
        dose = treatment_data.get("dose", {})
        dose_value = dose.get("value", 0)
        dose_unit = dose.get("unit", "l/ha")
        
        # Formatear dosis como el cuaderno (ej: "3 l", "2,35 l", "3 kg")
        if dose_value:
            dose_decimal = Decimal(str(dose_value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            dose_str = str(dose_decimal).replace(".", ",")
            # Simplificar unidad
            unit_short = "l" if "l" in dose_unit.lower() else "kg"
            dose_formatted = f"{dose_str} {unit_short}"
        else:
            dose_formatted = ""
        
        # Convertir fecha a datetime si es string
        if isinstance(resolved_date, str):
            try:
                date_obj = dt.strptime(resolved_date, "%d/%m/%Y")
            except:
                date_obj = resolved_date
        else:
            date_obj = resolved_date
        
        rows_written = 0
        
        for p in parcels:
            # Saltar filas con merged cells
            while write_row in merged_rows:
                write_row += 1
            
            # Escribir fila
            try:
                # A: Id. Parcelas (Nº Orden)
                cell_a = ws.cell(write_row, 1)
                if not isinstance(cell_a, MergedCell):
                    cell_a.value = p.nro_orden if p.nro_orden else ""
                
                # B: Especie (cultivo)
                cell_b = ws.cell(write_row, 2)
                if not isinstance(cell_b, MergedCell):
                    cell_b.value = p.crop
                
                # D: Superf. tratada
                cell_d = ws.cell(write_row, 4)
                if not isinstance(cell_d, MergedCell):
                    cell_d.value = p.surface_ha
                
                # E: Fecha
                cell_e = ws.cell(write_row, 5)
                if not isinstance(cell_e, MergedCell):
                    cell_e.value = date_obj
                
                # F: Problema fitosanitario
                cell_f = ws.cell(write_row, 6)
                if not isinstance(cell_f, MergedCell):
                    cell_f.value = treatment_data.get("pest") or "MALAS HIERBAS"
                
                # G: Aplicador
                cell_g = ws.cell(write_row, 7)
                if not isinstance(cell_g, MergedCell):
                    cell_g.value = 1
                
                # H: Equipo
                cell_h = ws.cell(write_row, 8)
                if not isinstance(cell_h, MergedCell):
                    cell_h.value = 1
                
                # I: Nombre comercial
                cell_i = ws.cell(write_row, 9)
                if not isinstance(cell_i, MergedCell):
                    cell_i.value = treatment_data.get("product", "")
                
                # J: Nº Registro
                cell_j = ws.cell(write_row, 10)
                if not isinstance(cell_j, MergedCell):
                    cell_j.value = treatment_data.get("registry_number") or ""
                
                # K: Dosis
                cell_k = ws.cell(write_row, 11)
                if not isinstance(cell_k, MergedCell):
                    cell_k.value = dose_formatted
                
                # L: Eficacia
                cell_l = ws.cell(write_row, 12)
                if not isinstance(cell_l, MergedCell):
                    cell_l.value = "BUENA"
                
                # M: Observaciones
                cell_m = ws.cell(write_row, 13)
                if not isinstance(cell_m, MergedCell):
                    cell_m.value = treatment_data.get("notes") or ""
                
                rows_written += 1
                write_row += 1
                
            except Exception as e:
                # Skip row on error, continue with next
                write_row += 1
                continue
        
        save_path = output_path or str(self.filepath).replace(".xlsx", "_RESUELTO.xlsx")
        wb.save(save_path)
        wb.close()
        
        return save_path, rows_written
    
    
    # ========================================
    # Métodos auxiliares privados
    # ========================================
    
    def _get_cell_value(self, sheet, row, col):
        """Obtiene el valor de una celda, manejando celdas combinadas"""
        cell = sheet.cell(row, col)
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return sheet.cell(merged_range.min_row, merged_range.min_col).value
            return None
        return cell.value
    
    def _find_sheet_by_pattern(self, wb, patterns: List[str]) -> Optional[Worksheet]:
        """Busca una hoja cuyo nombre coincida con algún patrón"""
        for sheet_name in wb.sheetnames:
            for pattern in patterns:
                if re.search(pattern, sheet_name.lower()):
                    return wb[sheet_name]
        return None
    
    def _find_headers_advanced(self, sheet: Worksheet) -> Tuple[int, Dict[str, int]]:
        """Busca la fila de encabezados buscando patrones específicos"""
        col_map = {}
        
        for row in range(1, min(20, sheet.max_row + 1)):
            for col in range(1, min(25, sheet.max_column + 1)):
                val = str(self._get_cell_value(sheet, row, col) or "").upper()
                val = val.replace("\n", " ").strip()
                
                if "POLÍGONO" in val or "POLIGONO" in val:
                    col_map["polygon"] = col
                elif "PARCELA" in val and "polygon" in col_map and "parcel" not in col_map:
                    col_map["parcel"] = col
                elif "RECINTO" in val and "recinto" not in col_map:
                    col_map["recinto"] = col
                elif "MUNICIPIO" in val or "TÉRMINO" in val or "TERMINO" in val:
                    col_map["municipio"] = col
                elif "ESPECIE" in val or "VARIEDAD" in val:
                    col_map["crop"] = col
                elif "CULTIVADA" in val or ("SUPERFICIE" in val and "SIGPAC" not in val):
                    col_map["surface"] = col
            
            if "polygon" in col_map and "parcel" in col_map:
                return row, col_map
        
        return 0, {}
    
    def _safe_int(self, value) -> Optional[int]:
        if value is None:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None
    
    def _safe_float(self, value) -> float:
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
