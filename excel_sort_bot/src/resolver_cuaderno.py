"""
RESOLVER CUADERNO v2.3 — Pipeline Enterprise (Production Ready)
Convierte un cuaderno vacío/sin resolver en uno RESUELTO
usando el Golden Template como base inmutable.

Reglas de Oro:
- Output SIEMPRE es copia del Golden Template
- NUNCA se crean hojas nuevas
- IA decide QUÉ escribir, código ejecuta DÓNDE
- DEDUPE GATE: No se permiten duplicados por business key
- DATA BLOCK: Inserción solo en el bloque de datos oficial
"""
import json
import shutil
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple, Set
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, PatternFill


class WorkbookDictionary:
    """Carga y gestiona el diccionario del template"""
    
    def __init__(self, dict_path: str = None):
        if dict_path is None:
            dict_path = Path(__file__).parent / "workbook_dictionary.json"
        
        with open(dict_path, "r", encoding="utf-8") as f:
            self.data = json.load(f)
        
        self.template_id = self.data.get("template_id")
        self.sheets = self.data.get("sheets", {})
        self.actions_schema = self.data.get("actions_schema", {})
        self.validation_rules = self.data.get("validation_rules", {})
    
    def get_sheet_config(self, sheet_name: str) -> Optional[Dict]:
        """Obtiene configuración de una hoja"""
        return self.sheets.get(sheet_name)
    
    def get_action_config(self, action: str) -> Optional[Dict]:
        """Obtiene configuración de una acción"""
        return self.actions_schema.get(action)


class GoldenTemplateWriter:
    """
    Escribe datos sobre una copia del Golden Template.
    NUNCA modifica el original. NUNCA crea hojas nuevas.
    
    Features v2.3:
    - DEDUPE GATE: Bloquea duplicados por business key
    - DATA BLOCK INSERTION: Escribe solo en el bloque de datos oficial
    """
    
    def __init__(self, golden_path: str, output_path: str, dict_path: str = None):
        self.golden_path = Path(golden_path)
        self.output_path = Path(output_path)
        self.dictionary = WorkbookDictionary(dict_path)
        
        if not self.golden_path.exists():
            raise FileNotFoundError(f"Golden template no encontrado: {golden_path}")
        
        # Copiar golden template al output
        shutil.copy(self.golden_path, self.output_path)
        
        self.wb = None
        self.parcels_index = {}  # (polygon, parcel, recinto) -> nro_orden
        self.write_stats = {}
        self.warnings = []
        self.errors = []
        
        # DEDUPE: Set de business keys existentes
        self._existing_keys: Set[str] = set()
        self._duplicates_blocked = 0
        
        # DATA BLOCK: Cache de filas de escritura
        self._next_write_row: Dict[str, int] = {}
    
    def open(self):
        """Abre el workbook para escritura"""
        self.wb = load_workbook(self.output_path)
        self._build_parcels_index()
        self._build_existing_keys()
    
    def close(self):
        """Guarda y cierra el workbook"""
        if self.wb:
            self.wb.save(self.output_path)
            self.wb.close()
            self.wb = None
    
    def _build_parcels_index(self):
        """Construye índice de parcelas para mapear pol/parc/rec -> nro_orden"""
        config = self.dictionary.get_sheet_config("2.1. DATOS PARCELAS")
        if not config:
            return
        
        sheet_name = "2.1. DATOS PARCELAS"
        if sheet_name not in self.wb.sheetnames:
            return
        
        ws = self.wb[sheet_name]
        cols = config["columns"]
        start_row = config["data_start_row"]
        
        for row in range(start_row, ws.max_row + 1):
            nro_orden = self._safe_int(ws.cell(row, cols["nro_orden"]).value)
            polygon = self._safe_int(ws.cell(row, cols["polygon"]).value)
            parcel = self._safe_int(ws.cell(row, cols["parcel"]).value)
            recinto = self._safe_int(ws.cell(row, cols["recinto"]).value) or 1
            crop = str(ws.cell(row, cols["crop"]).value or "")
            surface = self._safe_float(ws.cell(row, cols["superficie_cultivada"]).value)
            
            if polygon and parcel:
                key = (polygon, parcel, recinto)
                self.parcels_index[key] = {
                    "nro_orden": nro_orden,
                    "crop": crop,
                    "surface": surface
                }
    
    def _build_existing_keys(self):
        """Construye set de business keys existentes en inf.trat 1"""
        sheet_name = "inf.trat 1"
        if sheet_name not in self.wb.sheetnames:
            return
        
        config = self.dictionary.get_sheet_config(sheet_name)
        if not config:
            return
        
        ws = self.wb[sheet_name]
        cols = config["columns"]
        start_row = config.get("data_start_row", 11)
        
        for row in range(start_row, ws.max_row + 1):
            id_parcelas = ws.cell(row, cols["id_parcelas"]).value
            fecha = ws.cell(row, cols["fecha"]).value
            producto = ws.cell(row, cols["producto"]).value
            dosis = ws.cell(row, cols["dosis"]).value
            
            if id_parcelas and producto:
                key = self._make_business_key(id_parcelas, fecha, producto, dosis)
                self._existing_keys.add(key)
    
    def _make_business_key(self, id_parcelas, fecha, producto, dosis) -> str:
        """Genera business key para tratamiento"""
        # Normalizar fecha
        if isinstance(fecha, datetime):
            fecha_str = fecha.strftime("%Y-%m-%d")
        else:
            fecha_str = str(fecha or "")
        
        # Normalizar valores
        id_str = str(id_parcelas or "").strip()
        prod_str = str(producto or "").strip().upper()
        dosis_str = str(dosis or "").strip().replace(",", ".")
        
        return f"{id_str}|{fecha_str}|{prod_str}|{dosis_str}"
    
    def _is_duplicate(self, treatment: Dict[str, Any]) -> bool:
        """Verifica si un tratamiento ya existe"""
        id_parcelas = treatment.get("id_parcelas")
        if not id_parcelas and "polygon" in treatment:
            polygon = treatment.get("polygon")
            parcel = treatment.get("parcel")
            recinto = treatment.get("recinto", 1)
            id_parcelas = self.get_parcel_id(polygon, parcel, recinto)
        
        fecha = treatment.get("fecha") or treatment.get("date")
        producto = treatment.get("product") or treatment.get("producto")
        
        dose = treatment.get("dose", {})
        if isinstance(dose, dict):
            dosis = self._format_dose(dose.get("value", 0), dose.get("unit", "l/ha"))
        else:
            dosis = str(dose)
        
        key = self._make_business_key(id_parcelas, fecha, producto, dosis)
        return key in self._existing_keys
    
    def get_parcel_id(self, polygon: int, parcel: int, recinto: int = 1) -> Optional[int]:
        """Obtiene el Id. Parcelas (nro_orden) para una combinación pol/parc/rec"""
        key = (polygon, parcel, recinto)
        info = self.parcels_index.get(key)
        return info["nro_orden"] if info else None
    
    def write_treatment(self, treatment: Dict[str, Any], allow_duplicates: bool = False) -> bool:
        """
        Escribe un tratamiento en inf.trat 1
        
        Args:
            treatment: Datos del tratamiento
            allow_duplicates: Si True, permite duplicados (para migración)
        
        Returns:
            True si se escribió, False si fue bloqueado o falló
        """
        # DEDUPE GATE
        if not allow_duplicates and self._is_duplicate(treatment):
            self._duplicates_blocked += 1
            self.warnings.append(f"Duplicado bloqueado: {treatment.get('product')}")
            return False
        
        config = self.dictionary.get_sheet_config("inf.trat 1")
        if not config:
            self.errors.append("No hay configuración para inf.trat 1")
            return False
        
        sheet_name = "inf.trat 1"
        if sheet_name not in self.wb.sheetnames:
            self.errors.append(f"Hoja {sheet_name} no existe en el template")
            return False
        
        ws = self.wb[sheet_name]
        cols = config["columns"]
        defaults = config.get("defaults", {})
        
        # Resolver id_parcelas
        if "id_parcelas" in treatment:
            id_parcelas = treatment["id_parcelas"]
        else:
            polygon = treatment.get("polygon")
            parcel = treatment.get("parcel")
            recinto = treatment.get("recinto", 1)
            id_parcelas = self.get_parcel_id(polygon, parcel, recinto)
            
            if id_parcelas is None:
                # Si no existe el mapeo, crear id temporal
                id_parcelas = treatment.get("id_parcelas") or polygon
                self.warnings.append(f"Parcela no indexada: pol={polygon}, usando id={id_parcelas}")
        
        # Obtener datos de la parcela
        parcel_info = {}
        if "polygon" in treatment and "parcel" in treatment:
            key = (treatment["polygon"], treatment["parcel"], treatment.get("recinto", 1))
            parcel_info = self.parcels_index.get(key, {})
        
        # DATA BLOCK: Encontrar fila correcta
        write_row = self._find_data_block_row(ws, config, sheet_name)
        if write_row is None:
            self.errors.append("No se pudo encontrar fila en data block")
            return False
        
        # Formatear dosis
        dose = treatment.get("dose", {})
        if isinstance(dose, dict):
            dose_value = dose.get("value", 0)
            dose_unit = dose.get("unit", "l/ha")
        else:
            dose_value = dose
            dose_unit = "l/ha"
        
        dose_formatted = self._format_dose(dose_value, dose_unit)
        
        # Formatear fecha
        fecha = treatment.get("fecha") or treatment.get("date")
        if isinstance(fecha, str):
            try:
                fecha = datetime.strptime(fecha, "%d/%m/%Y")
            except:
                pass
        
        # Escribir celdas
        try:
            self._write_cell(ws, write_row, cols["id_parcelas"], id_parcelas)
            self._write_cell(ws, write_row, cols["especie"], 
                           treatment.get("crop") or parcel_info.get("crop", ""))
            self._write_cell(ws, write_row, cols["superficie_tratada"], 
                           treatment.get("surface") or parcel_info.get("surface", 0))
            self._write_cell(ws, write_row, cols["fecha"], fecha)
            self._write_cell(ws, write_row, cols["problema_fito"], 
                           treatment.get("pest") or treatment.get("problema_fito") or "MALAS HIERBAS")
            
            # Nº orden tratamiento y aplicación (cols 7 y 8 en el template real)
            nro_orden_trat = defaults.get("nro_orden_tratamiento", 1)
            nro_aplicacion = defaults.get("nro_aplicacion", 1)
            self._write_cell(ws, write_row, cols.get("nro_orden_tratamiento", 7), nro_orden_trat)
            self._write_cell(ws, write_row, cols.get("nro_aplicacion", 8), nro_aplicacion)
            
            self._write_cell(ws, write_row, cols["producto"], 
                           treatment.get("product") or treatment.get("producto", ""))
            self._write_cell(ws, write_row, cols["nro_registro"], 
                           treatment.get("registry_number") or treatment.get("nro_registro") or "")
            self._write_cell(ws, write_row, cols["dosis"], dose_formatted)
            self._write_cell(ws, write_row, cols["eficacia"], defaults.get("eficacia", "BUENA"))
            self._write_cell(ws, write_row, cols["observaciones"], 
                           treatment.get("notes") or treatment.get("observaciones") or "")
            
            # Registrar business key para evitar duplicados posteriores
            key = self._make_business_key(id_parcelas, fecha, 
                                         treatment.get("product") or treatment.get("producto"),
                                         dose_formatted)
            self._existing_keys.add(key)
            
            # Actualizar siguiente fila
            self._next_write_row[sheet_name] = write_row + 1
            
            # Actualizar stats
            self.write_stats[sheet_name] = self.write_stats.get(sheet_name, 0) + 1
            return True
            
        except Exception as e:
            self.errors.append(f"Error escribiendo tratamiento: {e}")
            return False
    
    def write_treatments_batch(self, treatments: List[Dict[str, Any]], 
                               allow_duplicates: bool = False) -> int:
        """Escribe múltiples tratamientos"""
        written = 0
        for t in treatments:
            if self.write_treatment(t, allow_duplicates=allow_duplicates):
                written += 1
        return written
    
    def _find_data_block_row(self, ws, config: Dict, sheet_name: str) -> Optional[int]:
        """
        Encuentra la primera fila vacía en el DATA BLOCK oficial.
        
        Reglas:
        1. Detecta header row (donde A contiene "Id. Parcelas")
        2. data_start = header_row + offset
        3. Busca primera fila donde col A vacío Y col I (producto) vacío
        4. No pasa del footer ([1], [2], etc.)
        """
        # Si ya tenemos cache, usarla
        if sheet_name in self._next_write_row:
            return self._next_write_row[sheet_name]
        
        start_row = config.get("data_start_row", 11)
        footer_marker = config.get("footer_marker", "[1]")
        
        # Encontrar donde empieza el footer (notas al pie)
        footer_row = ws.max_row + 1
        for row in range(start_row, min(ws.max_row + 1, start_row + 500)):
            val_a = str(ws.cell(row, 1).value or "")
            # Footer empieza con [1], [2] o es texto largo
            if val_a.startswith("[1]") or val_a.startswith("[2]") or len(val_a) > 100:
                footer_row = row
                break
        
        # Buscar primera fila vacía en el data block
        # Una fila vacía = col A vacía Y col I (producto) vacía
        for row in range(start_row, footer_row):
            val_a = ws.cell(row, 1).value
            val_i = ws.cell(row, 9).value  # col I = producto
            
            # Fila vacía encontrada
            if val_a is None and val_i is None:
                self._next_write_row[sheet_name] = row
                return row
            
            # Verificar si es merged cell que debemos saltar
            cell_a = ws.cell(row, 1)
            if isinstance(cell_a, MergedCell):
                continue
        
        # Si no hay filas vacías antes del footer, escribir justo antes
        write_row = footer_row
        self._next_write_row[sheet_name] = write_row
        return write_row
    
    def _write_cell(self, ws, row: int, col: int, value):
        """Escribe en una celda, evitando merged cells"""
        if col is None:
            return
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            cell.value = value
    
    def _format_dose(self, value, unit: str) -> str:
        """Formatea dosis como el cuaderno: '3,00 l' o '2,35 kg'"""
        if not value:
            return ""
        try:
            decimal_val = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            value_str = str(decimal_val).replace(".", ",")
            unit_short = "l" if "l" in unit.lower() else "kg"
            return f"{value_str} {unit_short}"
        except:
            return str(value)
    
    def _safe_int(self, value) -> Optional[int]:
        if value is None:
            return None
        try:
            return int(float(value))
        except:
            return None
    
    def _safe_float(self, value) -> float:
        if value is None:
            return 0.0
        try:
            return float(value)
        except:
            return 0.0
    
    def get_execution_report(self) -> Dict[str, Any]:
        """Devuelve reporte de ejecución"""
        return {
            "template_id": self.dictionary.template_id,
            "output_path": str(self.output_path),
            "write_stats": self.write_stats,
            "duplicates_blocked": self._duplicates_blocked,
            "warnings": self.warnings,
            "errors": self.errors,
            "success": len(self.errors) == 0
        }


class CuadernoResolver:
    """
    Pipeline completo para resolver un cuaderno.
    
    Flujo:
    1. Fingerprint del input (ACCEPT/WARN/REJECT)
    2. Extract contexto del input
    3. IA genera acciones (ADD_TREATMENT, etc.)
    4. Writer aplica acciones sobre Golden Template
    5. Validación y auditoría
    """
    
    def __init__(self, 
                 golden_template_path: str,
                 workbook_dict_path: str = None):
        self.golden_path = Path(golden_template_path)
        self.dict_path = workbook_dict_path
        
        if not self.golden_path.exists():
            raise FileNotFoundError(f"Golden template no encontrado: {golden_template_path}")
    
    def resolve(self, 
                input_path: str,
                output_path: str,
                actions: List[Dict[str, Any]] = None,
                context: Dict[str, Any] = None,
                allow_duplicates: bool = False) -> Dict[str, Any]:
        """
        Resuelve un cuaderno aplicando acciones sobre el golden template.
        
        Args:
            input_path: Cuaderno del cliente (para extraer contexto)
            output_path: Donde guardar el resultado
            actions: Lista de acciones a aplicar (generadas por IA)
            context: Contexto adicional (textos del usuario, etc.)
            allow_duplicates: Si True, permite duplicados (para migración)
        
        Returns:
            Reporte de ejecución
        """
        report = {
            "input_path": input_path,
            "output_path": output_path,
            "fingerprint_status": None,
            "context_extracted": {},
            "actions_applied": 0,
            "duplicates_blocked": 0,
            "write_stats": {},
            "warnings": [],
            "errors": [],
            "success": False
        }
        
        # 1. Fingerprint del input (opcional)
        try:
            from src.template_fingerprint import fingerprint_workbook, load_registry, match_template
            fp = fingerprint_workbook(input_path, parcels_sheet_hint_contains=["2.1.", "PARCELAS"])
            registry = load_registry()
            decision = match_template(fp, registry)
            report["fingerprint_status"] = decision["status"]
            
            if decision["status"] == "REJECT":
                report["errors"].append("Template no reconocido (REJECT)")
                return report
        except Exception as e:
            report["warnings"].append(f"Fingerprint falló: {e}")
        
        # 2. Extraer contexto del input (parcelas, año, etc.)
        try:
            from src.parcel_manager import ParcelManager
            pm = ParcelManager(input_path)
            parcels = pm.get_parcels()
            report["context_extracted"]["parcels_count"] = len(parcels)
            report["context_extracted"]["crops"] = pm.get_unique_crops()
            report["context_extracted"]["municipios"] = pm.get_unique_municipios()
        except Exception as e:
            report["warnings"].append(f"Extracción de contexto falló: {e}")
        
        # 3. Crear writer sobre golden template
        writer = GoldenTemplateWriter(
            str(self.golden_path),
            output_path,
            self.dict_path
        )
        writer.open()
        
        try:
            # 4. Aplicar acciones
            if actions:
                for action in actions:
                    action_type = action.get("action") or action.get("type")
                    
                    if action_type == "ADD_TREATMENT":
                        treatments = action.get("treatments", [action])
                        for t in treatments:
                            if writer.write_treatment(t, allow_duplicates=allow_duplicates):
                                report["actions_applied"] += 1
                    
                    # Aquí se añadirían más tipos de acciones:
                    # elif action_type == "ADD_FERTILIZATION": ...
                    # elif action_type == "ADD_HARVEST": ...
            
            report["write_stats"] = writer.write_stats
            report["duplicates_blocked"] = writer._duplicates_blocked
            report["warnings"].extend(writer.warnings)
            report["errors"].extend(writer.errors)
            report["success"] = len(writer.errors) == 0
            
        finally:
            writer.close()
        
        return report


# Función de conveniencia para uso directo
def resolver_cuaderno(input_path: str, 
                      output_path: str,
                      golden_path: str = None,
                      treatments: List[Dict] = None,
                      allow_duplicates: bool = False) -> Dict[str, Any]:
    """
    Resuelve un cuaderno con tratamientos.
    
    Ejemplo:
        treatments = [
            {"polygon": 8, "parcel": 148, "recinto": 1, 
             "product": "Glifosato", "dose": {"value": 3, "unit": "l/ha"},
             "fecha": "05/02/2026", "pest": "MALAS HIERBAS"}
        ]
        result = resolver_cuaderno(input, output, treatments=treatments)
    """
    if golden_path is None:
        # Default: buscar en la misma carpeta que el input
        golden_path = str(Path(input_path).parent / "PABLO PEREZ RUBIO 2025 RESUELTO.XLSX")
    
    resolver = CuadernoResolver(golden_path)
    
    actions = []
    if treatments:
        actions.append({
            "action": "ADD_TREATMENT",
            "treatments": treatments
        })
    
    return resolver.resolve(input_path, output_path, actions=actions, 
                           allow_duplicates=allow_duplicates)
