"""
RTE FUZZY RESOLVER v1.0
Resolver inteligente con fuzzy matching y catálogo de productos.

Features:
- Fuzzy match con umbral configurable
- Catálogo de productos/cultivos/municipios
- Sugerencias rankeadas con score
- requires_confirmation cuando hay ambigüedad
- Match exacto preferido sobre fuzzy

Ejemplo:
    "glifosato" → 
        match: "ROUNDUP ENERGY PRO" (score: 0.85)
        alternatives: ["GLIFOSATO 36%", "ROUNDUP CLASSIC"]
"""
import re
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from openpyxl import load_workbook


# ============================================
# DATA CLASSES
# ============================================

@dataclass
class FuzzyMatch:
    """Resultado de un fuzzy match"""
    term: str           # Término buscado
    match: str          # Match encontrado
    score: float        # 0.0 - 1.0
    match_type: str     # "exact", "fuzzy", "contains", "alias"
    category: str       # "product", "crop", "municipality"
    row: Optional[int] = None  # Fila donde está
    alternatives: List[Tuple[str, float]] = field(default_factory=list)
    requires_confirmation: bool = False
    
    def to_dict(self) -> Dict:
        return {
            "term": self.term,
            "match": self.match,
            "score": round(self.score, 3),
            "match_type": self.match_type,
            "category": self.category,
            "row": self.row,
            "alternatives": [(a, round(s, 3)) for a, s in self.alternatives],
            "requires_confirmation": self.requires_confirmation
        }


# ============================================
# PRODUCT CATALOG (Fitosanitarios)
# ============================================

# Mapa de aliases comunes en agricultura española
PRODUCT_ALIASES = {
    # Herbicidas
    "glifosato": ["ROUNDUP", "ROUNDUP ENERGY PRO", "GLIFOSATO 36%", "GLYFOS", "TOUCHDOWN"],
    "roundup": ["ROUNDUP ENERGY PRO", "ROUNDUP CLASSIC", "GLIFOSATO"],
    # Fungicidas
    "azufre": ["AZUFRE MOJABLE", "AZUFRE 80%", "AZUFRE WG"],
    "cobre": ["COBRE NORDOX", "OXICLORURO DE COBRE", "CALDO BORDELES"],
    # Insecticidas
    "decis": ["DECIS PROTECH", "DECIS EVO", "DELTAMETRINA"],
    "confidor": ["CONFIDOR 20 LS", "IMIDACLOPRID"],
    # Acaricidas
    "envidor": ["ENVIDOR 240 SC", "SPIRODICLOFEN"],
    # Genéricos
    "mancozeb": ["MANCOZEB 80%", "DITHANE", "MANCOZAN"],
    "fosetil": ["FOSETIL-AL", "ALIETTE"],
}

# Cultivos comunes
CROP_ALIASES = {
    "vid": ["VID", "VIÑA", "UVA", "VITIS"],
    "olivo": ["OLIVO", "OLIVA", "OLIVAR", "ACEITUNA"],
    "cereal": ["CEBADA", "TRIGO", "AVENA", "CENTENO"],
    "cebada": ["CEBADA", "HORDEUM"],
    "trigo": ["TRIGO", "TRITICUM"],
    "almendro": ["ALMENDRO", "ALMENDRA"],
    "frutales": ["FRUTALES", "MELOCOTONERO", "PERAL", "MANZANO"],
    "horticolas": ["HORTICOLAS", "HORTALIZAS", "TOMATE", "PIMIENTO"],
}


# ============================================
# FUZZY MATCHER
# ============================================

class FuzzyMatcher:
    """Motor de fuzzy matching con múltiples estrategias"""
    
    # Umbrales
    EXACT_THRESHOLD = 1.0
    HIGH_CONFIDENCE_THRESHOLD = 0.85
    MEDIUM_CONFIDENCE_THRESHOLD = 0.70
    LOW_CONFIDENCE_THRESHOLD = 0.50
    
    @staticmethod
    def similarity(s1: str, s2: str) -> float:
        """Calcula similitud entre dos strings (0-1)"""
        s1 = s1.upper().strip()
        s2 = s2.upper().strip()
        return SequenceMatcher(None, s1, s2).ratio()
    
    @staticmethod
    def contains_score(query: str, target: str) -> float:
        """Score basado en si query está contenido en target"""
        query = query.upper().strip()
        target = target.upper().strip()
        
        if query == target:
            return 1.0
        if query in target:
            # Score proporcional a cuánto del target es el query
            return len(query) / len(target) * 0.9
        if target in query:
            return len(target) / len(query) * 0.7
        return 0.0
    
    @staticmethod
    def word_match_score(query: str, target: str) -> float:
        """Score basado en palabras coincidentes"""
        query_words = set(query.upper().split())
        target_words = set(target.upper().split())
        
        if not query_words or not target_words:
            return 0.0
        
        intersection = query_words & target_words
        if not intersection:
            return 0.0
        
        return len(intersection) / max(len(query_words), len(target_words))
    
    @classmethod
    def combined_score(cls, query: str, target: str) -> float:
        """Combina múltiples estrategias de matching"""
        sim = cls.similarity(query, target)
        contains = cls.contains_score(query, target)
        words = cls.word_match_score(query, target)
        
        # Peso mayor a similitud directa, luego contains, luego palabras
        return max(sim, contains * 0.95, words * 0.8)


# ============================================
# FUZZY RESOLVER
# ============================================

class FuzzyResolver:
    """
    Resolver inteligente con fuzzy matching.
    
    Uso:
        resolver = FuzzyResolver(workbook_path)
        match = resolver.find_product("glifosato")
        # match.match = "ROUNDUP ENERGY PRO"
        # match.score = 0.85
        # match.alternatives = [("GLIFOSATO 36%", 0.72), ...]
    """
    
    def __init__(self, workbook_path: str = None, workbook = None):
        self.wb_path = workbook_path
        self.wb = workbook
        self._owns_workbook = False
        
        if workbook is None and workbook_path:
            self.wb = load_workbook(workbook_path, data_only=True)
            self._owns_workbook = True
        
        # Catálogos extraídos del workbook
        self._products: Dict[str, int] = {}  # nombre -> row
        self._crops: Dict[str, List[int]] = {}  # cultivo -> [rows]
        self._municipalities: Dict[str, List[int]] = {}  # municipio -> [rows]
        
        # Construir catálogos
        if self.wb:
            self._build_catalogs()
    
    def close(self):
        if self._owns_workbook and self.wb:
            self.wb.close()
    
    def _build_catalogs(self):
        """Construye catálogos desde el workbook"""
        self._extract_products()
        self._extract_crops()
        self._extract_municipalities()
    
    def _extract_products(self):
        """Extrae productos únicos de inf.trat 1-4"""
        for sheet_name in ["inf.trat 1", "inf.trat 2", "inf.trat 3", "inf.trat 4"]:
            if sheet_name not in self.wb.sheetnames:
                continue
            
            ws = self.wb[sheet_name]
            for row in range(11, ws.max_row + 1):
                producto = ws.cell(row, 9).value  # Col I
                if producto and isinstance(producto, str):
                    prod_clean = producto.strip().upper()
                    if prod_clean and prod_clean not in self._products:
                        self._products[prod_clean] = row
    
    def _extract_crops(self):
        """Extrae cultivos únicos"""
        for sheet_name in ["inf.trat 1", "inf.trat 2"]:
            if sheet_name not in self.wb.sheetnames:
                continue
            
            ws = self.wb[sheet_name]
            for row in range(11, ws.max_row + 1):
                cultivo = ws.cell(row, 2).value  # Col B
                if cultivo and isinstance(cultivo, str):
                    cult_clean = cultivo.strip().upper()
                    if cult_clean:
                        if cult_clean not in self._crops:
                            self._crops[cult_clean] = []
                        self._crops[cult_clean].append(row)
    
    def _extract_municipalities(self):
        """Extrae municipios de 2.1. DATOS PARCELAS"""
        if "2.1. DATOS PARCELAS" not in self.wb.sheetnames:
            return
        
        ws = self.wb["2.1. DATOS PARCELAS"]
        for row in range(14, ws.max_row + 1):
            municipio = ws.cell(row, 4).value  # Col D
            if municipio and isinstance(municipio, str):
                mun_clean = municipio.strip().upper()
                if mun_clean:
                    if mun_clean not in self._municipalities:
                        self._municipalities[mun_clean] = []
                    self._municipalities[mun_clean].append(row)
    
    # ============================================
    # PUBLIC API
    # ============================================
    
    def find_product(self, query: str, top_n: int = 3) -> FuzzyMatch:
        """
        Busca un producto por nombre con fuzzy matching.
        
        Args:
            query: Nombre del producto a buscar
            top_n: Número de alternativas a devolver
        
        Returns:
            FuzzyMatch con el mejor match y alternativas
        """
        query_clean = query.strip().upper()
        
        # 1. Match exacto
        if query_clean in self._products:
            return FuzzyMatch(
                term=query,
                match=query_clean,
                score=1.0,
                match_type="exact",
                category="product",
                row=self._products[query_clean],
                requires_confirmation=False
            )
        
        # 2. Buscar en aliases
        for alias_key, alias_values in PRODUCT_ALIASES.items():
            if query_clean == alias_key.upper() or alias_key.upper() in query_clean:
                # Buscar cuál de los alias values existe en el catálogo
                for alias_value in alias_values:
                    if alias_value.upper() in self._products:
                        return FuzzyMatch(
                            term=query,
                            match=alias_value.upper(),
                            score=0.95,
                            match_type="alias",
                            category="product",
                            row=self._products[alias_value.upper()],
                            requires_confirmation=False
                        )
        
        # 3. Fuzzy matching contra catálogo
        scores = []
        for product, row in self._products.items():
            score = FuzzyMatcher.combined_score(query_clean, product)
            if score > FuzzyMatcher.LOW_CONFIDENCE_THRESHOLD:
                scores.append((product, score, row))
        
        # Ordenar por score descendente
        scores.sort(key=lambda x: x[1], reverse=True)
        
        if not scores:
            # No match
            return FuzzyMatch(
                term=query,
                match="",
                score=0.0,
                match_type="none",
                category="product",
                requires_confirmation=True,
                alternatives=[]
            )
        
        best_match, best_score, best_row = scores[0]
        alternatives = [(p, s) for p, s, _ in scores[1:top_n+1]]
        
        # Determinar si requiere confirmación
        requires_confirmation = best_score < FuzzyMatcher.HIGH_CONFIDENCE_THRESHOLD
        
        # Si hay varias opciones con scores similares, pedir confirmación
        if len(scores) > 1:
            second_score = scores[1][1]
            if best_score - second_score < 0.1:  # Diferencia menor a 10%
                requires_confirmation = True
        
        return FuzzyMatch(
            term=query,
            match=best_match,
            score=best_score,
            match_type="fuzzy",
            category="product",
            row=best_row,
            alternatives=alternatives,
            requires_confirmation=requires_confirmation
        )
    
    def find_crop(self, query: str, top_n: int = 3) -> FuzzyMatch:
        """Busca un cultivo por nombre con fuzzy matching"""
        query_clean = query.strip().upper()
        
        # 1. Match exacto
        if query_clean in self._crops:
            return FuzzyMatch(
                term=query,
                match=query_clean,
                score=1.0,
                match_type="exact",
                category="crop",
                requires_confirmation=False
            )
        
        # 2. Buscar en aliases
        for alias_key, alias_values in CROP_ALIASES.items():
            if query_clean == alias_key.upper() or alias_key.upper() in query_clean:
                for alias_value in alias_values:
                    if alias_value.upper() in self._crops:
                        return FuzzyMatch(
                            term=query,
                            match=alias_value.upper(),
                            score=0.95,
                            match_type="alias",
                            category="crop",
                            requires_confirmation=False
                        )
        
        # 3. Fuzzy matching
        scores = []
        for crop in self._crops.keys():
            score = FuzzyMatcher.combined_score(query_clean, crop)
            if score > FuzzyMatcher.LOW_CONFIDENCE_THRESHOLD:
                scores.append((crop, score))
        
        scores.sort(key=lambda x: x[1], reverse=True)
        
        if not scores:
            return FuzzyMatch(
                term=query,
                match="",
                score=0.0,
                match_type="none",
                category="crop",
                requires_confirmation=True
            )
        
        best_match, best_score = scores[0]
        alternatives = scores[1:top_n+1]
        
        return FuzzyMatch(
            term=query,
            match=best_match,
            score=best_score,
            match_type="fuzzy",
            category="crop",
            alternatives=alternatives,
            requires_confirmation=best_score < FuzzyMatcher.HIGH_CONFIDENCE_THRESHOLD
        )
    
    def find_municipality(self, query: str, top_n: int = 3) -> FuzzyMatch:
        """Busca un municipio por nombre con fuzzy matching"""
        query_clean = query.strip().upper()
        
        # 1. Match exacto
        if query_clean in self._municipalities:
            return FuzzyMatch(
                term=query,
                match=query_clean,
                score=1.0,
                match_type="exact",
                category="municipality",
                requires_confirmation=False
            )
        
        # 2. Fuzzy matching
        scores = []
        for mun in self._municipalities.keys():
            score = FuzzyMatcher.combined_score(query_clean, mun)
            if score > FuzzyMatcher.LOW_CONFIDENCE_THRESHOLD:
                scores.append((mun, score))
        
        scores.sort(key=lambda x: x[1], reverse=True)
        
        if not scores:
            return FuzzyMatch(
                term=query,
                match="",
                score=0.0,
                match_type="none",
                category="municipality",
                requires_confirmation=True
            )
        
        best_match, best_score = scores[0]
        alternatives = scores[1:top_n+1]
        
        return FuzzyMatch(
            term=query,
            match=best_match,
            score=best_score,
            match_type="fuzzy",
            category="municipality",
            alternatives=alternatives,
            requires_confirmation=best_score < FuzzyMatcher.HIGH_CONFIDENCE_THRESHOLD
        )
    
    def resolve_any(self, query: str) -> FuzzyMatch:
        """
        Intenta resolver en todas las categorías y devuelve el mejor match.
        Orden de prioridad: product > crop > municipality
        """
        results = [
            self.find_product(query),
            self.find_crop(query),
            self.find_municipality(query)
        ]
        
        # Filtrar los que tienen match
        valid = [r for r in results if r.score > 0]
        
        if not valid:
            return FuzzyMatch(
                term=query,
                match="",
                score=0.0,
                match_type="none",
                category="unknown",
                requires_confirmation=True
            )
        
        # Devolver el de mayor score
        return max(valid, key=lambda x: x.score)
    
    def get_suggestions(self, query: str, category: str = "product", 
                        limit: int = 5) -> List[Tuple[str, float]]:
        """
        Devuelve sugerencias ordenadas por score para autocompletado.
        
        Args:
            query: Texto parcial
            category: "product", "crop", "municipality"
            limit: Máximo de sugerencias
        
        Returns:
            Lista de (nombre, score)
        """
        catalog = {
            "product": self._products.keys(),
            "crop": self._crops.keys(),
            "municipality": self._municipalities.keys()
        }.get(category, [])
        
        query_clean = query.strip().upper()
        
        scores = []
        for item in catalog:
            # Dar prioridad a los que empiezan con el query
            if item.startswith(query_clean):
                scores.append((item, 0.95 + len(query_clean)/len(item) * 0.05))
            else:
                score = FuzzyMatcher.combined_score(query_clean, item)
                if score > 0.3:
                    scores.append((item, score))
        
        scores.sort(key=lambda x: x[1], reverse=True)
        return scores[:limit]
    
    def get_catalog_stats(self) -> Dict[str, int]:
        """Devuelve estadísticas del catálogo"""
        return {
            "products": len(self._products),
            "crops": len(self._crops),
            "municipalities": len(self._municipalities)
        }
