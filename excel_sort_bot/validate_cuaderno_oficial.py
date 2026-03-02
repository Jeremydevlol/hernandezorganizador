#!/usr/bin/env python3
"""
✅ VALIDADOR DE CUADERNO OFICIAL
Verifica que un output tenga la MISMA estructura que el Golden Template.

Checks:
1. Orden de hojas idéntico
2. Headers críticos idénticos
3. Tratamientos: fecha col 5, producto col 9, id col 1
"""
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Configuración
GOLDEN_PATH = "/Volumes/Uniclick4TB/organizadorhndezbueno/PABLO PEREZ RUBIO 2025 RESUELTO.XLSX"


def norm(v):
    """Normaliza valor para comparación"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v).strip()


def rows(ws, rr, cols=40):
    """Extrae filas de un rango"""
    return [[norm(ws.cell(r, c).value) for c in range(1, cols + 1)] for r in rr]


def validate_cuaderno(output_path: str, verbose: bool = True) -> dict:
    """
    Valida que un output sea cuaderno oficial.
    
    Returns:
        dict con success, checks, errors
    """
    result = {
        "success": False,
        "output_path": output_path,
        "checks": {},
        "tratamientos_count": 0,
        "errors": []
    }
    
    if not Path(GOLDEN_PATH).exists():
        result["errors"].append(f"Golden no encontrado: {GOLDEN_PATH}")
        return result
    
    if not Path(output_path).exists():
        result["errors"].append(f"Output no encontrado: {output_path}")
        return result
    
    gwb = load_workbook(GOLDEN_PATH, data_only=True)
    owb = load_workbook(output_path, data_only=True)
    
    # 1) Orden de hojas idéntico
    sheets_ok = owb.sheetnames == gwb.sheetnames
    result["checks"]["sheets_order"] = sheets_ok
    if not sheets_ok:
        result["errors"].append(f"Orden de hojas distinto: {len(owb.sheetnames)} vs {len(gwb.sheetnames)}")
        # Mostrar diferencias
        extra = set(owb.sheetnames) - set(gwb.sheetnames)
        missing = set(gwb.sheetnames) - set(owb.sheetnames)
        if extra:
            result["errors"].append(f"Hojas extra: {extra}")
        if missing:
            result["errors"].append(f"Hojas faltantes: {missing}")
    
    # 2) Headers críticos idénticos
    checks = {
        "inf.trat 1": [8, 9, 10],
        "inf.trat 2": [8, 9, 10],
        "inf.trat 3": [8, 9, 10],
        "inf.trat 4": [8, 9, 10],
        "2.1. DATOS PARCELAS": [12, 13],
        "reg.prod": [8, 9, 10],
        "reg.fert.": [8, 9, 10],
        "reg. cosecha": [8, 9, 10],
    }
    
    headers_ok = True
    for sh, rr in checks.items():
        if sh not in owb.sheetnames or sh not in gwb.sheetnames:
            continue
        try:
            gold_rows = rows(gwb[sh], rr)
            out_rows = rows(owb[sh], rr)
            match = gold_rows == out_rows
            result["checks"][f"headers_{sh}"] = match
            if not match:
                headers_ok = False
                result["errors"].append(f"Headers distintos en {sh}")
        except Exception as e:
            result["errors"].append(f"Error verificando {sh}: {e}")
    
    result["checks"]["all_headers_match"] = headers_ok
    
    # 3) Tratamientos: fecha en col 5, producto en col 9, id en col 1
    if "inf.trat 1" in owb.sheetnames:
        ws = owb["inf.trat 1"]
        count = 0
        valid_rows = True
        
        for r in range(11, ws.max_row + 1):
            fecha = ws.cell(r, 5).value
            if not fecha:
                continue
            
            pid = ws.cell(r, 1).value
            prod = ws.cell(r, 9).value
            
            if pid is None or str(pid).strip() == "":
                result["errors"].append(f"Fila {r}: fecha sin id_parcelas")
                valid_rows = False
            if prod is None or str(prod).strip() == "":
                result["errors"].append(f"Fila {r}: fecha sin producto")
                valid_rows = False
            
            count += 1
        
        result["tratamientos_count"] = count
        result["checks"]["valid_treatment_rows"] = valid_rows and count > 0
    
    gwb.close()
    owb.close()
    
    # Éxito
    result["success"] = all(result["checks"].values()) and len(result["errors"]) == 0
    
    if verbose:
        status = "✅" if result["success"] else "❌"
        print(f"\n{status} {Path(output_path).name}")
        print(f"   Tratamientos: {result['tratamientos_count']}")
        for k, v in result["checks"].items():
            print(f"   {'✅' if v else '❌'} {k}")
        if result["errors"]:
            for e in result["errors"][:3]:
                print(f"   ⚠️ {e}")
    
    return result


def main():
    print("\n" + "="*60)
    print("  ✅ VALIDADOR DE CUADERNO OFICIAL")
    print("="*60)
    print(f"  Golden: {Path(GOLDEN_PATH).name}")
    
    # Validar los 3 outputs
    outputs = [
        "OUTPUT_cliente_sigpac_963trat.xlsx",
        "OUTPUT_cliente_hort_820trat.xlsx",
        "OUTPUT_cliente_coop_2226trat.xlsx",
    ]
    
    base_dir = Path(__file__).parent
    resultados = []
    
    for output in outputs:
        path = base_dir / output
        if path.exists():
            result = validate_cuaderno(str(path))
            resultados.append(result)
    
    # Resumen
    print("\n" + "="*60)
    print("  RESUMEN")
    print("="*60)
    
    all_passed = all(r["success"] for r in resultados)
    total_trat = sum(r["tratamientos_count"] for r in resultados)
    
    for r in resultados:
        status = "✅" if r["success"] else "❌"
        name = Path(r["output_path"]).name
        print(f"  {status} {name}: {r['tratamientos_count']} tratamientos")
    
    print(f"\n  Total tratamientos validados: {total_trat}")
    
    if all_passed:
        print("\n  🎉 TODOS LOS CUADERNOS SON OFICIALES 🎉")
        print("  Estructura = Golden Template")
    else:
        print("\n  ❌ Algunos cuadernos tienen errores")
    
    return 0 if all_passed else 1


if __name__ == "__main__":
    sys.exit(main())
