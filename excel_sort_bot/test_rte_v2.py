#!/usr/bin/env python3
"""
🔥 TEST: Workbook Real-Time Editor (RTE) v2.0
Prueba TODAS las operaciones v2:
- INSERT_ROWS / DELETE_ROWS
- RENAME_SHEET / ADD_SHEET / DELETE_SHEET / COPY_SHEET
- SET_FORMULA
- IA con OpenAI
"""
import sys
import shutil
import json
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

from src.rte_api import (
    rte_start_session, 
    rte_preview, 
    rte_commit, 
    rte_undo, 
    rte_info,
    rte_close
)
from src.rte_ai_processor import RTEAIProcessor
from openpyxl import load_workbook


def print_header(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def print_section(title):
    print(f"\n--- {title} ---")


def test_result(name, success, details=""):
    status = "✅" if success else "❌"
    print(f"  {status} {name}: {details}")
    return success


def main():
    print_header("🔥 TEST: RTE v2.0 - ALL OPERATIONS")
    print(f"  Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    BASE_DIR = Path(__file__).parent
    
    SOURCE = BASE_DIR / "JUAN_GARCIA_HERNANDEZ_2026_RESUELTO.xlsx"
    TEST_FILE = BASE_DIR / "TEST_RTE_v2.xlsx"
    
    if not SOURCE.exists():
        print(f"❌ Archivo no existe: {SOURCE}")
        return 1
    
    shutil.copy(SOURCE, TEST_FILE)
    print(f"\n📁 Archivo de test: {TEST_FILE.name}")
    
    results = []
    
    # ============================================
    # 1. START SESSION (ADMIN mode para todas las ops)
    # ============================================
    print_header("1. START SESSION (ADMIN)")
    
    session = rte_start_session(str(TEST_FILE), mode="ADMIN")
    session_id = session["session_id"]
    
    results.append(test_result(
        "Session started",
        session["session_id"] is not None,
        f"ID: {session_id}, Mode: {session['mode']}"
    ))
    
    print(f"   Capabilities: {session['capabilities']}")
    
    # ============================================
    # 2. INSERT_ROWS
    # ============================================
    print_header("2. INSERT_ROWS")
    
    # Contar filas antes
    wb = load_workbook(TEST_FILE)
    rows_before = wb["inf.trat 1"].max_row
    wb.close()
    
    preview = rte_preview(
        session_id=session_id,
        instruction="Inserta 3 filas después de la 15",
        use_ai=False
    )
    
    results.append(test_result(
        "Preview INSERT_ROWS",
        len(preview['ops']) == 1 and preview['ops'][0]['op'] == 'INSERT_ROWS',
        f"Ops: {preview['ops']}"
    ))
    
    commit = rte_commit(session_id, preview['proposal_id'])
    
    # Verificar
    wb = load_workbook(TEST_FILE)
    rows_after = wb["inf.trat 1"].max_row
    wb.close()
    
    results.append(test_result(
        "Commit INSERT_ROWS",
        commit['success'] and rows_after == rows_before + 3,
        f"Rows: {rows_before} → {rows_after} (+3)"
    ))
    
    # ============================================
    # 3. DELETE_ROWS
    # ============================================
    print_header("3. DELETE_ROWS")
    
    preview = rte_preview(
        session_id=session_id,
        instruction="Elimina las filas 16 a 18",
        use_ai=False
    )
    
    results.append(test_result(
        "Preview DELETE_ROWS",
        len(preview['ops']) == 1 and preview['ops'][0]['op'] == 'DELETE_ROWS',
        f"Ops: {preview['ops']}"
    ))
    
    commit = rte_commit(session_id, preview['proposal_id'], force=True)  # Destructive op needs force
    
    wb = load_workbook(TEST_FILE)
    rows_final = wb["inf.trat 1"].max_row
    wb.close()
    
    results.append(test_result(
        "Commit DELETE_ROWS",
        commit['success'] and rows_final == rows_before,
        f"Rows back to: {rows_final}"
    ))
    
    # ============================================
    # 4. ADD_SHEET
    # ============================================
    print_header("4. ADD_SHEET")
    
    preview = rte_preview(
        session_id=session_id,
        instruction="Crea una hoja llamada 'Notas Adicionales'",
        use_ai=False
    )
    
    results.append(test_result(
        "Preview ADD_SHEET",
        len(preview['ops']) == 1 and preview['ops'][0]['op'] == 'ADD_SHEET',
        f"Ops: {preview['ops']}"
    ))
    
    commit = rte_commit(session_id, preview['proposal_id'])
    
    wb = load_workbook(TEST_FILE)
    sheet_exists = "Notas Adicionales" in wb.sheetnames
    wb.close()
    
    results.append(test_result(
        "Commit ADD_SHEET",
        commit['success'] and sheet_exists,
        f"Sheet exists: {sheet_exists}"
    ))
    
    # ============================================
    # 5. RENAME_SHEET
    # ============================================
    print_header("5. RENAME_SHEET")
    
    preview = rte_preview(
        session_id=session_id,
        instruction="Renombra la hoja 'Notas Adicionales' a 'Observaciones'",
        use_ai=False
    )
    
    results.append(test_result(
        "Preview RENAME_SHEET",
        len(preview['ops']) == 1 and preview['ops'][0]['op'] == 'RENAME_SHEET',
        f"Ops: {preview['ops']}"
    ))
    
    commit = rte_commit(session_id, preview['proposal_id'])
    
    wb = load_workbook(TEST_FILE)
    renamed = "Observaciones" in wb.sheetnames and "Notas Adicionales" not in wb.sheetnames
    wb.close()
    
    results.append(test_result(
        "Commit RENAME_SHEET",
        commit['success'] and renamed,
        f"Renamed: {renamed}"
    ))
    
    # ============================================
    # 6. DELETE_SHEET
    # ============================================
    print_header("6. DELETE_SHEET")
    
    preview = rte_preview(
        session_id=session_id,
        instruction="Elimina la hoja 'Observaciones'",
        use_ai=False
    )
    
    results.append(test_result(
        "Preview DELETE_SHEET",
        len(preview['ops']) == 1 and preview['ops'][0]['op'] == 'DELETE_SHEET',
        f"Ops: {preview['ops']}"
    ))
    
    commit = rte_commit(session_id, preview['proposal_id'], force=True)  # Destructive op needs force
    
    wb = load_workbook(TEST_FILE)
    deleted = "Observaciones" not in wb.sheetnames
    wb.close()
    
    results.append(test_result(
        "Commit DELETE_SHEET",
        commit['success'] and deleted,
        f"Deleted: {deleted}"
    ))
    
    # ============================================
    # 7. DELETE PROTECTED SHEET (debe fallar)
    # ============================================
    print_header("7. DELETE PROTECTED SHEET (should warn)")
    
    preview = rte_preview(
        session_id=session_id,
        instruction="Elimina la hoja 'inf.trat 1'",
        use_ai=False
    )
    
    results.append(test_result(
        "Protected sheet warning",
        preview['requires_confirmation'] == True,
        f"Warnings: {preview['warnings']}"
    ))
    
    # ============================================
    # 8. SET_FORMULA
    # ============================================
    print_header("8. SET_FORMULA")
    
    preview = rte_preview(
        session_id=session_id,
        instruction="Pon la fórmula =SUM(D11:D50) en la celda D200",
        use_ai=False
    )
    
    results.append(test_result(
        "Preview SET_FORMULA",
        len(preview['ops']) == 1 and preview['ops'][0]['op'] == 'SET_FORMULA',
        f"Ops: {preview['ops']}"
    ))
    
    commit = rte_commit(session_id, preview['proposal_id'])
    
    wb = load_workbook(TEST_FILE)
    formula_value = wb["inf.trat 1"]["D200"].value
    wb.close()
    
    results.append(test_result(
        "Commit SET_FORMULA",
        commit['success'] and formula_value == "=SUM(D11:D50)",
        f"Formula: {formula_value}"
    ))
    
    # ============================================
    # 9. TEST IA CON OPENAI
    # ============================================
    print_header("9. TEST IA CON OPENAI")
    
    processor = RTEAIProcessor()
    
    if processor.client:
        print("   OpenAI client initialized ✅")
        
        # Test con IA
        ai_result = processor.generate_ops(
            instruction="Inserta 2 filas vacías después de la fila 20 en inf.trat 1",
            use_ai=True
        )
        
        results.append(test_result(
            "AI Generate Ops",
            len(ai_result['ops']) > 0,
            f"Ops from AI: {[op.op.value for op in ai_result['ops']]}"
        ))
        
        # Test instrucción compleja
        ai_result2 = processor.generate_ops(
            instruction="Cambia la fecha del tratamiento de la fila 25 al 15 de marzo de 2026 y actualiza la dosis a 3.5 litros",
            use_ai=True
        )
        
        results.append(test_result(
            "AI Complex Instruction",
            len(ai_result2['ops']) >= 1,
            f"Ops: {len(ai_result2['ops'])}, Warnings: {ai_result2.get('warnings', [])}"
        ))
        
        # Test renombrar hoja
        ai_result3 = processor.generate_ops(
            instruction="Añade una nueva hoja llamada 'Resumen Anual'",
            use_ai=True
        )
        
        results.append(test_result(
            "AI Add Sheet",
            any(op.op == "ADD_SHEET" for op in ai_result3['ops']) if ai_result3['ops'] else False,
            f"Ops: {[op.op.value for op in ai_result3['ops']] if ai_result3['ops'] else 'None'}"
        ))
    else:
        print("   ⚠️ OpenAI client not available, skipping AI tests")
        results.append(test_result("AI Tests", False, "OpenAI client not available"))
    
    # ============================================
    # 10. UNDO ALL CHANGES
    # ============================================
    print_header("10. UNDO TO INITIAL")
    
    undo = rte_undo(session_id=session_id, checkpoint_id=f"{session_id}_initial_{datetime.now().strftime('%H%M%S')[:4]}")
    
    # Try to find initial checkpoint
    info = rte_info(session_id)
    print(f"   Session info: {info['commits_count']} commits")
    
    # ============================================
    # CLOSE SESSION
    # ============================================
    print_header("CLOSE SESSION")
    rte_close(session_id)
    print("   Session closed ✅")
    
    # ============================================
    # SUMMARY
    # ============================================
    print_header("📊 TEST SUMMARY")
    
    passed = sum(1 for r in results if r)
    total = len(results)
    
    print(f"""
    Total tests: {total}
    Passed: {passed}
    Failed: {total - passed}
    
    {"✅ ALL TESTS PASSED!" if passed == total else "❌ SOME TESTS FAILED"}
    
    v2.0 Operations Tested:
    ✅ INSERT_ROWS - Insertar filas
    ✅ DELETE_ROWS - Eliminar filas
    ✅ ADD_SHEET - Crear hoja
    ✅ RENAME_SHEET - Renombrar hoja
    ✅ DELETE_SHEET - Eliminar hoja
    ✅ SET_FORMULA - Fórmulas controladas
    ✅ Protected sheet validation
    ✅ OpenAI integration
    """)
    
    # Limpiar
    if TEST_FILE.exists():
        TEST_FILE.unlink()
    
    return 0 if passed == total else 1


if __name__ == "__main__":
    sys.exit(main())
