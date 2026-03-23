#!/usr/bin/env python3
"""
Organiza el Excel EMPEÑO Y DOCUMENTACIÓN:
- Hojas en orden alfabético (A-Z)
- Separar COMPRAS de GV (intermediación) en la hoja AC GROUP
- Ordenar datos por fecha de entrada y alfabéticamente
"""
import openpyxl
from pathlib import Path
from datetime import datetime
import unicodedata
import copy


def normalize_for_sort(s):
    """Normaliza texto para ordenación alfabética (ignora acentos)"""
    if s is None or (isinstance(s, float) and str(s) == 'nan'):
        return ""
    s = str(s).strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.upper()


def parse_date(val):
    """Convierte valor a fecha para ordenar"""
    if val is None:
        return datetime.min
    if isinstance(val, datetime):
        return val
    if hasattr(val, 'date'):
        return val
    if isinstance(val, str):
        try:
            return datetime.strptime(val[:10], "%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    return datetime.min


def get_sort_key(row_data, date_col=0, abc_col=1):
    """Clave: fecha primero, luego alfabético"""
    date_val = row_data[date_col] if date_col < len(row_data) else None
    abc_val = row_data[abc_col] if abc_col < len(row_data) else ""
    return (parse_date(date_val), normalize_for_sort(abc_val))


def main():
    input_path = Path("/Users/jeremjeremydevy/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/9CF6101B-81F4-4212-B8D5-59F26EAFAEBD/EMPEÑO Y DOCUMENTACIÓN  .xlsx")
    output_path = Path("/Volumes/Uniclick4TB/organizadorhndezbueno/EMPEÑO Y DOCUMENTACIÓN - ORGANIZADO.xlsx")

    if not input_path.exists():
        print(f"ERROR: No se encuentra el archivo: {input_path}")
        return

    print("Cargando Excel...")
    wb = openpyxl.load_workbook(str(input_path), data_only=False)

    compras_gv_name = "COMPRAS Y GV CON AC GROUP LAS R"
    if compras_gv_name in wb.sheetnames:
        ws_orig = wb[compras_gv_name]
        compras_data = []
        gv_data = []
        compras_header_row = None
        gv_header_row = None
        state = None
        num_cols = ws_orig.max_column

        for r in range(1, ws_orig.max_row + 1):
            v1 = str(ws_orig.cell(r, 1).value or "").strip().upper()
            if v1 == "COMPRAS" or (v1.startswith("COMPRAS") and "GV" not in v1):
                state = "compras"
                continue
            if "GV CON AC GROUP" in v1 or "GV CON ACGROUP" in v1:
                state = "gv"
                continue

            if state == "compras":
                if "FECHA DE ENTRADA" in str(ws_orig.cell(r, 1).value or ""):
                    compras_header_row = r
                    continue
                if compras_header_row and r > compras_header_row:
                    row_vals = [ws_orig.cell(r, c).value for c in range(1, num_cols + 1)]
                    if any(x is not None and str(x).strip() for x in row_vals[:5]):
                        compras_data.append(row_vals)
            elif state == "gv":
                if "FECHA DE ENTRADA" in str(ws_orig.cell(r, 1).value or ""):
                    gv_header_row = r
                    continue
                if gv_header_row and r > gv_header_row:
                    row_vals = [ws_orig.cell(r, c).value for c in range(1, num_cols + 1)]
                    if any(x is not None and str(x).strip() for x in row_vals[:5]):
                        gv_data.append(row_vals)

        compras_data.sort(key=lambda row: get_sort_key(row, 0, 1))
        gv_data.sort(key=lambda row: get_sort_key(row, 0, 1))

        header_compras = [ws_orig.cell(compras_header_row, c).value for c in range(1, num_cols + 1)] if compras_header_row else []
        header_gv = [ws_orig.cell(gv_header_row, c).value for c in range(1, num_cols + 1)] if gv_header_row else []

        del wb[compras_gv_name]

        ws_compras = wb.create_sheet("COMPRAS AC GROUP LAS ROZAS")
        for c, val in enumerate(header_compras, 1):
            ws_compras.cell(1, c, val)
        for r_idx, row in enumerate(compras_data, 2):
            for c_idx, val in enumerate(row, 1):
                ws_compras.cell(r_idx, c_idx, val)

        ws_gv = wb.create_sheet("GV INTERMEDIACIÓN AC GROUP")
        for c, val in enumerate(header_gv, 1):
            ws_gv.cell(1, c, val)
        for r_idx, row in enumerate(gv_data, 2):
            for c_idx, val in enumerate(row, 1):
                ws_gv.cell(r_idx, c_idx, val)

        print(f"  - COMPRAS AC GROUP: {len(compras_data)} filas (orden: fecha, alfabético)")
        print(f"  - GV INTERMEDIACIÓN AC GROUP: {len(gv_data)} filas (orden: fecha, alfabético)")

    sheet_names = sorted(wb.sheetnames, key=normalize_for_sort)
    for i, name in enumerate(sheet_names):
        try:
            idx = wb.sheetnames.index(name)
            if idx != i:
                wb.move_sheet(name, offset=i - idx)
        except Exception:
            pass

    # Ordenar GV SIN ACGROUP por fecha (col B) y alfabético
    if "GV SIN ACGROUP" in wb.sheetnames:
        ws = wb["GV SIN ACGROUP"]
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(10, ws.max_column + 1)):
                if "FECHA" in str(ws.cell(r, c).value or ""):
                    header_row = r
                    date_col = c - 1
                    data_rows = []
                    for row in range(header_row + 1, ws.max_row + 1):
                        row_vals = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
                        if any(x is not None and str(x).strip() for x in row_vals[:7]):
                            data_rows.append(row_vals)
                    if data_rows:
                        data_rows.sort(key=lambda row: get_sort_key(row, date_col, date_col + 1))
                        for r_idx, row in enumerate(data_rows, header_row + 1):
                            for c_idx, val in enumerate(row, 1):
                                ws.cell(r_idx, c_idx, val)
                    break

    # Ordenar FACTURACIÓN por fecha
    for name in ["FACTURACIÓN AC GROUP LAS ROZAS ", "FACTURACIÓN AC GROUP LAS ROZAS"]:
        if name in wb.sheetnames:
            ws = wb[name]
            for r in range(1, min(5, ws.max_row + 1)):
                for c in range(1, min(15, ws.max_column + 1)):
                    if "FECHA" in str(ws.cell(r, c).value or ""):
                        header_row = r
                        date_col = c - 1
                        data_rows = []
                        for row in range(header_row + 1, ws.max_row + 1):
                            row_vals = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
                            if any(x is not None and str(x).strip() for x in row_vals[:5]):
                                data_rows.append(row_vals)
                        if data_rows:
                            data_rows.sort(key=lambda row: get_sort_key(row, date_col, date_col + 2))
                            for r_idx, row in enumerate(data_rows, header_row + 1):
                                for c_idx, val in enumerate(row, 1):
                                    ws.cell(r_idx, c_idx, val)
                        break

    print(f"Guardando en: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()
    print("¡Listo! Archivo organizado guardado.")


if __name__ == "__main__":
    main()
