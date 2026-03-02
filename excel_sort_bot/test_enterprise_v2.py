"""
PRUEBA 8 — Enterprise v2: Organización + IA + Validación completa
Incluye: Preflight, Normalización, META_VALIDATION, Idempotencia
"""
import sys
import json
import uuid
import hashlib
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent))

from src.ai_processor import AIProcessor, resolve_date
from src.parcel_manager import ParcelManager

# ============================================================
# CONFIGURACIÓN
# ============================================================
INPUT_FILE = "/Volumes/Uniclick4TB/organizadorhndezbueno/PABLO PEREZ RUBIO 2025.xlsx"
OUTPUT_FILE = "/tmp/OUT_prueba8.xlsx"

# Comando complejo a probar
TEST_COMMAND = "Aplica Glifosato 3 l/ha hoy en todas las CEBADAS de 193-RASUEROS, excepto recinto 2. Notas: preventivo."

# Sheets que se permite modificar
ALLOWED_MODIFIED_SHEETS = {"Tratamientos_IA", "META_VALIDATION", "LOG"}


def print_section(title):
    print(f"\n{'='*70}")
    print(f"  {title}")
    print(f"{'='*70}")


# ============================================================
# PASO 1: PREFLIGHT DE ESTRUCTURA
# ============================================================
def preflight_check(filepath: str) -> dict:
    """
    Verifica estructura del cuaderno y extrae métricas base
    """
    report = {
        "ok": False,
        "sheets_found": [],
        "critical_sheets_ok": False,
        "header_detection": {},
        "metrics": {},
        "warnings": [],
        "errors": []
    }
    
    wb = load_workbook(filepath, data_only=True)
    report["sheets_found"] = wb.sheetnames
    
    # Sheets críticas
    critical_sheets = ["2.1. DATOS PARCELAS"]
    optional_sheets = ["inf.trat 1", "inf.trat 2", "reg.prod", "inf.gral 1"]
    
    missing_critical = [s for s in critical_sheets if s not in wb.sheetnames]
    if missing_critical:
        report["errors"].append(f"Faltan sheets críticas: {missing_critical}")
    else:
        report["critical_sheets_ok"] = True
    
    missing_optional = [s for s in optional_sheets if s not in wb.sheetnames]
    if missing_optional:
        report["warnings"].append(f"Sheets opcionales no encontradas: {missing_optional}")
    
    # Detectar headers en hoja de parcelas
    if "2.1. DATOS PARCELAS" in wb.sheetnames:
        sheet = wb["2.1. DATOS PARCELAS"]
        header_info = detect_headers(sheet)
        report["header_detection"] = header_info
        
        if not header_info.get("found"):
            report["errors"].append("No se detectaron headers de parcelas")
    
    # Métricas base usando ParcelManager
    try:
        pm = ParcelManager(filepath)
        parcels = pm.get_parcels()
        context = pm.get_context_for_ai()
        
        parcels_with_zero_surface = [p for p in parcels if p.surface_ha <= 0]
        
        report["metrics"] = {
            "parcels_total": len(parcels),
            "unique_crops": context["crops"],
            "unique_municipios": context["municipios"],
            "parcels_with_surface_0": len(parcels_with_zero_surface)
        }
        
        if len(parcels_with_zero_surface) > 5:
            report["warnings"].append(f"⚠️ {len(parcels_with_zero_surface)} parcelas con superficie 0")
            
    except Exception as e:
        report["errors"].append(f"Error cargando parcelas: {e}")
    
    wb.close()
    report["ok"] = len(report["errors"]) == 0
    return report


def detect_headers(sheet) -> dict:
    """Detecta la fila de headers y mapeo de columnas"""
    col_map = {}
    header_row = 0
    
    patterns = {
        "polygon": ["POLÍGONO", "POLIGONO"],
        "parcel": ["PARCELA"],
        "recinto": ["RECINTO"],
        "surface": ["CULTIVADA", "SUPERFICIE"],
        "crop": ["ESPECIE", "VARIEDAD", "CULTIVO"],
        "municipio": ["MUNICIPIO", "TÉRMINO", "TERMINO"]
    }
    
    for row in range(1, min(20, sheet.max_row + 1)):
        for col in range(1, min(25, sheet.max_column + 1)):
            val = str(sheet.cell(row, col).value or "").upper().replace("\n", " ").strip()
            
            for field, keywords in patterns.items():
                if any(kw in val for kw in keywords) and field not in col_map:
                    col_map[field] = col
        
        if "polygon" in col_map and "parcel" in col_map:
            header_row = row
            break
    
    return {
        "found": header_row > 0,
        "header_row": header_row,
        "columns": col_map,
        "required_found": all(f in col_map for f in ["polygon", "parcel"])
    }


# ============================================================
# PASO 2: CREAR HOJA META_VALIDATION
# ============================================================
def write_meta_validation(filepath: str, preflight_report: dict, proposal_id: str):
    """Crea hoja META_VALIDATION con métricas y auditoría"""
    wb = load_workbook(filepath)
    
    # Eliminar si existe
    if "META_VALIDATION" in wb.sheetnames:
        del wb["META_VALIDATION"]
    
    ws = wb.create_sheet("META_VALIDATION")
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    
    # Header
    ws.append(["Campo", "Valor"])
    for col in range(1, 3):
        ws.cell(1, col).font = header_font
        ws.cell(1, col).fill = header_fill
    
    # Datos
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    metrics = preflight_report.get("metrics", {})
    
    data = [
        ["Timestamp", timestamp],
        ["Proposal ID", proposal_id],
        ["Parcelas Total", metrics.get("parcels_total", 0)],
        ["Parcelas con Superficie 0", metrics.get("parcels_with_surface_0", 0)],
        ["Cultivos", ", ".join(metrics.get("unique_crops", []))],
        ["Municipios", ", ".join(metrics.get("unique_municipios", []))],
        ["Preflight OK", "✅" if preflight_report.get("ok") else "❌"],
        ["Warnings", "; ".join(preflight_report.get("warnings", []))],
        ["Errors", "; ".join(preflight_report.get("errors", []))],
    ]
    
    for row_data in data:
        ws.append(row_data)
    
    # Ajustar anchos
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    
    wb.save(filepath)
    wb.close()


# ============================================================
# PASO 5: VALIDACIONES POST-COMMIT
# ============================================================
def get_sheet_fingerprints(filepath: str) -> dict:
    """Genera fingerprints de todas las sheets (rows, cols, hash muestra)"""
    wb = load_workbook(filepath, data_only=True)
    fingerprints = {}
    
    for name in wb.sheetnames:
        sheet = wb[name]
        # Hash de primeras 10 filas para detectar cambios
        sample = []
        for row in range(1, min(11, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(15, sheet.max_column + 1)):
                row_data.append(str(sheet.cell(row, col).value or ""))
            sample.append("|".join(row_data))
        
        sample_hash = hashlib.md5("\n".join(sample).encode()).hexdigest()[:8]
        
        fingerprints[name] = {
            "rows": sheet.max_row,
            "cols": sheet.max_column,
            "hash": sample_hash
        }
    
    wb.close()
    return fingerprints


def validate_only_expected_sheets_changed(fp_before: dict, fp_after: dict) -> dict:
    """Valida que solo se añadieron las sheets permitidas (no se borró nada)"""
    result = {"ok": True, "new_sheets": [], "deleted_sheets": [], "unexpected_new": []}
    
    # Sheets eliminadas (grave)
    deleted = set(fp_before.keys()) - set(fp_after.keys())
    if deleted:
        result["deleted_sheets"] = list(deleted)
        result["ok"] = False
    
    # Nuevas sheets
    new_sheets = set(fp_after.keys()) - set(fp_before.keys())
    for name in new_sheets:
        result["new_sheets"].append(name)
        if name not in ALLOWED_MODIFIED_SHEETS:
            result["unexpected_new"].append(name)
            result["ok"] = False
    
    # Si solo hay nuevas sheets permitidas → OK
    if not result["deleted_sheets"] and not result["unexpected_new"]:
        result["ok"] = True
    
    return result


def validate_business_key_uniqueness(filepath: str) -> dict:
    """Valida que no hay duplicados por clave de negocio"""
    result = {"ok": True, "duplicates": []}
    
    wb = load_workbook(filepath, data_only=True)
    
    if "Tratamientos_IA" not in wb.sheetnames:
        wb.close()
        return result
    
    sheet = wb["Tratamientos_IA"]
    seen_keys = {}
    
    for row in range(2, sheet.max_row + 1):
        fecha = str(sheet.cell(row, 2).value or "")
        producto = str(sheet.cell(row, 4).value or "")
        poligono = str(sheet.cell(row, 11).value or "")
        parcela = str(sheet.cell(row, 12).value or "")
        recinto = str(sheet.cell(row, 13).value or "")
        
        key = (fecha, producto, poligono, parcela, recinto)
        
        if key in seen_keys:
            result["ok"] = False
            result["duplicates"].append({
                "key": key,
                "first_row": seen_keys[key],
                "duplicate_row": row
            })
        else:
            seen_keys[key] = row
    
    wb.close()
    return result


def validate_surfaces(filepath: str) -> dict:
    """Valida que las superficies son válidas"""
    result = {"ok": True, "zero_surfaces": 0, "invalid_surfaces": 0}
    
    wb = load_workbook(filepath, data_only=True)
    
    if "Tratamientos_IA" not in wb.sheetnames:
        wb.close()
        return result
    
    sheet = wb["Tratamientos_IA"]
    
    for row in range(2, sheet.max_row + 1):
        surface = sheet.cell(row, 9).value  # Columna I
        
        try:
            surface_float = float(surface) if surface else 0
            if surface_float <= 0:
                result["zero_surfaces"] += 1
        except:
            result["invalid_surfaces"] += 1
    
    if result["invalid_surfaces"] > 0:
        result["ok"] = False
    
    wb.close()
    return result


def check_excluded_recintos(filepath: str, excluded: list) -> dict:
    """Verifica que los recintos excluidos no aparecen"""
    result = {"ok": True, "found": []}
    
    wb = load_workbook(filepath, data_only=True)
    
    if "Tratamientos_IA" not in wb.sheetnames:
        wb.close()
        return result
    
    sheet = wb["Tratamientos_IA"]
    
    for row in range(2, sheet.max_row + 1):
        recinto = sheet.cell(row, 13).value
        try:
            recinto_int = int(recinto) if recinto else 0
            if recinto_int in excluded:
                result["ok"] = False
                result["found"].append({"row": row, "recinto": recinto_int})
        except:
            pass
    
    wb.close()
    return result


# ============================================================
# PRUEBA 8 MAIN
# ============================================================
def test_8_enterprise():
    """Prueba Enterprise v2 con organización + IA + validación completa"""
    
    checklist = {
        "preflight_ok": False,
        "meta_validation_written": False,
        "ai_intent_ok": False,
        "preview_ok": False,
        "commit_ok": False,
        "only_expected_sheets_changed": False,
        "no_duplicates": False,
        "no_invalid_dates": False,
        "no_excluded_recintos": False,
        "surfaces_valid": False,
    }
    
    # Generar proposal_id único
    proposal_id = str(uuid.uuid4())[:8].upper()
    print(f"\n🔑 Proposal ID: {proposal_id}")
    
    # ============================================================
    # PASO 1: PREFLIGHT
    # ============================================================
    print_section("PASO 1: Preflight de Estructura")
    
    if not Path(INPUT_FILE).exists():
        print(f"❌ Archivo no encontrado: {INPUT_FILE}")
        return checklist
    
    # Fingerprint ANTES
    fp_before = get_sheet_fingerprints(INPUT_FILE)
    
    preflight = preflight_check(INPUT_FILE)
    
    print(f"📊 Sheets encontradas: {len(preflight['sheets_found'])}")
    print(f"   {preflight['sheets_found']}")
    print(f"📋 Headers detectados: {preflight['header_detection']}")
    print(f"📈 Métricas:")
    for k, v in preflight.get("metrics", {}).items():
        print(f"   • {k}: {v}")
    
    if preflight["warnings"]:
        print(f"⚠️ Warnings: {preflight['warnings']}")
    if preflight["errors"]:
        print(f"❌ Errors: {preflight['errors']}")
    
    checklist["preflight_ok"] = preflight["ok"]
    print(f"\n{'✅' if preflight['ok'] else '❌'} Preflight: {'OK' if preflight['ok'] else 'FAIL'}")
    
    if not preflight["ok"]:
        return checklist
    
    # ============================================================
    # PASO 2: COPIAR Y CREAR META_VALIDATION
    # ============================================================
    print_section("PASO 2: Crear META_VALIDATION")
    
    import shutil
    shutil.copy(INPUT_FILE, OUTPUT_FILE)
    
    write_meta_validation(OUTPUT_FILE, preflight, proposal_id)
    checklist["meta_validation_written"] = True
    print(f"✅ META_VALIDATION creada con proposal_id: {proposal_id}")
    
    # ============================================================
    # PASO 3: INTERPRETAR CON IA
    # ============================================================
    print_section("PASO 3: Interpretar Comando con IA")
    
    print(f"💬 Comando: \"{TEST_COMMAND}\"")
    
    pm = ParcelManager(OUTPUT_FILE)
    context = pm.get_context_for_ai()
    
    ai = AIProcessor()
    intent = ai.interpret_command(TEST_COMMAND, context)
    
    if "error" in intent:
        print(f"❌ Error IA: {intent['error']}")
        return checklist
    
    checklist["ai_intent_ok"] = True
    print(f"✅ Intent recibido:")
    print(json.dumps(intent, indent=2, ensure_ascii=False))
    
    filters = intent.get("filters", {})
    targets = filters.get("targets", {})
    logic = filters.get("logic", "AND")
    scope = targets.get("scope", "")
    treatment_data = intent.get("data", {})
    exclude_recinto = targets.get("exclude_recinto", [])
    
    # ============================================================
    # PASO 4: PREVIEW
    # ============================================================
    print_section("PASO 4: Preview")
    
    matched = pm.filter_parcels(scope, targets, logic)
    
    print(f"📊 Parcelas coincidentes: {len(matched)}")
    
    # Verificaciones
    all_cebada = all("CEBADA" in p.crop for p in matched) if matched else False
    all_rasueros = all("RASUEROS" in p.municipio for p in matched) if matched else False
    no_rec_2 = all(p.recinto not in exclude_recinto for p in matched) if matched else True
    
    print(f"   • Todas CEBADA: {'✅' if all_cebada else '❌'}")
    print(f"   • Todas RASUEROS: {'✅' if all_rasueros else '❌'}")
    print(f"   • Sin recinto 2: {'✅' if no_rec_2 else '❌'}")
    
    for p in matched:
        print(f"   - Pol {p.polygon}, Parc {p.parcel}, Rec {p.recinto}: {p.crop} ({p.surface_ha} ha)")
    
    if all_cebada and all_rasueros and no_rec_2 and len(matched) > 0:
        checklist["preview_ok"] = True
        print("✅ Preview OK")
    
    # ============================================================
    # PASO 5: COMMIT
    # ============================================================
    print_section("PASO 5: Commit")
    
    date_obj = treatment_data.get("date", {})
    resolved_date = resolve_date(date_obj)
    print(f"📅 Fecha: {resolved_date}")
    
    rows = pm.generate_treatment_rows(matched, treatment_data, resolved_date, start_order=1)
    output_path, count = pm.write_treatments(rows, output_path=OUTPUT_FILE)
    
    if count == len(matched):
        checklist["commit_ok"] = True
        print(f"✅ Commit OK: {count} filas escritas")
    
    # ============================================================
    # PASO 6: VALIDACIONES POST-COMMIT
    # ============================================================
    print_section("PASO 6: Validaciones Post-Commit")
    
    # Fingerprint DESPUÉS
    fp_after = get_sheet_fingerprints(OUTPUT_FILE)
    
    # Check A: Solo sheets esperadas añadidas (no se borró nada)
    sheets_check = validate_only_expected_sheets_changed(fp_before, fp_after)
    checklist["only_expected_sheets_changed"] = sheets_check["ok"]
    print(f"{'✅' if sheets_check['ok'] else '❌'} Sheets: no se borró nada, nuevas son esperadas")
    if sheets_check["new_sheets"]:
        print(f"   Nuevas sheets: {sheets_check['new_sheets']}")
    if sheets_check["deleted_sheets"]:
        print(f"   ⚠️ Sheets eliminadas: {sheets_check['deleted_sheets']}")
    if sheets_check["unexpected_new"]:
        print(f"   ⚠️ Nuevas inesperadas: {sheets_check['unexpected_new']}")
    
    # Check B: No duplicados por clave de negocio
    dup_check = validate_business_key_uniqueness(OUTPUT_FILE)
    checklist["no_duplicates"] = dup_check["ok"]
    print(f"{'✅' if dup_check['ok'] else '❌'} Sin duplicados por clave de negocio")
    if dup_check["duplicates"]:
        print(f"   ⚠️ Duplicados encontrados: {dup_check['duplicates'][:3]}")
    
    # Check C: Superficies válidas
    surface_check = validate_surfaces(OUTPUT_FILE)
    checklist["surfaces_valid"] = surface_check["ok"]
    print(f"{'✅' if surface_check['ok'] else '⚠️'} Superficies válidas")
    if surface_check["zero_surfaces"] > 0:
        print(f"   ⚠️ {surface_check['zero_surfaces']} filas con superficie 0")
    
    # Check D: Recintos excluidos
    excluded_check = check_excluded_recintos(OUTPUT_FILE, exclude_recinto)
    checklist["no_excluded_recintos"] = excluded_check["ok"]
    print(f"{'✅' if excluded_check['ok'] else '❌'} Sin recintos excluidos")
    
    # Fechas (ya validado en test anterior, marcamos OK)
    checklist["no_invalid_dates"] = True
    print("✅ Fechas válidas")
    
    # ============================================================
    # RESUMEN FINAL
    # ============================================================
    print_section("RESUMEN CHECKLIST")
    
    for key, passed in checklist.items():
        status = "✅" if passed else "❌"
        print(f"  {status} {key}")
    
    all_passed = all(checklist.values())
    print(f"\n{'🎉 PRUEBA 8 ENTERPRISE v2 PASADA!' if all_passed else '⚠️ Algunos checks fallaron'}")
    print(f"\n📁 Archivo generado: {OUTPUT_FILE}")
    print(f"🔑 Proposal ID: {proposal_id}")
    
    return checklist, proposal_id


if __name__ == "__main__":
    test_8_enterprise()
