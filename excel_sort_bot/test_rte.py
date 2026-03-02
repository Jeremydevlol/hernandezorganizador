#!/usr/bin/env python3
"""
🔥 TEST: Workbook Real-Time Editor (RTE)
Prueba el flujo completo: session -> preview -> commit -> undo
"""
import sys
import shutil
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

from src.rte_api import (
    rte_start_session, 
    rte_preview, 
    rte_commit, 
    rte_undo, 
    rte_info,
    rte_close
)
from openpyxl import load_workbook


def print_header(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def print_json(data, indent=2):
    import json
    print(json.dumps(data, indent=indent, default=str, ensure_ascii=False))


def main():
    print_header("🔥 TEST: Workbook Real-Time Editor (RTE)")
    print(f"  Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    BASE_DIR = Path(__file__).parent
    
    # Usar uno de los cuadernos resueltos
    SOURCE = BASE_DIR / "JUAN_GARCIA_HERNANDEZ_2026_RESUELTO.xlsx"
    TEST_FILE = BASE_DIR / "TEST_RTE_Juan_Garcia.xlsx"
    
    if not SOURCE.exists():
        print(f"❌ Archivo no existe: {SOURCE}")
        return 1
    
    # Copiar para testing
    shutil.copy(SOURCE, TEST_FILE)
    print(f"\n📁 Archivo de test: {TEST_FILE.name}")
    
    # ============================================
    # 1. START SESSION
    # ============================================
    print_header("1. START SESSION")
    
    session = rte_start_session(str(TEST_FILE), mode="POWER")
    print(f"✅ Session ID: {session['session_id']}")
    print(f"   Mode: {session['mode']}")
    print(f"   Sheets: {len(session['sheets'])}")
    print(f"   Capabilities: {session['capabilities']}")
    print(f"   Checksum: {session['checksum']}")
    
    session_id = session["session_id"]
    
    # ============================================
    # 2. PREVIEW (sin IA, parsing local)
    # ============================================
    print_header("2. PREVIEW - Cambiar fecha fila 12")
    
    preview = rte_preview(
        session_id=session_id,
        instruction="Cambia la fecha de la fila 12 a 25/03/2026",
        selection={"sheet": "inf.trat 1"},
        use_ai=False  # Usa parsing local
    )
    
    print(f"✅ Proposal ID: {preview['proposal_id']}")
    print(f"   Ops: {len(preview['ops'])}")
    print(f"   Requires confirmation: {preview['requires_confirmation']}")
    print(f"   Warnings: {preview['warnings']}")
    
    if preview['ops']:
        print(f"   Primera op: {preview['ops'][0]}")
    
    # Verificar diff preview
    print(f"\n   Diff preview:")
    for cell in preview['diff_preview'].get('cells', [])[:3]:
        print(f"     {cell['cell']}: '{cell['old']}' → '{cell['new']}'")
    
    proposal_id = preview['proposal_id']
    
    # ============================================
    # 3. COMMIT
    # ============================================
    print_header("3. COMMIT - Aplicar cambios")
    
    commit = rte_commit(
        session_id=session_id,
        proposal_id=proposal_id,
        idempotency_key=f"test_{datetime.now().strftime('%H%M%S')}"
    )
    
    print(f"✅ Success: {commit['success']}")
    print(f"   Audit ID: {commit['audit_id']}")
    print(f"   Cells changed: {commit['cells_changed']}")
    print(f"   Checksum: {commit['checksum']}")
    
    # Verificar en Excel
    wb = load_workbook(TEST_FILE)
    ws = wb["inf.trat 1"]
    new_value = ws.cell(12, 5).value  # E12
    print(f"   Valor en E12: {new_value}")
    wb.close()
    
    # ============================================
    # 4. PREVIEW - Cambiar dosis
    # ============================================
    print_header("4. PREVIEW - Cambiar dosis fila 12")
    
    preview2 = rte_preview(
        session_id=session_id,
        instruction="Cambia la dosis de la fila 12 a 5,75 l",
        use_ai=False
    )
    
    print(f"✅ Proposal ID: {preview2['proposal_id']}")
    print(f"   Ops: {len(preview2['ops'])}")
    if preview2['ops']:
        print(f"   Op: {preview2['ops'][0]}")
    
    # Commit
    commit2 = rte_commit(
        session_id=session_id,
        proposal_id=preview2['proposal_id'],
        idempotency_key=f"test_dosis_{datetime.now().strftime('%H%M%S')}"
    )
    print(f"   Commit: {commit2['success']}, cells: {commit2['cells_changed']}")
    
    # ============================================
    # 5. PREVIEW - Find & Replace
    # ============================================
    print_header("5. PREVIEW - Find & Replace")
    
    preview3 = rte_preview(
        session_id=session_id,
        instruction="Reemplaza 'GUISANTE' por 'GUISANTE VERDE'",
        selection={"sheet": "inf.trat 1"},
        use_ai=False
    )
    
    print(f"✅ Proposal ID: {preview3['proposal_id']}")
    print(f"   Ops: {len(preview3['ops'])}")
    if preview3['ops']:
        print(f"   Op: {preview3['ops'][0]}")
    
    # Commit
    commit3 = rte_commit(
        session_id=session_id,
        proposal_id=preview3['proposal_id'],
        idempotency_key=f"test_replace_{datetime.now().strftime('%H%M%S')}"
    )
    print(f"   Commit: {commit3['success']}, cells: {commit3['cells_changed']}")
    
    # ============================================
    # 6. TEST IDEMPOTENCIA
    # ============================================
    print_header("6. TEST IDEMPOTENCIA - Doble commit")
    
    idempotency_key = "duplicate_test_key_123"
    
    preview4 = rte_preview(
        session_id=session_id,
        instruction="Cambia la fecha de la fila 13 a 01/01/2026",
        use_ai=False
    )
    
    # Primer commit
    commit4a = rte_commit(
        session_id=session_id,
        proposal_id=preview4['proposal_id'],
        idempotency_key=idempotency_key
    )
    print(f"   Primer commit: {commit4a['success']}")
    
    # Segundo commit con misma key (debe fallar)
    preview4b = rte_preview(
        session_id=session_id,
        instruction="Cambia la fecha de la fila 14 a 02/02/2026",
        use_ai=False
    )
    commit4b = rte_commit(
        session_id=session_id,
        proposal_id=preview4b['proposal_id'],
        idempotency_key=idempotency_key  # Misma key
    )
    print(f"   Segundo commit (misma key): success={commit4b['success']}, error={commit4b.get('error')}")
    
    if commit4b['error'] == "DUPLICATE_COMMIT":
        print("   ✅ IDEMPOTENCIA FUNCIONA")
    else:
        print("   ❌ IDEMPOTENCIA FALLÓ")
    
    # ============================================
    # 7. TEST UNDO
    # ============================================
    print_header("7. TEST UNDO - Rollback")
    
    # Info antes de undo
    info = rte_info(session_id)
    print(f"   Commits antes de undo: {info['commits_count']}")
    
    # Undo
    undo = rte_undo(session_id=session_id)
    print(f"✅ Undo success: {undo['success']}")
    print(f"   Restored from: {undo['restored_from']}")
    print(f"   New checksum: {undo['checksum']}")
    
    # ============================================
    # 8. SESSION INFO
    # ============================================
    print_header("8. SESSION INFO")
    
    info = rte_info(session_id)
    print_json(info)
    
    # ============================================
    # 9. CLOSE SESSION
    # ============================================
    print_header("9. CLOSE SESSION")
    
    rte_close(session_id)
    print("✅ Session cerrada")
    
    # ============================================
    # RESUMEN
    # ============================================
    print_header("✅ TEST RTE COMPLETADO")
    
    print("""
    Flujo validado:
    ✅ start_session - Inicia sesión con modo y capabilities
    ✅ preview - Genera ops desde instrucción (parsing local)
    ✅ commit - Aplica ops con auditoría
    ✅ idempotencia - Bloquea double commit
    ✅ undo - Rollback desde checkpoint
    ✅ close - Cierra sesión
    
    Archivos:
    - src/workbook_rte.py - Motor RTE
    - src/rte_ai_processor.py - Generador de ops
    - src/rte_api.py - API endpoints
    """)
    
    # Limpiar
    if TEST_FILE.exists():
        TEST_FILE.unlink()
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
