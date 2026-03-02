#!/usr/bin/env python3
"""
🔥 STRESS TEST v2.3 — Con DEDUPE GATE
Procesa los 3 clientes ficticios y verifica:
- 0 duplicados
- Inserción correcta en data block
- Output = Golden Template
"""
import sys
import json
import time
from pathlib import Path
from datetime import datetime
from typing import Dict

sys.path.insert(0, str(Path(__file__).parent))

from src.resolver_cuaderno import resolver_cuaderno
from openpyxl import load_workbook


# ============================================
BASE_DIR = Path(__file__).parent
CLIENTES_DIR = BASE_DIR / "clientes_ficticios"
# ============================================


def print_header(title):
    print(f"\n{'🔥'*25}")
    print(f"  {title}")
    print(f"{'🔥'*25}\n")


def print_section(title):
    print(f"\n{'='*65}")
    print(f"  {title}")
    print(f"{'='*65}")


def count_sheets(filepath):
    wb = load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def count_data_rows(filepath, sheet_name, start_row=11):
    """Cuenta filas con datos reales"""
    wb = load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[sheet_name]
    count = 0
    for row in range(start_row, ws.max_row + 1):
        val_a = ws.cell(row, 1).value
        val_i = ws.cell(row, 9).value  # producto
        if val_a is not None and val_i is not None:
            count += 1
    wb.close()
    return count


def check_duplicates(filepath, sheet_name="inf.trat 1") -> int:
    """Cuenta duplicados por business key"""
    wb = load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    
    ws = wb[sheet_name]
    seen = set()
    duplicates = 0
    
    for row in range(11, ws.max_row + 1):
        id_p = ws.cell(row, 1).value
        fecha = ws.cell(row, 5).value
        prod = ws.cell(row, 9).value
        dosis = ws.cell(row, 11).value
        
        if id_p is not None and prod is not None:
            # Normalizar fecha
            if isinstance(fecha, datetime):
                fecha_str = fecha.strftime("%Y-%m-%d")
            else:
                fecha_str = str(fecha or "")
            
            key = f"{id_p}|{fecha_str}|{prod}|{dosis}"
            if key in seen:
                duplicates += 1
            seen.add(key)
    
    wb.close()
    return duplicates


def test_cliente(cliente_id: str) -> Dict:
    """Procesa un cliente con DEDUPE GATE"""
    client_dir = CLIENTES_DIR / cliente_id
    
    if not client_dir.exists():
        return {"cliente_id": cliente_id, "success": False, "errors": ["Directorio no existe"]}
    
    input_path = client_dir / "input.xlsx"
    golden_path = client_dir / "golden.xlsx"
    output_path = client_dir / "output.xlsx"
    tratamientos_path = client_dir / "tratamientos.json"
    config_path = client_dir / "config.json"
    
    result = {
        "cliente_id": cliente_id,
        "empresa": "",
        "titular": "",
        "success": False,
        "parcelas": 0,
        "tratamientos_input": 0,
        "tratamientos_escritos": 0,
        "duplicados_bloqueados": 0,
        "duplicados_en_output": 0,
        "filas_output": 0,
        "hojas_ok": False,
        "tiempo_seg": 0,
        "errors": []
    }
    
    # Cargar config
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
            result["empresa"] = config.get("empresa", "")
            result["titular"] = config.get("titular", "")
    
    # Cargar tratamientos
    if not tratamientos_path.exists():
        result["errors"].append("tratamientos.json no existe")
        return result
    
    with open(tratamientos_path, "r", encoding="utf-8") as f:
        tratamientos = json.load(f)
    
    result["tratamientos_input"] = len(tratamientos)
    
    # Contar parcelas únicas
    parcelas_unicas = set()
    for t in tratamientos:
        key = (t["polygon"], t["parcel"], t["recinto"])
        parcelas_unicas.add(key)
    result["parcelas"] = len(parcelas_unicas)
    
    print(f"\n  {result['empresa']}")
    print(f"  Titular: {result['titular']}")
    print(f"  📊 Input: {result['parcelas']} parcelas, {result['tratamientos_input']} tratamientos")
    
    # Ejecutar resolver con DEDUPE GATE
    start_time = time.time()
    
    try:
        resolver_result = resolver_cuaderno(
            input_path=str(input_path),
            output_path=str(output_path),
            golden_path=str(golden_path),
            treatments=tratamientos,
            allow_duplicates=False  # DEDUPE GATE ACTIVO
        )
        
        result["tratamientos_escritos"] = resolver_result.get("actions_applied", 0)
        result["duplicados_bloqueados"] = resolver_result.get("duplicates_blocked", 0)
        result["errors"].extend(resolver_result.get("errors", []))
        
    except Exception as e:
        result["errors"].append(f"Resolver falló: {e}")
        return result
    
    result["tiempo_seg"] = round(time.time() - start_time, 2)
    
    # Verificar output
    if not output_path.exists():
        result["errors"].append("Output no generado")
        return result
    
    # Contar filas
    result["filas_output"] = count_data_rows(output_path, "inf.trat 1")
    
    # Verificar duplicados en output
    result["duplicados_en_output"] = check_duplicates(output_path)
    
    # Verificar hojas
    golden_sheets = set(count_sheets(golden_path))
    output_sheets = set(count_sheets(output_path))
    extra_sheets = output_sheets - golden_sheets
    result["hojas_ok"] = len(extra_sheets) == 0
    
    if extra_sheets:
        result["errors"].append(f"Hojas extras: {extra_sheets}")
    
    # Éxito: 0 duplicados en output
    result["success"] = (
        len(result["errors"]) == 0 and
        result["hojas_ok"] and
        result["duplicados_en_output"] == 0  # CERO DUPLICADOS
    )
    
    return result


def main():
    print_header("STRESS TEST v2.3 — DEDUPE GATE")
    print(f"  Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Buscar clientes
    if not CLIENTES_DIR.exists():
        print(f"❌ Directorio no existe: {CLIENTES_DIR}")
        print("   Ejecuta primero: python3 generar_clientes_ficticios.py")
        return 1
    
    clientes = [d.name for d in CLIENTES_DIR.iterdir() if d.is_dir()]
    print(f"  Clientes encontrados: {len(clientes)}")
    
    resultados = []
    total_start = time.time()
    
    for cliente_id in clientes:
        print_section(f"Cliente: {cliente_id}")
        result = test_cliente(cliente_id)
        resultados.append(result)
        
        # Mostrar resultado
        status = "✅ PASSED" if result["success"] else "❌ FAILED"
        print(f"\n  Status: {status}")
        print(f"  📊 Tratamientos escritos: {result['tratamientos_escritos']}/{result['tratamientos_input']}")
        print(f"  📊 Duplicados bloqueados: {result['duplicados_bloqueados']}")
        print(f"  📊 Duplicados en output: {result['duplicados_en_output']}")
        print(f"  📊 Filas en output: {result['filas_output']}")
        print(f"  📊 Hojas OK: {'✅' if result['hojas_ok'] else '❌'}")
        print(f"  ⏱️ Tiempo: {result['tiempo_seg']}s")
        
        if result["errors"]:
            for e in result["errors"][:3]:
                print(f"  ❌ {e}")
    
    total_time = round(time.time() - total_start, 2)
    
    # Resumen
    print_section("RESUMEN FINAL")
    
    total_parcelas = sum(r["parcelas"] for r in resultados)
    total_tratamientos = sum(r["tratamientos_input"] for r in resultados)
    total_escritos = sum(r["tratamientos_escritos"] for r in resultados)
    total_bloqueados = sum(r["duplicados_bloqueados"] for r in resultados)
    total_duplicados_output = sum(r["duplicados_en_output"] for r in resultados)
    all_passed = all(r["success"] for r in resultados)
    
    print(f"\n  {'Empresa':<40} {'Parcelas':>8} {'Input':>8} {'Escritos':>8} {'Tiempo':>8}")
    print(f"  {'-'*75}")
    
    for r in resultados:
        status = "✅" if r["success"] else "❌"
        empresa = r["empresa"][:38] if r["empresa"] else r["cliente_id"]
        print(f"  {status} {empresa:<38} {r['parcelas']:>8} {r['tratamientos_input']:>8} {r['tratamientos_escritos']:>8} {r['tiempo_seg']:>7}s")
    
    print(f"  {'-'*75}")
    print(f"  {'TOTAL':<40} {total_parcelas:>8} {total_tratamientos:>8} {total_escritos:>8} {total_time:>7}s")
    
    print(f"\n  📊 Métricas de DEDUPE GATE:")
    print(f"     Duplicados bloqueados: {total_bloqueados}")
    print(f"     Duplicados en output: {total_duplicados_output}")
    print(f"     Velocidad: {round(total_escritos/total_time, 1) if total_time > 0 else 0} trat/seg")
    
    if all_passed and total_duplicados_output == 0:
        print(f"\n{'🎉'*20}")
        print(f"  STRESS TEST v2.3: ALL PASSED!")
        print(f"  ✅ DEDUPE GATE: 0 duplicados en output")
        print(f"  ✅ {total_parcelas:,} parcelas, {total_escritos:,} tratamientos")
        print(f"{'🎉'*20}\n")
    else:
        print(f"\n❌ Test fallido: {total_duplicados_output} duplicados en output")
    
    # Guardar reporte
    report_path = BASE_DIR / "stress_test_v23_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "total_time": total_time,
            "total_parcelas": total_parcelas,
            "total_tratamientos": total_tratamientos,
            "total_escritos": total_escritos,
            "total_bloqueados": total_bloqueados,
            "total_duplicados_output": total_duplicados_output,
            "all_passed": all_passed and total_duplicados_output == 0,
            "clientes": resultados
        }, f, indent=2, ensure_ascii=False)
    
    print(f"📁 Reporte: {report_path}\n")
    
    return 0 if (all_passed and total_duplicados_output == 0) else 1


if __name__ == "__main__":
    sys.exit(main())
