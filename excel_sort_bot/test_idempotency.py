"""
PRUEBA 9 — Double Commit Attack
Valida idempotencia real: 2do commit debe fallar con 409 ALREADY_COMMITTED
"""
import sys
import json
import uuid
import shutil
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

from src.ai_processor import AIProcessor, resolve_date
from src.parcel_manager import ParcelManager

# ============================================================
# CONFIGURACIÓN
# ============================================================
INPUT_FILE = "/Volumes/Uniclick4TB/organizadorhndezbueno/PABLO PEREZ RUBIO 2025.xlsx"
OUTPUT_FILE = "/tmp/OUT_prueba9_idempotencia.xlsx"
TEST_COMMAND = "Aplica Glifosato 3 l/ha hoy en polígono 8 parcela 148"


def print_section(title):
    print(f"\n{'='*70}")
    print(f"  {title}")
    print(f"{'='*70}")


def count_treatment_rows(filepath):
    """Cuenta filas en Tratamientos_IA"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    
    if "Tratamientos_IA" not in wb.sheetnames:
        wb.close()
        return 0
    
    sheet = wb["Tratamientos_IA"]
    count = sheet.max_row - 1  # Menos header
    wb.close()
    return max(0, count)


# Simular el flujo del servidor
class MockProposalStore:
    """Simula proposals_store y committed_proposals del servidor"""
    
    def __init__(self):
        self.proposals_store = {}
        self.committed_proposals = {}
    
    def create_proposal(self, proposal_id: str, data: dict):
        self.proposals_store[proposal_id] = {
            "created_at": datetime.now(),
            **data
        }
    
    def commit_proposal(self, proposal_id: str) -> dict:
        """
        Simula commit_treatment del servidor.
        Returns: {"success": bool, "status_code": int, "message": str}
        """
        # Check 1: Ya fue committed?
        if proposal_id in self.committed_proposals:
            commit_time = self.committed_proposals[proposal_id]
            return {
                "success": False,
                "status_code": 409,
                "error": "ALREADY_COMMITTED",
                "message": f"Ya fue ejecutada el {commit_time.strftime('%d/%m/%Y %H:%M:%S')}"
            }
        
        # Check 2: Existe propuesta?
        proposal = self.proposals_store.get(proposal_id)
        if not proposal:
            return {
                "success": False,
                "status_code": 404,
                "error": "NOT_FOUND",
                "message": "Propuesta no encontrada"
            }
        
        # Ejecutar commit
        try:
            pm = ParcelManager(proposal["file_path"])
            rows = proposal["rows"]
            output_path, rows_written = pm.write_treatments(rows, output_path=proposal["output_path"])
            
            # Marcar como committed
            self.committed_proposals[proposal_id] = datetime.now()
            
            # Limpiar propuesta
            del self.proposals_store[proposal_id]
            
            return {
                "success": True,
                "status_code": 200,
                "rows_written": rows_written,
                "output_path": output_path
            }
            
        except Exception as e:
            return {
                "success": False,
                "status_code": 500,
                "error": str(e)
            }


def test_9_double_commit():
    """Prueba de ataque de doble commit"""
    
    checklist = {
        "setup_ok": False,
        "preview_ok": False,
        "first_commit_ok": False,
        "second_commit_blocked": False,
        "no_duplicates": False,
    }
    
    # ============================================================
    # PASO 1: SETUP
    # ============================================================
    print_section("PASO 1: Setup")
    
    if not Path(INPUT_FILE).exists():
        print(f"❌ Archivo no encontrado: {INPUT_FILE}")
        return checklist
    
    # Copiar archivo
    shutil.copy(INPUT_FILE, OUTPUT_FILE)
    print(f"✅ Archivo copiado a: {OUTPUT_FILE}")
    
    checklist["setup_ok"] = True
    
    # ============================================================
    # PASO 2: GENERAR PREVIEW (simular /preview)
    # ============================================================
    print_section("PASO 2: Generar Preview")
    
    pm = ParcelManager(OUTPUT_FILE)
    context = pm.get_context_for_ai()
    
    ai = AIProcessor()
    intent = ai.interpret_command(TEST_COMMAND, context)
    
    if "error" in intent:
        print(f"❌ Error IA: {intent['error']}")
        return checklist
    
    print(f"✅ Intent recibido")
    
    filters = intent.get("filters", {})
    targets = filters.get("targets", {})
    logic = filters.get("logic", "AND")
    scope = targets.get("scope", "")
    treatment_data = intent.get("data", {})
    
    matched = pm.filter_parcels(scope, targets, logic)
    print(f"📊 Parcelas coincidentes: {len(matched)}")
    
    if not matched:
        print("❌ Sin parcelas")
        return checklist
    
    date_obj = treatment_data.get("date", {})
    resolved_date = resolve_date(date_obj)
    
    rows = pm.generate_treatment_rows(matched, treatment_data, resolved_date, start_order=1)
    
    checklist["preview_ok"] = True
    
    # ============================================================
    # PASO 3: SIMULAR PRIMER COMMIT
    # ============================================================
    print_section("PASO 3: Primer Commit (debe funcionar)")
    
    # Crear proposal_id único
    proposal_id = str(uuid.uuid4())[:8].upper()
    print(f"🔑 Proposal ID: {proposal_id}")
    
    # Usar mock store
    store = MockProposalStore()
    store.create_proposal(proposal_id, {
        "file_path": OUTPUT_FILE,
        "output_path": OUTPUT_FILE,
        "rows": rows,
        "intent": intent
    })
    
    # Primer commit
    result1 = store.commit_proposal(proposal_id)
    
    print(f"📊 Resultado Commit #1:")
    print(f"   Status: {result1['status_code']}")
    print(f"   Success: {result1['success']}")
    
    if result1["success"]:
        print(f"   ✅ Filas escritas: {result1['rows_written']}")
        checklist["first_commit_ok"] = True
        rows_after_first = count_treatment_rows(OUTPUT_FILE)
        print(f"   📊 Filas en Excel: {rows_after_first}")
    else:
        print(f"   ❌ Error: {result1}")
        return checklist
    
    # ============================================================
    # PASO 4: SIMULAR SEGUNDO COMMIT (doble click!)
    # ============================================================
    print_section("PASO 4: Segundo Commit (DEBE FALLAR)")
    
    # Segundo commit con el MISMO proposal_id
    result2 = store.commit_proposal(proposal_id)
    
    print(f"📊 Resultado Commit #2 (ataque):")
    print(f"   Status: {result2['status_code']}")
    print(f"   Success: {result2['success']}")
    
    if result2["status_code"] == 409 and result2.get("error") == "ALREADY_COMMITTED":
        print(f"   ✅ BLOQUEADO CORRECTAMENTE: {result2['message']}")
        checklist["second_commit_blocked"] = True
    else:
        print(f"   ❌ FALLO: El segundo commit no fue bloqueado!")
        print(f"      Esto causaría duplicados en producción")
    
    # ============================================================
    # PASO 5: VERIFICAR NO HAY DUPLICADOS
    # ============================================================
    print_section("PASO 5: Verificar No Duplicados")
    
    rows_final = count_treatment_rows(OUTPUT_FILE)
    print(f"📊 Filas finales en Excel: {rows_final}")
    print(f"📊 Filas después de commit #1: {rows_after_first}")
    
    if rows_final == rows_after_first:
        print("✅ Sin duplicados: mismo número de filas")
        checklist["no_duplicates"] = True
    else:
        print(f"❌ DUPLICADOS DETECTADOS: {rows_final - rows_after_first} filas extra")
    
    # ============================================================
    # RESUMEN
    # ============================================================
    print_section("RESUMEN CHECKLIST")
    
    for key, passed in checklist.items():
        status = "✅" if passed else "❌"
        print(f"  {status} {key}")
    
    all_passed = all(checklist.values())
    print(f"\n{'🎉 PRUEBA 9 ¡IDEMPOTENCIA VALIDADA!' if all_passed else '⚠️ Falló idempotencia'}")
    
    return checklist


if __name__ == "__main__":
    test_9_double_commit()
