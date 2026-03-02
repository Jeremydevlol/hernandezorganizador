#!/usr/bin/env python3
"""
🏭 GENERADOR DE CLIENTES FICTICIOS
Crea 3 clientes con datos realistas de empresas agrícolas españolas.

Clientes:
1. Juan García - Granja de Guisantes García S.L. (380 parcelas)
2. Pedro Pérez - Siembra de Maíz Pérez e Hijos S.A. (520 parcelas)
3. Enrique Iglesias - Granja Agrícola Iglesias Hnos. S.L. (650 parcelas)
"""
import sys
import json
import random
import shutil
from pathlib import Path
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

sys.path.insert(0, str(Path(__file__).parent))


# ============================================
# Configuración de Clientes Ficticios
# ============================================

CLIENTES = {
    "juan_garcia_guisantes": {
        "id": "juan_garcia_guisantes",
        "empresa": "Granja de Guisantes García S.L.",
        "titular": "Juan García Hernández",
        "nif": "12345678A",
        "direccion": "Ctra. Salamanca km 23, 37800 Alba de Tormes",
        "telefono": "923 456 789",
        "email": "juan@granjagarcia.es",
        "municipios": ["108-ALBA DE TORMES", "152-TERRADILLOS", "089-CALVARRASA", "178-MOZARBEZ"],
        "cultivos": [
            ("GUISANTE", 0.40),
            ("LENTEJA", 0.25),
            ("GARBANZO", 0.20),
            ("VEZA", 0.15)
        ],
        "productos": [
            {"nombre": "Glifosato 36%", "registro": "21001", "dosis": 3.0, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "Clorpirifos", "registro": "19887", "dosis": 1.5, "unidad": "l/ha", "plaga": "PULGON"},
            {"nombre": "Azufre Mojable", "registro": "15234", "dosis": 2.5, "unidad": "kg/ha", "plaga": "OIDIO"},
            {"nombre": "Deltametrina 2.5%", "registro": "22456", "dosis": 0.5, "unidad": "l/ha", "plaga": "GORGOJO"}
        ],
        "parcelas": 380,
        "aplicaciones_por_parcela": 2
    },
    
    "pedro_perez_maiz": {
        "id": "pedro_perez_maiz",
        "empresa": "Siembra de Maíz Pérez e Hijos S.A.",
        "titular": "Pedro Pérez Martínez",
        "nif": "87654321B",
        "direccion": "Avda. de la Agricultura 45, 47400 Medina del Campo",
        "telefono": "983 654 321",
        "email": "pedro@siembraperez.com",
        "municipios": ["201-MEDINA DEL CAMPO", "234-RUEDA", "189-LA SECA", "215-POZALDEZ", "223-RODILANA"],
        "cultivos": [
            ("MAIZ", 0.50),
            ("GIRASOL", 0.25),
            ("REMOLACHA", 0.15),
            ("ALFALFA", 0.10)
        ],
        "productos": [
            {"nombre": "Nicosulfuron 40%", "registro": "23567", "dosis": 1.2, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "Mesotriona", "registro": "24789", "dosis": 1.0, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "Cipermetrina 10%", "registro": "18234", "dosis": 0.4, "unidad": "l/ha", "plaga": "TALADRO"},
            {"nombre": "Tebuconazol 25%", "registro": "20567", "dosis": 1.0, "unidad": "l/ha", "plaga": "ROYA"}
        ],
        "parcelas": 520,
        "aplicaciones_por_parcela": 3
    },
    
    "enrique_iglesias_agricola": {
        "id": "enrique_iglesias_agricola",
        "empresa": "Granja Agrícola Iglesias Hnos. S.L.",
        "titular": "Enrique Iglesias López",
        "nif": "11223344C",
        "direccion": "Camino de la Vega s/n, 49800 Toro",
        "telefono": "980 123 456",
        "email": "enrique@granjaiglesias.es",
        "municipios": ["301-TORO", "298-MORALES DE TORO", "312-VENIALBO", "325-VILLABUENA", "287-PELEAGONZALO", "334-SANZOLES"],
        "cultivos": [
            ("VID", 0.35),
            ("CEBADA", 0.25),
            ("TRIGO BLANDO", 0.20),
            ("ALMENDRO", 0.12),
            ("OLIVO", 0.08)
        ],
        "productos": [
            {"nombre": "Glifosato 36%", "registro": "21001", "dosis": 3.0, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "Cobre 50% WP", "registro": "16890", "dosis": 2.0, "unidad": "kg/ha", "plaga": "MILDIU"},
            {"nombre": "Azufre 80% WG", "registro": "15234", "dosis": 3.0, "unidad": "kg/ha", "plaga": "OIDIO"},
            {"nombre": "Spirotetramat", "registro": "25123", "dosis": 0.6, "unidad": "l/ha", "plaga": "ARAÑA ROJA"},
            {"nombre": "Piriproxifen", "registro": "23890", "dosis": 0.5, "unidad": "l/ha", "plaga": "COCHINILLA"}
        ],
        "parcelas": 650,
        "aplicaciones_por_parcela": 3
    }
}


# ============================================
# Generadores
# ============================================

def generar_parcelas(config: dict, seed: int = None) -> list:
    """Genera parcelas realistas"""
    if seed:
        random.seed(seed)
    
    parcelas = []
    num_parcelas = config["parcelas"]
    municipios = config["municipios"]
    cultivos = config["cultivos"]
    
    poligono_actual = 1
    parcela_actual = 1
    
    for i in range(num_parcelas):
        municipio = municipios[i % len(municipios)]
        
        # Seleccionar cultivo según peso
        rand = random.random()
        cumulative = 0
        cultivo = cultivos[0][0]
        for c, prob in cultivos:
            cumulative += prob
            if rand <= cumulative:
                cultivo = c
                break
        
        # Recintos (70% tiene 1, 20% tiene 2, 10% tiene 3)
        num_recintos = random.choices([1, 2, 3], weights=[0.7, 0.2, 0.1])[0]
        
        for recinto in range(1, num_recintos + 1):
            superficie = round(random.uniform(0.5, 25.0), 2)
            
            parcelas.append({
                "nro_orden": len(parcelas) + 1,
                "municipio": municipio,
                "polygon": poligono_actual,
                "parcel": parcela_actual,
                "recinto": recinto,
                "superficie_sigpac": round(superficie * 1.03, 2),
                "superficie_cultivada": superficie,
                "crop": cultivo
            })
        
        parcela_actual += 1
        if parcela_actual > 300:
            parcela_actual = 1
            poligono_actual += 1
    
    return parcelas


def generar_tratamientos(parcelas: list, config: dict, seed: int = None) -> list:
    """Genera tratamientos únicos (sin duplicados por diseño)"""
    if seed:
        random.seed(seed + 1000)
    
    tratamientos = []
    productos = config["productos"]
    aplicaciones = config["aplicaciones_por_parcela"]
    
    fecha_base = datetime.now()
    tratamiento_id = 0
    
    for parcela in parcelas:
        # Número aleatorio de aplicaciones
        num_aplicaciones = random.randint(1, aplicaciones)
        
        # Seleccionar productos SIN repetir para esta parcela
        productos_seleccionados = random.sample(productos, min(num_aplicaciones, len(productos)))
        
        for producto in productos_seleccionados:
            # Fecha única para evitar duplicados
            dias_atras = random.randint(0, 180)
            hora_offset = tratamiento_id % 24  # Variar hora para unicidad
            fecha = (fecha_base - timedelta(days=dias_atras, hours=hora_offset)).strftime("%d/%m/%Y")
            
            tratamientos.append({
                "id_parcelas": parcela["nro_orden"],
                "polygon": parcela["polygon"],
                "parcel": parcela["parcel"],
                "recinto": parcela["recinto"],
                "crop": parcela["crop"],
                "surface": parcela["superficie_cultivada"],
                "product": producto["nombre"],
                "registry_number": producto["registro"],
                "dose": {"value": producto["dosis"], "unit": producto["unidad"]},
                "fecha": fecha,
                "pest": producto["plaga"],
                "notes": f"Tratamiento {tratamiento_id + 1}"
            })
            tratamiento_id += 1
    
    return tratamientos


def write_safe(ws, row, col, value):
    """Escribe en celda solo si no es merged"""
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def crear_workbook_cliente(parcelas: list, config: dict, 
                           output_path: str, golden_path: str):
    """Crea el workbook del cliente basado en el Golden"""
    shutil.copy(golden_path, output_path)
    
    wb = load_workbook(output_path)
    
    # Escribir datos del titular en inf.gral 1
    if "inf.gral 1" in wb.sheetnames:
        ws = wb["inf.gral 1"]
        # Buscar celda de titular y año (varía por template)
        # Por ahora dejamos los datos originales
    
    # Escribir parcelas en 2.1. DATOS PARCELAS
    if "2.1. DATOS PARCELAS" in wb.sheetnames:
        ws = wb["2.1. DATOS PARCELAS"]
        
        # Limpiar datos (desde fila 14)
        for row in range(14, min(ws.max_row + 1, 1500)):
            for col in range(1, 15):
                try:
                    write_safe(ws, row, col, None)
                except:
                    pass
        
        # Escribir parcelas
        for i, p in enumerate(parcelas):
            row = 14 + i
            write_safe(ws, row, 2, p["nro_orden"])
            write_safe(ws, row, 4, p["municipio"])
            write_safe(ws, row, 7, p["polygon"])
            write_safe(ws, row, 8, p["parcel"])
            write_safe(ws, row, 9, p["recinto"])
            write_safe(ws, row, 11, p["superficie_sigpac"])
            write_safe(ws, row, 12, p["superficie_cultivada"])
            write_safe(ws, row, 13, p["crop"])
    
    # Limpiar inf.trat 1 (tratamientos se escriben después)
    if "inf.trat 1" in wb.sheetnames:
        ws = wb["inf.trat 1"]
        for row in range(11, 300):
            for col in range(1, 14):
                try:
                    write_safe(ws, row, col, None)
                except:
                    pass
    
    wb.save(output_path)
    wb.close()


def generar_cliente_completo(cliente_id: str, base_dir: Path, golden_path: str) -> dict:
    """Genera todos los archivos para un cliente"""
    config = CLIENTES[cliente_id]
    
    print(f"\n{'='*60}")
    print(f"  🏭 {config['empresa']}")
    print(f"     Titular: {config['titular']}")
    print(f"{'='*60}")
    
    # Seed reproducible por cliente
    seed = hash(cliente_id) % 10000
    
    # Generar datos
    parcelas = generar_parcelas(config, seed)
    tratamientos = generar_tratamientos(parcelas, config, seed)
    
    print(f"📊 Parcelas: {len(parcelas)}")
    print(f"📊 Tratamientos: {len(tratamientos)}")
    print(f"📊 Municipios: {len(config['municipios'])}")
    print(f"📊 Cultivos: {[c[0] for c in config['cultivos']]}")
    
    # Crear directorio
    client_dir = base_dir / "clientes_ficticios" / cliente_id
    client_dir.mkdir(parents=True, exist_ok=True)
    
    # Copiar golden
    golden_dest = client_dir / "golden.xlsx"
    shutil.copy(golden_path, golden_dest)
    
    # Crear input
    input_path = client_dir / "input.xlsx"
    crear_workbook_cliente(parcelas, config, str(input_path), golden_path)
    
    # Guardar tratamientos
    trat_path = client_dir / "tratamientos.json"
    with open(trat_path, "w", encoding="utf-8") as f:
        json.dump(tratamientos, f, indent=2, ensure_ascii=False)
    
    # Guardar config
    config_path = client_dir / "config.json"
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False, default=str)
    
    # Copiar workbook dictionary
    dict_src = base_dir / "src" / "workbook_dictionary.json"
    dict_dest = client_dir / "workbook_dictionary.json"
    shutil.copy(dict_src, dict_dest)
    
    print(f"✅ Archivos creados en: {client_dir}")
    
    return {
        "cliente_id": cliente_id,
        "empresa": config["empresa"],
        "titular": config["titular"],
        "parcelas": len(parcelas),
        "tratamientos": len(tratamientos),
        "input_path": str(input_path),
        "golden_path": str(golden_dest),
        "tratamientos_path": str(trat_path)
    }


def main():
    print(f"\n{'🏭'*20}")
    print(f"  GENERADOR DE CLIENTES FICTICIOS")
    print(f"  3 empresas agrícolas con datos realistas")
    print(f"{'🏭'*20}")
    
    BASE_DIR = Path(__file__).parent
    GOLDEN_PATH = "/Volumes/Uniclick4TB/organizadorhndezbueno/PABLO PEREZ RUBIO 2025 RESUELTO.XLSX"
    
    if not Path(GOLDEN_PATH).exists():
        print(f"❌ Golden no encontrado: {GOLDEN_PATH}")
        return 1
    
    resultados = []
    
    for cliente_id in CLIENTES.keys():
        resultado = generar_cliente_completo(cliente_id, BASE_DIR, GOLDEN_PATH)
        resultados.append(resultado)
    
    # Resumen
    print(f"\n{'='*60}")
    print(f"  📊 RESUMEN CLIENTES FICTICIOS")
    print(f"{'='*60}")
    
    total_parcelas = 0
    total_tratamientos = 0
    
    for r in resultados:
        print(f"\n  {r['empresa']}:")
        print(f"    Titular: {r['titular']}")
        print(f"    Parcelas: {r['parcelas']}")
        print(f"    Tratamientos: {r['tratamientos']}")
        total_parcelas += r['parcelas']
        total_tratamientos += r['tratamientos']
    
    print(f"\n  {'='*40}")
    print(f"  TOTAL: {total_parcelas:,} parcelas, {total_tratamientos:,} tratamientos")
    print(f"  {'='*40}")
    
    print(f"\n✅ 3 clientes ficticios generados.\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
