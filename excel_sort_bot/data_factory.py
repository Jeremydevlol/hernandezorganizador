#!/usr/bin/env python3
"""
🏭 DATA FACTORY — Generador de Datos de Clientes
Genera parcelas y tratamientos realistas para stress test de producción.

NO usa IA. Genera datos determinísticos para validación.
"""
import random
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Any

from openpyxl import load_workbook


# ============================================
# Configuración de Clientes
# ============================================

CLIENTES = {
    "cliente_sigpac": {
        "nombre": "Agricultor SIGPAC Clásico",
        "parcelas": 450,
        "municipios": [
            "108-COCA DE ALBA",
            "175-MACOTERA", 
            "193-RASUEROS",
            "245-VILLORIA"
        ],
        "cultivos": [
            ("CEBADA", 0.35),
            ("TRIGO BLANDO", 0.30),
            ("GIRASOL", 0.20),
            ("BARBECHO TRADICIONAL", 0.15)
        ],
        "productos": [
            {"nombre": "Glifosato", "dosis": 3.0, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "Cobre", "dosis": 2.0, "unidad": "kg/ha", "plaga": "HONGOS"},
            {"nombre": "Azufre", "dosis": 2.5, "unidad": "kg/ha", "plaga": "OIDIO"}
        ],
        "aplicaciones_por_parcela": 2
    },
    
    "cliente_hort": {
        "nombre": "Finca Hortícola Intensiva",
        "parcelas": 300,
        "municipios": [
            "ALMERÍA CENTRO",
            "NÍJAR",
            "BERJA"
        ],
        "cultivos": [
            ("TOMATE", 0.40),
            ("PIMIENTO", 0.30),
            ("LECHUGA", 0.20),
            ("PEPINO", 0.10)
        ],
        "productos": [
            {"nombre": "Spinosad", "dosis": 0.3, "unidad": "l/ha", "plaga": "TRIPS"},
            {"nombre": "Bacillus thuringiensis", "dosis": 1.0, "unidad": "kg/ha", "plaga": "ORUGAS"},
            {"nombre": "Cobre", "dosis": 2.0, "unidad": "kg/ha", "plaga": "MILDIU"},
            {"nombre": "Azoxistrobin", "dosis": 1.5, "unidad": "l/ha", "plaga": "BOTRYTIS"}
        ],
        "aplicaciones_por_parcela": 3
    },
    
    "cliente_coop": {
        "nombre": "Cooperativa Multi-Explotación",
        "parcelas": 800,
        "municipios": [
            "TOLEDO NORTE",
            "TOLEDO SUR", 
            "TALAVERA",
            "ARANJUEZ",
            "OCAÑA",
            "MADRIDEJOS",
            "CONSUEGRA",
            "MORA"
        ],
        "cultivos": [
            ("CEBADA", 0.25),
            ("TRIGO BLANDO", 0.20),
            ("GIRASOL", 0.15),
            ("VID", 0.15),
            ("OLIVO", 0.10),
            ("ALMENDRO", 0.10),
            ("BARBECHO", 0.05)
        ],
        "productos": [
            {"nombre": "Glifosato", "dosis": 3.0, "unidad": "l/ha", "plaga": "MALAS HIERBAS"},
            {"nombre": "Cobre", "dosis": 2.0, "unidad": "kg/ha", "plaga": "HONGOS"},
            {"nombre": "Azufre", "dosis": 2.5, "unidad": "kg/ha", "plaga": "OIDIO"},
            {"nombre": "Deltametrina", "dosis": 0.5, "unidad": "l/ha", "plaga": "INSECTOS"},
            {"nombre": "Mancozeb", "dosis": 2.0, "unidad": "kg/ha", "plaga": "MILDIU"}
        ],
        "aplicaciones_por_parcela": 3
    }
}


# ============================================
# Generador de Parcelas
# ============================================

def generar_parcelas(config: Dict) -> List[Dict]:
    """Genera parcelas realistas para un cliente"""
    parcelas = []
    
    num_parcelas = config["parcelas"]
    municipios = config["municipios"]
    cultivos = config["cultivos"]
    
    # Generar polígonos y parcelas
    poligono_actual = 1
    parcela_actual = 1
    
    for i in range(num_parcelas):
        # Seleccionar municipio (distribución uniforme)
        municipio = municipios[i % len(municipios)]
        
        # Seleccionar cultivo (según probabilidades)
        rand = random.random()
        cumulative = 0
        cultivo = cultivos[0][0]
        for c, prob in cultivos:
            cumulative += prob
            if rand <= cumulative:
                cultivo = c
                break
        
        # Generar recinto (1-3 recintos por parcela)
        num_recintos = random.choices([1, 2, 3], weights=[0.7, 0.2, 0.1])[0]
        
        for recinto in range(1, num_recintos + 1):
            # Superficie realista (0.1 - 50 ha)
            superficie = round(random.uniform(0.1, 30.0), 2)
            
            parcelas.append({
                "nro_orden": len(parcelas) + 1,
                "municipio": municipio,
                "polygon": poligono_actual,
                "parcel": parcela_actual,
                "recinto": recinto,
                "superficie_sigpac": round(superficie * 1.05, 2),
                "superficie_cultivada": superficie,
                "crop": cultivo
            })
        
        # Avanzar polígono/parcela
        parcela_actual += 1
        if parcela_actual > 200:
            parcela_actual = 1
            poligono_actual += 1
    
    return parcelas


def generar_tratamientos(parcelas: List[Dict], config: Dict) -> List[Dict]:
    """Genera tratamientos para las parcelas"""
    tratamientos = []
    productos = config["productos"]
    aplicaciones = config["aplicaciones_por_parcela"]
    
    # Fechas de tratamiento (últimos 6 meses)
    fecha_base = datetime.now()
    
    for parcela in parcelas:
        # Seleccionar productos para esta parcela
        num_aplicaciones = random.randint(1, aplicaciones)
        productos_seleccionados = random.sample(productos, min(num_aplicaciones, len(productos)))
        
        for producto in productos_seleccionados:
            # Fecha aleatoria en los últimos 6 meses
            dias_atras = random.randint(0, 180)
            fecha = (fecha_base - timedelta(days=dias_atras)).strftime("%d/%m/%Y")
            
            tratamientos.append({
                "id_parcelas": parcela["nro_orden"],
                "polygon": parcela["polygon"],
                "parcel": parcela["parcel"],
                "recinto": parcela["recinto"],
                "crop": parcela["crop"],
                "surface": parcela["superficie_cultivada"],
                "product": producto["nombre"],
                "dose": {"value": producto["dosis"], "unit": producto["unidad"]},
                "fecha": fecha,
                "pest": producto["plaga"],
                "notes": f"Aplicación {producto['nombre']}"
            })
    
    return tratamientos


# ============================================
# Escritor de Excel (Input simulado)
# ============================================

def crear_input_excel(parcelas: List[Dict], output_path: str, golden_path: str):
    """Crea un Excel de input con las parcelas generadas"""
    from openpyxl.cell.cell import MergedCell
    
    def write_safe(ws, row, col, value):
        """Escribe en celda solo si no es merged"""
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            cell.value = value
    
    # Copiar el golden como base
    shutil.copy(golden_path, output_path)
    
    wb = load_workbook(output_path)
    
    # Escribir parcelas en 2.1. DATOS PARCELAS
    if "2.1. DATOS PARCELAS" in wb.sheetnames:
        ws = wb["2.1. DATOS PARCELAS"]
        
        # Limpiar datos existentes (desde fila 14)
        for row in range(14, min(ws.max_row + 1, 1000)):
            for col in range(1, 15):
                try:
                    write_safe(ws, row, col, None)
                except:
                    pass
        
        # Escribir nuevas parcelas
        for i, p in enumerate(parcelas):
            row = 14 + i
            write_safe(ws, row, 2, p["nro_orden"])
            write_safe(ws, row, 4, p["municipio"])
            write_safe(ws, row, 7, p["polygon"])
            write_safe(ws, row, 8, p["parcel"])
            write_safe(ws, row, 9, p["recinto"])
            write_safe(ws, row, 11, p["superficie_sigpac"])
            write_safe(ws, row, 12, p["superficie_cultivada"])
            write_safe(ws, row, 13, p["crop"])
    
    # Limpiar inf.trat 1 (solo headers)
    if "inf.trat 1" in wb.sheetnames:
        ws = wb["inf.trat 1"]
        for row in range(11, 185):  # Antes del footer
            for col in range(1, 14):
                try:
                    write_safe(ws, row, col, None)
                except:
                    pass
    
    wb.save(output_path)
    wb.close()


# ============================================
# Main
# ============================================

def generar_cliente(cliente_id: str, base_dir: Path, golden_path: str):
    """Genera todos los datos para un cliente"""
    config = CLIENTES[cliente_id]
    
    print(f"\n{'='*60}")
    print(f"  🏭 Generando: {config['nombre']}")
    print(f"{'='*60}")
    
    # 1. Generar parcelas
    random.seed(42 + hash(cliente_id) % 1000)  # Reproducible
    parcelas = generar_parcelas(config)
    print(f"📊 Parcelas generadas: {len(parcelas)}")
    
    # 2. Generar tratamientos
    tratamientos = generar_tratamientos(parcelas, config)
    print(f"📊 Tratamientos generados: {len(tratamientos)}")
    
    # 3. Crear directorio
    client_dir = base_dir / "clients" / cliente_id
    client_dir.mkdir(parents=True, exist_ok=True)
    
    # 4. Copiar golden
    golden_dest = client_dir / "golden.xlsx"
    shutil.copy(golden_path, golden_dest)
    print(f"✅ Golden copiado: {golden_dest.name}")
    
    # 5. Crear input
    input_path = client_dir / "input.xlsx"
    crear_input_excel(parcelas, str(input_path), golden_path)
    print(f"✅ Input creado: {input_path.name}")
    
    # 6. Guardar tratamientos como JSON
    import json
    tratamientos_path = client_dir / "tratamientos.json"
    with open(tratamientos_path, "w", encoding="utf-8") as f:
        json.dump(tratamientos, f, indent=2, ensure_ascii=False)
    print(f"✅ Tratamientos guardados: {tratamientos_path.name}")
    
    # 7. Copiar workbook dictionary
    dict_src = base_dir / "src" / "workbook_dictionary.json"
    dict_dest = client_dir / "workbook_dictionary.json"
    shutil.copy(dict_src, dict_dest)
    print(f"✅ Dictionary copiado: {dict_dest.name}")
    
    return {
        "cliente_id": cliente_id,
        "nombre": config["nombre"],
        "parcelas": len(parcelas),
        "tratamientos": len(tratamientos),
        "input_path": str(input_path),
        "golden_path": str(golden_dest),
        "tratamientos_path": str(tratamientos_path)
    }


def main():
    print(f"\n{'🏭'*20}")
    print(f"  DATA FACTORY — Generador de Clientes para Stress Test")
    print(f"{'🏭'*20}\n")
    
    BASE_DIR = Path(__file__).parent
    GOLDEN_PATH = "/Volumes/Uniclick4TB/organizadorhndezbueno/PABLO PEREZ RUBIO 2025 RESUELTO.XLSX"
    
    if not Path(GOLDEN_PATH).exists():
        print(f"❌ Golden no encontrado: {GOLDEN_PATH}")
        return 1
    
    resultados = []
    
    for cliente_id in CLIENTES.keys():
        resultado = generar_cliente(cliente_id, BASE_DIR, GOLDEN_PATH)
        resultados.append(resultado)
    
    # Resumen
    print(f"\n{'='*60}")
    print(f"  📊 RESUMEN")
    print(f"{'='*60}")
    
    total_parcelas = 0
    total_tratamientos = 0
    
    for r in resultados:
        print(f"\n  {r['nombre']}:")
        print(f"    Parcelas: {r['parcelas']}")
        print(f"    Tratamientos: {r['tratamientos']}")
        total_parcelas += r['parcelas']
        total_tratamientos += r['tratamientos']
    
    print(f"\n  {'='*40}")
    print(f"  TOTAL: {total_parcelas:,} parcelas, {total_tratamientos:,} tratamientos")
    print(f"  {'='*40}")
    
    print(f"\n✅ Data factory completado. Listo para stress test.\n")
    return 0


if __name__ == "__main__":
    exit(main())
