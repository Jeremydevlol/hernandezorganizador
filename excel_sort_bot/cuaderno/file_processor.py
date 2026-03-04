"""
CUADERNO FILE PROCESSOR v2.0
Procesa archivos subidos (Excel, PDF, imágenes) usando GPT-5.2 Vision
Extrae datos de cuadernos de explotación agrícola automáticamente
con edición en tiempo real
"""

import os
import io
import base64
import json
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
import tempfile

# Ruta al workbook_dictionary (columnas y filas predeterminadas SIGPAC)
_WORKBOOK_DICT_PATH = Path(__file__).resolve().parent.parent / "src" / "workbook_dictionary.json"
_MAPPING_PATH = Path(__file__).resolve().parent.parent / "config" / "mapping.json"

# Hojas tipo formulario (inf.gral 1, inf.gral 2): se extraen como Campo | Valor ordenado
INF_GRAL_SHEET_NAMES = ("inf.gral 1", "inf.gral 2", "inf.gral")

# Mapeo celda -> etiqueta para inf.gral 1 (estructura RTE)
INF_GRAL_1_CELLS = [
    ("E6", "Fecha apertura"), ("M6", "Año"),
    ("D9", "Nombre titular"), ("L9", "NIF/CIF"), ("L10", "Registro autonómico"),
    ("B11", "Dirección"), ("G11", "Localidad"), ("J11", "Código postal"), ("M11", "Provincia"),
    ("B12", "Teléfono fijo"), ("F12", "Teléfono móvil"), ("I12", "Email"),
]

# inf.gral 2: secciones con celdas (formato RTE)
INF_GRAL_2_HEADER = [("C1", "Explotación / Titular"), ("I1", "Año")]
INF_GRAL_2_PERSONAS = [  # fila 10
    ("A10", "Nº orden"), ("B10", "Nombre"), ("C10", "NIF"), ("D10", "Nº ROPO"),
    ("E10", "Básico"), ("F10", "Cualificado"), ("G10", "Fumigador"), ("H10", "Piloto"), ("I10", "Asesor"),
]
INF_GRAL_2_EQUIPOS = [  # fila 17
    ("A17", "Nº orden"), ("B17", "Descripción equipo"), ("D17", "Nº ROMA"),
    ("E17", "Fecha adquisición"), ("G17", "Fecha inspección"),
]
INF_GRAL_2_ASESOR = [  # fila 24
    ("A24", "Nombre o razón social"), ("C24", "NIF"), ("E24", "Nº identificación"), ("G24", "Tipo explotación"),
]

# Excel processing
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PDF processing
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

# Image processing
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# OpenAI
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False


class FileProcessor:
    """Procesa archivos y extrae datos de cuadernos agrícolas usando GPT-5.2 Vision"""
    
    # Modelo GPT a usar (GPT-5.2 - el más avanzado)
    GPT_MODEL = "gpt-5.2"
    
    # Tipos de archivo soportados
    SUPPORTED_EXTENSIONS = {
        'excel': ['.xlsx', '.xls', '.xlsm'],
        'pdf': ['.pdf'],
        'image': ['.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp'],
        'csv': ['.csv']
    }
    
    # Prompt mejorado para GPT-5.2
    EXTRACTION_PROMPT = """Eres un experto en cuadernos de explotación agrícola. Analiza el documento/archivo subido y extrae TODOS los datos de forma estructurada.

DEBES extraer la información en formato JSON estructurado EXACTO:

{
  "tipo_documento": "cuaderno_explotacion|albarán|factura|inventory|otro",
  "datos_explotacion": {
    "nombre": "nombre de la finca/explotación (OBLIGATORIO si existe)",
    "titular": "nombre del titular",
    "nif": "NIF/CIF del titular",
    "direccion": "dirección completa",
    "codigo_explotacion": "código REGA, SIGPAC u otro identificador"
  },
  "parcelas": [
    {
      "nombre": "nombre o número de parcela",
      "referencia_catastral": "polígono-parcela o ref completa",
      "superficie_ha": 0.0,
      "cultivo": "tipo de cultivo",
      "variedad": "variedad específica",
      "municipio": "municipio",
      "provincia": "provincia"
    }
  ],
  "productos": [
    {
      "nombre_comercial": "nombre comercial del producto",
      "numero_registro": "nº registro fitosanitario (ES-XXXXX)",
      "materia_activa": "sustancia activa principal",
      "formulacion": "tipo de formulación",
      "lote": "número de lote",
      "cantidad": 0.0,
      "unidad": "L|Kg|g",
      "fecha_adquisicion": "YYYY-MM-DD",
      "proveedor": "nombre del proveedor"
    }
  ],
  "tratamientos": [
    {
      "fecha": "YYYY-MM-DD",
      "parcela": "nombre de la parcela tratada",
      "producto": "nombre del producto aplicado",
      "dosis": 0.0,
      "unidad_dosis": "L/Ha|Kg/Ha|cc/L|g/L",
      "plaga_enfermedad": "problema tratado",
      "metodo_aplicacion": "pulverización|goteo|etc",
      "operador": "nombre del aplicador"
    }
  ]
}

INSTRUCCIONES CRÍTICAS:
1. Extrae TODOS los datos visibles, no omitas ninguna fila o columna
2. Convierte fechas a formato YYYY-MM-DD
3. Números con punto decimal (ej: 15.5, no 15,5)
4. Si un Excel tiene múltiples hojas, procesa TODAS
5. Para imágenes/PDFs, usa tu capacidad OCR para leer TODO el texto
6. Mantén la estructura exacta del JSON
7. Si no hay datos de algún campo, omítelo del JSON
8. Los arrays pueden tener múltiples elementos

Responde SOLO con el JSON, sin markdown, sin explicaciones."""

    def __init__(self, openai_api_key: Optional[str] = None):
        self.api_key = openai_api_key or os.getenv("OPENAI_API_KEY")
        self.client = None
        if OPENAI_AVAILABLE and self.api_key:
            self.client = OpenAI(api_key=self.api_key)
    
    def get_file_type(self, filename: str) -> Optional[str]:
        """Determina el tipo de archivo por su extensión"""
        ext = Path(filename).suffix.lower()
        for file_type, extensions in self.SUPPORTED_EXTENSIONS.items():
            if ext in extensions:
                return file_type
        return None
    
    def is_supported(self, filename: str) -> bool:
        """Verifica si el archivo está soportado"""
        return self.get_file_type(filename) is not None
    
    async def process_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Procesa un archivo y extrae datos estructurados"""
        file_type = self.get_file_type(filename)
        
        if not file_type:
            return {
                "success": False,
                "error": f"Tipo de archivo no soportado: {filename}",
                "supported_types": list(self.SUPPORTED_EXTENSIONS.keys())
            }
        
        try:
            if file_type == 'excel':
                return await self._process_excel(file_content, filename)
            elif file_type == 'pdf':
                return await self._process_pdf(file_content, filename)
            elif file_type == 'image':
                return await self._process_image(file_content, filename)
            elif file_type == 'csv':
                return await self._process_csv(file_content, filename)
            else:
                return {"success": False, "error": "Tipo no implementado"}
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "file_type": file_type
            }
    
    def _load_workbook_dict(self) -> Dict[str, Any]:
        """Carga el diccionario de hojas predeterminadas SIGPAC"""
        if _WORKBOOK_DICT_PATH.exists():
            try:
                with open(_WORKBOOK_DICT_PATH, "r", encoding="utf-8") as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError):
                pass
        return {}

    def _format_column_label(self, key: str) -> str:
        """Convierte 'nro_orden' -> 'Nro orden', 'codigo_provincia' -> 'Codigo provincia'"""
        return key.replace("_", " ").title()

    def _apply_sheet_config(
        self,
        ws,
        sheet_name: str,
        config: Dict[str, Any],
        doc_name: str,
    ) -> Optional[Dict[str, Any]]:
        """
        Aplica la configuración del workbook_dictionary: títulos oficiales en fila 1,
        datos desde data_start_row. Devuelve la estructura de la hoja o None si no aplica.
        """
        if config.get("type") != "table" or "columns" not in config:
            return None
        columns_cfg = config.get("columns", {})
        data_start = config.get("data_start_row", 11)
        if not columns_cfg:
            return None

        # Filas completas (1-based con Excel: row 1 = index 0)
        max_row = max(ws.max_row, data_start)
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
            rows.append([self._format_cell(c) for c in row])

        if not rows:
            return None

        # Títulos de columna desde el diccionario (fila 1 = primera fila)
        sorted_cols = sorted(columns_cfg.items(), key=lambda x: x[1])
        headers = [self._format_column_label(k) for k, _ in sorted_cols]

        # Datos desde data_start_row (sin filas de título antes)
        start_idx = max(0, data_start - 1)
        data_rows = rows[start_idx:]
        max_col = max(columns_cfg.values()) if columns_cfg else 0

        # Extraer valores de las columnas correctas
        datos = []
        for row in data_rows:
            extracted = []
            for _, col_idx in sorted_cols:
                idx = col_idx - 1
                val = row[idx] if idx < len(row) else None
                extracted.append(val)
            datos.append(extracted)

        return {
            "nombre": sheet_name,
            "columnas": headers,
            "datos": datos,
            "num_filas": len(datos),
            "num_columnas": len(headers)
        }

    def _is_header_like(self, cell: Any) -> bool:
        """True si la celda parece un nombre de columna (texto no numérico, no vacío)."""
        if cell is None:
            return False
        s = str(cell).strip()
        if len(s) < 2:
            return False
        # Evitar filas que son solo números (datos)
        try:
            float(s.replace(",", "."))
            return False
        except ValueError:
            pass
        return True

    def _find_header_row(self, rows: List[List[Any]], max_scan: int = 12) -> int:
        """
        Busca la fila que más parece encabezado en las primeras max_scan filas.
        En cuadernos oficiales la fila 0 suele ser título; la 2–4 suele tener los nombres de columna.
        """
        if not rows:
            return 0
        best_idx = 0
        best_count = 0
        for idx in range(min(max_scan, len(rows))):
            row = rows[idx]
            count = sum(1 for c in row if self._is_header_like(c))
            if count > best_count:
                best_count = count
                best_idx = idx
        return best_idx

    async def _process_excel(self, content: bytes, filename: str) -> Dict[str, Any]:
        """Procesa archivo Excel extrayendo datos de todas las hojas"""
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "openpyxl no instalado"}
        
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
            workbook_dict = self._load_workbook_dict()
            sheets_config = workbook_dict.get("sheets", {})
            
            sheets_data = []
            all_data = []
            
            doc_name = Path(filename).stem or "Documento"
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Hojas inf.gral 1 / inf.gral 2: extraer como formulario ordenado (Campo | Valor)
                form_extracted = self._extract_form_sheet(ws, sheet_name)
                if form_extracted:
                    sheets_data.append(form_extracted)
                    for row in form_extracted.get("datos", []):
                        if len(row) >= 2:
                            all_data.append({"sheet": sheet_name, "Campo": row[0], "Valor": row[1]})
                    continue

                # Si la hoja coincide con el workbook_dictionary, usar títulos oficiales y fila 1
                # Matching exacto primero, luego por prefijo (e.g. "2.1. DATOS PARCELAS (2)")
                sheet_config = sheets_config.get(sheet_name)
                if not sheet_config:
                    for cfg_name, cfg_val in sheets_config.items():
                        if sheet_name.startswith(cfg_name) and cfg_val.get("type") == "table":
                            sheet_config = cfg_val
                            break
                if sheet_config:
                    parsed = self._apply_sheet_config(ws, sheet_name, sheet_config, doc_name)
                    if parsed:
                        sheets_data.append(parsed)
                        headers = parsed["columnas"]
                        datos = parsed["datos"]
                        if headers and datos:
                            for row in datos:
                                row_dict = {}
                                for i, header in enumerate(headers):
                                    row_dict[header] = row[i] if i < len(row) else None
                                all_data.append({"sheet": sheet_name, **row_dict})
                        continue
                
                # Fallback: extraer filas (solo con contenido)
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows.append([self._format_cell(cell) for cell in row])
                
                if not rows:
                    sheets_data.append({
                        "nombre": sheet_name,
                        "columnas": [],
                        "datos": [],
                        "num_filas": 0,
                        "num_columnas": 0
                    })
                    continue

                # Detectar header y usar nombres detectados (legacy)
                header_idx = self._find_header_row(rows)
                header_row = rows[header_idx]
                max_cols = max(len(r) for r in rows)
                headers = []
                for i in range(max_cols):
                    val = header_row[i] if i < len(header_row) else None
                    if val is not None and str(val).strip():
                        headers.append(str(val).strip())
                    else:
                        headers.append(f"{doc_name} - Columna {i + 1}")
                # Datos: solo filas DESPUÉS del header (no incluir filas de título antes)
                datos = rows[header_idx + 1:]
                
                sheets_data.append({
                    "nombre": sheet_name,
                    "columnas": headers,
                    "datos": datos,
                    "num_filas": len(datos),
                    "num_columnas": len(headers)
                })
                    
                if headers and datos:
                    for row in datos:
                        row_dict = {}
                        for i, header in enumerate(headers):
                            row_dict[header] = row[i] if i < len(row) else None
                        all_data.append({"sheet": sheet_name, **row_dict})
            
            result = {
                "success": True,
                "file_type": "excel",
                "filename": filename,
                "hojas": sheets_data,
                "total_hojas": len(sheets_data),
                "datos_combinados": all_data
            }
            
            # Si hay OpenAI disponible, usar GPT para análisis inteligente
            if self.client and sheets_data:
                analysis = await self._analyze_with_gpt(
                    json.dumps(sheets_data[:3], ensure_ascii=False, indent=2),  # Limitar a 3 hojas
                    is_image=False
                )
                if analysis:
                    result["analisis_ia"] = analysis
            
            return result
            
        except Exception as e:
            return {"success": False, "error": f"Error procesando Excel: {str(e)}"}
    
    async def _process_pdf(self, content: bytes, filename: str) -> Dict[str, Any]:
        """
        Procesa PDF extrayendo tablas como hojas editables (formato 'hojas' estándar)
        y texto como contenido adicional. Usa PyMuPDF find_tables() para extracción
        estructurada de tablas.
        """
        if not PYMUPDF_AVAILABLE:
            return await self._process_pdf_as_images(content, filename)
        
        try:
            doc = fitz.open(stream=content, filetype="pdf")
            doc_name = Path(filename).stem or "PDF"
            
            sheets_data = []       # Hojas editables (formato estándar)
            pages_data = []        # Info por página (texto, metadatos)
            images_for_ocr = []    # Páginas escaneadas que necesitan OCR
            all_data = []          # Datos combinados para análisis IA
            table_counter = 0      # Contador global de tablas
            
            for page_num, page in enumerate(doc):
                page_label = f"Pag. {page_num + 1}"
                text = page.get_text()
                
                # --- Extraer tablas de la página ---
                tables_found = []
                try:
                    tab_finder = page.find_tables()
                    if tab_finder and tab_finder.tables:
                        for table in tab_finder.tables:
                            table_counter += 1
                            raw_data = table.extract()
                            if not raw_data:
                                continue
                            
                            # Limpiar celdas: None -> "", strip strings
                            cleaned = []
                            for row in raw_data:
                                cleaned.append([
                                    str(c).strip() if c is not None else ""
                                    for c in row
                                ])
                            
                            if not cleaned:
                                continue
                            
                            # Detectar cabecera: primera fila con texto
                            header_row = cleaned[0]
                            data_rows = cleaned[1:]
                            
                            # Si la primera fila está vacía, buscar la siguiente con datos
                            header_idx = 0
                            for idx, row in enumerate(cleaned):
                                non_empty = sum(1 for c in row if c.strip())
                                if non_empty >= 2:  # Al menos 2 celdas con datos
                                    header_row = row
                                    header_idx = idx
                                    data_rows = cleaned[idx + 1:]
                                    break
                            
                            # Construir nombres de columna
                            num_cols = len(header_row)
                            headers = []
                            for i, h in enumerate(header_row):
                                if h.strip():
                                    # Limpiar saltos de línea en cabeceras
                                    clean_h = " ".join(h.split())
                                    headers.append(clean_h)
                                else:
                                    headers.append(f"Columna {i + 1}")
                            
                            # Filtrar filas completamente vacías
                            valid_rows = [
                                row for row in data_rows
                                if any(c.strip() for c in row)
                            ]
                            
                            if not valid_rows and not any(h.strip() for h in headers):
                                continue  # Tabla totalmente vacía
                            
                            # Normalizar: asegurar que todas las filas tengan el mismo nro de columnas
                            for row in valid_rows:
                                while len(row) < num_cols:
                                    row.append("")
                                while len(row) > num_cols:
                                    row.pop()
                            
                            sheet_name = f"{page_label} - Tabla {table_counter}"
                            tables_found.append({
                                "nombre": sheet_name,
                                "columnas": headers,
                                "datos": valid_rows,
                                "num_filas": len(valid_rows),
                                "num_columnas": num_cols,
                            })
                            
                            # Añadir a datos combinados
                            for row in valid_rows:
                                row_dict = {"_pagina": page_num + 1, "_tabla": table_counter}
                                for i, h in enumerate(headers):
                                    row_dict[h] = row[i] if i < len(row) else ""
                                all_data.append(row_dict)
                except Exception:
                    pass  # find_tables puede fallar en páginas sin tablas
                
                # --- Si hay tablas, añadirlas como hojas ---
                sheets_data.extend(tables_found)
                
                # --- Texto de la página (para páginas sin tablas) ---
                has_tables = len(tables_found) > 0
                text_content = text.strip()
                
                # Si la página tiene poco texto y no tiene tablas -> necesita OCR
                needs_ocr = len(text_content) < 50 and not has_tables
                
                if needs_ocr:
                    try:
                        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                        img_bytes = pix.tobytes("png")
                        img_base64 = base64.b64encode(img_bytes).decode()
                        images_for_ocr.append({
                            "page": page_num + 1,
                            "image_base64": img_base64
                        })
                    except Exception:
                        pass
                
                # Si la página tiene texto pero no tablas, crear una hoja de texto
                if text_content and not has_tables and len(text_content) > 20:
                    # Convertir texto en filas (una por línea)
                    text_lines = [l.strip() for l in text_content.split("\n") if l.strip()]
                    if text_lines:
                        text_rows = [[line] for line in text_lines]
                        sheets_data.append({
                            "nombre": f"{page_label} - Texto",
                            "columnas": ["Contenido"],
                            "datos": text_rows,
                            "num_filas": len(text_rows),
                            "num_columnas": 1,
                        })
                
                pages_data.append({
                    "pagina": page_num + 1,
                    "texto": text_content,
                    "tiene_tablas": has_tables,
                    "num_tablas": len(tables_found),
                    "tiene_imagen": needs_ocr,
                })
            
            doc.close()
            
            result = {
                "success": True,
                "file_type": "pdf",
                "filename": filename,
                "total_paginas": len(pages_data),
                "total_tablas": table_counter,
                "paginas": pages_data,
                "hojas": sheets_data,             # ← Formato estándar editable
                "total_hojas": len(sheets_data),
                "datos_combinados": all_data,
            }
            
            # --- Análisis con IA (si disponible) ---
            if self.client:
                if images_for_ocr:
                    analysis = await self._analyze_with_gpt(
                        images_for_ocr[0]["image_base64"],
                        is_image=True
                    )
                    if analysis:
                        result["analisis_ia"] = analysis
                elif sheets_data:
                    # Enviar muestra de datos tabulares para análisis
                    sample = json.dumps(sheets_data[:3], ensure_ascii=False, indent=2)
                    analysis = await self._analyze_with_gpt(sample, is_image=False)
                    if analysis:
                        result["analisis_ia"] = analysis
                elif pages_data:
                    combined_text = "\n\n".join([p["texto"] for p in pages_data[:5]])
                    if combined_text.strip():
                        analysis = await self._analyze_with_gpt(combined_text, is_image=False)
                        if analysis:
                            result["analisis_ia"] = analysis
            
            return result
            
        except Exception as e:
            return {"success": False, "error": f"Error procesando PDF: {str(e)}"}
    
    async def _process_pdf_as_images(self, content: bytes, filename: str) -> Dict[str, Any]:
        """
        Fallback cuando PyMuPDF no está disponible:
        Intenta usar GPT Vision para OCR del PDF convertido a imagen.
        """
        doc_name = Path(filename).stem or "PDF"
        
        result = {
            "success": True,
            "file_type": "pdf",
            "filename": filename,
            "hojas": [],
            "total_hojas": 0,
            "warning": "PyMuPDF no disponible - instala con: pip install PyMuPDF",
        }
        
        # Si hay GPT disponible, enviar como texto crudo
        if self.client:
            try:
                # Intentar extraer texto básico sin PyMuPDF
                text = content.decode('latin-1', errors='replace')
                # Buscar texto entre streams de PDF (básico)
                import re
                text_chunks = re.findall(r'\((.*?)\)', text)
                extracted = " ".join(t for t in text_chunks if len(t) > 3 and t.isascii())
                if extracted and len(extracted) > 50:
                    analysis = await self._analyze_with_gpt(extracted[:5000], is_image=False)
                    if analysis:
                        result["analisis_ia"] = analysis
            except Exception:
                pass
        
        return result
    
    async def _process_image(self, content: bytes, filename: str) -> Dict[str, Any]:
        """Procesa imagen usando GPT-4o Vision para OCR"""
        try:
            # Convertir a base64
            img_base64 = base64.b64encode(content).decode()
            
            # Detectar tipo MIME
            ext = Path(filename).suffix.lower()
            mime_types = {
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.webp': 'image/webp',
                '.gif': 'image/gif'
            }
            mime_type = mime_types.get(ext, 'image/png')
            
            result = {
                "success": True,
                "file_type": "image",
                "filename": filename,
                "mime_type": mime_type
            }
            
            # Usar GPT Vision para OCR y extracción
            if self.client:
                analysis = await self._analyze_with_gpt(
                    img_base64,
                    is_image=True,
                    mime_type=mime_type
                )
                if analysis:
                    result["analisis_ia"] = analysis
                    result["datos_extraidos"] = analysis
            else:
                result["warning"] = "OpenAI no configurado - OCR no disponible"
            
            return result
            
        except Exception as e:
            return {"success": False, "error": f"Error procesando imagen: {str(e)}"}
    
    async def _process_csv(self, content: bytes, filename: str) -> Dict[str, Any]:
        """Procesa archivo CSV"""
        import csv
        
        try:
            # Decodificar contenido
            text = content.decode('utf-8', errors='replace')
            reader = csv.reader(io.StringIO(text))
            
            rows = list(reader)
            raw_headers = rows[0] if rows else []
            doc_name = Path(filename).stem or "Documento"
            # Rellenar headers vacíos con nombre real del documento
            headers = [
                h.strip() if h and str(h).strip() else f"{doc_name} - Columna {i + 1}"
                for i, h in enumerate(raw_headers)
            ] if raw_headers else [f"{doc_name} - Columna {i + 1}" for i in range(max(len(r) for r in rows) if rows else 0)]
            
            result = {
                "success": True,
                "file_type": "csv",
                "filename": filename,
                "hojas": [{
                    "nombre": doc_name,
                    "columnas": headers,
                    "datos": rows,
                    "num_filas": len(rows),
                    "num_columnas": len(headers)
                }]
            }
            
            # Análisis con GPT si está disponible
            if self.client and rows:
                sample = rows[:20]  # Primeras 20 filas
                analysis = await self._analyze_with_gpt(
                    json.dumps(sample, ensure_ascii=False),
                    is_image=False
                )
                if analysis:
                    result["analisis_ia"] = analysis
            
            return result
            
        except Exception as e:
            return {"success": False, "error": f"Error procesando CSV: {str(e)}"}
    
    async def _analyze_with_gpt(
        self, 
        content: str, 
        is_image: bool = False,
        mime_type: str = "image/png"
    ) -> Optional[Dict[str, Any]]:
        """Usa GPT-4o para analizar contenido y extraer datos estructurados"""
        if not self.client:
            return None
        
        try:
            messages = [{"role": "system", "content": self.EXTRACTION_PROMPT}]
            
            if is_image:
                # Usar Vision API
                messages.append({
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{content}",
                                "detail": "high"
                            }
                        },
                        {
                            "type": "text",
                            "text": "Analiza esta imagen y extrae los datos según el formato indicado."
                        }
                    ]
                })
            else:
                # Texto normal
                messages.append({
                    "role": "user",
                    "content": f"Analiza estos datos y extrae la información según el formato indicado:\n\n{content}"
                })
            
            # Llamar a GPT-5.2 (el modelo más avanzado con vision y OCR)
            response = self.client.chat.completions.create(
                model=self.GPT_MODEL,  # GPT-5.2
                messages=messages,
                max_tokens=8192,  # Mayor capacidad para archivos grandes
                temperature=0.05  # Muy baja para máxima precisión
            )
            
            result_text = response.choices[0].message.content
            
            # Intentar parsear JSON
            try:
                # Limpiar posibles markdown code blocks
                if "```json" in result_text:
                    result_text = result_text.split("```json")[1].split("```")[0]
                elif "```" in result_text:
                    result_text = result_text.split("```")[1].split("```")[0]
                
                return json.loads(result_text.strip())
            except json.JSONDecodeError:
                # Devolver como texto si no es JSON válido
                return {"raw_response": result_text}
                
        except Exception as e:
            return {"error": str(e)}
    
    def _cell_ref_to_coord(self, cell_ref: str) -> Tuple[int, int]:
        """Convierte 'E6' -> (6, 5), 'B11' -> (11, 2). openpyxl usa 1-based."""
        col_str = ""
        row_str = ""
        for c in cell_ref.upper():
            if c.isalpha():
                col_str += c
            else:
                row_str += c
        col = 0
        for c in col_str:
            col = col * 26 + (ord(c) - ord("A") + 1)
        row = int(row_str) if row_str else 1
        return (row, col)

    def _extract_form_sheet(self, ws, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Extrae hojas tipo formulario (inf.gral 1, inf.gral 2) como tabla ordenada Campo | Valor.
        Así se ven bien en el editor en lugar del grid desordenado original.
        """
        sheet_lower = sheet_name.lower().strip()
        is_inf_gral_1 = "inf.gral 1" in sheet_lower or (sheet_lower.startswith("inf.gral") and "2" not in sheet_lower)
        is_inf_gral_2 = "inf.gral 2" in sheet_lower

        if is_inf_gral_1:
            # inf.gral 1: mapeo conocido de celdas
            pairs = []
            for cell_ref, label in INF_GRAL_1_CELLS:
                try:
                    row, col = self._cell_ref_to_coord(cell_ref)
                    val = ws.cell(row=row, column=col).value
                    val_fmt = self._format_cell(val)
                    if val_fmt is not None and str(val_fmt).strip():
                        pairs.append([label, str(val_fmt).strip()])
                except Exception:
                    pass
            if pairs:
                return {
                    "nombre": sheet_name,
                    "columnas": ["Campo", "Valor"],
                    "datos": pairs,
                    "num_filas": len(pairs),
                    "num_columnas": 2,
                }

        if is_inf_gral_2:
            return self._extract_inf_gral_2(ws, sheet_name)

        # Cualquier hoja que empiece por inf.gral: intentar escaneo genérico
        if "inf.gral" in sheet_lower:
            return self._extract_form_by_scan(ws, sheet_name)
        return None

    def _extract_inf_gral_2(self, ws, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Extrae inf.gral 2 como tabla ordenada Campo | Valor.
        Secciones: 1) Cabecera | 2) Personas | 3) Equipos | 4) Asesor
        """
        pairs: List[List[str]] = []

        def _add_cell(cell_ref: str, label: str) -> None:
            row, col = self._cell_ref_to_coord(cell_ref)
            val = ws.cell(row=row, column=col).value
            val_fmt = self._format_cell(val)
            if val_fmt is not None and str(val_fmt).strip():
                pairs.append([label, str(val_fmt).strip()])

        for cell_ref, label in INF_GRAL_2_HEADER:
            _add_cell(cell_ref, label)

        pairs.append(["--- 1.2 PERSONAS ---", ""])
        # Personas (1.2): fila 10
        for cell_ref, label in INF_GRAL_2_PERSONAS:
            row, col = self._cell_ref_to_coord(cell_ref)
            val = ws.cell(row=row, column=col).value
            val_fmt = self._format_cell(val)
            if val_fmt is not None and str(val_fmt).strip():
                pairs.append([label, str(val_fmt).strip()])
        # Más filas de personas (11, 12, 13, 14)
        for extra_row in range(11, 15):
            a_val = ws.cell(row=extra_row, column=1).value
            if a_val is None or not str(a_val).strip():
                continue
            pairs.append([f"--- Persona {extra_row - 9} ---", ""])
            for cell_ref, label in INF_GRAL_2_PERSONAS:
                _, c = self._cell_ref_to_coord(cell_ref)
                val = ws.cell(row=extra_row, column=c).value
                val_fmt = self._format_cell(val)
                if val_fmt is not None and str(val_fmt).strip():
                    pairs.append([label, str(val_fmt).strip()])

        pairs.append(["--- 1.3 EQUIPOS ---", ""])
        # Equipos (1.3): fila 17 + posibles filas 18+
        for cell_ref, label in INF_GRAL_2_EQUIPOS:
            row, col = self._cell_ref_to_coord(cell_ref)
            val = ws.cell(row=row, column=col).value
            val_fmt = self._format_cell(val)
            if val_fmt is not None and str(val_fmt).strip():
                pairs.append([label, str(val_fmt).strip()])
        for extra_row in range(18, 22):
            a_val = ws.cell(row=extra_row, column=1).value
            if a_val is None or not str(a_val).strip():
                continue
            pairs.append([f"--- Equipo {extra_row - 16} ---", ""])
            for cell_ref, label in INF_GRAL_2_EQUIPOS:
                _, c = self._cell_ref_to_coord(cell_ref)
                val = ws.cell(row=extra_row, column=c).value
                val_fmt = self._format_cell(val)
                if val_fmt is not None and str(val_fmt).strip():
                    pairs.append([label, str(val_fmt).strip()])

        pairs.append(["--- 1.4 ASESOR ---", ""])
        # Asesor (1.4): fila 24
        for cell_ref, label in INF_GRAL_2_ASESOR:
            _add_cell(cell_ref, label)

        if pairs:
            return {
                "nombre": sheet_name,
                "columnas": ["Campo", "Valor"],
                "datos": pairs,
                "num_filas": len(pairs),
                "num_columnas": 2,
            }
        return self._extract_form_by_scan(ws, sheet_name)

    def _extract_form_by_scan(self, ws, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Escanea la hoja buscando patrones label: valor.
        Labels: texto que termina en : o coincide con etiquetas conocidas.
        Valor: celda a la derecha o en la misma celda tras ':'.
        """
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([self._format_cell(c) for c in row])

        pairs = []
        seen_labels = set()
        label_patterns = (
            "dirección", "direccion", "teléfono", "telefono", "titular", "nif", "nombre",
            "localidad", "c. postal", "codigo postal", "provincia", "e-mail", "email",
            "fecha", "año", "ano", "registro", "representante", "firma", "fecha"
        )

        for r_idx, row in enumerate(rows):
            for c_idx, cell in enumerate(row):
                if cell is None or not str(cell).strip():
                    continue
                s = str(cell).strip()
                if len(s) < 3:
                    continue
                s_lower = s.lower()

                # Caso 1: "Label: Valor" en la misma celda
                if ":" in s and not s.strip().endswith(":"):
                    parts = s.split(":", 1)
                    label = parts[0].strip()
                    value = parts[1].strip() if len(parts) > 1 else ""
                    if label and len(label) > 1 and label.lower() not in seen_labels:
                        seen_labels.add(label.lower())
                        if value or any(x in label.lower() for x in label_patterns):
                            pairs.append([label, value or "-"])

                # Caso 2: Celda parece label (termina en : o es etiqueta conocida)
                elif s.endswith(":") or any(p in s_lower for p in label_patterns):
                    label = s.rstrip(":").strip()
                    if not label or label.lower() in seen_labels:
                        continue
                    value = ""
                    for dc in range(1, 6):
                        if c_idx + dc < len(row):
                            v = row[c_idx + dc]
                            if v is not None and str(v).strip():
                                value = str(v).strip()
                                break
                    seen_labels.add(label.lower())
                    pairs.append([label, value or "-"])

        if pairs:
            return {
                "nombre": sheet_name,
                "columnas": ["Campo", "Valor"],
                "datos": pairs,
                "num_filas": len(pairs),
                "num_columnas": 2,
            }
        return None

    def _format_cell(self, value: Any) -> Any:
        """Formatea valores de celda para JSON"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, (int, float)):
            return value
        return str(value)


# Singleton para uso global
_processor: Optional[FileProcessor] = None

def get_processor() -> FileProcessor:
    """Obtiene la instancia global del procesador"""
    global _processor
    if _processor is None:
        _processor = FileProcessor()
    return _processor
