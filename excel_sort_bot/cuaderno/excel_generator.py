"""
Generador de Excel con formato oficial del Cuaderno de Explotación Agrícola.
Replica fielmente el diseño del Ministerio de Agricultura de España:
  inf.gral 1, inf.gral 2, 2.1. DATOS PARCELAS, id.parc 2,
  inf.trat 1, inf.trat 2, inf.trat 3, inf.trat 4,
  + vista agrupada por cultivo.
"""

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Optional, Dict, Any, TYPE_CHECKING, Iterable
import re

_log = logging.getLogger(__name__)

from .pdf_generator import PDFGenerator
from .models import ProductoAplicado

if TYPE_CHECKING:
    from .models import CuadernoExplotacion, Parcela, Tratamiento

# =====================================================================
# CONSTANTS  (referencia: diseño cuaderno RESUELTO / script oficial)
# =====================================================================
HEADER_GRAY = "C0C0C0"   # cabeceras de bloque / tablas
LIGHT_GRAY = "D9D9D9"    # bandas de sección 1.1, 2.1, 3.1, etc.
DARK_HEADER = "595959"   # hojas tipo editor (Productos, importadas)
BLACK = "000000"
ZEBRA_BASE = "FFFFFF"
ZEBRA_ALT = "E8E8E8"
_PARCELAS_NCOL = 22
_PARCELAS_DATA_START = 11

_THIN_SIDE = Side(style="thin")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_MEDIUM_SIDE = Side(style="medium")
_MEDIUM_BORDER = Border(
    left=_MEDIUM_SIDE, right=_MEDIUM_SIDE, top=_MEDIUM_SIDE, bottom=_MEDIUM_SIDE
)


# =====================================================================
# HELPERS  (reutilizables para cualquier hoja)
# =====================================================================

def _sc(ws, row, col, value="", bold=False, size=9, bg=None, align_h="left",
        align_v="center", wrap=False, border=True, font_color=BLACK,
        italic=False, underline=None):
    """Set a single cell with full styling."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    fk: Dict[str, Any] = {"bold": bold, "size": size, "color": font_color, "name": "Arial"}
    if italic:
        fk["italic"] = True
    if underline:
        fk["underline"] = underline
    cell.font = Font(**fk)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if border:
        cell.border = _THIN_BORDER
    return cell


def _mc(ws, r1, c1, r2, c2, value="", bold=False, size=9, bg=None,
        align_h="center", align_v="center", wrap=False, font_color=BLACK,
        italic=False, border=True, underline=None):
    """Merge cells and style the top-left cell."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    cell.value = value
    fk: Dict[str, Any] = {"bold": bold, "size": size, "color": font_color, "name": "Arial"}
    if italic:
        fk["italic"] = True
    if underline:
        fk["underline"] = underline
    cell.font = Font(**fk)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if border:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = _THIN_BORDER
    return cell


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _set_row_heights(ws, start, end, height=15):
    for r in range(start, end + 1):
        ws.row_dimensions[r].height = height


def _unique_sheet_title(wb, base: str) -> str:
    """Evita colisión con nombres ya usados en el libro (p. ej. hoja importada 'Productos')."""
    safe = re.sub(r"[\[\]:*?/\\]", "", base)[:31].strip() or "Hoja"
    used = {ws.title for ws in wb.worksheets}
    if safe not in used:
        return safe
    n = 2
    while True:
        suffix = f" ({n})"
        cand = (safe[: 31 - len(suffix)] + suffix)[:31]
        if cand not in used:
            return cand
        n += 1


def _fmt_date(s):
    """YYYY-MM-DD -> dd/mm/yyyy."""
    if not s:
        return ""
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", str(s).strip())
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return str(s).strip()


def _safe_float(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _fmt_ha_cell(v) -> Any:
    """Superficies en ha con dos decimales como en la plantilla oficial."""
    if v is None or v == "":
        return ""
    try:
        x = float(v)
        if x == 0.0:
            return ""
        return round(x, 2)
    except (TypeError, ValueError):
        return v


def _fmt_superficie_tratada(v) -> Any:
    """Número con 2 decimales para columna Superf. tratada; vacío si no aplica."""
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        if f == 0.0:
            return ""
        return round(f, 2)
    except (TypeError, ValueError):
        return v


def _trat_parcela_group_key(t) -> str:
    """Misma lógica que api.export / editor para agrupar por parcela."""
    if t.parcela_nombres:
        return ",".join(t.parcela_nombres).lower()
    return (t.num_orden_parcelas or "").lower()


def _tratamiento_exportable(pdf: PDFGenerator, t) -> bool:
    """
    Incluye los mismos tratamientos que el PDF, más filas editadas con producto
    sin nombre pero con dosis o registro (evita filas 'vacías' en el Excel).
    """
    if pdf._tratamiento_tiene_datos(t):
        return True
    fa = (t.fecha_aplicacion or "").strip()
    if not fa:
        return False
    prods = list(t.productos or [])
    if not prods:
        return False
    for p in prods:
        if (p.nombre_comercial or "").strip():
            return True
        try:
            d = float(p.dosis) if p.dosis is not None else 0.0
        except (TypeError, ValueError):
            d = 0.0
        if d != 0.0:
            return True
        if (p.numero_registro or "").strip():
            return True
    return False


def _hoja_tiene_datos(h) -> bool:
    """Igual que PDFGenerator._hoja_tiene_datos (hojas importadas con filas no vacías)."""
    if not h.datos or not h.columnas:
        return False
    return any(
        any(c is not None and str(c).strip() for c in fila)
        for fila in h.datos
    )


def _header_row(ws, cuaderno, row, label_c1, label_c2, name_c1, name_c2,
                year_label_c1, year_label_c2, year_c):
    """Standard 'Explotación/Titular ...' + 'AÑO:' header that appears on most sheets."""
    _mc(ws, row, label_c1, row, label_c2,
        "Explotación/ Titular de la explotación:", bold=True, size=9, align_h="left")
    _mc(ws, row, name_c1, row, name_c2,
        (cuaderno.titular or cuaderno.nombre_explotacion).upper(), bold=True, size=9, align_h="left")
    _mc(ws, row, year_label_c1, row, year_label_c2,
        "AÑO:", bold=True, size=9, align_h="right")
    _sc(ws, row, year_c, str(cuaderno.año), bold=True, size=9, align_h="center")


def _page_footer(ws, row, col_label_start, col_label_end, col_num, col_suffix_start, col_suffix_end, page):
    _mc(ws, row, col_label_start, row, col_label_end,
        "Hoja nº", size=9, align_h="right", border=False)
    _sc(ws, row, col_num, str(page), bold=True, size=9, align_h="center", border=False)
    _mc(ws, row, col_suffix_start, row, col_suffix_end,
        "de la sección nº", size=9, align_h="left", border=False)


# =====================================================================
# SHEET 1: inf.gral 1
# =====================================================================

def _build_inf_gral_1(wb, cuaderno):
    ws = wb.active
    ws.title = "inf.gral 1"
    _set_col_widths(ws, [3, 8, 8, 8, 8, 8, 10, 10, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8])
    for r in range(1, 40):
        ws.row_dimensions[r].height = 15
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[10].height = 18

    name = (cuaderno.titular or cuaderno.nombre_explotacion or "").upper()
    nif = (cuaderno.nif_titular or "").upper()
    domicilio = cuaderno.domicilio or ""
    codigo = cuaderno.codigo_explotacion or ""
    año = str(cuaderno.año)
    apertura = _fmt_date(getattr(cuaderno, "fecha_apertura", "")) or ""

    _mc(ws, 2, 3, 4, 17, "1. INFORMACIÓN GENERAL", bold=True, size=14, align_h="center")

    _mc(ws, 5, 1, 5, 4, "FECHA DE APERTURA DEL CUADERNO", bold=True, size=9, align_h="left")
    _mc(ws, 5, 5, 5, 7, apertura, size=9, align_h="center")
    _mc(ws, 5, 14, 5, 16, "AÑO", bold=True, size=9, align_h="right")
    _mc(ws, 5, 17, 5, 20, año, bold=True, size=11, align_h="center")

    _mc(ws, 7, 1, 7, 20, "1.1 DATOS GENERALES DE LA EXPLOTACIÓN",
        bold=True, size=10, bg=LIGHT_GRAY, align_h="center")

    _mc(ws, 8, 1, 8, 4, "Nombre y apellidos o razón social:", bold=True, size=9, align_h="left")
    _mc(ws, 8, 5, 8, 13, name, bold=True, size=9, align_h="left")
    _sc(ws, 8, 14, "NIF:", bold=True, size=9, align_h="right")
    _mc(ws, 8, 15, 8, 20, nif, bold=True, size=9, align_h="center")

    _mc(ws, 9, 1, 9, 5, "Nº Registro de Explotaciones Nacional:", bold=True, size=9, align_h="left")
    _mc(ws, 9, 6, 9, 11, "", size=9)
    _mc(ws, 9, 12, 9, 16, "Nº Registro de Explotaciones Autonómico:",
        bold=True, size=9, align_h="center", wrap=True)
    _mc(ws, 9, 17, 9, 20, codigo, bold=True, size=9, align_h="center")

    _sc(ws, 10, 1, "Dirección:", bold=True, size=9)
    _mc(ws, 10, 2, 10, 5, domicilio, bold=True, size=9, align_h="left")
    _sc(ws, 10, 6, "Localidad:", bold=True, size=9)
    _mc(ws, 10, 7, 10, 10, "", bold=True, size=9, align_h="center")
    _sc(ws, 10, 11, "C. Postal:", bold=True, size=9)
    _mc(ws, 10, 12, 10, 13, "", bold=True, size=9, align_h="center")
    _sc(ws, 10, 14, "Provincia:", bold=True, size=9)
    _mc(ws, 10, 15, 10, 20, "", bold=True, size=9, align_h="center")

    _mc(ws, 11, 1, 11, 2, "Teléfono fijo:", bold=True, size=9, align_h="left")
    _mc(ws, 11, 3, 11, 5, "", size=9, align_h="center")
    _mc(ws, 11, 6, 11, 7, "Teléfono móvil:", bold=True, size=9, align_h="center")
    _mc(ws, 11, 8, 11, 10, "", size=9, align_h="center")
    _sc(ws, 11, 11, "e-mail:", bold=True, size=9)
    _mc(ws, 11, 12, 11, 20, "", size=9, align_h="center")

    _mc(ws, 13, 1, 13, 20, "TITULAR O REPRESENTANTE DE LA EXPLOTACIÓN",
        bold=True, size=10, align_h="center")

    _mc(ws, 14, 1, 14, 2, "Nombre y apellidos:", bold=True, size=9, align_h="left")
    _mc(ws, 14, 3, 14, 11, name, bold=True, size=9, align_h="left")
    _sc(ws, 14, 12, "NIF:", bold=True, size=9, align_h="right")
    _mc(ws, 14, 13, 14, 20, nif, bold=True, size=9, align_h="center")

    _sc(ws, 15, 1, "Dirección:", bold=True, size=9)
    _mc(ws, 15, 2, 15, 5, domicilio, bold=True, size=9, align_h="left")
    _sc(ws, 15, 6, "Localidad:", bold=True, size=9)
    _mc(ws, 15, 7, 15, 10, "", bold=True, size=9, align_h="center")
    _sc(ws, 15, 11, "C. Postal:", bold=True, size=9)
    _mc(ws, 15, 12, 15, 13, "", bold=True, size=9, align_h="center")
    _sc(ws, 15, 14, "Provincia:", bold=True, size=9)
    _mc(ws, 15, 15, 15, 20, "", bold=True, size=9, align_h="center")

    _mc(ws, 16, 1, 16, 3, "Tipo de representación:", bold=True, size=9, align_h="left")
    _mc(ws, 16, 4, 16, 7, "", size=9)
    _sc(ws, 16, 8, "Teléfono:", bold=True, size=9)
    _mc(ws, 16, 9, 16, 11, "", size=9, align_h="center")
    _sc(ws, 16, 12, "e-mail:", bold=True, size=9)
    _mc(ws, 16, 13, 16, 20, "", size=9)

    _mc(ws, 22, 12, 22, 20,
        "Firma del titular o representante de la explotación (1)",
        size=8, align_h="left", border=False)
    _mc(ws, 26, 12, 26, 14, "Fecha:", size=9, align_h="left", border=False)

    _mc(ws, 34, 1, 34, 15,
        "(1) La persona firmante se hace responsable de la veracidad de los datos "
        "consignados en el presente cuaderno de explotación",
        size=8, align_h="left", border=False, italic=True)

    _page_footer(ws, 36, 14, 16, 17, 18, 20, 1)


# =====================================================================
# SHEET 2: inf.gral 2
# =====================================================================

def _build_inf_gral_2(wb, cuaderno, tratamientos: Iterable):
    """tratamientos: mismos que van a inf.trat 1 (mismo criterio que el PDF)."""
    ws = wb.create_sheet("inf.gral 2")
    _set_col_widths(ws, [4, 12, 20, 12, 14, 8, 8, 8, 8, 10, 14])
    _set_row_heights(ws, 1, 55)

    _header_row(ws, cuaderno, 1, 1, 3, 4, 8, 9, 10, 11)

    _mc(ws, 3, 2, 5, 10, "1. INFORMACIÓN GENERAL", bold=True, size=14, align_h="center")

    # --- Section 1.2: Personas ---
    _mc(ws, 7, 1, 7, 11,
        "1.2 PERSONAS O EMPRESAS QUE INTERVIENEN EN EL TRATAMIENTO "
        "CON PRODUCTOS FITOSANITARIOS (1)",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center", wrap=True)
    ws.row_dimensions[7].height = 25

    _mc(ws, 8, 1, 9, 1, "Nº de orden", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 2, 9, 3, "Nombre y apellidos/ Empresas de servicios",
        bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 4, 9, 4, "NIF", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 8, 5, 9, 5, "Nº Inscripción ROPO", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 6, 8, 9, "Tipo de carné (2)", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 9, 6, "Básico", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 9, 7, "Cualif.", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 9, 8, "Fumig", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 9, 9, "Piloto", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 8, 10, 9, 11, "Asesor (2)", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 20

    aplicadores = []
    seen = set()
    for t in tratamientos:
        name = (t.aplicador or t.operador or "").strip()
        if name and name not in seen:
            seen.add(name)
            aplicadores.append(name)

    for i in range(6):
        r = 10 + i
        nombre = aplicadores[i] if i < len(aplicadores) else ""
        _sc(ws, r, 1, str(i + 1) if nombre else "", size=9, align_h="center")
        _mc(ws, r, 2, r, 3, nombre, size=9, align_h="left")
        for c in range(4, 10):
            _sc(ws, r, c, "", size=9)
        _mc(ws, r, 10, r, 11, "", size=9)

    # --- Section 1.3: Equipos ---
    _mc(ws, 17, 1, 17, 11,
        "1.3 EQUIPOS DE APLICACIÓN DE PRODUCTOS FITOSANITARIOS PROPIOS DE LA EXPLOTACIÓN",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center")
    ws.row_dimensions[17].height = 20

    _mc(ws, 18, 1, 19, 1, "Nº de orden", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 18, 2, 19, 5, "Descripción del equipo (3)", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 18, 6, 19, 6, "Nº Inscripción ROMA (4)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 18, 7, 19, 8, "Fecha de adquisición", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 18, 9, 19, 11, "Fecha de la ultima inspección", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    ws.row_dimensions[18].height = 20
    ws.row_dimensions[19].height = 20

    equipos = []
    seen_eq = set()
    for t in tratamientos:
        eq = (t.equipo or "").strip()
        if eq and eq not in seen_eq:
            seen_eq.add(eq)
            equipos.append(eq)

    for i in range(6):
        r = 20 + i
        desc = equipos[i] if i < len(equipos) else ""
        _sc(ws, r, 1, str(i + 1) if desc else "", size=9, align_h="center")
        _mc(ws, r, 2, r, 5, desc, size=9, align_h="left")
        _sc(ws, r, 6, "", size=9, align_h="center")
        _mc(ws, r, 7, r, 8, "", size=9)
        _mc(ws, r, 9, r, 11, "", size=9)

    # --- Section 1.4: Asesor ---
    _mc(ws, 27, 1, 27, 11,
        "1.4 ASESOR, AGRUPACIÓN O ENTIDAD DE ASESORAMIENTO A LA QUE PERTENECE LA EXPLOTACIÓN",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center", wrap=True)
    ws.row_dimensions[27].height = 25

    _mc(ws, 28, 1, 29, 4, "Nombre o razón social", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 28, 5, 29, 6, "NIF", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 28, 7, 29, 9, "Nº de identificación", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 28, 10, 29, 11, "Tipo de explotación (5)",
        bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY, underline="single")

    for r in range(30, 34):
        _mc(ws, r, 1, r, 4, "", size=9)
        _mc(ws, r, 5, r, 6, "", size=9)
        _mc(ws, r, 7, r, 9, "", size=9)
        _mc(ws, r, 10, r, 11, "", size=9)

    notes = [
        "1 - Rellenar lo que proceda.",
        "2 - Marcar con una cruz",
        "3 - Indicar el tipo de equipo o máquina, marca y modelo.",
        "4 - En equipos en los que no sea obligatoria la inscripción en el ROMA, indicar el número de referencia en el censo correspondiente, en su caso.",
        "5 - Tipo de explotación en cuanto a gestión integrada de plagas: (AE) Agricultura Ecológica, (PI) Producción Integrada, (CP) Certificación Privada, (Atrias) Agrupación de Tratamiento Integrado en Agricultura, (AS) Asistida de un asesor, (NO) Sin obligación de aplicar la GIP.",
    ]
    for i, note in enumerate(notes):
        _mc(ws, 35 + i, 1, 35 + i, 11, note, size=7, align_h="left", border=False, italic=True, wrap=True)
        ws.row_dimensions[35 + i].height = 18 if i == 4 else 13

    _page_footer(ws, 41, 7, 9, 10, 11, 11, 2)

    return aplicadores, equipos


# =====================================================================
# SHEET 3: 2.1. DATOS PARCELAS
# =====================================================================

def _parcelas_header_block(ws, cuaderno):
    """
    Cabecera alineada al diseño RESUELTO de referencia (22 columnas A–V; año en T).
    Primera fila de datos = 11 (filas 9–10 reservadas si se amplía cabecera).
    """
    nc = _PARCELAS_NCOL
    _mc(ws, 1, 1, 1, 3, "Explotación/Titular de la explotación:", bold=True, size=9, align_h="left")
    _mc(ws, 1, 4, 1, 15,
        (cuaderno.titular or cuaderno.nombre_explotacion).upper(), bold=True, size=9, align_h="left")
    _mc(ws, 1, 18, 1, 19, "AÑO", bold=True, size=9, align_h="right")
    _sc(ws, 1, 20, str(cuaderno.año), bold=True, size=9, align_h="center")

    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 8

    _mc(ws, 4, 1, 4, nc, "2. IDENTIFICACIÓN DE LAS PARCELAS",
        bold=True, size=12, bg=LIGHT_GRAY, align_h="center")
    _mc(ws, 5, 1, 5, nc, "2.1 DATOS IDENTIFICATIVOS Y AGRONÓMICOS DE LAS PARCELAS",
        bold=True, size=10, bg=LIGHT_GRAY, align_h="center")

    _mc(ws, 6, 1, 6, 1, "", bg=LIGHT_GRAY)
    _mc(ws, 6, 2, 6, 8, "REFERENCIAS SIGPAC", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 6, 9, 6, nc, "DATOS AGRONÓMICOS", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)

    cols = [
        (1, "Nº DE\nORDEN (1)"), (2, "Código\nProvincia"), (3, "Término municipal\n(código y nombre)"),
        (4, "Código\nAgregado"), (5, "Zona"), (6, "Nº de\nPolígono"),
        (7, "Nº de\nParcela"), (8, "Nº de\nRecinto"), (9, "Uso SIGPAC"),
        (10, "Superficie\nSIGPAC (ha)"), (11, "Superficie\nCultivada\n(ha)"),
        (12, "Especie/\nVariedad"), (13, "Ecoregimen/\nPráctica\n(4)"),
        (14, "Secano/\nRegadío"), (15, "Cultivo\nprincipal/\nsecundario"),
        (16, "Fecha inicio\ncultivo"), (17, "Fecha fin\ncultivo"),
        (18, "Aire libre o\nprotegido (3)"), (19, "Sistema\ngestor (4)"),
    ]
    for c_idx, label in cols:
        _mc(ws, 7, c_idx, 8, c_idx, label, bold=True, size=7, align_h="center", wrap=True, bg=HEADER_GRAY)

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 30
    ws.row_dimensions[9].height = 14
    ws.row_dimensions[10].height = 14


def _write_parcela_row(ws, r, p, zebra: bool = False):
    """Una fila de parcela; zebra alterna fondo como en la plantilla de referencia."""
    bg = ZEBRA_ALT if zebra else ZEBRA_BASE
    _sc(ws, r, 1, p.num_orden or "", size=8, align_h="center", bg=bg)
    _sc(ws, r, 2, p.codigo_provincia or p.provincia or "", size=8, align_h="center", bg=bg)
    _sc(ws, r, 3, p.termino_municipal or p.municipio or "", size=8, align_h="left", wrap=True, bg=bg)
    _sc(ws, r, 4, p.codigo_agregado or 0, size=8, align_h="center", bg=bg)
    _sc(ws, r, 5, p.zona or 0, size=8, align_h="center", bg=bg)
    _sc(ws, r, 6, p.num_poligono or "", size=8, align_h="center", bg=bg)
    _sc(ws, r, 7, p.num_parcela or "", size=8, align_h="center", bg=bg)
    _sc(ws, r, 8, p.num_recinto or "", size=8, align_h="center", bg=bg)
    _sc(ws, r, 9, p.uso_sigpac or "", size=8, align_h="center", bg=bg)
    _sc(ws, r, 10, _fmt_ha_cell(p.superficie_sigpac), size=8, align_h="center", bg=bg)
    sup_cult = _safe_float(p.superficie_cultivada) or _safe_float(p.superficie_ha)
    _sc(ws, r, 11, _fmt_ha_cell(sup_cult), size=8, align_h="center", bg=bg)
    _sc(ws, r, 12, p.especie or p.cultivo or "", size=8, align_h="left", wrap=True, bg=bg)
    _sc(ws, r, 13, p.ecoregimen or "", size=8, align_h="center", bg=bg)
    _sc(ws, r, 14, p.secano_regadio or "S", size=8, align_h="center", bg=bg)
    _sc(ws, r, 15, p.cultivo_tipo or "Principal", size=8, align_h="center", bg=bg)
    _sc(ws, r, 16, _fmt_date(p.fecha_inicio_cultivo) if p.fecha_inicio_cultivo else "", size=8, bg=bg)
    _sc(ws, r, 17, _fmt_date(p.fecha_fin_cultivo) if p.fecha_fin_cultivo else "", size=8, bg=bg)
    _sc(ws, r, 18, p.aire_libre_protegido or "", size=8, bg=bg)
    _sc(ws, r, 19, p.sistema_asesoramiento or "", size=8, bg=bg)
    for c in range(20, 23):
        _sc(ws, r, c, "", size=8, bg=bg)


def _build_datos_parcelas(wb, cuaderno, parcelas):
    ws = wb.create_sheet("2.1. DATOS PARCELAS")
    _set_col_widths(ws, [4, 5, 18, 7, 6, 7, 8, 7, 7, 8, 8, 8, 20, 10, 8, 12, 10, 10, 10, 10, 10, 10])
    for r in range(1, 100):
        ws.row_dimensions[r].height = 14

    _parcelas_header_block(ws, cuaderno)

    for i, p in enumerate(parcelas):
        _write_parcela_row(ws, _PARCELAS_DATA_START + i, p, zebra=(i % 2 == 1))


# =====================================================================
# SHEET 4: id.parc 2
# =====================================================================

def _build_id_parc_2(wb, cuaderno, parcelas):
    ws = wb.create_sheet("id.parc 2")
    _set_col_widths(ws, [8, 8, 8, 8, 12, 10, 14, 12, 8, 12, 12])
    _set_row_heights(ws, 1, 50)

    _header_row(ws, cuaderno, 1, 1, 3, 4, 7, 8, 9, 10)

    _mc(ws, 3, 2, 5, 9, "2. IDENTIFICACIÓN DE LAS PARCELAS DE LA EXPLOTACIÓN",
        bold=True, size=13, align_h="center")

    _mc(ws, 7, 1, 7, 11,
        "2.2. DATOS IDENTIFICATIVOS MEDIOAMBIENTALES DE LAS PARCELAS",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center")

    _mc(ws, 8, 1, 10, 1, "Id. parcelas (1)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 2, 8, 3, "Cultivo", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 9, 2, "Especie", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _sc(ws, 9, 3, "Variedad", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 4, 8, 7,
        "Puntos de captación de agua procedente de pozos y masas de agua\n"
        "utilizadas para consumo humano",
        bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 4, 10, 4, "Incluido en la parcela (SI/NO)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 5, 10, 5, "Distancia (m) (3)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 6, 10, 6, "Coordenadas UTM(4)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 7, 10, 7, "Denominación(5)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 8, 8, 11, "Parcelas en zonas específicas(2)", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 9, 8, 10, 9, "Totalmente (SI/ NO)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 10, 10, 11, "Parcialmente (SI/ NO)(6)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    ws.row_dimensions[8].height = 30
    ws.row_dimensions[9].height = 25

    for i, p in enumerate(parcelas):
        r = 11 + i
        _sc(ws, r, 1, p.num_orden or "", size=8, align_h="center")
        _sc(ws, r, 2, p.especie or p.cultivo or "", size=8)
        _sc(ws, r, 3, p.variedad or "", size=8)
        for c in range(4, 12):
            _sc(ws, r, c, "", size=8)

    foot_r = max(11 + len(parcelas) + 2, 26)
    notes = [
        "[1] Identificar el número de orden de identificación de las parcelas tratadas.",
        "[2] Zonas específicas según el artículo 34 del RD 1311/2012.",
        "[3] Indicar la distancia en metros si el punto de captación está fuera de las parcelas.",
        "[4] Campo voluntario.",
        "[5] Identificar los pozos y las masas de agua superficial.",
        "[6] En caso afirmativo, indicar las hectáreas de parcela afectada.",
    ]
    for i, note in enumerate(notes):
        _mc(ws, foot_r + i, 1, foot_r + i, 11, note, size=7, align_h="left", border=False, italic=True, wrap=True)

    _page_footer(ws, foot_r + len(notes) + 2, 7, 9, 10, 11, 11, 4)


# =====================================================================
# SHEET 5: inf.trat 1  (tratamientos principales)
# =====================================================================

def _build_inf_trat_1(wb, cuaderno, _parcelas, tratamientos, aplicadores, equipos,
                      orden_tratamientos_modo=None):
    ws = wb.create_sheet("inf.trat 1")
    _set_col_widths(ws, [5, 8, 8, 8, 12, 14, 6, 6, 12, 8, 8, 12])
    _set_row_heights(ws, 1, 120)

    _header_row(ws, cuaderno, 1, 1, 3, 4, 8, 9, 10, 11)

    _mc(ws, 3, 2, 5, 10,
        "3. INFORMACIÓN SOBRE TRATAMIENTOS FITOSANITARIOS",
        bold=True, size=13, align_h="center")

    _mc(ws, 6, 1, 6, 12,
        "3.1. REGISTRO DE ACTUACIONES FITOSANITARIAS DE LA PARCELA",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center")

    # Fila de grupos (diseño RESUELTO: col 1 vacía; «Cultivo» solo B–C; D–H reservadas a filas 8–9)
    _sc(ws, 7, 1, "", bg=LIGHT_GRAY, align_h="center")
    _mc(ws, 7, 2, 7, 3, "Cultivo", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 7, 4, 7, 8, "", bg=LIGHT_GRAY)
    _mc(ws, 7, 9, 7, 11, "Producto Fitosanitario", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 7, 12, "", bg=LIGHT_GRAY, align_h="center")
    _mc(ws, 8, 1, 9, 1, "Id. Parcelas (1)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 2, 9, 2, "Especie", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 8, 3, 9, 3, "Variedad", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 8, 4, 9, 4, "Superf. tratada", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 5, 9, 5, "Intervalo de fechas (2)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 6, 9, 6, "Problema fitosanitario", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 7, 9, 7, "Aplicador (3)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 8, 9, 8, "Equipo (4)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 9, 9, 9, "Nombre Comercial", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 10, 9, 10, "Nº Registro", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 11, 9, 11, "Dosis (kg/ ha o l/ha)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 12, 9, 12, "Eficacia (5)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 25
    ws.row_dimensions[9].height = 25

    aplicador_map = {name: str(i + 1) for i, name in enumerate(aplicadores)}
    equipo_map = {name: str(i + 1) for i, name in enumerate(equipos)}

    r = 10
    prev_parcela_key: Optional[str] = None
    parcela_mode = (orden_tratamientos_modo or "").strip().lower() == "parcela"

    for t in tratamientos:
        if parcela_mode and prev_parcela_key is not None:
            if _trat_parcela_group_key(t) != prev_parcela_key:
                for c in range(1, 13):
                    _sc(ws, r, c, "", size=8)
                ws.row_dimensions[r].height = 10
                r += 1
        prev_parcela_key = _trat_parcela_group_key(t)

        productos = t.productos if t.productos else [ProductoAplicado()]
        parcela_ref = t.num_orden_parcelas or ", ".join(t.parcela_nombres) or ""
        aplicador_ref = aplicador_map.get((t.aplicador or t.operador or "").strip(), t.aplicador or t.operador or "")
        equipo_ref = equipo_map.get((t.equipo or "").strip(), t.equipo or "")

        for pi, prod in enumerate(productos):
            dosis_str = ""
            if prod.dosis:
                dosis_str = f"{prod.dosis} {prod.unidad_dosis or ''}".strip()

            _sc(ws, r, 1, parcela_ref if pi == 0 else "", size=8, align_h="center", wrap=True)
            _sc(ws, r, 2, (t.cultivo_especie or "") if pi == 0 else "", size=8, align_h="left", wrap=True)
            _sc(ws, r, 3, (t.cultivo_variedad or "") if pi == 0 else "", size=8, align_h="center", wrap=True)
            sup_txt = _fmt_superficie_tratada(t.superficie_tratada) if pi == 0 else ""
            _sc(ws, r, 4, sup_txt, size=8, align_h="center")
            _sc(ws, r, 5, _fmt_date(t.fecha_aplicacion) if pi == 0 else "", size=8, align_h="center")
            _sc(ws, r, 6, (t.problema_fitosanitario or t.plaga_enfermedad or "") if pi == 0 else "",
                size=8, align_h="left", wrap=True)
            _sc(ws, r, 7, aplicador_ref if pi == 0 else "", size=8, align_h="center", wrap=True)
            _sc(ws, r, 8, equipo_ref if pi == 0 else "", size=8, align_h="center", wrap=True)
            _sc(ws, r, 9, prod.nombre_comercial or "", size=8, align_h="left", wrap=True)
            _sc(ws, r, 10, prod.numero_registro or "", size=8, align_h="center")
            _sc(ws, r, 11, dosis_str, size=8, align_h="center", wrap=True)
            _sc(ws, r, 12, (t.eficacia or "") if pi == 0 else "", size=8, align_h="center")
            # Altura suficiente para texto envuelto (evita solapar bordes)
            texto_largo = " ".join(
                str(x) for x in (
                    parcela_ref if pi == 0 else "",
                    (t.cultivo_especie or "") if pi == 0 else "",
                    (t.problema_fitosanitario or t.plaga_enfermedad or "") if pi == 0 else "",
                    prod.nombre_comercial or "",
                ) if x
            )
            extra = max(0, len(texto_largo) - 40)
            ws.row_dimensions[r].height = min(96, 28 + max(0, (extra // 35)) * 14)
            r += 1


# =====================================================================
# SHEET 6: inf.trat 2  (asesoramiento)
# =====================================================================

def _build_inf_trat_2(wb, cuaderno):
    ws = wb.create_sheet("inf.trat 2")
    _set_col_widths(ws, [5, 8, 6, 8, 8, 12, 10, 8, 10, 8, 10, 8, 10, 12, 12])
    _set_row_heights(ws, 1, 60)

    _mc(ws, 1, 1, 1, 4, "Explotación/ Titular de la explotación:", bold=True, size=9, align_h="left")
    _mc(ws, 1, 5, 1, 10,
        (cuaderno.titular or cuaderno.nombre_explotacion).upper(), bold=True, size=9, align_h="left")
    _mc(ws, 1, 11, 1, 13, "AÑO:", bold=True, size=9, align_h="right")
    _sc(ws, 1, 14, str(cuaderno.año), bold=True, size=9, align_h="center")

    _mc(ws, 3, 2, 5, 13,
        "3. INFORMACIÓN SOBRE TRATAMIENTOS FITOSANITARIOS",
        bold=True, size=13, align_h="center")

    _mc(ws, 6, 1, 7, 15,
        "3.1. bis REGISTRO DE ACTUACIONES FITOSANITARIAS POR PARCELA\n"
        "(SOLAMENTE PARA CULTIVOS Y SUPERFICIES OBJETO DE ASESORAMIENTO)",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center", wrap=True)
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 20

    _mc(ws, 8, 1, 10, 1, "Especie", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 8, 2, 10, 2, "Variedad", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 8, 3, 8, 5, "DATOS DE LA PARCELA", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 9, 3, 10, 3, "Id. Parcelas (1)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 4, 10, 4, "Superficie cultivada (ha)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 5, 10, 5, "Superficie tratada (ha)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 6, 10, 6, "PLAGA A CONTROLAR\nPlaga", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 7, 8, 10,
        "Justificación de la actuación\n(Superación de umbrales, condiciones meteorológicas, etc.)",
        bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 7, 10, 8,
        "ALTERNATIVAS NO QUIMICAS DE INTERVENCIÓN\nTipo de medida",
        bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 9, 10, 9,
        "Intensidad de la medida (Nº de trampas, nº de difusores, etc.)",
        bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 10, 10, 10, "Fecha de actuación", bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 11, 8, 14, "ALTERNATIVAS QUIMICAS DE INTERVENCIÓN",
        bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 9, 11, 10, 11, "Nombre comercial", bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 12, 10, 12, "Número de registro", bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 13, 10, 13, "Dosis utilizada (l ó kg. / ha)", bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 9, 14, 10, 14, "Fecha de actuación", bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 15, 10, 15, "Eficacia de la intervención (Buena, regular o mala)",
        bold=True, size=7, align_h="center", wrap=True, bg=LIGHT_GRAY)
    ws.row_dimensions[8].height = 30
    ws.row_dimensions[9].height = 30

    for r in range(11, 24):
        for c in range(1, 16):
            _sc(ws, r, c, "", size=8)

    _mc(ws, 25, 2, 25, 8, "VALIDACIÓN INTERMEDIA", bold=True, size=9, align_h="center", underline="single")
    _mc(ws, 26, 2, 26, 4, "Firma", size=9, align_h="left")
    _mc(ws, 28, 2, 28, 4, "Asesor:", size=9, align_h="left")
    _mc(ws, 29, 2, 29, 4, "Nº Inscripción ROPO:", size=9, align_h="left")
    _mc(ws, 30, 2, 30, 4, "Fecha:", size=9, align_h="left")

    _mc(ws, 25, 9, 25, 15, "VALIDACIÓN FINAL", bold=True, size=9, align_h="center", underline="single")
    _mc(ws, 26, 9, 26, 11, "Firma", size=9, align_h="left")
    _mc(ws, 28, 9, 28, 11, "Asesor:", size=9, align_h="left")
    _mc(ws, 29, 9, 29, 11, "Nº Inscripción ROPO:", size=9, align_h="left")
    _mc(ws, 30, 9, 30, 11, "Fecha fin de campaña:", size=9, align_h="left")

    _mc(ws, 34, 1, 34, 12,
        '(1) Indicar el número de orden de identificación de las parcelas tratadas '
        '(Si se trata a todas las parcelas indicar "TODAS").',
        size=7, align_h="left", border=False, italic=True)
    _page_footer(ws, 36, 9, 12, 13, 14, 15, 6)


# =====================================================================
# SHEET 7: inf.trat 3  (semilla tratada + postcosecha)
# =====================================================================

def _build_inf_trat_3(wb, cuaderno):
    ws = wb.create_sheet("inf.trat 3")
    _set_col_widths(ws, [6, 8, 8, 10, 8, 10, 12, 14, 10])
    _set_row_heights(ws, 1, 55)

    _mc(ws, 1, 1, 1, 3, "Explotación/ Titular de la explotación:", bold=True, size=9, align_h="left")
    _mc(ws, 1, 4, 1, 7,
        (cuaderno.titular or cuaderno.nombre_explotacion).upper(), bold=True, size=9, align_h="left")
    _sc(ws, 1, 8, "AÑO:", bold=True, size=9, align_h="right")
    _sc(ws, 1, 9, str(cuaderno.año), bold=True, size=9, align_h="center")

    _mc(ws, 3, 2, 5, 8,
        "3. INFORMACIÓN SOBRE TRATAMIENTOS FITOSANITARIOS",
        bold=True, size=13, align_h="center")

    # 3.2 Semilla tratada
    _mc(ws, 6, 1, 6, 9, "3.2 REGISTRO DE USO DE SEMILLA TRATADA",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center")
    _mc(ws, 7, 1, 7, 9, "Aplica tratamiento: ✓ si  ✓ no (1)",
        size=9, align_h="center", italic=True)

    _mc(ws, 8, 1, 9, 1, "Fecha de siembra", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 2, 9, 2, "Id. parcelas (2)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 3, 8, 4, "Cultivo", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 9, 3, "Especie", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 9, 4, "Variedad", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 8, 5, 9, 5, "Superficie sembrada (Ha)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 6, 9, 6, "Cantidad de semilla (Kg)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 7, 8, 9, "Producto fitosanitario", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 9, 7, 9, 8, "Materia activa / Nombre comercial", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _sc(ws, 9, 9, "Nº registro", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 20

    for r in range(10, 17):
        for c in range(1, 10):
            _sc(ws, r, c, "", size=8)

    # 3.3 Postcosecha
    _mc(ws, 18, 1, 18, 9, "3.3 REGISTRO DE TRATAMIENTOS POSTCOSECHA (en producto vegetal)",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center")
    _mc(ws, 19, 1, 19, 9, "Aplica tratamiento: ✓ si  ✓ no (1)",
        size=9, align_h="center", italic=True)

    _mc(ws, 20, 1, 21, 1, "Fecha", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 20, 2, 21, 2, "Producto vegetal tratado", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 20, 3, 21, 3, "Problemática fitosanitaria", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 20, 4, 21, 4, "Cantidad de prod. veg. tratado (Tm)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 20, 5, 20, 9, "Producto fitosanitario", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 21, 5, 21, 6, "Nombre comercial", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 21, 7, 21, 8, "Nº Registro", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _sc(ws, 21, 9, "Cantidad utilizada (Kg o l)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    ws.row_dimensions[20].height = 20
    ws.row_dimensions[21].height = 20

    for r in range(22, 30):
        for c in range(1, 10):
            _sc(ws, r, c, "", size=8)

    _mc(ws, 31, 1, 31, 5, "(1) Marcar con una cruz.", size=7, align_h="left", border=False, italic=True)
    _mc(ws, 32, 1, 32, 8,
        '(2) Identificar el número de orden de las parcelas tratadas '
        '(Si se trata a todas indicar "TODAS").',
        size=7, align_h="left", border=False, italic=True)
    _page_footer(ws, 34, 5, 7, 8, 9, 9, 7)


# =====================================================================
# SHEET 8: inf.trat 4  (almacenamiento + transporte)
# =====================================================================

def _build_inf_trat_4(wb, cuaderno):
    ws = wb.create_sheet("inf.trat 4")
    _set_col_widths(ws, [6, 16, 12, 8, 10, 8, 12])
    _set_row_heights(ws, 1, 50)

    _mc(ws, 1, 1, 1, 2, "Explotación/ Titular de la explotación:", bold=True, size=9, align_h="left")
    _mc(ws, 1, 3, 1, 5,
        (cuaderno.titular or cuaderno.nombre_explotacion).upper(), bold=True, size=9, align_h="left")
    _sc(ws, 1, 6, "AÑO:", bold=True, size=9, align_h="right")
    _sc(ws, 1, 7, str(cuaderno.año), bold=True, size=9, align_h="center")

    _mc(ws, 3, 2, 5, 6, "3. INFORMACIÓN SOBRE TRATAMIENTOS", bold=True, size=13, align_h="center")

    # 3.4 Almacenamiento
    _mc(ws, 6, 1, 6, 7,
        "3.4 REGISTRO DE TRATAMIENTOS DE LOS LOCALES DE ALMACENAMIENTO",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center")
    _mc(ws, 7, 1, 7, 7, "Aplica tratamiento: ✓ si  ✓ no (1)", size=9, align_h="center", italic=True)

    _mc(ws, 8, 1, 9, 1, "Fecha", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 8, 2, 9, 2, "Local tratado (tipo y dirección)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 3, 9, 3, "Problemática Fitosanitaria", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 4, 9, 4, "Volumen tratado (m³)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 8, 5, 8, 7, "Producto", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 9, 5, 9, 5, "Nombre comercial", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _sc(ws, 9, 6, "Nº Registro", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _sc(ws, 9, 7, "Cantidad utilizada (kg o l)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 20

    for r in range(10, 16):
        for c in range(1, 8):
            _sc(ws, r, c, "", size=8)

    # 3.5 Transporte
    _mc(ws, 17, 1, 17, 7,
        "3.5 REGISTRO DE TRATAMIENTOS DE LOS MEDIOS DE TRANSPORTE",
        bold=True, size=9, bg=LIGHT_GRAY, align_h="center")
    _mc(ws, 18, 1, 18, 7, "Aplica tratamiento: ✓ si  ✓ no (1)", size=9, align_h="center", italic=True)

    _mc(ws, 19, 1, 20, 1, "Fecha", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 19, 2, 20, 2, "Vehículo tratado (tipo, modelo y matrícula)",
        bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 19, 3, 20, 3, "Problemática Fitosanitaria", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 19, 4, 20, 4, "Volumen tratado (m³)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _mc(ws, 19, 5, 19, 7, "Producto", bold=True, size=8, align_h="center", bg=LIGHT_GRAY)
    _mc(ws, 20, 5, 20, 5, "Nombre comercial", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _sc(ws, 20, 6, "Nº Registro", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    _sc(ws, 20, 7, "Cantidad utilizada (kg o l)", bold=True, size=8, align_h="center", wrap=True, bg=LIGHT_GRAY)
    ws.row_dimensions[19].height = 20
    ws.row_dimensions[20].height = 20

    for r in range(21, 28):
        for c in range(1, 8):
            _sc(ws, r, c, "", size=8)

    _mc(ws, 29, 1, 29, 3, "(1) Marcar con una cruz.", size=7, align_h="left", border=False, italic=True)
    _page_footer(ws, 31, 4, 5, 6, 7, 7, 8)


# =====================================================================
# SHEET 9: Parcelas agrupadas por cultivo (con subtotales)
# =====================================================================

def _build_parcelas_por_cultivo(wb, cuaderno, parcelas):
    if not parcelas:
        return
    ws = wb.create_sheet("2.1. DATOS PARCELAS (2)")
    _set_col_widths(ws, [4, 5, 18, 7, 6, 7, 8, 7, 7, 8, 8, 8, 20, 10, 8, 12, 10, 10, 10, 10, 10, 10])
    for r in range(1, 100):
        ws.row_dimensions[r].height = 14

    _parcelas_header_block(ws, cuaderno)

    grouped: Dict[str, list] = {}
    for p in parcelas:
        key = (p.especie or p.cultivo or "SIN CULTIVO").strip().upper()
        grouped.setdefault(key, []).append(p)

    r = _PARCELAS_DATA_START
    idx = 0
    for cultivo in sorted(grouped.keys()):
        grupo = grouped[cultivo]
        total_sup = 0.0
        for p in grupo:
            _write_parcela_row(ws, r, p, zebra=(idx % 2 == 1))
            idx += 1
            total_sup += _safe_float(p.superficie_cultivada) or _safe_float(p.superficie_ha) or _safe_float(p.superficie_sigpac)
            r += 1
        _mc(ws, r, 1, r, 8, f"Subtotal superficie (ha) — {cultivo}",
            bold=True, size=8, align_h="right", bg=LIGHT_GRAY)
        _mc(ws, r, 9, r, 10, round(total_sup, 2),
            bold=True, size=9, align_h="center", font_color="FF0000", bg=LIGHT_GRAY)
        for c in range(11, 23):
            _sc(ws, r, c, "", size=8, bg=LIGHT_GRAY)
        r += 2


# =====================================================================
# DATOS DEL EDITOR: productos, fertilizaciones, cosechas (como en el PDF)
# =====================================================================

def _build_hoja_productos_editor(wb, productos):
    from openpyxl.styles import Font as _F, Alignment as _A

    ws = wb.create_sheet(_unique_sheet_title(wb, "Productos"))
    headers = [
        "Nombre comercial", "Nº registro", "Materia activa", "Formulación",
        "Cantidad", "Unidad", "Fecha adquisición", "Proveedor", "Caducidad", "Lote",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _F(name="Arial", bold=True, size=8, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="595959")
        cell.alignment = _A(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    for i, p in enumerate(productos, 2):
        row = [
            p.nombre_comercial, p.numero_registro, p.materia_activa, p.formulacion,
            p.cantidad_adquirida, p.unidad, p.fecha_adquisicion, p.proveedor,
            p.fecha_caducidad, p.numero_lote,
        ]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = _F(name="Arial", size=8)
            cell.border = _THIN_BORDER
            cell.alignment = _A(vertical="top", wrap_text=True)
    _set_col_widths(ws, [22, 12, 14, 12, 10, 6, 14, 18, 12, 12])


def _build_hoja_fertilizaciones_editor(wb, filas):
    from openpyxl.styles import Font as _F, Alignment as _A

    ws = wb.create_sheet(_unique_sheet_title(wb, "Fertilizaciones"))
    headers = [
        "Fecha inicio", "Fecha fin", "Nº parcelas", "Especie", "Variedad",
        "Tipo abono", "Albarán", "NPK", "Dosis", "Tipo fertilización", "Observaciones",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _F(name="Arial", bold=True, size=8, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="595959")
        cell.alignment = _A(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    for i, f in enumerate(filas, 2):
        row = [
            f.fecha_inicio, f.fecha_fin, f.num_orden_parcelas, f.cultivo_especie, f.cultivo_variedad,
            f.tipo_abono, f.num_albaran, f.riqueza_npk, f.dosis, f.tipo_fertilizacion, f.observaciones,
        ]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = _F(name="Arial", size=8)
            cell.border = _THIN_BORDER
            cell.alignment = _A(vertical="top", wrap_text=True)
    _set_col_widths(ws, [12, 12, 12, 14, 12, 14, 10, 10, 12, 16, 20])


def _build_hoja_cosechas_editor(wb, filas):
    from openpyxl.styles import Font as _F, Alignment as _A

    ws = wb.create_sheet(_unique_sheet_title(wb, "Cosecha"))
    headers = [
        "Fecha", "Producto", "Cantidad (kg)", "Nº parcelas", "Albarán", "Lote",
        "Cliente", "NIF", "Dirección", "RGSEAA",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _F(name="Arial", bold=True, size=8, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="595959")
        cell.alignment = _A(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    for i, cose in enumerate(filas, 2):
        row = [
            cose.fecha, cose.producto, cose.cantidad_kg, cose.num_orden_parcelas,
            cose.num_albaran, cose.num_lote, cose.cliente_nombre, cose.cliente_nif,
            cose.cliente_direccion, cose.cliente_rgseaa,
        ]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = _F(name="Arial", size=8)
            cell.border = _THIN_BORDER
            cell.alignment = _A(vertical="top", wrap_text=True)
    _set_col_widths(ws, [12, 18, 12, 12, 12, 12, 20, 12, 24, 12])


# =====================================================================
# HOJAS IMPORTADAS (misma regla que el PDF: IDs + filas con datos)
# =====================================================================

def _build_hojas_importadas(wb, cuaderno, incluir_hojas_csv: str):
    from openpyxl.styles import Font as _F, Alignment as _A

    if not (incluir_hojas_csv and incluir_hojas_csv.strip()):
        return

    ordered_ids = [s.strip() for s in incluir_hojas_csv.split(",") if s.strip()]
    seen: set = set()
    id_order = []
    for sid in ordered_ids:
        if sid not in seen:
            seen.add(sid)
            id_order.append(sid)
    by_id = {
        h.sheet_id: h
        for h in (cuaderno.hojas_originales or [])
        if _hoja_tiene_datos(h)
    }
    hojas = [by_id[i] for i in id_order if i in by_id]

    used = {ws.title for ws in wb.worksheets}
    for hoja in hojas:
        has_cols = bool(hoja.columnas)
        has_data = bool(hoja.datos and any(
            any(c is not None and str(c).strip() for c in fila)
            for fila in hoja.datos
        ))
        if not has_cols and not has_data:
            continue
        safe = re.sub(r'[\[\]:*?/\\]', '', hoja.nombre or "Hoja")[:31].strip() or "Hoja"
        ctr = 2
        orig = safe
        while safe in used:
            safe = f"{orig[:27]} ({ctr})"[:31]
            ctr += 1
        used.add(safe)

        ws_imp = wb.create_sheet(safe)
        start = 1
        if has_cols:
            for c, cn in enumerate(hoja.columnas, 1):
                cell = ws_imp.cell(row=1, column=c, value=cn)
                cell.font = _F(name="Arial", bold=True, size=8, color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="595959")
                cell.alignment = _A(horizontal="center", vertical="center", wrap_text=True)
                cell.border = _THIN_BORDER
            start = 2
        if hoja.datos:
            for ri, fila in enumerate(hoja.datos, start):
                for ci, v in enumerate(fila, 1):
                    cell = ws_imp.cell(row=ri, column=ci, value=v)
                    cell.border = _THIN_BORDER
                    cell.font = _F(name="Arial", size=8)
                    cell.alignment = _A(vertical="top", wrap_text=True)
                ws_imp.row_dimensions[ri].height = max(
                    15.0,
                    float(ws_imp.row_dimensions[ri].height or 15)
                )


# =====================================================================
# ENTRY POINT
# =====================================================================

def _datasets_filtrados_pdf(cuaderno, parcelas_ordenadas, tratamientos_ordenados, desde, hasta):
    """Parcelas y tratamientos con los mismos filtros que el PDF."""
    pdf_gen = PDFGenerator()
    parcelas = list(parcelas_ordenadas if parcelas_ordenadas is not None else cuaderno.parcelas)
    parcelas = [p for p in parcelas if p.activa and pdf_gen._parcela_tiene_datos(p)]
    tratamientos = list(tratamientos_ordenados if tratamientos_ordenados is not None else cuaderno.tratamientos)
    if desde:
        tratamientos = [t for t in tratamientos if (t.fecha_aplicacion or "") >= desde]
    if hasta:
        tratamientos = [t for t in tratamientos if (t.fecha_aplicacion or "") <= hasta]
    tratamientos = [t for t in tratamientos if _tratamiento_exportable(pdf_gen, t)]
    return parcelas, tratamientos


def _append_hojas_editor_si_faltan_en_plantilla(wb, cuaderno):
    """
    Si el Excel oficial ya trae reg.prod / reg.fert. / reg. cosecha, los datos
    van ahí (template_export). Solo añadimos hojas 'planas' si no existen en la plantilla.
    """
    productos_v = [
        p for p in (cuaderno.productos or [])
        if p.nombre_comercial and str(p.nombre_comercial).strip()
    ]
    if productos_v and "reg.prod" not in wb.sheetnames and "Productos" not in wb.sheetnames:
        _build_hoja_productos_editor(wb, productos_v)

    fertil_v = list(cuaderno.fertilizaciones or [])
    if fertil_v and "reg.fert." not in wb.sheetnames and "Fertilizaciones" not in wb.sheetnames:
        _build_hoja_fertilizaciones_editor(wb, fertil_v)

    cosecha_v = list(cuaderno.cosechas or [])
    if cosecha_v and "reg. cosecha" not in wb.sheetnames and "Cosecha" not in wb.sheetnames:
        _build_hoja_cosechas_editor(wb, cosecha_v)


def generar_excel_oficial(cuaderno, parcelas_ordenadas=None, tratamientos_ordenados=None,
                          desde=None, hasta=None, incluir_hojas_csv: str = "",
                          orden_tratamientos_modo=None):
    """
    Genera un Workbook del Cuaderno de Explotación.

    Si existe la plantilla física RESUELTO (misma que usa `resolver_cuaderno` /
    `workbook_dictionary.json`), se **copia ese Excel** y solo se rellenan datos,
    conservando diseño, estilos y hojas oficiales.

    Si no hay plantilla, se genera con openpyxl como hasta ahora (fallback).

    incluir_hojas_csv: IDs de hojas importadas (vacío = ninguna).
    """
    parcelas, tratamientos = _datasets_filtrados_pdf(
        cuaderno, parcelas_ordenadas, tratamientos_ordenados, desde, hasta
    )

    try:
        from .template_export import resolve_golden_template_path, build_workbook_from_golden

        golden = resolve_golden_template_path()
        if golden is not None:
            wb = build_workbook_from_golden(
                golden,
                cuaderno,
                parcelas,
                tratamientos,
                orden_tratamientos_modo=orden_tratamientos_modo,
            )
            _append_hojas_editor_si_faltan_en_plantilla(wb, cuaderno)
            _build_hojas_importadas(wb, cuaderno, incluir_hojas_csv)
            return wb
    except Exception as e:
        _log.warning(
            "Excel: no se pudo usar la plantilla RESUELTO (%s). Se usa generación programática.",
            e,
            exc_info=_log.isEnabledFor(logging.DEBUG),
        )

    wb = Workbook()
    _build_inf_gral_1(wb, cuaderno)
    aplicadores, equipos = _build_inf_gral_2(wb, cuaderno, tratamientos)
    _build_datos_parcelas(wb, cuaderno, parcelas)
    _build_id_parc_2(wb, cuaderno, parcelas)
    _build_inf_trat_1(wb, cuaderno, parcelas, tratamientos, aplicadores, equipos,
                      orden_tratamientos_modo)
    _build_inf_trat_2(wb, cuaderno)
    _build_inf_trat_3(wb, cuaderno)
    _build_inf_trat_4(wb, cuaderno)
    _build_parcelas_por_cultivo(wb, cuaderno, parcelas)

    productos_v = [
        p for p in (cuaderno.productos or [])
        if p.nombre_comercial and str(p.nombre_comercial).strip()
    ]
    if productos_v:
        _build_hoja_productos_editor(wb, productos_v)

    fertil_v = list(cuaderno.fertilizaciones or [])
    if fertil_v:
        _build_hoja_fertilizaciones_editor(wb, fertil_v)

    cosecha_v = list(cuaderno.cosechas or [])
    if cosecha_v:
        _build_hoja_cosechas_editor(wb, cosecha_v)

    _build_hojas_importadas(wb, cuaderno, incluir_hojas_csv)

    return wb
