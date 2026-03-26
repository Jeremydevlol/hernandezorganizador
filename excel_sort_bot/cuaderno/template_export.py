"""
Exportación Excel usando la MISMA plantilla física que el resto del proyecto
(resolver_cuaderno.GoldenTemplateWriter, validate_cuaderno_oficial, ParcelManager).

Si existe el archivo RESUELTO (golden), se copia y solo se rellenan celdas de datos
preservando estilos, fusionados y estructura del Excel oficial.

Ruta de la plantilla (primera que exista):
  - Variable de entorno CUADERNO_EXCEL_GOLDEN o CUADERNO_GOLDEN_XLSX
  - cuaderno/templates/cuaderno_resuelto.xlsx
  - excel_sort_bot/PABLO PEREZ RUBIO 2025 RESUELTO.XLSX (desarrollo local)
"""
from __future__ import annotations

import json
import os
import re
import shutil
import tempfile
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.workbook import Workbook

from .models import ProductoAplicado


def _dict_path() -> Path:
    return Path(__file__).resolve().parent.parent / "src" / "workbook_dictionary.json"


def load_workbook_dictionary() -> Dict[str, Any]:
    p = _dict_path()
    if not p.exists():
        raise FileNotFoundError(f"No se encuentra el diccionario de plantilla: {p}")
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def resolve_golden_template_path() -> Optional[Path]:
    for key in ("CUADERNO_EXCEL_GOLDEN", "CUADERNO_GOLDEN_XLSX"):
        v = os.environ.get(key)
        if v and Path(v).is_file():
            return Path(v).resolve()
    here = Path(__file__).resolve().parent
    bot = here.parent
    candidates = [
        here / "templates" / "cuaderno_resuelto.xlsx",
        bot / "PABLO PEREZ RUBIO 2025 RESUELTO.XLSX",
        bot.parent / "PABLO PEREZ RUBIO 2025 RESUELTO.XLSX",
    ]
    for p in candidates:
        if p.is_file():
            return p.resolve()
    return None


def _write_cell(ws, row: int, col: int, value: Any) -> None:
    if col is None or row < 1:
        return
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _format_dosis_es(val: Any, unidad: str) -> str:
    if val is None or val == "" or val == 0:
        return ""
    try:
        d = Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        s = str(d).replace(".", ",")
        u = (unidad or "L/Ha").lower()
        unit_short = "l" if "l" in u or "litro" in u else "kg"
        return f"{s} {unit_short}"
    except Exception:
        return f"{val} {unidad or ''}".strip()


def _fecha_aplicacion_to_value(fecha: str):
    if not fecha:
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", str(fecha).strip())
    if m:
        from datetime import datetime

        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m2 = re.match(r"^(\d{2})/(\d{2})/(\d{4})", str(fecha).strip())
    if m2:
        from datetime import datetime

        return datetime(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
    return fecha


def _find_inf_trat_footer(ws, start_row: int) -> int:
    for row in range(start_row, min(ws.max_row + 1, start_row + 2000)):
        val_a = str(ws.cell(row=row, column=1).value or "")
        if val_a.startswith("[1]") or val_a.startswith("[2]") or (len(val_a) > 80 and "[" in val_a):
            return row
    return ws.max_row + 1


def _find_parcelas_footer(ws, start_row: int) -> int:
    for row in range(start_row, min(ws.max_row + 1, start_row + 2000)):
        val = str(ws.cell(row=row, column=1).value or "") + str(ws.cell(row=row, column=2).value or "")
        if val.strip().startswith("[1]") or val.strip().startswith("[2]"):
            return row
    return ws.max_row + 1


def _clear_block(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            _write_cell(ws, r, c, None)


def _fill_inf_gral_1(wb: Workbook, cuaderno, d: Dict[str, Any]) -> None:
    cfg = d.get("sheets", {}).get("inf.gral 1") or {}
    fields = cfg.get("fields") or {}
    if "inf.gral 1" not in wb.sheetnames:
        return
    ws = wb["inf.gral 1"]
    name = (cuaderno.titular or cuaderno.nombre_explotacion or "").strip()
    if "titular" in fields:
        ref = fields["titular"].get("cell")
        if ref:
            ws[ref] = name.upper() if name else ""
    if "explotacion" in fields and fields["explotacion"].get("cell") != fields.get("titular", {}).get("cell"):
        ref = fields["explotacion"].get("cell")
        if ref:
            ws[ref] = name.upper() if name else ""
    if "year" in fields:
        ref = fields["year"].get("cell")
        if ref:
            ws[ref] = cuaderno.año


def _fill_parcelas(wb: Workbook, parcelas: List, d: Dict[str, Any]) -> None:
    cfg = d.get("sheets", {}).get("2.1. DATOS PARCELAS")
    if not cfg or "2.1. DATOS PARCELAS" not in wb.sheetnames:
        return
    ws = wb["2.1. DATOS PARCELAS"]
    cols = cfg["columns"]
    start = int(cfg.get("data_start_row", 14))
    footer = _find_parcelas_footer(ws, start)
    ckeys = list(cols.keys())
    cmin = min(cols[k] for k in ckeys)
    cmax = max(cols[k] for k in ckeys)
    _clear_block(ws, start, footer - 1, cmin, cmax)

    row = start
    for p in parcelas:
        _write_cell(ws, row, cols["nro_orden"], p.num_orden or "")
        _write_cell(ws, row, cols["codigo_provincia"], p.codigo_provincia or p.provincia or "")
        _write_cell(ws, row, cols["municipio"], p.termino_municipal or p.municipio or "")
        _write_cell(ws, row, cols["codigo_agregado"], p.codigo_agregado or 0)
        _write_cell(ws, row, cols["zona"], p.zona or 0)
        _write_cell(ws, row, cols["polygon"], p.num_poligono or "")
        _write_cell(ws, row, cols["parcel"], p.num_parcela or "")
        _write_cell(ws, row, cols["recinto"], p.num_recinto or "")
        _write_cell(ws, row, cols["uso_sigpac"], p.uso_sigpac or "")
        _write_cell(ws, row, cols["superficie_sigpac"], p.superficie_sigpac or "")
        sup = p.superficie_cultivada or p.superficie_ha
        _write_cell(ws, row, cols["superficie_cultivada"], sup or "")
        _write_cell(ws, row, cols["crop"], (p.especie or p.cultivo or "").strip())
        _write_cell(ws, row, cols["ecoregimen"], p.ecoregimen or "")
        _write_cell(ws, row, cols["secano_regadio"], p.secano_regadio or "S")
        row += 1


def _fill_id_parc_2(wb: Workbook, parcelas: List, d: Dict[str, Any]) -> None:
    cfg = d.get("sheets", {}).get("id.parc 2")
    if not cfg or "id.parc 2" not in wb.sheetnames:
        return
    ws = wb["id.parc 2"]
    cols = cfg["columns"]
    start = int(cfg.get("data_start_row", 10))
    footer = _find_inf_trat_footer(ws, start)
    cmin, cmax = 1, max(cols.values())
    _clear_block(ws, start, footer - 1, cmin, cmax)
    row = start
    for p in parcelas:
        _write_cell(ws, row, cols["id_parcelas"], p.num_orden or "")
        _write_cell(ws, row, cols["especie"], p.especie or p.cultivo or "")
        _write_cell(ws, row, cols["variedad"], p.variedad or "")
        row += 1


def _fill_inf_gral_2_listas(wb: Workbook, tratamientos: List, d: Dict[str, Any]) -> Tuple[List[str], List[str]]:
    cfg = d.get("sheets", {}).get("inf.gral 2")
    aplicadores: List[str] = []
    equipos: List[str] = []
    seen_a, seen_e = set(), set()
    for t in tratamientos:
        a = (t.aplicador or t.operador or "").strip()
        if a and a not in seen_a:
            seen_a.add(a)
            aplicadores.append(a)
        e = (t.equipo or "").strip()
        if e and e not in seen_e:
            seen_e.add(e)
            equipos.append(e)

    if not cfg or "inf.gral 2" not in wb.sheetnames:
        return aplicadores, equipos
    ws = wb["inf.gral 2"]
    cols = cfg.get("columns") or {}
    # Tabla 1.2: filas 10–15 (índices del generador programático)
    base = int(cfg.get("data_start_row", 10))
    for i in range(6):
        r = base + i
        nombre = aplicadores[i] if i < len(aplicadores) else ""
        _write_cell(ws, r, cols.get("nro_orden", 1), str(i + 1) if nombre else "")
        _write_cell(ws, r, cols.get("nombre", 2), nombre)
    # Tabla 1.3 equipos: filas 20–25
    eq_base = 20
    for i in range(6):
        r = eq_base + i
        desc = equipos[i] if i < len(equipos) else ""
        _write_cell(ws, r, 1, str(i + 1) if desc else "")
        _write_cell(ws, r, 2, desc)
    return aplicadores, equipos


def _trat_parcela_group_key(t) -> str:
    if t.parcela_nombres:
        return ",".join(t.parcela_nombres).lower()
    return (t.num_orden_parcelas or "").lower()


def _fill_inf_trat_1(
    wb: Workbook,
    tratamientos: List,
    aplicadores: List[str],
    equipos: List[str],
    d: Dict[str, Any],
    orden_tratamientos_modo: Optional[str],
) -> None:
    cfg = d.get("sheets", {}).get("inf.trat 1")
    if not cfg or "inf.trat 1" not in wb.sheetnames:
        return
    ws = wb["inf.trat 1"]
    cols = cfg["columns"]
    start = int(cfg.get("data_start_row", 11))
    footer = _find_inf_trat_footer(ws, start)
    cmax = max(cols.values())
    _clear_block(ws, start, footer - 1, 1, cmax)

    aplicador_map = {n: str(i + 1) for i, n in enumerate(aplicadores)}
    equipo_map = {n: str(i + 1) for i, n in enumerate(equipos)}
    parcela_mode = (orden_tratamientos_modo or "").strip().lower() == "parcela"

    row = start
    prev_key: Optional[str] = None
    for t in tratamientos:
        if parcela_mode and prev_key is not None and _trat_parcela_group_key(t) != prev_key:
            row += 1
        prev_key = _trat_parcela_group_key(t)

        productos = t.productos if t.productos else [ProductoAplicado()]
        parcela_ref = t.num_orden_parcelas or ", ".join(t.parcela_nombres) or ""
        aplicador_ref = aplicador_map.get((t.aplicador or t.operador or "").strip(), t.aplicador or t.operador or "")
        equipo_ref = equipo_map.get((t.equipo or "").strip(), t.equipo or "")

        for pi, prod in enumerate(productos):
            _write_cell(ws, row, cols["id_parcelas"], parcela_ref if pi == 0 else "")
            _write_cell(ws, row, cols["especie"], (t.cultivo_especie or "") if pi == 0 else "")
            _write_cell(ws, row, cols["variedad"], (t.cultivo_variedad or "") if pi == 0 else "")
            _write_cell(ws, row, cols["superficie_tratada"], t.superficie_tratada if pi == 0 else "")
            _write_cell(ws, row, cols["fecha"], _fecha_aplicacion_to_value(t.fecha_aplicacion) if pi == 0 else None)
            _write_cell(
                ws,
                row,
                cols["problema_fito"],
                (t.problema_fitosanitario or t.plaga_enfermedad or "") if pi == 0 else "",
            )
            _write_cell(ws, row, cols["aplicador"], aplicador_ref if pi == 0 else "")
            _write_cell(ws, row, cols["equipo"], equipo_ref if pi == 0 else "")
            _write_cell(ws, row, cols["producto"], prod.nombre_comercial or "")
            _write_cell(ws, row, cols["nro_registro"], prod.numero_registro or "")
            _write_cell(ws, row, cols["dosis"], _format_dosis_es(prod.dosis, prod.unidad_dosis or "L/Ha"))
            _write_cell(ws, row, cols["eficacia"], (t.eficacia or "") if pi == 0 else "")
            if "observaciones" in cols:
                _write_cell(ws, row, cols["observaciones"], (t.observaciones or "") if pi == 0 else "")
            row += 1


def _fill_reg_prod(wb: Workbook, cuaderno, d: Dict[str, Any]) -> None:
    cfg = d.get("sheets", {}).get("reg.prod")
    if not cfg or "reg.prod" not in wb.sheetnames:
        return
    productos = [p for p in (cuaderno.productos or []) if (p.nombre_comercial or "").strip()]
    if not productos:
        return
    ws = wb["reg.prod"]
    cols = cfg["columns"]
    start = int(cfg.get("data_start_row", 11))
    footer = _find_inf_trat_footer(ws, start)
    _clear_block(ws, start, footer - 1, 1, max(cols.values()))
    row = start
    for pr in productos:
        _write_cell(ws, row, cols.get("fecha", 1), pr.fecha_adquisicion or "")
        _write_cell(ws, row, cols.get("material_analizado", 2), pr.nombre_comercial or "")
        _write_cell(ws, row, cols.get("cultivo_muestreado", 3), "")
        _write_cell(ws, row, cols.get("nro_boletin", 4), pr.numero_registro or "")
        _write_cell(ws, row, cols.get("laboratorio_nombre", 5), pr.proveedor or "")
        _write_cell(ws, row, cols.get("laboratorio_direccion", 6), "")
        _write_cell(ws, row, cols.get("sustancias_detectadas", 7), pr.materia_activa or "")
        row += 1


def _fill_reg_fert(wb: Workbook, cuaderno, d: Dict[str, Any]) -> None:
    cfg = d.get("sheets", {}).get("reg.fert.")
    if not cfg or "reg.fert." not in wb.sheetnames:
        return
    filas = list(cuaderno.fertilizaciones or [])
    if not filas:
        return
    ws = wb["reg.fert."]
    cols = cfg["columns"]
    start = int(cfg.get("data_start_row", 11))
    footer = _find_inf_trat_footer(ws, start)
    _clear_block(ws, start, footer - 1, 1, max(cols.values()))
    row = start
    for f in filas:
        _write_cell(ws, row, cols.get("fecha_inicio", 1), f.fecha_inicio or "")
        _write_cell(ws, row, cols.get("fecha_fin", 2), f.fecha_fin or "")
        _write_cell(ws, row, cols.get("parcelas", 3), f.num_orden_parcelas or "")
        _write_cell(ws, row, cols.get("especie", 4), f.cultivo_especie or "")
        _write_cell(ws, row, cols.get("variedad", 5), f.cultivo_variedad or "")
        _write_cell(ws, row, cols.get("tipo_abono", 6), f.tipo_abono or "")
        _write_cell(ws, row, cols.get("albaran", 7), f.num_albaran or "")
        _write_cell(ws, row, cols.get("dosis", 11), f.dosis or "")
        _write_cell(ws, row, cols.get("tipo_fertilizacion", 12), f.tipo_fertilizacion or "")
        _write_cell(ws, row, cols.get("observaciones", 13), f.observaciones or "")
        row += 1


def _fill_reg_cosecha(wb: Workbook, cuaderno, d: Dict[str, Any]) -> None:
    cfg = d.get("sheets", {}).get("reg. cosecha")
    if not cfg or "reg. cosecha" not in wb.sheetnames:
        return
    filas = list(cuaderno.cosechas or [])
    if not filas:
        return
    ws = wb["reg. cosecha"]
    cols = cfg["columns"]
    start = int(cfg.get("data_start_row", 11))
    footer = _find_inf_trat_footer(ws, start)
    _clear_block(ws, start, footer - 1, 1, max(cols.values()))
    row = start
    for c in filas:
        _write_cell(ws, row, cols.get("fecha", 1), c.fecha or "")
        _write_cell(ws, row, cols.get("producto", 2), c.producto or "")
        _write_cell(ws, row, cols.get("cantidad_kg", 3), c.cantidad_kg or "")
        _write_cell(ws, row, cols.get("parcelas_origen", 4), c.num_orden_parcelas or "")
        _write_cell(ws, row, cols.get("albaran", 5), c.num_albaran or "")
        _write_cell(ws, row, cols.get("lote", 6), c.num_lote or "")
        _write_cell(ws, row, cols.get("cliente_nombre", 7), c.cliente_nombre or "")
        _write_cell(ws, row, cols.get("cliente_nif", 8), c.cliente_nif or "")
        _write_cell(ws, row, cols.get("cliente_direccion", 9), c.cliente_direccion or "")
        _write_cell(ws, row, cols.get("rgseaa", 10), c.cliente_rgseaa or "")
        row += 1


def build_workbook_from_golden(
    golden_path: Path,
    cuaderno,
    parcelas: List,
    tratamientos: List,
    orden_tratamientos_modo: Optional[str] = None,
) -> Workbook:
    """
    Copia el Excel RESUELTO y rellena datos según workbook_dictionary.json.
    """
    d = load_workbook_dictionary()
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        shutil.copy(golden_path, tmp_path)
        wb = load_workbook(tmp_path)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    _fill_inf_gral_1(wb, cuaderno, d)
    _fill_parcelas(wb, parcelas, d)
    _fill_id_parc_2(wb, parcelas, d)
    aplicadores, equipos = _fill_inf_gral_2_listas(wb, tratamientos, d)
    _fill_inf_trat_1(wb, tratamientos, aplicadores, equipos, d, orden_tratamientos_modo)
    _fill_reg_prod(wb, cuaderno, d)
    _fill_reg_fert(wb, cuaderno, d)
    _fill_reg_cosecha(wb, cuaderno, d)
    return wb
