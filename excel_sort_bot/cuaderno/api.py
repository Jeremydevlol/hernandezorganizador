"""
CUADERNO DE EXPLOTACIÓN - API REST v2.0
Endpoints para gestión de cuadernos, parcelas, productos y tratamientos.
Soporte para upload y procesamiento de archivos con GPT-4o Vision.
"""
import os
import re
import json
import time
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime, date
import tempfile


from fastapi import APIRouter, BackgroundTasks, HTTPException, Query, File, UploadFile, Form, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field

from .models import (
    CuadernoExplotacion, Parcela, ProductoFitosanitario,
    ProductoAplicado, Tratamiento, EstadoTratamiento, TipoProducto, HojaExcel,
    Fertilizacion, Cosecha, Asesoramiento
)
from .storage import get_storage
from .productos_catalogo import get_catalogo
from .pdf_generator import PDFGenerator, _sort_key_parcela
from .file_processor import FileProcessor, get_processor


# ============================================
# ROUTER
# ============================================

router = APIRouter(prefix="/api/cuaderno", tags=["Cuaderno de Explotación"])


# ============================================
# WEBSOCKET HUB (stock realtime)
# ============================================

class _StockWsHub:
    def __init__(self):
        self._global: List[WebSocket] = []
        self._by_cuaderno: Dict[str, List[WebSocket]] = {}

    async def connect_global(self, ws: WebSocket):
        await ws.accept()
        self._global.append(ws)

    async def connect_cuaderno(self, ws: WebSocket, cuaderno_id: str):
        await ws.accept()
        self._by_cuaderno.setdefault(cuaderno_id, []).append(ws)

    def disconnect_global(self, ws: WebSocket):
        self._global = [c for c in self._global if c is not ws]

    def disconnect_cuaderno(self, ws: WebSocket, cuaderno_id: str):
        conns = self._by_cuaderno.get(cuaderno_id, [])
        self._by_cuaderno[cuaderno_id] = [c for c in conns if c is not ws]

    async def _broadcast(self, conns: List[WebSocket], payload: Dict[str, Any]):
        for ws in list(conns):
            try:
                await ws.send_json(payload)
            except Exception:
                # La conexión murió; se limpiará en el siguiente ciclo.
                continue

    async def publish_stock_changed(self, cuaderno_id: str):
        payload = {"event": "stock_changed", "cuaderno_id": cuaderno_id, "ts": datetime.now().isoformat()}
        await self._broadcast(self._global, payload)
        await self._broadcast(self._by_cuaderno.get(cuaderno_id, []), payload)


_stock_ws_hub = _StockWsHub()

# Cache en memoria para acelerar endpoints costosos (catálogo/stock global).
_FAST_CACHE: Dict[str, Dict[str, Any]] = {}
_LAST_CATALOGO_SYNC_TS: float = 0.0


def _cache_get(key: str) -> Optional[Any]:
    item = _FAST_CACHE.get(key)
    if not item:
        return None
    if time.time() >= float(item.get("expires_at", 0)):
        _FAST_CACHE.pop(key, None)
        return None
    return item.get("value")


def _cache_set(key: str, value: Any, ttl_sec: int = 8) -> None:
    _FAST_CACHE[key] = {"value": value, "expires_at": time.time() + max(1, ttl_sec)}


def _cache_invalidate(prefixes: List[str]) -> None:
    keys = list(_FAST_CACHE.keys())
    for k in keys:
        if any(k.startswith(p) for p in prefixes):
            _FAST_CACHE.pop(k, None)


def _invalidar_cuaderno(cuaderno_id: str) -> None:
    """Invalida la caché del cuaderno y de la lista. Llamar tras cualquier escritura."""
    _cache_invalidate([f"cuaderno::{cuaderno_id}", "cuadernos_list::"])


def _guardar_cuaderno(storage, cuaderno) -> None:
    """Guarda y limpia caché automáticamente. Usar en lugar de storage.guardar()."""
    storage.guardar(cuaderno)
    _invalidar_cuaderno(cuaderno.id)


CULTIVOS_ASESORAMIENTO_OBLIGATORIO = ("PATATA", "REMOLACHA")
TIPO_ASESORAMIENTO_AUTO = "AUTO_SUPERFICIE_ESPECIAL"


def _parcelas_especiales_y_superficie(cuaderno: CuadernoExplotacion) -> tuple[List[Parcela], float]:
    parcelas_especiales: List[Parcela] = []
    total_ha = 0.0
    for parcela in (cuaderno.parcelas or []):
        cultivo = (parcela.especie or parcela.cultivo or "").upper().strip()
        if any(c in cultivo for c in CULTIVOS_ASESORAMIENTO_OBLIGATORIO):
            parcelas_especiales.append(parcela)
            try:
                total_ha += float(parcela.superficie_cultivada or parcela.superficie_ha or parcela.superficie_sigpac or 0)
            except (TypeError, ValueError):
                pass
    return parcelas_especiales, round(total_ha, 2)


def _normalizar_cultivo_variante_parcela(parcela: Parcela) -> bool:
    """
    Normaliza cultivos tipo 'REMOLACHA.ROSADONNA' a:
      - cultivo/especie: 'REMOLACHA'
      - variedad: 'ROSADONNA'
    """
    raw_cultivo = (parcela.especie or parcela.cultivo or "").strip()
    if not raw_cultivo:
        return False
    if "." not in raw_cultivo:
        return False

    base, variante = raw_cultivo.split(".", 1)
    base = base.strip()
    variante = variante.strip()
    if not base:
        return False

    changed = False
    if (parcela.especie or "").strip() != base:
        parcela.especie = base
        changed = True
    if (parcela.cultivo or "").strip() != base:
        parcela.cultivo = base
        changed = True
    if variante and not (parcela.variedad or "").strip():
        parcela.variedad = variante
        changed = True
    return changed


def _normalizar_parcelas_cultivo_variante(cuaderno: CuadernoExplotacion) -> bool:
    changed = False
    for parcela in (cuaderno.parcelas or []):
        if _normalizar_cultivo_variante_parcela(parcela):
            changed = True
    if changed:
        cuaderno._actualizar_modificacion()
    return changed


def _desglosar_asesoramientos_multi_parcela(cuaderno: CuadernoExplotacion) -> bool:
    """
    Normaliza asesoramientos históricos con "num_orden_parcelas" agrupado (ej: "45,46,58")
    en filas individuales (una por Nº de orden), igual que el resto de hojas.
    """
    ases = list(cuaderno.asesoramientos or [])
    if not ases:
        return False

    orden_to_parcela: Dict[str, Parcela] = {}
    for p in (cuaderno.parcelas or []):
        if isinstance(p.num_orden, int) and p.num_orden > 0:
            orden_to_parcela[str(p.num_orden)] = p

    changed = False
    nuevos: List[Asesoramiento] = []
    for a in ases:
        raw = (a.num_orden_parcelas or "").strip()
        tokens = [t.strip() for t in raw.replace(";", ",").split(",") if t.strip()]
        if len(tokens) <= 1:
            nuevos.append(a)
            continue

        changed = True
        divisor = max(1, len(tokens))
        sup_base = float(a.superficie_ha or 0)
        for tok in tokens:
            parcela = orden_to_parcela.get(tok)
            sup = float(parcela.superficie_cultivada or parcela.superficie_ha or parcela.superficie_sigpac or 0) if parcela else 0.0
            if sup <= 0 and sup_base > 0:
                sup = round(sup_base / divisor, 2)
            cultivo_parcela = ((parcela.especie or parcela.cultivo) if parcela else "") or ""
            cultivo_origen = (a.cultivo_especie or "").strip()
            cultivo_final = cultivo_origen
            # Si el histórico quedó con etiqueta combinada (PATATA/REMOLACHA), sustituimos por cultivo real de la parcela.
            if not cultivo_final or "/" in cultivo_final or "," in cultivo_final or cultivo_final.upper() == "PATATA/REMOLACHA":
                cultivo_final = cultivo_parcela
            nuevos.append(
                Asesoramiento(
                    fecha=a.fecha,
                    num_orden_parcelas=tok,
                    cultivo_especie=cultivo_final,
                    cultivo_variedad=a.cultivo_variedad or ((parcela.variedad or "") if parcela else ""),
                    superficie_ha=round(sup, 2) if sup else 0.0,
                    nombre_asesor=a.nombre_asesor,
                    num_habilitacion=a.num_habilitacion,
                    tipo_asesoramiento=a.tipo_asesoramiento,
                    recomendacion=a.recomendacion,
                    observaciones=a.observaciones,
                    color_fila=a.color_fila,
                )
            )

    if changed:
        cuaderno.asesoramientos = nuevos
        cuaderno._actualizar_modificacion()
    return changed


def _normalizar_cultivo_asesoramiento_individual(cuaderno: CuadernoExplotacion) -> bool:
    """
    Corrige filas de asesoramiento ya desglosadas que aún conservan cultivo combinado
    (p.ej. PATATA/REMOLACHA), sustituyéndolo por el cultivo real de su parcela.
    """
    orden_to_parcela: Dict[str, Parcela] = {}
    for p in (cuaderno.parcelas or []):
        if isinstance(p.num_orden, int) and p.num_orden > 0:
            orden_to_parcela[str(p.num_orden)] = p

    changed = False
    for a in (cuaderno.asesoramientos or []):
        cultivo = (a.cultivo_especie or "").strip()
        if not cultivo:
            continue
        if "/" not in cultivo and "," not in cultivo and cultivo.upper() != "PATATA/REMOLACHA":
            continue

        tokens = [t.strip() for t in str(a.num_orden_parcelas or "").replace(";", ",").split(",") if t.strip()]
        if len(tokens) != 1:
            continue
        parcela = orden_to_parcela.get(tokens[0])
        if not parcela:
            continue
        cultivo_real = (parcela.especie or parcela.cultivo or "").strip()
        if cultivo_real and cultivo_real != cultivo:
            a.cultivo_especie = cultivo_real
            if not (a.cultivo_variedad or "").strip():
                a.cultivo_variedad = (parcela.variedad or "").strip()
            changed = True

    if changed:
        cuaderno._actualizar_modificacion()
    return changed


def _asegurar_asesoramiento_automatico(cuaderno: CuadernoExplotacion) -> bool:
    """
    Si PATATA+REMOLACHA del cuaderno supera 5 ha (sumadas), crea un asesoramiento automático.
    Devuelve True si el cuaderno fue modificado.
    """
    parcelas_especiales, total_ha = _parcelas_especiales_y_superficie(cuaderno)
    if total_ha <= 5 or not parcelas_especiales:
        return False

    existentes_por_orden = {
        (a.num_orden_parcelas or "").strip()
        for a in (cuaderno.asesoramientos or [])
        if (a.num_orden_parcelas or "").strip()
    }

    created = False
    for parcela in parcelas_especiales:
        if not isinstance(parcela.num_orden, int) or parcela.num_orden <= 0:
            continue
        orden = str(parcela.num_orden)
        if orden in existentes_por_orden:
            continue
        sup = float(parcela.superficie_cultivada or parcela.superficie_ha or parcela.superficie_sigpac or 0)
        cultivo_base = (parcela.especie or parcela.cultivo or "").strip().upper()
        asesoramiento = Asesoramiento(
            fecha=datetime.now().date().isoformat(),
            num_orden_parcelas=orden,
            cultivo_especie=cultivo_base or "PATATA/REMOLACHA",
            cultivo_variedad=(parcela.variedad or "").strip(),
            superficie_ha=round(sup, 2) if sup else 0.0,
            tipo_asesoramiento=TIPO_ASESORAMIENTO_AUTO,
            recomendacion=f"Asesoramiento automático: superficie conjunta de cultivos especiales = {total_ha:.2f} ha (>5 ha).",
            observaciones="Creado automáticamente por regla de Patata+Remolacha.",
        )
        cuaderno.agregar_asesoramiento(asesoramiento)
        existentes_por_orden.add(orden)
        created = True
    return created


# ============================================
# PYDANTIC REQUEST/RESPONSE MODELS
# ============================================

class ParcelaCreate(BaseModel):
    nombre: str
    num_orden: int = 0
    referencia_catastral: str = ""
    superficie_ha: float = 0.0
    cultivo: str = ""
    variedad: str = ""
    municipio: str = ""
    provincia: str = ""
    notas: str = ""


class ParcelaUpdate(BaseModel):
    nombre: Optional[str] = None
    referencia_catastral: Optional[str] = None
    superficie_ha: Optional[float] = None
    cultivo: Optional[str] = None
    variedad: Optional[str] = None
    municipio: Optional[str] = None
    provincia: Optional[str] = None
    notas: Optional[str] = None
    activa: Optional[bool] = None


class ProductoCreate(BaseModel):
    nombre_comercial: str
    numero_registro: str
    materia_activa: str = ""
    formulacion: str = ""
    tipo: str = "fitosanitario"
    numero_lote: str = ""
    cantidad_adquirida: float = 0.0
    unidad: str = "L"
    fecha_adquisicion: str = ""
    proveedor: str = ""
    fecha_caducidad: str = ""
    notas: str = ""


class CatalogoProductoCreate(BaseModel):
    """Producto en el catálogo global (compartido entre todos los cuadernos)."""
    nombre_comercial: str
    numero_registro: str = ""
    materia_activa: str = ""
    formulacion: str = ""
    tipo: str = "fitosanitario"
    unidad: str = "L"
    proveedor: str = ""
    notas: str = ""


class CatalogoProductoUpdate(BaseModel):
    nombre_comercial: Optional[str] = None
    numero_registro: Optional[str] = None
    materia_activa: Optional[str] = None
    formulacion: Optional[str] = None
    tipo: Optional[str] = None
    unidad: Optional[str] = None
    proveedor: Optional[str] = None
    notas: Optional[str] = None


class ImportarDesdeCatalogoRequest(BaseModel):
    """Importa un producto del catálogo global al inventario del cuaderno."""
    catalogo_id: str


class ProductoAplicadoCreate(BaseModel):
    """Línea de producto aplicado. Snapshot (nombre, registro, lote) se rellena desde inventario si producto_id existe; si no, se usan estos valores."""
    producto_id: str = ""
    nombre_comercial: str = ""
    numero_registro: str = ""
    numero_lote: str = ""
    plaga_enfermedad: str = ""   # Problemática específica de este producto (si varios productos en el tratamiento)
    dosis: float = 0.0
    unidad_dosis: str = "L/Ha"
    caldo_hectarea: float = 0.0
    notas: str = ""


class TratamientoCreate(BaseModel):
    fecha_aplicacion: str
    parcela_ids: List[str]
    productos: List[ProductoAplicadoCreate]
    plaga_enfermedad: str = ""
    justificacion: str = ""
    metodo_aplicacion: str = ""
    condiciones_climaticas: str = ""
    operador: str = "1"
    equipo: str = "1"
    eficacia: str = "BUENA"
    hora_inicio: str = ""
    hora_fin: str = ""
    observaciones: str = ""
    # Asesoramiento
    asesorado: bool = False
    nombre_asesor_trat: str = ""
    num_colegiado_asesor: str = ""
    fecha_recomendacion_asesor: str = ""
    firma_asesor: str = ""
    firma_cliente: str = ""


class FertilizacionCreate(BaseModel):
    fecha_inicio: str = ""
    fecha_fin: str = ""
    parcela_ids: List[str] = Field(default_factory=list)  # Se convierte a num_orden_parcelas
    cultivo_especie: str = ""
    cultivo_variedad: str = ""
    tipo_abono: str = ""
    riqueza_npk: str = ""
    dosis: str = ""
    tipo_fertilizacion: str = ""
    observaciones: str = ""


class CosechaCreate(BaseModel):
    fecha: str = ""
    producto: str = ""
    cantidad_kg: float = 0.0
    parcela_ids: List[str] = Field(default_factory=list)  # Se convierte a num_orden_parcelas
    num_albaran: str = ""
    num_lote: str = ""
    cliente_nombre: str = ""
    cliente_nif: str = ""
    cliente_direccion: str = ""
    cliente_rgseaa: str = ""


class AsesoramientoCreate(BaseModel):
    fecha: str = ""
    parcela_ids: List[str] = Field(default_factory=list)
    cultivo_especie: str = ""
    cultivo_variedad: str = ""
    superficie_ha: float = 0.0
    nombre_asesor: str = ""
    num_habilitacion: str = ""
    tipo_asesoramiento: str = ""
    recomendacion: str = ""
    observaciones: str = ""


class TratamientoUpdate(BaseModel):
    """Campos opcionales para actualizar un tratamiento."""
    fecha_aplicacion: Optional[str] = None
    parcela_ids: Optional[List[str]] = None
    productos: Optional[List[ProductoAplicadoCreate]] = None
    plaga_enfermedad: Optional[str] = None
    justificacion: Optional[str] = None
    metodo_aplicacion: Optional[str] = None
    condiciones_climaticas: Optional[str] = None
    operador: Optional[str] = None
    equipo: Optional[str] = None
    eficacia: Optional[str] = None
    hora_inicio: Optional[str] = None
    hora_fin: Optional[str] = None
    observaciones: Optional[str] = None
    # Asesoramiento
    asesorado: Optional[bool] = None
    nombre_asesor_trat: Optional[str] = None
    num_colegiado_asesor: Optional[str] = None
    fecha_recomendacion_asesor: Optional[str] = None
    firma_asesor: Optional[str] = None
    firma_cliente: Optional[str] = None
    color_fila: Optional[str] = None


class CuadernoCreate(BaseModel):
    nombre_explotacion: str
    titular: str = ""
    nif_titular: str = ""
    domicilio: str = ""
    codigo_explotacion: str = ""
    año: int = Field(default_factory=lambda: datetime.now().year)


class CuadernoResponse(BaseModel):
    id: str
    nombre_explotacion: str
    titular: str
    año: int
    num_parcelas: int
    num_productos: int
    num_tratamientos: int
    fecha_modificacion: str


class HojaUpdate(BaseModel):
    """Actualización parcial de una hoja importada. Todo lo que entra se conserva."""
    datos: Optional[List[List[Any]]] = None
    columnas: Optional[List[str]] = None
    nombre: Optional[str] = None


class CellPatch(BaseModel):
    """Delta atómico: una celda. Edición en tiempo real sin enviar la hoja entera."""
    sheet_id: str  # "parcelas" | "productos" | "tratamientos" | uuid de hoja importada
    row: Any  # int (índice en importadas) o str (entity_id en base)
    column: Any  # int (índice col en importadas) o str (field key en base)
    value: Any = None


# ============================================
# CUADERNO ENDPOINTS
# ============================================

async def _listar_cuadernos_payload():
    cache_key = "cuadernos_list::"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    storage = get_storage()
    payload = {"cuadernos": storage.listar()}
    _cache_set(cache_key, payload, ttl_sec=20)
    return payload


@router.get("/catalog/cuadernos")
async def listar_cuadernos_catalog():
    """Lista cuadernos (ruta de varios segmentos: evita que /{cuaderno_id} capture 'list' en algunos despliegues)."""
    return await _listar_cuadernos_payload()


@router.get("/list")
async def listar_cuadernos():
    """Lista todos los cuadernos disponibles (legacy; preferir /catalog/cuadernos en clientes nuevos)."""
    return await _listar_cuadernos_payload()


@router.post("/create")
async def crear_cuaderno(data: CuadernoCreate):
    """Crea un nuevo cuaderno de explotación"""
    cuaderno = CuadernoExplotacion(
        nombre_explotacion=data.nombre_explotacion,
        titular=data.titular,
        nif_titular=data.nif_titular,
        domicilio=data.domicilio,
        codigo_explotacion=data.codigo_explotacion,
        año=data.año
    )
    
    storage = get_storage()
    storage.crear(cuaderno)
    _cache_invalidate(["cuadernos_list::"])

    return {
        "success": True,
        "cuaderno_id": cuaderno.id,
        "message": "Cuaderno creado correctamente"
    }


def _reparar_y_normalizar(cuaderno, storage) -> bool:
    """
    Ejecuta TODAS las reparaciones y normalizaciones sobre un objeto cuaderno ya cargado.
    Devuelve True si hubo algún cambio (el llamador debe guardar).
    """
    changed = False
    # --- Reparaciones de datos (antes vivían en storage.cargar, bloqueando cada petición) ---
    try:
        r1 = cuaderno.reparar_tratamientos_multi_cultivo()
        r2 = cuaderno.reparar_tratamientos_num_orden_multi_parcela()
        r3 = cuaderno.reestablecer_num_orden_individual_tratamientos()
        if r1 + r2 + r3 > 0:
            print(f"[repair_bg] {cuaderno.id}: multi-cultivo={r1} num_orden={r2} individual={r3}")
            changed = True
    except Exception as e:
        print(f"[repair_bg] Error reparando {cuaderno.id}: {e}")
    # --- Normalizaciones de la API ---
    if _normalizar_parcelas_cultivo_variante(cuaderno):
        changed = True
    if _desglosar_asesoramientos_multi_parcela(cuaderno):
        changed = True
    if _normalizar_cultivo_asesoramiento_individual(cuaderno):
        changed = True
    if _asegurar_asesoramiento_automatico(cuaderno):
        changed = True
    return changed


def _normalizar_y_guardar_bg(cuaderno_id: str) -> None:
    """
    Tarea de fondo que recibe solo el ID. Usada cuando no tenemos el objeto en memoria.
    Carga, repara/normaliza y guarda si hubo cambios.
    """
    try:
        storage = get_storage()
        cuaderno = storage.cargar(cuaderno_id)
        if not cuaderno:
            return
        if _reparar_y_normalizar(cuaderno, storage):
            _guardar_cuaderno(storage, cuaderno)
    except Exception as e:
        print(f"[normalizar_bg] Error en cuaderno {cuaderno_id}: {e}")


def _normalizar_y_guardar_bg_obj(cuaderno, storage) -> None:
    """
    Tarea de fondo que recibe el objeto ya cargado (evita un segundo storage.cargar).
    Usada desde GET /{cuaderno_id} donde el objeto ya está en memoria.
    """
    try:
        if _reparar_y_normalizar(cuaderno, storage):
            _guardar_cuaderno(storage, cuaderno)
    except Exception as e:
        print(f"[normalizar_bg_obj] Error en cuaderno {cuaderno.id}: {e}")


@router.get("/{cuaderno_id}")
async def obtener_cuaderno(cuaderno_id: str, background_tasks: BackgroundTasks):
    """Obtiene un cuaderno completo. Sirve desde caché en-memoria; normaliza en segundo plano."""
    cache_key = f"cuaderno::{cuaderno_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)

    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    payload = {"cuaderno": cuaderno.to_dict()}
    _cache_set(cache_key, payload, ttl_sec=120)

    # Normalización en segundo plano: pasamos el objeto ya cargado (evita segundo storage.cargar)
    background_tasks.add_task(_normalizar_y_guardar_bg_obj, cuaderno, storage)

    return payload


@router.delete("/{cuaderno_id}")
async def eliminar_cuaderno(cuaderno_id: str):
    """Elimina un cuaderno (mueve a backup)"""
    storage = get_storage()
    success = storage.eliminar(cuaderno_id)

    if not success:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    _invalidar_cuaderno(cuaderno_id)
    return {"success": True, "message": "Cuaderno eliminado (backup creado)"}


# ============================================
# HOJAS IMPORTADAS (edición en tiempo real)
# ============================================

@router.put("/{cuaderno_id}/hojas/{sheet_id}")
async def actualizar_hoja_importada(cuaderno_id: str, sheet_id: str, data: HojaUpdate):
    """
    Actualiza una hoja importada (datos, columnas, nombre).
    Cada cambio se persiste en el JSON y la hoja pasa a origen importado_editable.
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    updates = data.model_dump(exclude_unset=True)
    hoja = cuaderno.actualizar_hoja(
        sheet_id,
        datos=updates.get("datos"),
        columnas=updates.get("columnas"),
        nombre=updates.get("nombre"),
    )
    if not hoja:
        raise HTTPException(status_code=404, detail="Hoja no encontrada")
    storage.guardar(cuaderno)
    return {"success": True, "hoja": hoja.to_dict(), "message": "Hoja actualizada"}


@router.delete("/{cuaderno_id}/hojas/{sheet_id}")
async def eliminar_hoja_importada(cuaderno_id: str, sheet_id: str):
    """Elimina una hoja importada del cuaderno. Requiere confirmación en el cliente."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    if not cuaderno.eliminar_hoja(sheet_id):
        raise HTTPException(status_code=404, detail="Hoja no encontrada")
    storage.guardar(cuaderno)
    return {"success": True, "message": "Hoja eliminada"}


class RenameSheet(BaseModel):
    nombre: str


@router.patch("/{cuaderno_id}/hojas/{sheet_id}/rename")
async def renombrar_hoja(cuaderno_id: str, sheet_id: str, data: RenameSheet):
    """Renombra una hoja importada del cuaderno."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    hoja = cuaderno.obtener_hoja(sheet_id)
    if not hoja:
        raise HTTPException(status_code=404, detail="Hoja no encontrada")
    hoja.nombre = data.nombre.strip()
    cuaderno._actualizar_modificacion()
    storage.guardar(cuaderno)
    return {"success": True, "message": f"Hoja renombrada a '{hoja.nombre}'", "nombre": hoja.nombre}


@router.patch("/{cuaderno_id}/cell")
async def actualizar_celda(cuaderno_id: str, data: CellPatch):
    """
    Edición atómica de una celda. Cada cambio se aplica en memoria, persiste en JSON y genera backup.
    - Hojas base (parcelas, productos, tratamientos): sheet_id = tipo, row = entity_id, column = field key.
    - Hojas importadas: sheet_id = uuid de la hoja, row = índice fila, column = índice columna.
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    sheet_id = data.sheet_id
    if sheet_id in ("parcelas", "productos", "tratamientos", "fertilizantes", "cosecha", "asesoramiento"):
        raw_row = data.row
        entity_id = str(raw_row)
        column = str(data.column).strip()

        # Tolerancia: si row viene como índice visual, resolver a entity_id real.
        if sheet_id == "parcelas":
            items = cuaderno.parcelas
        elif sheet_id == "productos":
            items = cuaderno.productos
        elif sheet_id == "tratamientos":
            items = cuaderno.tratamientos
        elif sheet_id == "fertilizantes":
            items = getattr(cuaderno, "fertilizaciones", [])
        elif sheet_id == "cosecha":
            items = getattr(cuaderno, "cosechas", [])
        else:  # asesoramiento
            items = getattr(cuaderno, "asesoramientos", [])
        if items and entity_id and not any(getattr(it, "id", None) == entity_id for it in items):
            try:
                idx = int(raw_row)
                if 0 <= idx < len(items):
                    entity_id = str(getattr(items[idx], "id", entity_id))
            except (TypeError, ValueError):
                pass

        # Tolerancia: alias de columnas visibles en UI -> campo real de modelo
        aliases = {
            # parcelas
            "termino municipal": "termino_municipal",
            "término municipal": "termino_municipal",
            # comunes de tablas estructurales
            "parcela_nombres": "num_orden_parcelas",
            "parcelas": "num_orden_parcelas",
        }
        if sheet_id == "tratamientos":
            aliases.update({
                "fecha": "fecha_aplicacion",
                "problema": "problema_fitosanitario",
                "operador": "aplicador",
            })
        col_norm = column.lower()
        mapped_column = aliases.get(col_norm, column)

        ok = cuaderno.aplicar_celda_estructural(sheet_id, entity_id, mapped_column, data.value)
    else:
        try:
            row_idx = int(data.row)
            col_idx = int(data.column)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="En hojas importadas row y column deben ser índices numéricos")
        ok = cuaderno.aplicar_celda(sheet_id, row_idx, col_idx, data.value)
    if not ok:
        raise HTTPException(status_code=400, detail="No se pudo aplicar el cambio (entidad o campo no encontrado)")

    if sheet_id == "parcelas":
        _normalizar_parcelas_cultivo_variante(cuaderno)
        _asegurar_asesoramiento_automatico(cuaderno)

    _guardar_cuaderno(storage, cuaderno)
    return {"success": True, "timestamp": datetime.now().isoformat()}


# ============================================
# PARCELA ENDPOINTS
# ============================================

@router.get("/{cuaderno_id}/parcelas")
async def listar_parcelas(cuaderno_id: str, activas_only: bool = True):
    """Lista las parcelas de un cuaderno"""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    parcelas = cuaderno.parcelas
    if activas_only:
        parcelas = [p for p in parcelas if p.activa]
    
    return {"parcelas": [p.to_dict() for p in parcelas]}


@router.post("/{cuaderno_id}/parcelas")
async def crear_parcela(cuaderno_id: str, data: ParcelaCreate):
    """Añade una parcela al cuaderno"""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    # Auto-assign num_orden único si no se especifica o ya existe
    existing_ordenes = {p.num_orden for p in cuaderno.parcelas if isinstance(p.num_orden, int) and p.num_orden > 0}
    num_orden = data.num_orden
    if num_orden <= 0 or num_orden in existing_ordenes:
        num_orden = max(existing_ordenes, default=0) + 1
        while num_orden in existing_ordenes:
            num_orden += 1

    parcela = Parcela(
        nombre=data.nombre,
        num_orden=num_orden,
        referencia_catastral=data.referencia_catastral,
        superficie_ha=data.superficie_ha,
        cultivo=data.cultivo,
        variedad=data.variedad,
        municipio=data.municipio,
        provincia=data.provincia,
        notas=data.notas
    )
    _normalizar_cultivo_variante_parcela(parcela)

    cuaderno.agregar_parcela(parcela)
    _asegurar_asesoramiento_automatico(cuaderno)
    _guardar_cuaderno(storage, cuaderno)

    return {
        "success": True,
        "parcela": parcela.to_dict(),
        "message": "Parcela añadida correctamente"
    }


@router.put("/{cuaderno_id}/parcelas/{parcela_id}")
async def actualizar_parcela(cuaderno_id: str, parcela_id: str, data: ParcelaUpdate):
    """Actualiza una parcela"""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    parcela = cuaderno.obtener_parcela(parcela_id)
    if not parcela:
        raise HTTPException(status_code=404, detail="Parcela no encontrada")
    
    # Actualizar campos
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if hasattr(parcela, key):
            setattr(parcela, key, value)
    _normalizar_cultivo_variante_parcela(parcela)

    _asegurar_asesoramiento_automatico(cuaderno)
    cuaderno._actualizar_modificacion()
    _guardar_cuaderno(storage, cuaderno)

    return {
        "success": True,
        "parcela": parcela.to_dict(),
        "message": "Parcela actualizada"
    }


@router.delete("/{cuaderno_id}/parcelas/{parcela_id}")
async def eliminar_parcela(cuaderno_id: str, parcela_id: str):
    """Elimina una parcela del cuaderno."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    if not cuaderno.eliminar_parcela(parcela_id):
        raise HTTPException(status_code=404, detail="Parcela no encontrada")
    _asegurar_asesoramiento_automatico(cuaderno)
    _guardar_cuaderno(storage, cuaderno)
    return {"success": True, "message": "Parcela eliminada"}


@router.get("/{cuaderno_id}/parcelas/{parcela_id}/historico")
async def obtener_historico_parcela(cuaderno_id: str, parcela_id: str):
    """Obtiene el histórico de tratamientos de una parcela"""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    parcela = cuaderno.obtener_parcela(parcela_id)
    if not parcela:
        raise HTTPException(status_code=404, detail="Parcela no encontrada")
    
    tratamientos = cuaderno.obtener_tratamientos_parcela(parcela_id)
    
    return {
        "parcela": parcela.to_dict(),
        "tratamientos": [t.to_dict() for t in tratamientos],
        "total": len(tratamientos)
    }


# ============================================
# PRODUCTO ENDPOINTS
# ============================================

@router.get("/{cuaderno_id}/productos")
async def listar_productos(cuaderno_id: str):
    """Lista los productos de un cuaderno"""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    return {"productos": [p.to_dict() for p in cuaderno.productos]}


@router.post("/{cuaderno_id}/productos")
async def crear_producto(cuaderno_id: str, data: ProductoCreate):
    """Añade un producto al inventario"""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    producto = ProductoFitosanitario(
        nombre_comercial=data.nombre_comercial,
        numero_registro=data.numero_registro,
        materia_activa=data.materia_activa,
        formulacion=data.formulacion,
        tipo=TipoProducto(data.tipo) if data.tipo else TipoProducto.FITOSANITARIO,
        numero_lote=data.numero_lote,
        cantidad_adquirida=data.cantidad_adquirida,
        unidad=data.unidad,
        fecha_adquisicion=data.fecha_adquisicion,
        proveedor=data.proveedor,
        fecha_caducidad=data.fecha_caducidad,
        notas=data.notas
    )
    
    cuaderno.agregar_producto(producto)
    _guardar_cuaderno(storage, cuaderno)
    # Publicar automáticamente en catálogo global (best-effort).
    try:
        get_catalogo().upsert({
            "nombre_comercial": producto.nombre_comercial,
            "numero_registro": producto.numero_registro,
            "materia_activa": producto.materia_activa,
            "formulacion": producto.formulacion,
            "tipo": producto.tipo.value if hasattr(producto.tipo, "value") else str(producto.tipo or "fitosanitario"),
            "unidad": producto.unidad or "L",
            "proveedor": producto.proveedor,
            "notas": producto.notas,
        })
    except Exception:
        pass
    _cache_invalidate(["catalog_global::", "stock_global::all"])
    await _stock_ws_hub.publish_stock_changed(cuaderno_id)
    
    return {
        "success": True,
        "producto": producto.to_dict(),
        "message": "Producto añadido correctamente"
    }


@router.delete("/{cuaderno_id}/productos/{producto_id}")
async def eliminar_producto(cuaderno_id: str, producto_id: str):
    """Elimina un producto del inventario del cuaderno."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    if not cuaderno.eliminar_producto(producto_id):
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    _guardar_cuaderno(storage, cuaderno)
    _cache_invalidate(["catalog_global::", "stock_global::all"])
    return {"success": True, "message": "Producto eliminado"}


# ============================================
# CATÁLOGO GLOBAL DE PRODUCTOS (compartido)
# ============================================

def _sync_catalogo_desde_cuadernos() -> None:
    """
    Sincroniza productos ya existentes en cuadernos hacia el catálogo global.
    Idempotente por upsert (nombre+registro).
    """
    try:
        storage = get_storage()
        catalogo = get_catalogo()
        for meta in storage.listar():
            cid = meta.get("id")
            if not cid:
                continue
            cuaderno = storage.cargar(cid)
            if not cuaderno:
                continue
            for prod in (cuaderno.productos or []):
                try:
                    catalogo.upsert({
                        "nombre_comercial": prod.nombre_comercial,
                        "numero_registro": prod.numero_registro,
                        "materia_activa": prod.materia_activa,
                        "formulacion": prod.formulacion,
                        "tipo": prod.tipo.value if hasattr(prod.tipo, "value") else str(prod.tipo or "fitosanitario"),
                        "unidad": prod.unidad or "L",
                        "proveedor": prod.proveedor,
                        "notas": prod.notas,
                    })
                except Exception:
                    continue
    except Exception:
        pass


def _collect_productos_desde_cuadernos(q: str = "") -> List[Dict[str, Any]]:
    """
    Recolecta productos directamente desde todos los cuadernos (sin depender del
    backend del catálogo global), deduplicando por nombre+registro+unidad.
    Usa caché en memoria cuando está disponible para evitar round-trips a Supabase.
    """
    storage = get_storage()
    q_norm = (q or "").strip().lower()
    agg: Dict[str, Dict[str, Any]] = {}

    for meta in storage.listar():
        cid = meta.get("id")
        if not cid:
            continue
        # Intentar servir desde caché en memoria antes de ir a Supabase
        cached = _cache_get(f"cuaderno::{cid}")
        if cached is not None:
            productos_raw = cached.get("cuaderno", {}).get("productos", [])
            for p in productos_raw:
                nombre = (p.get("nombre_comercial") or "").strip()
                registro = (p.get("numero_registro") or "").strip()
                materia = (p.get("materia_activa") or "").strip()
                formulacion = (p.get("formulacion") or "").strip()
                unidad = (p.get("unidad") or "L").strip()
                proveedor = (p.get("proveedor") or "").strip()
                notas = (p.get("notas") or "").strip()
                if q_norm and not any(q_norm in (v or "").lower() for v in (nombre, registro, materia, formulacion, proveedor, notas)):
                    continue
                key = f"{nombre.lower()}||{registro.lower()}||{unidad.lower()}"
                if key not in agg:
                    agg[key] = {
                        "id": f"cuaderno::{key}",
                        "nombre_comercial": nombre,
                        "numero_registro": registro,
                        "materia_activa": materia,
                        "formulacion": formulacion,
                        "tipo": p.get("tipo") or "fitosanitario",
                        "unidad": unidad,
                        "proveedor": proveedor,
                        "notas": notas,
                        "created_at": "",
                        "updated_at": "",
                    }
            continue
        cuaderno = storage.cargar(cid)
        if not cuaderno:
            continue
        for p in (cuaderno.productos or []):
            nombre = (p.nombre_comercial or "").strip()
            registro = (p.numero_registro or "").strip()
            materia = (p.materia_activa or "").strip()
            formulacion = (p.formulacion or "").strip()
            unidad = (p.unidad or "L").strip()
            proveedor = (p.proveedor or "").strip()
            notas = (p.notas or "").strip()
            if q_norm:
                hay_match = any(
                    q_norm in (v or "").lower()
                    for v in (nombre, registro, materia, formulacion, proveedor, notas)
                )
                if not hay_match:
                    continue

            key = f"{nombre.lower()}||{registro.lower()}||{unidad.lower()}"
            if key not in agg:
                agg[key] = {
                    "id": f"cuaderno::{key}",
                    "nombre_comercial": nombre,
                    "numero_registro": registro,
                    "materia_activa": materia,
                    "formulacion": formulacion,
                    "tipo": p.tipo.value if hasattr(p.tipo, "value") else str(p.tipo or "fitosanitario"),
                    "unidad": unidad,
                    "proveedor": proveedor,
                    "notas": notas,
                    "created_at": "",
                    "updated_at": "",
                }
            else:
                row = agg[key]
                if not row.get("materia_activa") and materia:
                    row["materia_activa"] = materia
                if not row.get("formulacion") and formulacion:
                    row["formulacion"] = formulacion
                if not row.get("proveedor") and proveedor:
                    row["proveedor"] = proveedor
                if not row.get("notas") and notas:
                    row["notas"] = notas

    return sorted(agg.values(), key=lambda r: (r.get("nombre_comercial") or "").lower())


@router.get("/catalog/global/productos")
async def listar_catalogo_productos(background_tasks: BackgroundTasks, q: str = "", limit: int = Query(default=50, le=200)):
    """Busca productos en el catálogo GLOBAL (compartido entre todos los cuadernos)."""
    global _LAST_CATALOGO_SYNC_TS
    cache_key = f"catalog_global::{(q or '').strip().lower()}::{int(limit or 50)}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    try:
        # Sync en background (no bloquea la petición): se ejecuta después de responder.
        now = time.time()
        if now - _LAST_CATALOGO_SYNC_TS > 60:
            _LAST_CATALOGO_SYNC_TS = now
            background_tasks.add_task(_sync_catalogo_desde_cuadernos)
        productos_catalogo = get_catalogo().listar(q=q, limit=limit)
    except Exception:
        productos_catalogo = []

    try:
        productos_cuadernos = _collect_productos_desde_cuadernos(q=q)
        merged: Dict[str, Dict[str, Any]] = {}
        for p in productos_catalogo:
            key = f"{(p.get('nombre_comercial') or '').strip().lower()}||{(p.get('numero_registro') or '').strip().lower()}||{(p.get('unidad') or 'L').strip().lower()}"
            merged[key] = p
        for p in productos_cuadernos:
            key = f"{(p.get('nombre_comercial') or '').strip().lower()}||{(p.get('numero_registro') or '').strip().lower()}||{(p.get('unidad') or 'L').strip().lower()}"
            if key not in merged:
                merged[key] = p
        productos = sorted(merged.values(), key=lambda r: (r.get("nombre_comercial") or "").lower())[: max(1, int(limit or 50))]
        payload = {"productos": productos, "total": len(productos)}
        _cache_set(cache_key, payload, ttl_sec=30)
        return payload
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error listando catálogo: {e}")


@router.post("/catalog/global/productos")
async def crear_catalogo_producto(data: CatalogoProductoCreate):
    """Crea o actualiza (upsert por nombre+registro) un producto en el catálogo global."""
    if not (data.nombre_comercial or "").strip():
        raise HTTPException(status_code=400, detail="nombre_comercial es obligatorio")
    try:
        producto = get_catalogo().upsert(data.model_dump())
        _cache_invalidate(["catalog_global::"])
        return {"success": True, "producto": producto}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creando producto en catálogo: {e}")


@router.put("/catalog/global/productos/{producto_id}")
async def actualizar_catalogo_producto(producto_id: str, data: CatalogoProductoUpdate):
    """Actualiza un producto del catálogo global."""
    patch = {k: v for k, v in data.model_dump().items() if v is not None}
    try:
        actualizado = get_catalogo().actualizar(producto_id, patch)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error actualizando catálogo: {e}")
    if not actualizado:
        raise HTTPException(status_code=404, detail="Producto no encontrado en catálogo")
    _cache_invalidate(["catalog_global::"])
    return {"success": True, "producto": actualizado}


@router.delete("/catalog/global/productos/{producto_id}")
async def eliminar_catalogo_producto(producto_id: str):
    """Elimina un producto del catálogo global."""
    try:
        ok = get_catalogo().eliminar(producto_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error eliminando del catálogo: {e}")
    if not ok:
        raise HTTPException(status_code=404, detail="Producto no encontrado en catálogo")
    _cache_invalidate(["catalog_global::"])
    return {"success": True, "message": "Producto eliminado del catálogo global"}


@router.post("/{cuaderno_id}/productos/{producto_id}/publicar-catalogo")
async def publicar_producto_en_catalogo(cuaderno_id: str, producto_id: str):
    """Publica un producto del inventario del cuaderno en el catálogo global (upsert)."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    prod = cuaderno.obtener_producto(producto_id)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado en el cuaderno")
    try:
        publicado = get_catalogo().upsert({
            "nombre_comercial": prod.nombre_comercial,
            "numero_registro": prod.numero_registro,
            "materia_activa": prod.materia_activa,
            "formulacion": prod.formulacion,
            "tipo": prod.tipo.value if hasattr(prod.tipo, "value") else str(prod.tipo or "fitosanitario"),
            "unidad": prod.unidad or "L",
            "proveedor": prod.proveedor,
            "notas": prod.notas,
        })
        return {"success": True, "producto": publicado, "message": "Publicado en el catálogo global"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error publicando en catálogo: {e}")


@router.post("/{cuaderno_id}/catalogo/importar")
async def importar_desde_catalogo(cuaderno_id: str, data: ImportarDesdeCatalogoRequest):
    """
    Importa un producto del catálogo global al inventario del cuaderno.
    Si ya existe uno equivalente (mismo nombre+registro), lo reusa en vez de duplicar.
    Devuelve el producto resultante dentro del cuaderno.
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    cat = get_catalogo().obtener(data.catalogo_id)
    if not cat:
        raise HTTPException(status_code=404, detail="Producto no encontrado en el catálogo")

    def _clave(nombre: str, registro: str) -> str:
        return f"{(nombre or '').strip().lower()}||{(registro or '').strip().lower()}"

    clave_cat = _clave(cat.get("nombre_comercial", ""), cat.get("numero_registro", ""))
    existente = next(
        (p for p in cuaderno.productos if _clave(p.nombre_comercial, p.numero_registro) == clave_cat),
        None,
    )
    if existente:
        return {
            "success": True,
            "producto": existente.to_dict(),
            "reused": True,
            "message": "Ya existía en el cuaderno; reutilizado.",
        }

    try:
        tipo_val = cat.get("tipo") or "fitosanitario"
        producto = ProductoFitosanitario(
            nombre_comercial=cat.get("nombre_comercial", ""),
            numero_registro=cat.get("numero_registro", ""),
            materia_activa=cat.get("materia_activa", ""),
            formulacion=cat.get("formulacion", ""),
            tipo=TipoProducto(tipo_val) if tipo_val in [t.value for t in TipoProducto] else TipoProducto.FITOSANITARIO,
            unidad=cat.get("unidad", "L") or "L",
            proveedor=cat.get("proveedor", ""),
            notas=cat.get("notas", ""),
        )
        cuaderno.agregar_producto(producto)
        storage.guardar(cuaderno)
        return {
            "success": True,
            "producto": producto.to_dict(),
            "reused": False,
            "message": "Producto importado del catálogo global",
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error importando del catálogo: {e}")


# ============================================
# TRATAMIENTO ENDPOINTS
# ============================================

@router.get("/{cuaderno_id}/tratamientos")
async def listar_tratamientos(cuaderno_id: str, 
                               limit: int = Query(default=50, le=500),
                               offset: int = 0):
    """Lista los tratamientos de un cuaderno"""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    def _orden_trat_list(t: Tratamiento):
        try:
            o = int(str(t.num_orden_parcelas).split(",")[0].strip()) if str(t.num_orden_parcelas or "").strip() else 9999
        except (ValueError, TypeError, IndexError):
            o = 9999
        return (
            (t.cultivo_especie or "").strip().lower(),
            o,
            t.fecha_aplicacion or "",
            t.id or "",
        )

    tratamientos = sorted(cuaderno.tratamientos, key=_orden_trat_list)
    
    # Paginación
    total = len(tratamientos)
    tratamientos = tratamientos[offset:offset + limit]
    
    return {
        "tratamientos": [t.to_dict() for t in tratamientos],
        "total": total,
        "limit": limit,
        "offset": offset
    }


def _validar_tratamiento_create(data: TratamientoCreate, cuaderno: CuadernoExplotacion) -> None:
    """Valida modelo legal mínimo: fecha, parcelas, productos con registro y dosis (snapshot defendible en inspección)."""
    if not (data.fecha_aplicacion or "").strip():
        raise HTTPException(status_code=400, detail="La fecha de aplicación es obligatoria.")
    if not data.parcela_ids:
        raise HTTPException(status_code=400, detail="Debe indicar al menos una parcela.")
    if not data.productos:
        raise HTTPException(status_code=400, detail="Debe indicar al menos un producto con dosis.")
    for i, p in enumerate(data.productos):
        if p.dosis is None or (isinstance(p.dosis, (int, float)) and p.dosis <= 0):
            raise HTTPException(
                status_code=400,
                detail=f"El producto {i + 1} debe tener una dosis mayor que 0."
            )
        # Snapshot obligatorio: o viene de inventario (producto_id) o se envía nombre + registro
        if p.producto_id:
            if not cuaderno.obtener_producto(p.producto_id):
                raise HTTPException(status_code=400, detail=f"Producto {i + 1}: ID no encontrado en el inventario.")
        else:
            if not (p.nombre_comercial or "").strip():
                raise HTTPException(status_code=400, detail=f"El producto {i + 1} debe tener nombre o estar seleccionado del inventario.")
            if not (p.numero_registro or "").strip():
                raise HTTPException(status_code=400, detail=f"El producto {i + 1} debe tener Nº registro fitosanitario.")
    for parcela_id in data.parcela_ids:
        if not cuaderno.obtener_parcela(parcela_id):
            raise HTTPException(status_code=400, detail=f"Parcela no encontrada: {parcela_id}")


@router.post("/{cuaderno_id}/tratamientos")
async def crear_tratamiento(cuaderno_id: str, data: TratamientoCreate, background_tasks: BackgroundTasks):
    """
    Registra un nuevo tratamiento desglosado: 1 línea por parcela × 1 línea por producto.
    Cache-first load + persist en background → respuesta inmediata.
    """
    storage = get_storage()

    # 1. Intentar cargar desde caché para evitar round-trip a Supabase (~8s)
    cache_key = f"cuaderno::{cuaderno_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        cuaderno = CuadernoExplotacion.from_dict(cached["cuaderno"])
    else:
        cuaderno = storage.cargar(cuaderno_id)

    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    _validar_tratamiento_create(data, cuaderno)
    
    productos_aplicados = []
    for prod_data in data.productos:
        prod = ProductoAplicado(
            producto_id=prod_data.producto_id or "",
            nombre_comercial=prod_data.nombre_comercial or "",
            numero_registro=prod_data.numero_registro or "",
            numero_lote=prod_data.numero_lote or "",
            problema_fitosanitario=(prod_data.plaga_enfermedad or "").strip(),
            dosis=prod_data.dosis,
            unidad_dosis=prod_data.unidad_dosis,
            caldo_hectarea=prod_data.caldo_hectarea,
            notas=prod_data.notas or ""
        )
        productos_aplicados.append(prod)
    
    plantilla = Tratamiento(
        fecha_aplicacion=data.fecha_aplicacion.strip(),
        parcela_ids=data.parcela_ids,
        productos=productos_aplicados,
        problema_fitosanitario=(data.plaga_enfermedad or "").strip(),
        plaga_enfermedad=data.plaga_enfermedad,
        justificacion=data.justificacion,
        metodo_aplicacion=data.metodo_aplicacion,
        condiciones_climaticas=data.condiciones_climaticas,
        operador=(data.operador or "1").strip(),
        equipo=(data.equipo or "1").strip(),
        eficacia=(data.eficacia or "BUENA").strip(),
        hora_inicio=data.hora_inicio,
        hora_fin=data.hora_fin,
        observaciones=data.observaciones,
        asesorado=data.asesorado,
        nombre_asesor_trat=data.nombre_asesor_trat or "",
        num_colegiado_asesor=data.num_colegiado_asesor or "",
        fecha_recomendacion_asesor=data.fecha_recomendacion_asesor or "",
        firma_asesor=data.firma_asesor or "",
        firma_cliente=data.firma_cliente or "",
    )
    
    creados = cuaderno.agregar_tratamiento_desglosado(plantilla)

    # Deducir stock de productos aplicados
    # Índice O(P) construido una sola vez → lookups O(1) por parcela_id
    parcelas_idx = {p.id: p for p in cuaderno.parcelas}
    superficie_total = sum(
        (parcelas_idx[pid].superficie_cultivada or parcelas_idx[pid].superficie_sigpac or 0.0)
        for pid in data.parcela_ids if pid in parcelas_idx
    )
    if superficie_total <= 0:
        superficie_total = float(data.superficie_tratada or 1.0)
    for prod_data in data.productos:
        if prod_data.producto_id:
            prod_obj = cuaderno.obtener_producto(prod_data.producto_id)
            if prod_obj and prod_obj.cantidad_adquirida > 0:
                consumo = float(prod_data.dosis or 0) * superficie_total
                prod_obj.cantidad_adquirida = max(0.0, round(prod_obj.cantidad_adquirida - consumo, 4))

    # 2. Auto-publicar productos en catálogo global si no existen ya
    #    Se hace best-effort en background para no añadir latencia
    def _sync_prods_al_catalogo_global():
        catalogo = get_catalogo()
        for prod in productos_aplicados:
            nombre = (prod.nombre_comercial or "").strip()
            if not nombre:
                continue
            try:
                catalogo.upsert({
                    "nombre_comercial": nombre,
                    "numero_registro": prod.numero_registro or "",
                    "numero_lote": prod.numero_lote or "",
                    "tipo": "fitosanitario",
                })
            except Exception:
                pass

    background_tasks.add_task(_sync_prods_al_catalogo_global)

    # 3. Actualizar caché con el estado nuevo ANTES de responder
    #    → el próximo GET sirve datos frescos sin ir a Supabase
    _cache_set(cache_key, {"cuaderno": cuaderno.to_dict()}, ttl_sec=120)

    # 4. Persistir en Supabase en segundo plano (no bloquea la respuesta)
    background_tasks.add_task(storage.guardar, cuaderno)

    return {
        "success": True,
        "tratamientos": [t.to_dict() for t in creados],
        "message": f"{len(creados)} línea(s) creada(s) ({len(data.parcela_ids)} parcela(s) × {len(productos_aplicados)} producto(s))"
    }


@router.get("/{cuaderno_id}/tratamientos/{tratamiento_id}")
async def obtener_tratamiento(cuaderno_id: str, tratamiento_id: str):
    """Obtiene un tratamiento por ID."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    t = cuaderno.obtener_tratamiento(tratamiento_id)
    if not t:
        raise HTTPException(status_code=404, detail="Tratamiento no encontrado")
    return {"tratamiento": t.to_dict()}


@router.put("/{cuaderno_id}/tratamientos/{tratamiento_id}")
async def actualizar_tratamiento(cuaderno_id: str, tratamiento_id: str, data: TratamientoUpdate):
    """Actualiza un tratamiento existente."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    t = cuaderno.obtener_tratamiento(tratamiento_id)
    if not t:
        raise HTTPException(status_code=404, detail="Tratamiento no encontrado")
    
    updates = data.model_dump(exclude_unset=True)
    if "productos" in updates:
        raw = updates["productos"]
        productos_aplicados = []
        for p in raw:
            d = p if isinstance(p, dict) else p.model_dump()
            pa = ProductoAplicado(
                producto_id=d.get("producto_id", ""),
                nombre_comercial=d.get("nombre_comercial", ""),
                numero_registro=d.get("numero_registro", ""),
                numero_lote=d.get("numero_lote", ""),
                problema_fitosanitario=d.get("plaga_enfermedad", d.get("problema_fitosanitario", "")),
                dosis=d.get("dosis", 0),
                unidad_dosis=d.get("unidad_dosis", "L/Ha"),
                caldo_hectarea=d.get("caldo_hectarea", 0),
                notas=d.get("notas", "")
            )
            productos_aplicados.append(pa)
        updates["productos"] = productos_aplicados
    cuaderno.actualizar_tratamiento(tratamiento_id, **updates)
    storage.guardar(cuaderno)
    t = cuaderno.obtener_tratamiento(tratamiento_id)
    return {"success": True, "tratamiento": t.to_dict()}


@router.delete("/{cuaderno_id}/tratamientos/{tratamiento_id}")
async def eliminar_tratamiento(cuaderno_id: str, tratamiento_id: str):
    """Elimina un tratamiento."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    if not cuaderno.eliminar_tratamiento(tratamiento_id):
        raise HTTPException(status_code=404, detail="Tratamiento no encontrado")
    storage.guardar(cuaderno)
    return {"success": True, "message": "Tratamiento eliminado"}


# ============================================
# FERTILIZACIONES
# ============================================

@router.post("/{cuaderno_id}/fertilizaciones")
async def crear_fertilizacion(cuaderno_id: str, data: FertilizacionCreate):
    """Añade un registro de fertilización."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    # Convertir parcela_ids a num_orden_parcelas
    num_orden_parcelas = ""
    if data.parcela_ids:
        ordenes = []
        for pid in data.parcela_ids:
            p = cuaderno.obtener_parcela(pid)
            if p and p.num_orden:
                ordenes.append(str(p.num_orden))
        num_orden_parcelas = ",".join(ordenes) if ordenes else ",".join(data.parcela_ids)

    fertilizacion = Fertilizacion(
        fecha_inicio=data.fecha_inicio,
        fecha_fin=data.fecha_fin,
        num_orden_parcelas=num_orden_parcelas,
        cultivo_especie=data.cultivo_especie,
        cultivo_variedad=data.cultivo_variedad,
        tipo_abono=data.tipo_abono,
        riqueza_npk=data.riqueza_npk,
        dosis=data.dosis,
        tipo_fertilizacion=data.tipo_fertilizacion,
        observaciones=data.observaciones,
    )
    cuaderno.agregar_fertilizacion(fertilizacion)
    storage.guardar(cuaderno)
    return {"fertilizacion": fertilizacion.to_dict()}


# ============================================
# COSECHAS
# ============================================

@router.post("/{cuaderno_id}/cosechas")
async def crear_cosecha(cuaderno_id: str, data: CosechaCreate):
    """Añade un registro de cosecha."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    # Convertir parcela_ids a num_orden_parcelas
    num_orden_parcelas = ""
    if data.parcela_ids:
        ordenes = []
        for pid in data.parcela_ids:
            p = cuaderno.obtener_parcela(pid)
            if p and p.num_orden:
                ordenes.append(str(p.num_orden))
        num_orden_parcelas = ",".join(ordenes) if ordenes else ",".join(data.parcela_ids)

    cosecha = Cosecha(
        fecha=data.fecha,
        producto=data.producto,
        cantidad_kg=data.cantidad_kg,
        num_orden_parcelas=num_orden_parcelas,
        num_albaran=data.num_albaran,
        num_lote=data.num_lote,
        cliente_nombre=data.cliente_nombre,
        cliente_nif=data.cliente_nif,
        cliente_direccion=data.cliente_direccion,
        cliente_rgseaa=data.cliente_rgseaa,
    )
    cuaderno.agregar_cosecha(cosecha)
    storage.guardar(cuaderno)
    return {"cosecha": cosecha.to_dict()}


# ============================================
# ASESORAMIENTO (Hoja 3.2)
# ============================================

@router.post("/{cuaderno_id}/asesoramientos")
async def crear_asesoramiento(cuaderno_id: str, data: AsesoramientoCreate):
    """Añade un registro de asesoramiento fitosanitario (Hoja 3.2)."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    asesoramientos_creados: List[Asesoramiento] = []
    if data.parcela_ids:
        for pid in data.parcela_ids:
            parcela = cuaderno.obtener_parcela(pid)
            if not parcela:
                continue
            orden = str(parcela.num_orden) if isinstance(parcela.num_orden, int) and parcela.num_orden > 0 else pid
            sup_parcela = float(parcela.superficie_cultivada or parcela.superficie_ha or parcela.superficie_sigpac or 0)
            asesoramiento = Asesoramiento(
                fecha=data.fecha,
                num_orden_parcelas=orden,
                cultivo_especie=data.cultivo_especie or (parcela.especie or parcela.cultivo or ""),
                cultivo_variedad=data.cultivo_variedad or (parcela.variedad or ""),
                superficie_ha=round(sup_parcela, 2),
                nombre_asesor=data.nombre_asesor,
                num_habilitacion=data.num_habilitacion,
                tipo_asesoramiento=data.tipo_asesoramiento,
                recomendacion=data.recomendacion,
                observaciones=data.observaciones,
            )
            cuaderno.agregar_asesoramiento(asesoramiento)
            asesoramientos_creados.append(asesoramiento)
    else:
        asesoramiento = Asesoramiento(
            fecha=data.fecha,
            num_orden_parcelas="",
            cultivo_especie=data.cultivo_especie,
            cultivo_variedad=data.cultivo_variedad,
            superficie_ha=round(float(data.superficie_ha or 0), 2),
            nombre_asesor=data.nombre_asesor,
            num_habilitacion=data.num_habilitacion,
            tipo_asesoramiento=data.tipo_asesoramiento,
            recomendacion=data.recomendacion,
            observaciones=data.observaciones,
        )
        cuaderno.agregar_asesoramiento(asesoramiento)
        asesoramientos_creados.append(asesoramiento)

    storage.guardar(cuaderno)
    first = asesoramientos_creados[0] if asesoramientos_creados else None
    return {
        "asesoramiento": first.to_dict() if first else None,
        "asesoramientos": [a.to_dict() for a in asesoramientos_creados],
        "total": len(asesoramientos_creados),
    }


@router.delete("/{cuaderno_id}/asesoramientos/{asesoramiento_id}")
async def eliminar_asesoramiento(cuaderno_id: str, asesoramiento_id: str):
    """Elimina un registro de asesoramiento."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    ok = cuaderno.eliminar_asesoramiento(asesoramiento_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Asesoramiento no encontrado")
    storage.guardar(cuaderno)
    return {"success": True}


class StockEntradaCreate(BaseModel):
    producto_id: Optional[str] = None
    nombre_comercial: str = ""
    cantidad: float
    unidad: str = "L"
    fecha: str = ""
    proveedor: str = ""
    num_albaran: str = ""
    num_lote: str = ""
    precio_unidad: float = 0.0
    notas: str = ""

class StockEntradaUpdate(BaseModel):
    cantidad: Optional[float] = None
    fecha: Optional[str] = None
    proveedor: Optional[str] = None
    num_albaran: Optional[str] = None
    num_lote: Optional[str] = None
    precio_unidad: Optional[float] = None
    notas: Optional[str] = None


@router.websocket("/ws/stock/global")
async def ws_stock_global(websocket: WebSocket):
    await _stock_ws_hub.connect_global(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        _stock_ws_hub.disconnect_global(websocket)
    except Exception:
        _stock_ws_hub.disconnect_global(websocket)


@router.websocket("/ws/stock/{cuaderno_id}")
async def ws_stock_cuaderno(websocket: WebSocket, cuaderno_id: str):
    await _stock_ws_hub.connect_cuaderno(websocket, cuaderno_id)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        _stock_ws_hub.disconnect_cuaderno(websocket, cuaderno_id)
    except Exception:
        _stock_ws_hub.disconnect_cuaderno(websocket, cuaderno_id)


@router.get("/stock/global")
@router.get("/catalog/stock-global")
async def get_stock_global():
    """Resumen global de stock agregando todos los cuadernos."""
    cache_key = "stock_global::all"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    storage = get_storage()
    from collections import defaultdict
    import re

    def _to_float_dosis(value: Any) -> float:
        """Convierte dosis numérica o texto ('1.6 L/Ha', '1,6', etc.) a float seguro."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s:
            return 0.0
        # Extraer primer número con coma o punto decimal.
        m = re.search(r"[-+]?\d+(?:[.,]\d+)?", s)
        if not m:
            return 0.0
        try:
            return float(m.group(0).replace(",", "."))
        except Exception:
            return 0.0

    agg = {}

    for meta in storage.listar():
        cid = meta.get("id")
        if not cid:
            continue
        # Usar caché en memoria primero para evitar round-trips a Supabase
        cached_c = _cache_get(f"cuaderno::{cid}")
        if cached_c is not None:
            cdata = cached_c.get("cuaderno", {})
            from .models import CuadernoExplotacion
            try:
                cuaderno = CuadernoExplotacion.from_dict(cdata)
            except Exception:
                cuaderno = storage.cargar(cid)
        else:
            cuaderno = storage.cargar(cid)
        if not cuaderno:
            continue

        uso_por_producto: dict = defaultdict(float)
        for t in cuaderno.tratamientos:
            sup = t.superficie_tratada or 0.0
            for pa in t.productos:
                if pa.producto_id:
                    uso_por_producto[pa.producto_id] += _to_float_dosis(pa.dosis) * sup

        entradas_por_producto = defaultdict(float)
        for e in cuaderno.stock_entradas:
            if e.producto_id:
                entradas_por_producto[e.producto_id] += float(e.cantidad or 0)

        for p in cuaderno.productos:
            key = ((p.nombre_comercial or "").strip().lower(), (p.numero_registro or "").strip().lower(), (p.unidad or "L").strip())
            if key not in agg:
                agg[key] = {
                    "producto_id": p.id,
                    "nombre_comercial": p.nombre_comercial,
                    "numero_registro": p.numero_registro,
                    "unidad": p.unidad or "L",
                    "stock_actual": 0.0,
                    "total_entradas": 0.0,
                    "total_consumido": 0.0,
                    "cuadernos_ids": set(),
                }
            row = agg[key]
            row["stock_actual"] += float(p.cantidad_adquirida or 0)
            row["total_entradas"] += float(entradas_por_producto.get(p.id, 0.0))
            row["total_consumido"] += float(uso_por_producto.get(p.id, 0.0))
            row["cuadernos_ids"].add(cid)

    productos = []
    for row in agg.values():
        total_entradas = round(row["total_entradas"], 4)
        stock_actual = round(row["stock_actual"], 4)
        total_consumido = round(row["total_consumido"], 4)
        if total_entradas > 0:
            semaforo = "rojo" if stock_actual <= 0 else "amarillo" if stock_actual < total_entradas * 0.2 else "verde"
        else:
            semaforo = "rojo" if stock_actual <= 0 else "amarillo" if stock_actual < 2 else "verde"
        productos.append({
            "producto_id": row["producto_id"],
            "nombre_comercial": row["nombre_comercial"],
            "numero_registro": row["numero_registro"],
            "unidad": row["unidad"],
            "stock_actual": stock_actual,
            "total_entradas": total_entradas,
            "total_consumido": total_consumido,
            "cuadernos_count": len(row["cuadernos_ids"]),
            "semaforo": semaforo,
        })

    productos.sort(key=lambda x: (x["nombre_comercial"] or "").lower())
    payload = {"productos": productos, "total": len(productos)}
    _cache_set(cache_key, payload, ttl_sec=300)
    return payload


@router.get("/{cuaderno_id}/stock")
async def get_stock(cuaderno_id: str):
    """Resumen de stock: entradas + stock actual por producto + movimientos de tratamientos."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    # Calcular total usado por producto a través de tratamientos
    from collections import defaultdict
    import re

    def _to_float_dosis(value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s:
            return 0.0
        m = re.search(r"[-+]?\d+(?:[.,]\d+)?", s)
        if not m:
            return 0.0
        try:
            return float(m.group(0).replace(",", "."))
        except Exception:
            return 0.0

    uso_por_producto: dict = defaultdict(float)
    for t in cuaderno.tratamientos:
        sup = t.superficie_tratada or 0.0
        for pa in t.productos:
            if pa.producto_id:
                uso_por_producto[pa.producto_id] += _to_float_dosis(pa.dosis) * sup

    # Construir índice entradas por producto en O(E) antes del bucle — evita O(P×E)
    entradas_por_producto: dict = defaultdict(float)
    for e in cuaderno.stock_entradas:
        if e.producto_id:
            entradas_por_producto[e.producto_id] += e.cantidad

    productos_resumen = []
    for p in cuaderno.productos:
        total_entradas = entradas_por_producto.get(p.id, 0.0)
        total_usado = round(uso_por_producto.get(p.id, 0.0), 4)
        stock_actual = p.cantidad_adquirida
        if total_entradas > 0:
            semaforo = "rojo" if stock_actual <= 0 else "amarillo" if stock_actual < total_entradas * 0.2 else "verde"
        else:
            semaforo = "rojo" if stock_actual <= 0 else "amarillo" if stock_actual < 2 else "verde"
        productos_resumen.append({
            "producto_id": p.id,
            "nombre_comercial": p.nombre_comercial,
            "unidad": p.unidad,
            "proveedor": p.proveedor,
            "stock_actual": stock_actual,
            "total_entradas": total_entradas,
            "total_consumido": total_usado,
            "semaforo": semaforo,
        })

    entradas = [e.to_dict() for e in sorted(cuaderno.stock_entradas, key=lambda x: x.fecha or x.fecha_creacion, reverse=True)]
    return {"productos": productos_resumen, "entradas": entradas}


@router.post("/{cuaderno_id}/stock/entrada")
async def crear_stock_entrada(cuaderno_id: str, data: StockEntradaCreate):
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    prod = None
    if data.producto_id:
        prod = cuaderno.obtener_producto(data.producto_id)

    # Si no viene producto_id, intentar resolver por nombre.
    if not prod and (data.nombre_comercial or "").strip():
        nom = (data.nombre_comercial or "").strip().lower()
        prod = next((p for p in cuaderno.productos if (p.nombre_comercial or "").strip().lower() == nom), None)

    # Si no existe, crearlo automáticamente para permitir flujo rápido desde Stock.
    if not prod:
        nombre = (data.nombre_comercial or "").strip()
        if not nombre:
            raise HTTPException(status_code=400, detail="Debes indicar producto_id o nombre_comercial")
        prod = ProductoFitosanitario(
            nombre_comercial=nombre,
            numero_registro="",
            materia_activa="",
            formulacion="",
            tipo=TipoProducto.FITOSANITARIO,
            unidad=(data.unidad or "L"),
            proveedor=data.proveedor or "",
            notas="Creado automáticamente desde entrada de stock",
        )
        cuaderno.agregar_producto(prod)

    from .models import StockEntrada
    entrada = StockEntrada(
        producto_id=prod.id,
        nombre_comercial=prod.nombre_comercial,
        cantidad=data.cantidad,
        unidad=prod.unidad or data.unidad or "L",
        fecha=data.fecha,
        proveedor=data.proveedor or prod.proveedor,
        num_albaran=data.num_albaran,
        num_lote=data.num_lote,
        precio_unidad=data.precio_unidad,
        notas=data.notas,
    )
    cuaderno.agregar_stock_entrada(entrada)
    storage.guardar(cuaderno)
    _cache_invalidate(["stock_global::all"])
    await _stock_ws_hub.publish_stock_changed(cuaderno_id)
    return {"success": True, "entrada": entrada.to_dict(), "stock_actual": prod.cantidad_adquirida}


@router.put("/{cuaderno_id}/stock/entrada/{entrada_id}")
async def actualizar_stock_entrada(cuaderno_id: str, entrada_id: str, data: StockEntradaUpdate):
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    entrada = cuaderno.obtener_stock_entrada(entrada_id)
    if not entrada:
        raise HTTPException(status_code=404, detail="Entrada no encontrada")

    prod = cuaderno.obtener_producto(entrada.producto_id) if entrada.producto_id else None

    # Ajustar stock del producto si cambia cantidad
    if data.cantidad is not None and prod:
        nueva_cantidad = float(data.cantidad or 0.0)
        delta = nueva_cantidad - float(entrada.cantidad or 0.0)
        prod.cantidad_adquirida = max(0.0, round(float(prod.cantidad_adquirida or 0.0) + delta, 4))
        entrada.cantidad = nueva_cantidad

    if data.fecha is not None:
        entrada.fecha = data.fecha
    if data.proveedor is not None:
        entrada.proveedor = data.proveedor
    if data.num_albaran is not None:
        entrada.num_albaran = data.num_albaran
    if data.num_lote is not None:
        entrada.num_lote = data.num_lote
    if data.precio_unidad is not None:
        entrada.precio_unidad = float(data.precio_unidad or 0.0)
    if data.notas is not None:
        entrada.notas = data.notas

    cuaderno._actualizar_modificacion()
    storage.guardar(cuaderno)
    _cache_invalidate(["stock_global::all"])
    await _stock_ws_hub.publish_stock_changed(cuaderno_id)
    return {"success": True, "entrada": entrada.to_dict(), "stock_actual": (prod.cantidad_adquirida if prod else None)}


@router.delete("/{cuaderno_id}/stock/entrada/{entrada_id}")
async def eliminar_stock_entrada(cuaderno_id: str, entrada_id: str):
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    ok = cuaderno.eliminar_stock_entrada(entrada_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Entrada no encontrada")
    storage.guardar(cuaderno)
    _cache_invalidate(["stock_global::all"])
    await _stock_ws_hub.publish_stock_changed(cuaderno_id)
    return {"success": True}


@router.get("/{cuaderno_id}/alertas")
async def get_alertas(cuaderno_id: str):
    """Analiza el cuaderno y devuelve alertas inteligentes."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    alertas = []
    from datetime import datetime, timedelta
    hoy = datetime.now()

    CULTIVOS_ESPECIALES = ["PATATA", "REMOLACHA"]

    # 1. Stock bajo o agotado
    for p in cuaderno.productos:
        if p.cantidad_adquirida <= 0:
            alertas.append({"tipo": "stock_agotado", "severidad": "alta", "mensaje": f"Stock agotado: {p.nombre_comercial}", "producto_id": p.id, "accion": "ver_producto"})
        elif p.cantidad_adquirida < 2:
            alertas.append({"tipo": "stock_bajo", "severidad": "media", "mensaje": f"Stock bajo ({p.cantidad_adquirida:.1f} {p.unidad}): {p.nombre_comercial}", "producto_id": p.id, "accion": "ver_producto"})

    # 2. Productos con caducidad próxima (<30 días) o vencida
    for p in cuaderno.productos:
        if p.fecha_caducidad:
            try:
                fc = datetime.strptime(p.fecha_caducidad[:10], "%Y-%m-%d")
                dias = (fc - hoy).days
                if dias < 0:
                    alertas.append({"tipo": "caducado", "severidad": "alta", "mensaje": f"Caducado hace {abs(dias)} días: {p.nombre_comercial}", "producto_id": p.id, "accion": "ver_producto"})
                elif dias <= 30:
                    alertas.append({"tipo": "caducidad_proxima", "severidad": "media", "mensaje": f"Caduca en {dias} días: {p.nombre_comercial}", "producto_id": p.id, "accion": "ver_producto"})
            except Exception:
                pass

    # 3. Patata/Remolacha con superficie conjunta >5 ha sin asesoramiento
    parcelas_especiales, total_especial_ha = _parcelas_especiales_y_superficie(cuaderno)
    if total_especial_ha > 5:
        ordenes_requeridas = {
            str(p.num_orden)
            for p in parcelas_especiales
            if isinstance(p.num_orden, int) and p.num_orden > 0
        }
        ordenes_cubiertas = set()
        for a in cuaderno.asesoramientos:
            tokens = {
                tok.strip()
                for tok in str(a.num_orden_parcelas or "").replace(";", ",").split(",")
                if tok.strip()
            }
            ordenes_cubiertas.update(tokens)
        tiene_asesoramiento = (not ordenes_requeridas) or ordenes_requeridas.issubset(ordenes_cubiertas)
        if not tiene_asesoramiento:
            alertas.append({
                "tipo": "sin_asesoramiento",
                "severidad": "alta",
                "mensaje": f"Patata/Remolacha suma {total_especial_ha:.2f} ha (>5 ha) y requiere asesoramiento fitosanitario",
                "accion": "crear_asesoramiento",
            })

    # 4. Parcelas con Patata/Remolacha con >5 tratamientos
    from collections import defaultdict
    trats_por_parcela = defaultdict(int)
    for t in cuaderno.tratamientos:
        for pid in (t.parcela_ids or []):
            trats_por_parcela[pid] += 1
    for parcela in cuaderno.parcelas:
        cultivo = (parcela.especie or parcela.cultivo or "").upper()
        if any(c in cultivo for c in CULTIVOS_ESPECIALES) and trats_por_parcela[parcela.id] > 5:
            alertas.append({"tipo": "muchos_tratamientos_especiales", "severidad": "info", "mensaje": f"Parcela {parcela.nombre or parcela.num_orden} ({cultivo}) tiene {trats_por_parcela[parcela.id]} tratamientos — revisa el registro de Asesoramiento", "parcela_id": parcela.id, "accion": "ver_asesoramiento"})

    # 5. Parcelas sin tratamiento en >60 días
    # Construir índice parcela_id → fecha_max en O(T) antes del bucle de parcelas
    ultima_fecha_por_parcela: dict = {}
    for t in cuaderno.tratamientos:
        if not t.fecha_aplicacion:
            continue
        fecha_str = t.fecha_aplicacion[:10]
        for pid in (t.parcela_ids or []):
            if pid not in ultima_fecha_por_parcela or fecha_str > ultima_fecha_por_parcela[pid]:
                ultima_fecha_por_parcela[pid] = fecha_str

    for parcela in cuaderno.parcelas:
        if not parcela.activa:
            continue
        ultima_fecha_str = ultima_fecha_por_parcela.get(parcela.id)
        if ultima_fecha_str:
            try:
                ultima = datetime.strptime(ultima_fecha_str, "%Y-%m-%d")
                if (hoy - ultima).days > 60:
                    alertas.append({"tipo": "sin_tratamiento_reciente", "severidad": "baja", "mensaje": f"Parcela '{parcela.nombre or parcela.num_orden}' sin tratamiento hace {(hoy - ultima).days} días", "parcela_id": parcela.id, "accion": "crear_tratamiento"})
            except Exception:
                pass

    return {"alertas": alertas, "total": len(alertas)}


@router.post("/{cuaderno_id}/tratamientos/{tratamiento_id}/duplicar")
async def duplicar_tratamiento(cuaderno_id: str, tratamiento_id: str):
    """Duplica un tratamiento (nuevo ID, mismo resto de datos)."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    t = cuaderno.obtener_tratamiento(tratamiento_id)
    if not t:
        raise HTTPException(status_code=404, detail="Tratamiento no encontrado")
    
    productos_nuevos = [ProductoAplicado(
        producto_id=p.producto_id,
        nombre_comercial=p.nombre_comercial,
        numero_registro=p.numero_registro,
        numero_lote=p.numero_lote or "",
        dosis=p.dosis,
        unidad_dosis=p.unidad_dosis,
        caldo_hectarea=p.caldo_hectarea,
        notas=p.notas
    ) for p in t.productos]
    
    nuevo = Tratamiento(
        parcela_ids=list(t.parcela_ids),
        parcela_nombres=list(t.parcela_nombres),
        fecha_aplicacion=t.fecha_aplicacion,
        productos=productos_nuevos,
        plaga_enfermedad=t.plaga_enfermedad,
        justificacion=t.justificacion,
        metodo_aplicacion=t.metodo_aplicacion,
        condiciones_climaticas=t.condiciones_climaticas,
        operador=t.operador,
        hora_inicio=t.hora_inicio,
        hora_fin=t.hora_fin,
        observaciones=t.observaciones,
        color_fila="",
    )
    cuaderno.agregar_tratamiento(nuevo)
    storage.guardar(cuaderno)
    return {"success": True, "tratamiento": nuevo.to_dict(), "message": "Tratamiento duplicado"}


class CopiarAParcelas(BaseModel):
    parcela_ids: List[str]

@router.post("/{cuaderno_id}/tratamientos/{tratamiento_id}/copiar-a-parcelas")
async def copiar_tratamiento_a_parcelas(cuaderno_id: str, tratamiento_id: str, data: CopiarAParcelas):
    """Copia un tratamiento existente a otras parcelas (1 línea por parcela por producto)."""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    if not data.parcela_ids:
        raise HTTPException(status_code=400, detail="Debes seleccionar al menos una parcela destino")
    for pid in data.parcela_ids:
        if not cuaderno.obtener_parcela(pid):
            raise HTTPException(status_code=400, detail=f"Parcela no encontrada: {pid}")
    creados = cuaderno.copiar_tratamiento_a_parcelas(tratamiento_id, data.parcela_ids)
    if not creados:
        raise HTTPException(status_code=404, detail="Tratamiento origen no encontrado")
    storage.guardar(cuaderno)
    return {
        "success": True,
        "tratamientos": [t.to_dict() for t in creados],
        "message": f"Tratamiento copiado a {len(data.parcela_ids)} parcela(s) ({len(creados)} línea(s))"
    }


@router.post("/{cuaderno_id}/tratamientos/repair")
async def reparar_tratamientos_existentes(cuaderno_id: str):
    """
    Ejecuta reparación manual de tratamientos en cuaderno existente.
    Útil para forzar el arreglo sin reiniciar.
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    reparados_multi_cultivo = 0
    reparados_num_orden = 0
    try:
        reparados_multi_cultivo = cuaderno.reparar_tratamientos_multi_cultivo()
    except Exception:
        reparados_multi_cultivo = 0
    try:
        reparados_num_orden = cuaderno.reparar_tratamientos_num_orden_multi_parcela()
    except Exception:
        reparados_num_orden = 0
    try:
        restablecidos_individual = cuaderno.reestablecer_num_orden_individual_tratamientos()
    except Exception:
        restablecidos_individual = 0

    reparadas_total = int(reparados_multi_cultivo or 0) + int(reparados_num_orden or 0) + int(restablecidos_individual or 0)
    if reparadas_total > 0:
        storage.guardar(cuaderno)

    return {
        "success": True,
        "reparadas_total": reparadas_total,
        "reparadas_multi_cultivo": int(reparados_multi_cultivo or 0),
        "reparadas_num_orden": int(reparados_num_orden or 0),
        "restablecidas_individual": int(restablecidos_individual or 0),
        "total_tratamientos": len(cuaderno.tratamientos or []),
    }


@router.get("/{cuaderno_id}/historico")
async def obtener_historico_completo(
    cuaderno_id: str,
    parcela_id: Optional[str] = Query(None, description="Filtrar por parcela"),
    date_from: Optional[str] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    product_id: Optional[str] = Query(None, description="Filtrar por producto"),
    num_lote: Optional[str] = Query(None, description="Filtrar por número de lote"),
):
    """
    Histórico canónico: tabla desde tratamientos + líneas producto (snapshot).
    Filtros: parcela_id, date_from, date_to, product_id, num_lote.
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    historico = cuaderno.obtener_historico_completo_filtrado(
        parcela_id=parcela_id,
        date_from=date_from,
        date_to=date_to,
        product_id=product_id,
        num_lote=num_lote,
    )
    
    return {
        "historico": historico,
        "total": len(historico),
        "columnas": [
            "fecha", "parcelas", "producto", "num_registro", "num_lote",
            "dosis", "plaga", "operador", "observaciones", "estado"
        ]
    }


# ============================================
# EXPORTACIÓN PDF
# ============================================

def _sanitize_filename(name: str, max_len: int = 80) -> str:
    """Nombre seguro para archivo: sin caracteres inválidos, longitud limitada."""
    if not name or not name.strip():
        return "cuaderno"
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", name.strip())
    s = re.sub(r"\s+", " ", s)
    return s[:max_len].strip() or "cuaderno"


@router.get("/{cuaderno_id}/export/pdf")
async def exportar_pdf_cuaderno(
    cuaderno_id: str,
    desde: Optional[str] = Query(None, description="Solo tratamientos desde (YYYY-MM-DD)"),
    hasta: Optional[str] = Query(None, description="Solo tratamientos hasta (YYYY-MM-DD)"),
    incluir_hojas: Optional[str] = Query(None, description="IDs de hojas a incluir separados por coma (ej: 'uuid1,uuid2')"),
    orden_parcelas: Optional[str] = Query(None, description="IDs de parcelas en el orden deseado (separados por coma)"),
    orden_tratamientos: Optional[str] = Query(None, description="IDs de tratamientos en el orden deseado (separados por coma)"),
    orden_parcelas_modo: Optional[str] = Query(None, description="Modo de orden: num_orden, cultivo, alfabetico, etc."),
    check_hojas_editadas: Optional[bool] = Query(False, description="Solo verificar si hay hojas editadas sin exportar"),
):
    """
    Exporta el cuaderno a PDF. 
    
    Si check_hojas_editadas=true, solo devuelve info sin generar PDF.
    Si hay hojas editadas y no se especifica incluir_hojas, se incluyen todas por defecto.
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    disponibles = _hojas_disponibles_para_exportar(cuaderno)

    if check_hojas_editadas:
        todas = disponibles["hojas_base"] + disponibles["hojas_importadas"]
        return JSONResponse({
            "tiene_hojas_editadas": True,
            "hojas_editadas": todas,
        })

    # Procesar lista de hojas a incluir.
    todas_ids = {h["sheet_id"] for h in disponibles["hojas_base"] + disponibles["hojas_importadas"]}
    ids_incluir_set: set = set()
    if incluir_hojas is not None:
        ids_incluir_set = {s.strip() for s in incluir_hojas.split(",") if s.strip()}
    else:
        ids_incluir_set = todas_ids

    # IDs de hojas importadas a pasar al generador PDF
    sheet_ids_a_incluir: List[str] = [
        h["sheet_id"] for h in disponibles["hojas_importadas"]
        if h["sheet_id"] in ids_incluir_set
    ]
    
    import tempfile
    output_dir = Path(tempfile.gettempdir()) / "cuaderno_exports"
    output_dir.mkdir(exist_ok=True)
    
    base_name = _sanitize_filename(cuaderno.nombre_explotacion)
    if desde or hasta:
        sufijo = "_" + (desde or "") + "_" + (hasta or "")
        filename = f"{base_name}_Cuaderno_{cuaderno.año}{sufijo}.pdf"
    else:
        filename = f"{base_name}_Cuaderno_{cuaderno.año}.pdf"
    output_path = output_dir / filename
    
    # Determinar qué secciones base incluir
    incluir_base = {
        k: (BASE_SHEET_IDS[k] in ids_incluir_set)
        for k in BASE_SHEET_IDS
    }

    try:
        generator = PDFGenerator()
        orden_parcelas_list = [s.strip() for s in orden_parcelas.split(",")] if orden_parcelas else None
        orden_tratamientos_list = [s.strip() for s in orden_tratamientos.split(",")] if orden_tratamientos else None
        generator.generar_cuaderno_completo(
            cuaderno, str(output_path),
            date_desde=desde, date_hasta=hasta,
            hojas_a_incluir=sheet_ids_a_incluir if sheet_ids_a_incluir else None,
            orden_parcelas=orden_parcelas_list,
            orden_tratamientos=orden_tratamientos_list,
            orden_parcelas_modo=orden_parcelas_modo,
            incluir_base=incluir_base,
        )
        
        return FileResponse(
            path=str(output_path),
            filename=filename,
            media_type="application/pdf"
        )
    except ImportError as e:
        raise HTTPException(
            status_code=500, 
            detail="fpdf2 no instalado. Ejecuta: pip install fpdf2"
        )


@router.get("/{cuaderno_id}/parcelas/{parcela_id}/export/pdf")
async def exportar_pdf_parcela(cuaderno_id: str, parcela_id: str):
    """Exporta el histórico de una parcela a PDF"""
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    parcela = cuaderno.obtener_parcela(parcela_id)
    if not parcela:
        raise HTTPException(status_code=404, detail="Parcela no encontrada")
    
    # Generar PDF
    import tempfile
    output_dir = Path(tempfile.gettempdir()) / "cuaderno_exports"
    output_dir.mkdir(exist_ok=True)
    
    base_name = _sanitize_filename(parcela.nombre)
    filename = f"{base_name}_Historico_{cuaderno.año}.pdf"
    output_path = output_dir / filename
    
    try:
        generator = PDFGenerator()
        generator.generar_historico_parcela(cuaderno, parcela_id, str(output_path))
        
        return FileResponse(
            path=str(output_path),
            filename=filename,
            media_type="application/pdf"
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="fpdf2 no instalado. Ejecuta: pip install fpdf2"
        )


# ============================================
# EXPORTACIÓN EXCEL
# ============================================

def _parse_date_for_excel(val: Any) -> Any:
    """Convierte strings habituales a date para Excel; si no se reconoce, devuelve el valor original."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except ValueError:
            pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return val


def _trat_parcela_group_key(t: Tratamiento) -> str:
    """Misma lógica que el editor (parcela_nombres o num_orden_parcelas), en minúsculas."""
    if t.parcela_nombres:
        return ",".join(t.parcela_nombres).lower()
    return (t.num_orden_parcelas or "").lower()


# IDs fijos para las hojas base del editor (usados en incluir_hojas)
BASE_SHEET_IDS = {
    "info_general": "__info_general__",
    "parcelas": "__parcelas__",
    "productos": "__productos__",
    "tratamientos": "__tratamientos__",
    "fertilizantes": "__fertilizantes__",
    "cosecha": "__cosecha__",
}


def _hojas_disponibles_para_exportar(cuaderno) -> dict:
    """Devuelve las hojas base del editor y las importadas editadas disponibles."""
    parcelas_validas = [p for p in cuaderno.parcelas if p.activa]
    productos_validos = [p for p in cuaderno.productos if p.nombre_comercial and p.nombre_comercial.strip()]
    tratamientos_validos = [t for t in cuaderno.tratamientos if t.fecha_aplicacion or t.cultivo_especie or t.productos]
    fertilizantes = getattr(cuaderno, "fertilizaciones", None) or []
    cosechas = getattr(cuaderno, "cosechas", None) or []

    hojas_base = [
        {"sheet_id": BASE_SHEET_IDS["info_general"], "nombre": "1. Información General", "num_filas": 1, "tipo": "base"},
    ]
    if parcelas_validas:
        hojas_base.append({"sheet_id": BASE_SHEET_IDS["parcelas"], "nombre": "2.1. Datos Parcelas", "num_filas": len(parcelas_validas), "tipo": "base"})
    if productos_validos:
        hojas_base.append({"sheet_id": BASE_SHEET_IDS["productos"], "nombre": "Productos Fitosanitarios", "num_filas": len(productos_validos), "tipo": "base"})
    if tratamientos_validos:
        hojas_base.append({"sheet_id": BASE_SHEET_IDS["tratamientos"], "nombre": "3.1. Reg. Tratamientos", "num_filas": len(tratamientos_validos), "tipo": "base"})
    if fertilizantes:
        hojas_base.append({"sheet_id": BASE_SHEET_IDS["fertilizantes"], "nombre": "Reg. Fertilizantes", "num_filas": len(fertilizantes), "tipo": "base"})
    if cosechas:
        hojas_base.append({"sheet_id": BASE_SHEET_IDS["cosecha"], "nombre": "Reg. Cosecha", "num_filas": len(cosechas), "tipo": "base"})

    def _hoja_tiene_datos(h):
        if not h.datos or not h.columnas:
            return False
        return any(any(c is not None and str(c).strip() for c in fila) for fila in h.datos)

    hojas_importadas = [
        {
            "sheet_id": h.sheet_id,
            "nombre": h.nombre,
            "num_filas": sum(1 for fila in h.datos if any(c is not None and str(c).strip() for c in fila)),
            "tipo": "importada",
        }
        for h in cuaderno.hojas_originales
        if h.origen == "importado_editable" and _hoja_tiene_datos(h)
    ]

    return {"hojas_base": hojas_base, "hojas_importadas": hojas_importadas}


@router.get("/{cuaderno_id}/export/excel")
async def exportar_excel_cuaderno(
    cuaderno_id: str,
    desde: Optional[str] = Query(None, description="Solo tratamientos desde (YYYY-MM-DD)"),
    hasta: Optional[str] = Query(None, description="Solo tratamientos hasta (YYYY-MM-DD)"),
    incluir_hojas: Optional[str] = Query(None, description="IDs de hojas a incluir separados por coma; '' = ninguna"),
    orden_parcelas: Optional[str] = Query(None, description="IDs de parcelas en el orden deseado (separados por coma)"),
    orden_tratamientos: Optional[str] = Query(None, description="IDs de tratamientos en el orden deseado (separados por coma)"),
    check_hojas_editadas: Optional[bool] = Query(False, description="Solo verificar hojas disponibles sin exportar"),
    orden_parcelas_modo: Optional[str] = Query(None, description="Modo de orden: num_orden, cultivo, alfabetico, etc."),
    orden_tratamientos_modo: Optional[str] = Query(
        None,
        description="Modo orden tratamientos del editor (ej. parcela): inserta fila en blanco entre grupos de parcela en la hoja Tratamientos",
    ),
):
    """
    Exporta el cuaderno completo a un archivo Excel (.xlsx).
    Si check_hojas_editadas=true, devuelve JSON con las hojas disponibles sin generar Excel.
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")

    disponibles = _hojas_disponibles_para_exportar(cuaderno)

    if check_hojas_editadas:
        todas = disponibles["hojas_base"] + disponibles["hojas_importadas"]
        return JSONResponse({
            "tiene_hojas_editadas": True,
            "hojas_editadas": todas,
        })

    # Determinar qué hojas base incluir
    todas_ids = {h["sheet_id"] for h in disponibles["hojas_base"] + disponibles["hojas_importadas"]}
    ids_incluir_set: set = set()
    if incluir_hojas is not None:
        ids_incluir_set = {s.strip() for s in incluir_hojas.split(",") if s.strip()}
    else:
        ids_incluir_set = todas_ids

    incluir_base = {k: (v in ids_incluir_set) for k, v in BASE_SHEET_IDS.items()}

    # Ordenar parcelas y tratamientos según el orden del editor
    parcelas_ordenadas = list(cuaderno.parcelas)
    if orden_parcelas_modo:
        parcelas_ordenadas = sorted(
            parcelas_ordenadas,
            key=lambda p: _sort_key_parcela(p, orden_parcelas_modo)
        )
    elif orden_parcelas:
        orden_ids = [s.strip() for s in orden_parcelas.split(",") if s.strip()]
        if orden_ids:
            id_to_idx = {pid: i for i, pid in enumerate(orden_ids)}
            parcelas_ordenadas = sorted(
                parcelas_ordenadas,
                key=lambda p: (id_to_idx.get(p.id, 99999), p.num_orden or 0)
            )

    tratamientos_ordenados = list(cuaderno.tratamientos)
    if orden_tratamientos:
        orden_ids = [s.strip() for s in orden_tratamientos.split(",") if s.strip()]
        if orden_ids:
            id_to_idx = {tid: i for i, tid in enumerate(orden_ids)}
            tratamientos_ordenados = sorted(
                tratamientos_ordenados,
                key=lambda t: (id_to_idx.get(t.id, 99999), t.fecha_aplicacion or "")
            )
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado. Ejecuta: pip install openpyxl")

    wb = Workbook()

    # ================================================================
    # ESTILOS — Paleta oficial SIGPAC (azul marino + fondos claros)
    # ================================================================
    NAVY       = "1F3864"
    NAVY_MID   = "2E4A7A"
    NAVY_LIGHT = "D6E4F0"
    WHITE      = "FFFFFF"

    navy_fill       = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    navy_mid_fill   = PatternFill(start_color=NAVY_MID, end_color=NAVY_MID, fill_type="solid")
    navy_light_fill = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
    white_fill      = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    hdr_font       = Font(name="Calibri", bold=True, size=10, color=WHITE)
    hdr_font_sm    = Font(name="Calibri", bold=True, size=9,  color=WHITE)
    title_font     = Font(name="Calibri", bold=True, size=13, color=NAVY)
    section_font   = Font(name="Calibri", bold=True, size=11, color=NAVY)
    info_bold_font = Font(name="Calibri", bold=True, size=10)
    info_font      = Font(name="Calibri", size=10)
    label_font     = Font(name="Calibri", bold=True, size=11)
    value_font     = Font(name="Calibri", size=11)
    total_font     = Font(name="Calibri", bold=True, size=11, color=NAVY)
    year_font      = Font(name="Calibri", bold=True, size=12, color=NAVY)

    center_align     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align       = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    left_align_wrap  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    num_align        = Alignment(horizontal="right",  vertical="center")
    data_align       = Alignment(vertical="center", wrap_text=False)

    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"),
    )

    num_format_2d  = "#,##0.00"
    date_format_es = "dd/mm/yyyy"
    ROW_HEIGHT     = 18

    def _fill_from_color_hex(hex_str: Optional[str]):
        if not hex_str or not str(hex_str).strip():
            return None
        h = str(hex_str).strip().lstrip("#").upper()
        if len(h) == 6 and all(c in "0123456789ABCDEF" for c in h):
            return PatternFill(start_color=h, end_color=h, fill_type="solid")
        return None

    def _style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            c = ws.cell(row=row_num, column=col)
            c.font = Font(name="Calibri", bold=True, size=9)
            c.alignment = center_align
            c.border = thin_border
        ws.row_dimensions[row_num].height = 36

    def _write_data_row(ws, row_num, values, col_types=None, *, row_fill_hex: Optional[str] = None, use_zebra: bool = True):
        row_fill = _fill_from_color_hex(row_fill_hex)
        for col_idx, raw_val in enumerate(values, 1):
            ct = col_types[col_idx - 1] if col_types and col_idx - 1 < len(col_types) else None
            val = _parse_date_for_excel(raw_val) if ct == "date" else raw_val
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = info_font
            if ct == "num":
                cell.alignment = num_align
                if isinstance(val, (int, float)):
                    cell.number_format = num_format_2d
            elif ct == "date":
                cell.alignment = data_align
                if isinstance(val, date):
                    cell.number_format = date_format_es
            else:
                cell.alignment = data_align
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_num].height = ROW_HEIGHT

    def _write_blank_separator_row(ws, row_num: int, num_cols: int):
        for col in range(1, num_cols + 1):
            c = ws.cell(row=row_num, column=col)
            c.value = None
            c.border = thin_border

    def _auto_width(ws, num_cols, data_start_row=2, max_width=55, min_width=12):
        for col in range(1, num_cols + 1):
            letter = get_column_letter(col)
            header_len = len(str(ws.cell(row=data_start_row - 1, column=col).value or ""))
            max_data_len = header_len
            for row in range(data_start_row, min(ws.max_row + 1, data_start_row + 100)):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_data_len = max(max_data_len, min(len(str(val)), max_width))
            ws.column_dimensions[letter].width = max(min_width, min(max_data_len + 5, max_width))

    def _add_autofilter(ws, num_cols, header_row=1):
        last_col = get_column_letter(num_cols)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

    def _freeze(ws, cell="A2"):
        ws.freeze_panes = cell

    def _build_oficial_sheet(ws, *, sheet_title: str, section_title: str, section_subtitle: str,
                             group_headers: list, col_headers: list, col_types: list,
                             col_widths: list, data_rows: list, total_col: int = 0,
                             total_values: dict = None, row_colors: list = None):
        """
        Generic builder for all official-style sheets.
        group_headers: [(label, start_col, end_col), ...]
        """
        num_cols = len(col_headers)

        # Row 1: Titular + Año
        titular_text = f"Explotación/Titular de la explotación: {cuaderno.nombre_explotacion or cuaderno.titular or ''}"
        ws.cell(row=1, column=1, value=titular_text).font = info_bold_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols - 2, 1))
        ws.cell(row=1, column=num_cols - 1, value="AÑO").font = year_font
        ws.cell(row=1, column=num_cols - 1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=1, column=num_cols, value=cuaderno.año).font = year_font
        ws.cell(row=1, column=num_cols).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        # Row 3: Título
        ws.cell(row=3, column=1, value=section_title).font = title_font
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
        ws.cell(row=3, column=1).alignment = center_align
        ws.row_dimensions[3].height = 24

        # Row 4: Subtítulo
        ws.cell(row=4, column=1, value=section_subtitle).font = section_font
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=num_cols)
        ws.cell(row=4, column=1).alignment = center_align
        ws.row_dimensions[4].height = 22

        # Row 5: Group headers (merged) — sin color, solo bordes y negrita
        if group_headers:
            for label, start_c, end_c in group_headers:
                ws.merge_cells(start_row=5, start_column=start_c, end_row=5, end_column=end_c)
                c = ws.cell(row=5, column=start_c, value=label)
                c.font = Font(name="Calibri", bold=True, size=10)
                c.alignment = center_align
                c.border = thin_border
                for cc in range(start_c + 1, end_c + 1):
                    cx = ws.cell(row=5, column=cc)
                    cx.alignment = center_align
                    cx.border = thin_border
            ws.row_dimensions[5].height = 22
            header_row = 6
        else:
            header_row = 5

        # Column headers — sin color, solo bordes y negrita
        for c, h in enumerate(col_headers, 1):
            cell = ws.cell(row=header_row, column=c, value=h)
            cell.font = Font(name="Calibri", bold=True, size=9)
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 48

        # Column widths
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        data_start = header_row + 1

        # Data
        for r_idx, row_vals in enumerate(data_rows):
            row_num = data_start + r_idx
            _cf = row_colors[r_idx] if row_colors and r_idx < len(row_colors) else None
            _write_data_row(ws, row_num, row_vals, col_types, row_fill_hex=_cf, use_zebra=True)

        # Totals
        if total_values:
            total_r = data_start + len(data_rows)
            for col_idx, val in total_values.items():
                c = ws.cell(row=total_r, column=col_idx, value=val)
                c.font = total_font
                if isinstance(val, (int, float)):
                    c.number_format = num_format_2d
                    c.alignment = num_align

        _add_autofilter(ws, num_cols, header_row=header_row)
        ws.freeze_panes = f"A{data_start}"

        return data_start

    # ================================================================
    # 1. INFORMACIÓN GENERAL  — Formato oficial del cuaderno
    # ================================================================
    ws_info = wb.active
    ws_info.title = "inf.gral 1"

    INFO_COLS = 8
    info_lbl_font  = Font(name="Calibri", bold=True, size=9)
    info_val_font  = Font(name="Calibri", size=9)
    info_title_font = Font(name="Calibri", bold=True, size=12)
    info_sub_font   = Font(name="Calibri", bold=True, size=10)
    info_note_font  = Font(name="Calibri", italic=True, size=8)
    info_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    lbl_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    val_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for c in range(1, INFO_COLS + 1):
        ws_info.column_dimensions[get_column_letter(c)].width = 18
    ws_info.column_dimensions["A"].width = 28
    ws_info.column_dimensions["B"].width = 30
    ws_info.column_dimensions["E"].width = 12
    ws_info.column_dimensions["G"].width = 12

    def _info_cell(r, c, val, font=info_val_font, merge_end_c=None, alignment=val_align):
        cell = ws_info.cell(row=r, column=c, value=val)
        cell.font = font
        cell.border = info_border
        cell.alignment = alignment
        if merge_end_c and merge_end_c > c:
            ws_info.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_end_c)
            for cc in range(c + 1, merge_end_c + 1):
                cx = ws_info.cell(row=r, column=cc)
                cx.border = info_border

    # --- Row 2: Título centrado ---
    ws_info.merge_cells(start_row=2, start_column=1, end_row=2, end_column=INFO_COLS)
    ws_info.cell(row=2, column=1, value="1. INFORMACIÓN GENERAL").font = info_title_font
    ws_info.cell(row=2, column=1).alignment = center_align
    ws_info.row_dimensions[2].height = 28

    # --- Row 4: Fecha apertura + Año ---
    _info_cell(4, 1, "FECHA DE APERTURA DEL CUADERNO:", info_lbl_font, 2)
    _info_cell(4, 3, cuaderno.fecha_apertura if hasattr(cuaderno, 'fecha_apertura') and cuaderno.fecha_apertura else "", info_val_font, 6)
    _info_cell(4, 7, "AÑO", info_lbl_font)
    _info_cell(4, 8, cuaderno.año, Font(name="Calibri", bold=True, size=11))
    ws_info.row_dimensions[4].height = 20

    # --- Row 6: Subtítulo datos generales ---
    ws_info.merge_cells(start_row=6, start_column=1, end_row=6, end_column=INFO_COLS)
    ws_info.cell(row=6, column=1, value="1.1 DATOS GENERALES DE LA EXPLOTACIÓN").font = info_sub_font
    ws_info.cell(row=6, column=1).alignment = center_align
    for cc in range(1, INFO_COLS + 1):
        ws_info.cell(row=6, column=cc).border = info_border
    ws_info.row_dimensions[6].height = 20

    # --- Row 7: Nombre + NIF ---
    _info_cell(7, 1, "Nombre y apellidos o razón social:", info_lbl_font)
    _info_cell(7, 2, cuaderno.nombre_explotacion or cuaderno.titular or "", info_val_font, 6)
    _info_cell(7, 7, "NIF:", info_lbl_font)
    _info_cell(7, 8, cuaderno.nif_titular or "")

    # --- Row 8: Registros ---
    _info_cell(8, 1, "Nº Registro de Explotaciones Nacional:", info_lbl_font, 2)
    _info_cell(8, 3, cuaderno.codigo_explotacion or "", info_val_font, 4)
    _info_cell(8, 5, "Nº Registro de Explotaciones Autonómico:", info_lbl_font, 6)
    _info_cell(8, 7, "", info_val_font, 8)

    # --- Row 9: Dirección ---
    _info_cell(9, 1, "Dir:", info_lbl_font)
    _info_cell(9, 2, cuaderno.domicilio or "", info_val_font, 3)
    _info_cell(9, 4, "Localidad:", info_lbl_font)
    _info_cell(9, 5, "", info_val_font)
    _info_cell(9, 6, "C. Postal:", info_lbl_font)
    _info_cell(9, 7, "Provincia:", info_lbl_font)
    _info_cell(9, 8, "")

    # --- Row 10: Teléfono ---
    _info_cell(10, 1, "Teléfono fijo:", info_lbl_font)
    _info_cell(10, 2, "")
    _info_cell(10, 3, "Teléfono móvil:", info_lbl_font)
    _info_cell(10, 4, "", info_val_font, 5)
    _info_cell(10, 6, "e-mail:", info_lbl_font)
    _info_cell(10, 7, "", info_val_font, 8)

    # --- Row 12: TITULAR O REPRESENTANTE ---
    ws_info.merge_cells(start_row=12, start_column=1, end_row=12, end_column=INFO_COLS)
    ws_info.cell(row=12, column=1, value="TITULAR O REPRESENTANTE DE LA EXPLOTACIÓN").font = info_sub_font
    ws_info.cell(row=12, column=1).alignment = center_align
    for cc in range(1, INFO_COLS + 1):
        ws_info.cell(row=12, column=cc).border = info_border
    ws_info.row_dimensions[12].height = 20

    # --- Row 13: Nombre titular + NIF ---
    _info_cell(13, 1, "Nombre y apellidos:", info_lbl_font)
    _info_cell(13, 2, cuaderno.titular or cuaderno.nombre_explotacion or "", info_val_font, 6)
    _info_cell(13, 7, "NIF:", info_lbl_font)
    _info_cell(13, 8, cuaderno.nif_titular or "")

    # --- Row 14: Dir titular ---
    _info_cell(14, 1, "Dir:", info_lbl_font)
    _info_cell(14, 2, cuaderno.domicilio or "", info_val_font, 3)
    _info_cell(14, 4, "Localidad:", info_lbl_font)
    _info_cell(14, 5, "", info_val_font)
    _info_cell(14, 6, "C. Postal:", info_lbl_font)
    _info_cell(14, 7, "Provincia:", info_lbl_font)
    _info_cell(14, 8, "")

    # --- Row 15: Tipo representación ---
    _info_cell(15, 1, "Tipo de representación:", info_lbl_font)
    _info_cell(15, 2, "", info_val_font, 3)
    _info_cell(15, 4, "Teléfono:", info_lbl_font)
    _info_cell(15, 5, "", info_val_font)
    _info_cell(15, 6, "e-mail:", info_lbl_font)
    _info_cell(15, 7, "", info_val_font, 8)

    # --- Rows 18-19: Firma ---
    ws_info.merge_cells(start_row=18, start_column=4, end_row=18, end_column=INFO_COLS)
    ws_info.cell(row=18, column=4, value="Firma del titular o representante de la explotación (1)").font = info_note_font
    ws_info.cell(row=18, column=4).alignment = Alignment(horizontal="center", vertical="center")

    ws_info.merge_cells(start_row=20, start_column=4, end_row=20, end_column=5)
    ws_info.cell(row=20, column=4, value="Fecha:").font = info_lbl_font

    # --- Row 23: Nota legal ---
    ws_info.merge_cells(start_row=23, start_column=1, end_row=23, end_column=INFO_COLS)
    ws_info.cell(row=23, column=1,
        value="(1) La persona firmante se hace responsable de la veracidad de los datos consignados en el presente cuaderno de explotación"
    ).font = info_note_font

    # --- Row 25: Hoja nº ---
    _info_cell(25, 6, "Hoja nº", info_lbl_font)
    _info_cell(25, 7, 1, Font(name="Calibri", bold=True, size=10))
    _info_cell(25, 8, "de la sección nº", info_lbl_font)

    # ================================================================
    # 2. PARCELAS — 2.1 DATOS PARCELAS (columnas del editor)
    # ================================================================
    ws_parc = wb.create_sheet("2.1. DATOS PARCELAS")
    ws_parc.sheet_properties.tabColor = "4CAF50"

    parc_col_headers = [
        "Nº Orden",                            # 1
        "Código\nProvincia",                   # 2
        "Nombre",                              # 3
        "Término Municipal",                   # 4
        "Nº Polígono",                         # 5
        "Nº Parcela",                          # 6
        "Nº Recinto",                          # 7
        "Uso SIGPAC",                          # 8
        "Superficie\nSIGPAC (ha)",             # 9
        "Superficie\nCultivada (ha)",          # 10
        "Cultivo / Especie",                   # 11
        "Ecorégimen",                          # 12
        "Secano /\nRegadío",                   # 13
        "Zonas\nvulnerables",                  # 14
    ]
    parc_col_types  = ["num", "str", "str", "str", "str", "str", "str", "str", "num", "num", "str", "str", "str", "str"]
    parc_col_widths = [10, 10, 22, 30, 12, 12, 10, 10, 14, 14, 20, 12, 12, 14]

    parc_data_rows = []
    parc_row_colors = []
    for p in parcelas_ordenadas:
        zn = getattr(p, "zona_nitratos", False)
        parc_data_rows.append([
            p.num_orden,
            p.codigo_provincia or getattr(p, "provincia", "") or "",
            p.nombre,
            p.termino_municipal or p.municipio or "",
            p.num_poligono,
            p.num_parcela,
            p.num_recinto,
            p.uso_sigpac,
            p.superficie_sigpac or None,
            p.superficie_cultivada or p.superficie_ha or None,
            p.especie or p.cultivo or "",
            p.ecoregimen,
            p.secano_regadio,
            "Sí" if zn else "No",
        ])
        parc_row_colors.append((getattr(p, "color_fila", None) or "").strip() or None)

    total_sup_sigpac = round(sum(p.superficie_sigpac or 0 for p in parcelas_ordenadas), 2)
    total_sup_cult = round(sum(p.superficie_cultivada or p.superficie_ha or p.superficie_sigpac or 0 for p in parcelas_ordenadas), 2)

    _build_oficial_sheet(
        ws_parc,
        sheet_title="2.1. DATOS PARCELAS",
        section_title="2. IDENTIFICACIÓN DE LAS PARCELAS",
        section_subtitle="2.1 DATOS IDENTIFICATIVOS Y AGRONÓMICOS DE LAS PARCELAS",
        group_headers=[
            ("REFERENCIAS SIGPAC", 1, 8),
            ("DATOS AGRONÓMICOS", 9, 14),
        ],
        col_headers=parc_col_headers,
        col_types=parc_col_types,
        col_widths=parc_col_widths,
        data_rows=parc_data_rows,
        row_colors=parc_row_colors,
        total_values={8: "TOTAL:", 9: total_sup_sigpac, 10: total_sup_cult},
    )

    # ================================================================
    # 3. PRODUCTOS FITOSANITARIOS
    # ================================================================
    ws_prod = wb.create_sheet("Productos Fitosanitarios")
    ws_prod.sheet_properties.tabColor = "FF9800"

    prod_col_headers = [
        "Nombre Comercial",   # 1
        "Nº Registro",        # 2
        "Materia Activa",     # 3
        "Nº Lote",            # 4
        "Cantidad",           # 5
        "Unidad",             # 6
        "F. Adquisición",     # 7
    ]
    prod_col_types  = ["str", "str", "str", "str", "num", "str", "date"]
    prod_col_widths = [28, 16, 26, 16, 12, 10, 16]

    prod_data_rows = []
    prod_row_colors = []
    for p in sorted(cuaderno.productos, key=lambda x: (x.nombre_comercial or "").upper()):
        prod_data_rows.append([
            p.nombre_comercial, p.numero_registro, p.materia_activa,
            p.numero_lote, p.cantidad_adquirida or None, p.unidad,
            p.fecha_adquisicion,
        ])
        prod_row_colors.append((getattr(p, "color_fila", None) or "").strip() or None)

    _build_oficial_sheet(
        ws_prod,
        sheet_title="Productos Fitosanitarios",
        section_title="PRODUCTOS FITOSANITARIOS",
        section_subtitle="Registro de productos adquiridos",
        group_headers=[],
        col_headers=prod_col_headers,
        col_types=prod_col_types,
        col_widths=prod_col_widths,
        data_rows=prod_data_rows,
        row_colors=prod_row_colors,
    )

    # ================================================================
    # 4. TRATAMIENTOS FITOSANITARIOS — 3.1 Registro Tratamientos
    # ================================================================
    ws_trat = wb.create_sheet("3.1. Reg. Tratamientos")
    ws_trat.sheet_properties.tabColor = "F44336"

    trat_col_headers = [
        "Nº Parcela",           # 1
        "Cultivo",              # 2
        "Sup. Tratada\n(ha)",   # 3
        "Fecha\nAplicación",    # 4
        "Problemática",         # 5
        "Aplicador",            # 6
        "Equipo",               # 7
        "Producto",             # 8
        "Nº Registro",          # 9
        "Dosis",                # 10
        "Eficacia",             # 11
    ]
    trat_col_types  = ["str", "str", "num", "date", "str", "str", "str", "str", "str", "str", "str"]
    trat_col_widths = [14, 18, 14, 14, 20, 14, 14, 24, 14, 12, 12]

    num_trat_cols = len(trat_col_headers)

    # Build title/header portion via helper
    _build_oficial_sheet(
        ws_trat,
        sheet_title="3.1. Reg. Tratamientos",
        section_title="3. TRATAMIENTOS FITOSANITARIOS",
        section_subtitle="3.1 REGISTRO DE TRATAMIENTOS FITOSANITARIOS",
        group_headers=[
            ("IDENTIFICACIÓN PARCELA", 1, 2),
            ("TRATAMIENTO APLICADO", 3, 11),
        ],
        col_headers=trat_col_headers,
        col_types=trat_col_types,
        col_widths=trat_col_widths,
        data_rows=[],
    )

    tratamientos = tratamientos_ordenados
    if desde or hasta:
        if desde:
            tratamientos = [t for t in tratamientos if (t.fecha_aplicacion or "") >= desde]
        if hasta:
            tratamientos = [t for t in tratamientos if (t.fecha_aplicacion or "") <= hasta]

    modo_parcela_export = (orden_tratamientos_modo or "").strip().lower() == "parcela"
    trat_data_start = 7
    row = trat_data_start
    prev_parcela_key: Optional[str] = None
    for t in tratamientos:
        parcela_key = _trat_parcela_group_key(t)
        if modo_parcela_export and prev_parcela_key is not None and parcela_key != prev_parcela_key:
            _write_blank_separator_row(ws_trat, row, num_trat_cols)
            row += 1
        prev_parcela_key = parcela_key

        parcela_ref = t.num_orden_parcelas or ", ".join(t.parcela_nombres) or ""
        productos = t.productos if t.productos else [ProductoAplicado()]
        color_hex = (getattr(t, "color_fila", None) or "").strip() or None
        for pi, prod in enumerate(productos):
            _write_data_row(ws_trat, row, [
                parcela_ref if pi == 0 else "",
                t.cultivo_especie if pi == 0 else "",
                t.superficie_tratada if pi == 0 else None,
                t.fecha_aplicacion if pi == 0 else "",
                (t.problema_fitosanitario or t.plaga_enfermedad) if pi == 0 else "",
                (t.aplicador or t.operador) if pi == 0 else "",
                t.equipo if pi == 0 else "",
                prod.nombre_comercial,
                prod.numero_registro,
                (f"{prod.dosis} {prod.unidad_dosis}".strip() if prod.dosis else None),
                t.eficacia if pi == 0 else "",
            ], trat_col_types, row_fill_hex=color_hex, use_zebra=True)
            row += 1

    # ================================================================
    # 5. FERTILIZACIONES — Reg. Fertilizantes
    # ================================================================
    fertilizaciones = getattr(cuaderno, 'fertilizaciones', None) or []
    if fertilizaciones:
        ws_fert = wb.create_sheet("Reg. Fertilizantes")
        ws_fert.sheet_properties.tabColor = "795548"

        fert_col_headers = [
            "Mes Inicio",      # 1
            "Mes Fin",          # 2
            "Parcelas",         # 3
            "Cultivo",          # 4
            "Tipo Abono",       # 5
            "N/P/K",            # 6
            "Dosis",            # 7
            "Tipo",             # 8
        ]
        fert_col_types  = ["str", "str", "str", "str", "str", "str", "str", "str"]
        fert_col_widths = [14, 14, 14, 18, 24, 12, 16, 16]

        fert_data_rows = []
        fert_row_colors = []
        for f in fertilizaciones:
            fert_data_rows.append([
                f.fecha_inicio, f.fecha_fin, f.num_orden_parcelas, f.cultivo_especie,
                f.tipo_abono, f.riqueza_npk, f.dosis, f.tipo_fertilizacion,
            ])
            fert_row_colors.append((getattr(f, "color_fila", None) or "").strip() or None)

        _build_oficial_sheet(
            ws_fert,
            sheet_title="Reg. Fertilizantes",
            section_title="REGISTRO DE FERTILIZACIÓN",
            section_subtitle="Abonado y fertilizaciones realizadas",
            group_headers=[],
            col_headers=fert_col_headers,
            col_types=fert_col_types,
            col_widths=fert_col_widths,
            data_rows=fert_data_rows,
            row_colors=fert_row_colors,
        )

    # ================================================================
    # 6. COSECHAS — Reg. Cosecha
    # ================================================================
    cosechas = getattr(cuaderno, 'cosechas', None) or []
    if cosechas:
        ws_cos = wb.create_sheet("Reg. Cosecha")
        ws_cos.sheet_properties.tabColor = "607D8B"

        cos_col_headers = [
            "Fecha",             # 1
            "Producto",          # 2
            "Cantidad (kg)",     # 3
            "Parcelas",          # 4
            "Nº Albarán",        # 5
            "Nº Lote",           # 6
            "Cliente",           # 7
        ]
        cos_col_types  = ["date", "str", "num", "str", "str", "str", "str"]
        cos_col_widths = [14, 18, 14, 14, 16, 14, 24]

        cos_data_rows = []
        cos_row_colors = []
        for co in cosechas:
            cos_data_rows.append([
                co.fecha, co.producto, co.cantidad_kg or None, co.num_orden_parcelas,
                co.num_albaran, co.num_lote, co.cliente_nombre,
            ])
            cos_row_colors.append((getattr(co, "color_fila", None) or "").strip() or None)

        total_kg = round(sum(co.cantidad_kg or 0 for co in cosechas), 2)

        _build_oficial_sheet(
            ws_cos,
            sheet_title="Reg. Cosecha",
            section_title="REGISTRO DE COSECHA",
            section_subtitle="Producción y entregas realizadas",
            group_headers=[],
            col_headers=cos_col_headers,
            col_types=cos_col_types,
            col_widths=cos_col_widths,
            data_rows=cos_data_rows,
            row_colors=cos_row_colors,
            total_values={2: "TOTAL kg:", 3: total_kg},
        )

    # ================================================================
    # 7. HOJAS IMPORTADAS (con estilo oficial)
    # ================================================================
    hojas_a_exportar = [
        h for h in (cuaderno.hojas_originales or [])
        if h.sheet_id in ids_incluir_set
    ]

    def _imported_sheet_meta(hoja):
        """Detecta si la hoja importada es de parcelas o tratamientos para aplicar estilo oficial."""
        nombre_lc = (hoja.nombre or "").lower().strip()
        tipo_lc = (hoja.tipo or "").lower().strip()
        if tipo_lc == "tratamientos" or nombre_lc in {n.lower() for n in _KNOWN_TRATAMIENTOS_SHEETS}:
            return "tratamientos"
        if tipo_lc == "parcelas" or nombre_lc in {n.lower() for n in _KNOWN_PARCELAS_SHEETS}:
            return "parcelas"
        return "custom"

    used_names = {ws.title for ws in wb.worksheets}
    for hoja in hojas_a_exportar:
        has_cols = bool(hoja.columnas)
        has_data = bool(hoja.datos and any(
            any(c is not None and str(c).strip() for c in fila)
            for fila in hoja.datos
        ))
        if not has_cols and not has_data:
            continue
        safe_name = re.sub(r'[\[\]:*?/\\]', '', hoja.nombre or "Hoja")[:31]
        if not safe_name.strip():
            safe_name = "Hoja"
        counter = 2
        original_name = safe_name
        while safe_name in used_names:
            safe_name = f"{original_name[:27]} ({counter})"[:31]
            counter += 1
        used_names.add(safe_name)

        ws_imp = wb.create_sheet(safe_name)
        meta_tipo = _imported_sheet_meta(hoja)

        # --- Extraer datos reales (sin filas vacías) ---
        filas_reales = [
            fila for fila in (hoja.datos or [])
            if any(c is not None and str(c).strip() for c in fila)
        ]
        col_names = list(hoja.columnas) if has_cols else [f"Col {i+1}" for i in range(len(filas_reales[0]) if filas_reales else 0)]
        num_cols_imp = len(col_names)
        if num_cols_imp == 0:
            continue

        if meta_tipo == "tratamientos":
            ws_imp.sheet_properties.tabColor = "F44336"
            imp_col_types = ["str"] * num_cols_imp
            imp_col_widths = [max(14, min(len(c) + 6, 30)) for c in col_names]
            imp_data_rows = []
            for fila in filas_reales:
                row_vals = []
                for i in range(num_cols_imp):
                    row_vals.append(fila[i] if i < len(fila) else None)
                imp_data_rows.append(row_vals)

            _build_oficial_sheet(
                ws_imp,
                sheet_title=safe_name,
                section_title="3. TRATAMIENTOS FITOSANITARIOS",
                section_subtitle=f"3.1 REGISTRO DE TRATAMIENTOS — {safe_name}",
                group_headers=[],
                col_headers=col_names,
                col_types=imp_col_types,
                col_widths=imp_col_widths,
                data_rows=imp_data_rows,
            )

        elif meta_tipo == "parcelas":
            ws_imp.sheet_properties.tabColor = "4CAF50"
            imp_col_types = ["str"] * num_cols_imp
            imp_col_widths = [max(14, min(len(c) + 6, 30)) for c in col_names]
            imp_data_rows = []
            for fila in filas_reales:
                row_vals = []
                for i in range(num_cols_imp):
                    row_vals.append(fila[i] if i < len(fila) else None)
                imp_data_rows.append(row_vals)

            _build_oficial_sheet(
                ws_imp,
                sheet_title=safe_name,
                section_title="2. IDENTIFICACIÓN DE LAS PARCELAS",
                section_subtitle=f"2.1 DATOS IDENTIFICATIVOS Y AGRONÓMICOS — {safe_name}",
                group_headers=[],
                col_headers=col_names,
                col_types=imp_col_types,
                col_widths=imp_col_widths,
                data_rows=imp_data_rows,
            )

        else:
            ws_imp.sheet_properties.tabColor = "7B1FA2"
            imp_col_types = ["str"] * num_cols_imp
            imp_col_widths = [max(14, min(len(c) + 6, 30)) for c in col_names]
            imp_data_rows = []
            for fila in filas_reales:
                row_vals = []
                for i in range(num_cols_imp):
                    row_vals.append(fila[i] if i < len(fila) else None)
                imp_data_rows.append(row_vals)

            _build_oficial_sheet(
                ws_imp,
                sheet_title=safe_name,
                section_title=safe_name.upper(),
                section_subtitle=f"Datos importados — {safe_name}",
                group_headers=[],
                col_headers=col_names,
                col_types=imp_col_types,
                col_widths=imp_col_widths,
                data_rows=imp_data_rows,
            )

    # ================================================================
    # Eliminar hojas base que el usuario no seleccionó
    # ================================================================
    _base_sheet_names = {
        "info_general": "inf.gral 1",
        "parcelas": "2.1. DATOS PARCELAS",
        "productos": "Productos Fitosanitarios",
        "tratamientos": "3.1. Reg. Tratamientos",
        "fertilizantes": "Reg. Fertilizantes",
        "cosecha": "Reg. Cosecha",
    }
    for key, sheet_name in _base_sheet_names.items():
        if not incluir_base.get(key, True) and sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
    if not wb.sheetnames:
        wb.create_sheet("Vacío")

    # ================================================================
    # GUARDAR Y ENVIAR
    # ================================================================
    output_dir = Path(tempfile.gettempdir()) / "cuaderno_exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    base_name = _sanitize_filename(cuaderno.nombre_explotacion)
    suffix = ""
    if desde or hasta:
        suffix = f"_{desde or ''}_a_{hasta or ''}"
    filename = f"{base_name}_Cuaderno_{cuaderno.año}{suffix}.xlsx"
    output_path = output_dir / filename
    try:
        wb.save(str(output_path))
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error guardando Excel: {str(e)}"
        )

    return FileResponse(
        path=str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================
# UPLOAD Y PROCESAMIENTO DE ARCHIVOS
# ============================================

def _max_upload_bytes() -> int:
    """Tamaño máximo de subida (Excel/PDF/etc.); configurable con MAX_UPLOAD_BYTES."""
    return int(os.environ.get("MAX_UPLOAD_BYTES", str(250 * 1024 * 1024)))


async def _spool_upload_to_temp(file: UploadFile) -> Path:
    """
    Escribe el UploadFile en un temporal por trozos (no carga todo el archivo en RAM).
    """
    suffix = Path(file.filename or "upload").suffix or ".bin"
    fd, raw_path = tempfile.mkstemp(prefix="cuaderno_upload_", suffix=suffix)
    os.close(fd)
    path = Path(raw_path)
    total = 0
    max_b = _max_upload_bytes()
    chunk_size = 1024 * 1024  # 1 MiB
    try:
        with open(path, "wb") as out:
            while True:
                chunk = await file.read(chunk_size)
                if not chunk:
                    break
                total += len(chunk)
                if total > max_b:
                    raise HTTPException(
                        status_code=413,
                        detail=f"El archivo supera el tamaño máximo permitido ({max_b // (1024 * 1024)} MB). "
                        "Configura MAX_UPLOAD_BYTES en el servidor si necesitas un límite mayor.",
                    )
                out.write(chunk)
        return path
    except HTTPException:
        path.unlink(missing_ok=True)
        raise
    except Exception:
        path.unlink(missing_ok=True)
        raise


@router.post("/upload/analyze")
async def analizar_archivo(file: UploadFile = File(...)):
    """
    Analiza un archivo (Excel, PDF, imagen) y extrae datos estructurados.
    Usa GPT-4o Vision para OCR y extracción inteligente.
    
    Formatos soportados:
    - Excel: .xlsx, .xls, .xlsm
    - PDF: .pdf
    - Imágenes: .png, .jpg, .jpeg, .webp
    - CSV: .csv
    
    Returns:
        Datos estructurados extraídos del archivo
    """
    processor = get_processor()
    
    if not processor.is_supported(file.filename):
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de archivo no soportado: {file.filename}. Formatos: Excel, PDF, imágenes, CSV"
        )
    
    tmp: Optional[Path] = None
    try:
        tmp = await _spool_upload_to_temp(file)
        result = await processor.process_file(tmp, file.filename)
        
        # Resumen de hojas para el modal (nombre, vacía, num_filas)
        hojas_resumen = []
        for i, h in enumerate(result.get("hojas", [])):
            datos = h.get("datos", [])
            num_filas = len(datos) if datos else (h.get("num_filas", 0) or 0)
            hojas_resumen.append({
                "indice": i,
                "nombre": h.get("nombre", f"Hoja {i + 1}"),
                "vacia": num_filas == 0,
                "num_filas": num_filas,
            })
        
        return {
            "success": result.get("success", False),
            "filename": file.filename,
            "file_type": result.get("file_type"),
            "data": result,
            "hojas_resumen": hojas_resumen,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando archivo: {str(e)}"
        )
    finally:
        if tmp is not None:
            tmp.unlink(missing_ok=True)


# ============================================
# MAPEO DIRECTO: Hojas conocidas → Modelos
# Sin depender de GPT para datos estructurados
# ============================================

_PARCELAS_COLUMN_MAP = {
    # From workbook_dictionary (formatted keys)
    "Nro Orden": "num_orden",
    "Codigo Provincia": "codigo_provincia",
    "Municipio": "termino_municipal",
    "Codigo Agregado": "codigo_agregado",
    "Zona": "zona",
    "Polygon": "num_poligono",
    "Parcel": "num_parcela",
    "Recinto": "num_recinto",
    "Uso Sigpac": "uso_sigpac",
    "Superficie Sigpac": "superficie_sigpac",
    "Superficie Cultivada": "superficie_cultivada",
    "Crop": "especie",
    "Ecoregimen": "ecoregimen",
    "Secano Regadio": "secano_regadio",
    "Parcela Ubicada En Zonas Vulnerables": "zona_vulnerables_raw",
    "Parcela ubicada en zonas vulnerables": "zona_vulnerables_raw",
}

_TRATAMIENTOS_COLUMN_MAP = {
    "Id Parcelas": "id_parcelas",
    "Especie": "especie",
    "Variedad": "variedad",
    "Superficie Tratada": "superficie_tratada",
    "Fecha": "fecha",
    "Problema Fito": "problema_fito",
    "Aplicador": "aplicador",
    "Equipo": "equipo",
    "Producto": "producto",
    "Nro Registro": "nro_registro",
    "Dosis": "dosis",
    "Eficacia": "eficacia",
    "Observaciones": "observaciones",
}

_FERTILIZANTES_COLUMN_MAP = {
    "Fecha Inicio": "fecha_inicio",
    "Fecha Fin": "fecha_fin",
    "Parcelas": "parcelas",
    "Especie": "especie",
    "Variedad": "variedad",
    "Tipo Abono": "tipo_abono",
    "Albaran": "albaran",
    "Riqueza N": "riqueza_n",
    "Riqueza P": "riqueza_p",
    "Riqueza K": "riqueza_k",
    "Dosis": "dosis",
    "Tipo Fertilizacion": "tipo_fertilizacion",
    "Observaciones": "observaciones",
}

_COSECHA_COLUMN_MAP = {
    "Fecha": "fecha",
    "Producto": "producto",
    "Cantidad Kg": "cantidad_kg",
    "Parcelas Origen": "parcelas_origen",
    "Albaran": "albaran",
    "Lote": "lote",
    "Cliente Nombre": "cliente_nombre",
    "Cliente Nif": "cliente_nif",
    "Cliente Direccion": "cliente_direccion",
    "Rgseaa": "rgseaa",
}

_KNOWN_PARCELAS_SHEETS = {"2.1. DATOS PARCELAS", "2.1. DATOS PARCELAS (2)"}
_KNOWN_TRATAMIENTOS_SHEETS = {"inf.trat 1"}

# Aliases para columnas de tratamientos (nombres que pueden venir del Excel)
_TRATAMIENTOS_RAW_ALIASES = {
    "id parcelas": "id_parcelas", "id_parcelas": "id_parcelas", "nº orden parcelas": "id_parcelas",
    "especie": "especie", "cultivo": "especie",
    "variedad": "variedad",
    "superficie tratada": "superficie_tratada", "superficie tratada (ha)": "superficie_tratada",
    "fecha": "fecha", "fecha aplicación": "fecha", "fecha aplicacion": "fecha",
    "problema fito": "problema_fito", "plaga": "problema_fito", "plaga/enfermedad": "problema_fito",
    "aplicador": "aplicador", "operador": "aplicador",
    "equipo": "equipo",
    "producto": "producto", "nombre comercial": "producto",
    "nro registro": "nro_registro", "nº registro": "nro_registro", "numero registro": "nro_registro",
    "dosis": "dosis", "dose": "dosis", "dosis (l/ha)": "dosis", "dosis (l/ha.)": "dosis",
    "dosis (kg/ha)": "dosis", "dosis (l)": "dosis", "dosis (kg)": "dosis",
    "eficacia": "eficacia",
    "observaciones": "observaciones",
}


def _parse_si_no_zona(val) -> bool:
    """Interpreta celdas S/N, Sí/No, 1/0 para 'parcela en zonas vulnerables'."""
    if val is None:
        return False
    s = str(val).strip().lower()
    if not s:
        return False
    if s in ("sí", "si", "s", "yes", "true", "1", "x", "v", "bien"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return False


def _parse_dosis_y_unidad(val) -> tuple:
    """(valor numérico, unidad). Unidades: L/Ha, Kg/Ha, ml/H, g/Ha."""
    if val is None:
        return 0.0, "L/Ha"
    raw = str(val).strip()
    if not raw:
        return 0.0, "L/Ha"
    s_compact = re.sub(r"\s+", "", raw.lower())
    unidad = "L/Ha"
    if "ml/h" in s_compact:
        unidad = "ml/H"
    elif "kg/ha" in s_compact:
        unidad = "Kg/Ha"
    elif "g/ha" in s_compact:
        unidad = "g/Ha"
    elif "l/ha" in s_compact:
        unidad = "L/Ha"
    return _parse_dosis(val), unidad


def _parse_dosis(val) -> float:
    """Extrae valor numérico de dosis desde formatos como '3,50 l', '3.5 L/Ha', '3'."""
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s:
        return 0.0
    # Quitar unidades (primero compuestas para no dejar restos tipo "/ha")
    s = re.sub(
        r'\s*(l/ha\.?|kg/ha\.?|ml/h|g/ha)\s*$',
        '',
        s,
        flags=re.IGNORECASE,
    )
    s = re.sub(
        r'\s*(l|kg|litros|kilos|ml|g)\s*$',
        '',
        s,
        flags=re.IGNORECASE,
    )
    s = s.strip()
    s = s.replace(",", ".")
    # Extraer primer número (p.ej. "3.5 - 4" -> 3.5)
    match = re.search(r'[\d.]+', s)
    if match:
        try:
            return float(match.group(0))
        except (ValueError, TypeError):
            pass
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0
_KNOWN_FERTILIZANTES_SHEETS = {"reg.fert."}
_KNOWN_COSECHA_SHEETS = {"reg. cosecha"}


_PARCELAS_RAW_ALIASES = {
    "nº de\norden": "num_orden",
    "código\nprovincia": "codigo_provincia",
    "término municipal": "termino_municipal",
    "término municipal\n(código y nombre)": "termino_municipal",
    "código\nagregado": "codigo_agregado",
    "nº de\npolígono": "num_poligono",
    "nº de\nparcela": "num_parcela",
    "nº de\nrecinto": "num_recinto",
    "uso sigpac": "uso_sigpac",
    "superficie\nsigpac (ha)": "superficie_sigpac",
    "superficie\ncultivada\n(ha)": "superficie_cultivada",
    "especie/\nvariedad": "especie",
    "ecoregimen/\npráctica\n(4)": "ecoregimen",
    "secano/\nregadío": "secano_regadio",
    "parcela ubicada en zonas vulnerables": "zona_vulnerables_raw",
    "parcela ubicada en zonas vulnerables (s/n)": "zona_vulnerables_raw",
}


def _row_to_dict(columnas: List[str], row: list, col_map: dict, raw_aliases: dict = None) -> dict:
    """Convierte una fila de hoja con sus columnas a un dict usando el col_map y aliases opcionales."""
    d = {}
    for i, col_name in enumerate(columnas):
        field = col_map.get(col_name)
        if not field and raw_aliases:
            norm = re.sub(r'\s+', ' ', str(col_name or "").lower().strip()
                         .replace("\n", " ").replace("\r", ""))
            field = raw_aliases.get(norm)
            if not field and "dosis" in norm:
                field = "dosis"
            elif not field and "dose" in norm:
                field = "dosis"
        if field and i < len(row):
            val = row[i]
            if val is not None and str(val).strip():
                d[field] = val
    return d


def _norm_termino_municipal(tm: str) -> str:
    t = (tm or "").strip()
    if "-" in t:
        parts = t.split("-", 1)
        if parts[0].strip().isdigit():
            return parts[1].strip()
    return t


def _parcela_clave_sigpac(p) -> tuple:
    """Componente SIGPAC normalizado (sin Nº orden)."""
    tm = getattr(p, "termino_municipal", "") or getattr(p, "municipio", "")
    return (
        _norm_termino_municipal(tm),
        str(getattr(p, "num_poligono", "") or "").strip(),
        str(getattr(p, "num_parcela", "") or "").strip(),
        str(getattr(p, "num_recinto", "") or "").strip(),
    )


def _parcela_clave_importacion(p) -> tuple:
    """
    Clave al importar: (num_orden, SIGPAC). Permite varias filas con la misma referencia SIGPAC
    (p. ej. distinto cultivo o líneas duplicadas en el Excel oficial) si el Nº de orden difiere.
    """
    try:
        no = int(getattr(p, "num_orden", 0) or 0)
    except (TypeError, ValueError):
        no = 0
    return (no, _parcela_clave_sigpac(p))


def _extraer_parcelas_directas(hojas: List[dict]) -> List[Parcela]:
    """Extrae parcelas directamente de hojas conocidas sin GPT.
    Clave de deduplicación: (num_orden, poligono, parcela, recinto) → dos filas con el mismo
    num_orden pero distinto recinto se importan como parcelas independientes.
    """
    parcelas = []
    vistos: set = set()  # clave SIGPAC compuesta (ver abajo)
    for hoja in hojas:
        nombre_hoja = hoja.get("nombre", "")
        es_parcelas = any(known in nombre_hoja for known in _KNOWN_PARCELAS_SHEETS)
        if not es_parcelas:
            continue
        columnas = hoja.get("columnas", [])
        for row in hoja.get("datos", []):
            d = _row_to_dict(columnas, row, _PARCELAS_COLUMN_MAP, _PARCELAS_RAW_ALIASES)
            if not d:
                continue
            num_orden_raw = d.get("num_orden")
            try:
                num_orden = int(float(str(num_orden_raw))) if num_orden_raw else 0
            except (ValueError, TypeError):
                num_orden = 0
            if num_orden == 0:
                continue
            sup_sigpac = 0.0
            try:
                sup_sigpac = float(str(d.get("superficie_sigpac", 0)).replace(",", "."))
            except (ValueError, TypeError):
                pass
            sup_cultivada = 0.0
            try:
                sup_cultivada = float(str(d.get("superficie_cultivada", 0)).replace(",", "."))
            except (ValueError, TypeError):
                pass
            especie = str(d.get("especie", "")).strip()
            termino = str(d.get("termino_municipal", "")).strip()
            nombre_gen = f"Parcela {num_orden}"
            if termino:
                nombre_gen = f"{termino} P{d.get('num_poligono', '')}-{d.get('num_parcela', '')}-{d.get('num_recinto', '')}"

            parcela = Parcela(
                num_orden=num_orden,
                nombre=nombre_gen,
                codigo_provincia=str(d.get("codigo_provincia", "")).strip(),
                termino_municipal=termino,
                codigo_agregado=str(d.get("codigo_agregado", "")).strip(),
                zona=str(d.get("zona", "")).strip(),
                num_poligono=str(d.get("num_poligono", "")).strip(),
                num_parcela=str(d.get("num_parcela", "")).strip(),
                num_recinto=str(d.get("num_recinto", "")).strip(),
                uso_sigpac=str(d.get("uso_sigpac", "")).strip(),
                superficie_sigpac=sup_sigpac,
                superficie_cultivada=sup_cultivada,
                superficie_ha=sup_cultivada or sup_sigpac,
                especie=especie,
                cultivo=especie,
                ecoregimen=str(d.get("ecoregimen", "")).strip(),
                secano_regadio=str(d.get("secano_regadio", "")).strip(),
                zona_nitratos=_parse_si_no_zona(d.get("zona_vulnerables_raw"))
                if "zona_vulnerables_raw" in d
                else False,
            )
            # Clave compuesta: mismo num_orden + mismo SIGPAC = duplicado exacto; distinto recinto = fila nueva
            clave_vista = (num_orden, d.get("num_poligono", ""), d.get("num_parcela", ""), d.get("num_recinto", ""))
            if clave_vista in vistos:
                continue
            vistos.add(clave_vista)
            parcelas.append(parcela)
    return parcelas


def _extraer_tratamientos_directos(hojas: List[dict], parcelas: List[Parcela]) -> List[Tratamiento]:
    """
    Extrae tratamientos desde hojas conocidas sin GPT.
    Si una fila lista varios Nº de parcela (p. ej. '3,4'), genera un registro por parcela
    con el cultivo tomado de la parcela (evita mezclar cultivos en una sola línea).
    """
    tratamientos = []
    # Un Nº orden puede corresponder a varias filas SIGPAC (p. ej. recintos distintos).
    orden_to_parcelas: Dict[int, List[Parcela]] = {}
    for p in parcelas:
        try:
            no = int(float(getattr(p, "num_orden", 0) or 0))
        except (TypeError, ValueError):
            continue
        if no > 0:
            orden_to_parcelas.setdefault(no, []).append(p)
    for hoja in hojas:
        nombre_hoja = hoja.get("nombre", "")
        es_trat = any(known in nombre_hoja for known in _KNOWN_TRATAMIENTOS_SHEETS)
        if not es_trat:
            continue
        columnas = hoja.get("columnas", [])
        for row in hoja.get("datos", []):
            d = _row_to_dict(columnas, row, _TRATAMIENTOS_COLUMN_MAP, _TRATAMIENTOS_RAW_ALIASES)
            if not d:
                continue
            fecha_raw = str(d.get("fecha", "")).strip()
            producto = str(d.get("producto", "")).strip()
            if not fecha_raw and not producto:
                continue
            fecha = ""
            if fecha_raw:
                try:
                    if "/" in fecha_raw:
                        parts = fecha_raw.split("/")
                        if len(parts) == 3:
                            dd, mm, yy = parts
                            if len(yy) == 2:
                                yy = "20" + yy
                            fecha = f"{yy}-{mm.zfill(2)}-{dd.zfill(2)}"
                    elif "-" in fecha_raw and len(fecha_raw) >= 10:
                        fecha = fecha_raw[:10]
                    else:
                        fecha = fecha_raw
                except Exception:
                    fecha = fecha_raw

            id_parcelas_str = str(d.get("id_parcelas", "")).strip()
            sup = 0.0
            try:
                sup = float(str(d.get("superficie_tratada", 0)).replace(",", "."))
            except (ValueError, TypeError):
                pass
            dosis, unidad_dosis = _parse_dosis_y_unidad(d.get("dosis"))
            nro_reg = str(d.get("nro_registro", "")).strip()

            especie_excel = str(d.get("especie", "")).strip()

            def _append_uno(
                pid_list: List[str],
                pnombres: List[str],
                num_ord_str: str,
                cultivo: str,
                sup_trat: float,
            ) -> None:
                prod_aplicado = ProductoAplicado(
                    nombre_comercial=producto,
                    numero_registro=nro_reg,
                    dosis=dosis,
                    unidad_dosis=unidad_dosis,
                )
                tratamientos.append(
                    Tratamiento(
                        fecha_aplicacion=fecha,
                        parcela_ids=pid_list,
                        parcela_nombres=pnombres,
                        num_orden_parcelas=num_ord_str,
                        cultivo_especie=cultivo,
                        superficie_tratada=sup_trat,
                        problema_fitosanitario=str(d.get("problema_fito", "")).strip(),
                        aplicador=str(d.get("aplicador", "")).strip(),
                        equipo=str(d.get("equipo", "")).strip(),
                        productos=[prod_aplicado],
                        eficacia=str(d.get("eficacia", "")).strip(),
                        observaciones=str(d.get("observaciones", "")).strip(),
                        estado=EstadoTratamiento.APLICADO,
                    )
                )

            if id_parcelas_str:
                ordenes = [x.strip() for x in id_parcelas_str.replace(";", ",").split(",") if x.strip()]
                filas_parcela = []
                for o in ordenes:
                    try:
                        orden_num = int(float(o))
                    except (ValueError, TypeError):
                        continue
                    parc_list = orden_to_parcelas.get(orden_num, [])
                    for parc in parc_list:
                        cultivo_p = (parc.especie or parc.cultivo or "").strip().upper()
                        cultivo = cultivo_p or especie_excel
                        sup_p = sup
                        try:
                            sup_c = float(
                                parc.superficie_cultivada
                                or parc.superficie_ha
                                or parc.superficie_sigpac
                                or 0
                            )
                            if sup_c > 0:
                                sup_p = round(sup_c, 2)
                        except (TypeError, ValueError):
                            pass
                        filas_parcela.append(
                            ([parc.id], [parc.nombre], str(orden_num), cultivo, sup_p)
                        )
                if filas_parcela:
                    for pid_list, pnombres, num_ord_str, cultivo, sup_trat in filas_parcela:
                        _append_uno(pid_list, pnombres, num_ord_str, cultivo, sup_trat)
                    continue

            parcela_ids = []
            parcela_nombres = []
            if id_parcelas_str:
                ordenes = [x.strip() for x in id_parcelas_str.replace(";", ",").split(",") if x.strip()]
                for o in ordenes:
                    try:
                        orden_num = int(float(o))
                    except (ValueError, TypeError):
                        continue
                    for p in orden_to_parcelas.get(orden_num, []):
                        parcela_ids.append(p.id)
                        parcela_nombres.append(p.nombre)

            _append_uno(parcela_ids, parcela_nombres, id_parcelas_str, especie_excel, sup)
    return tratamientos


@router.post("/upload/create-from-file")
async def crear_cuaderno_desde_archivo(
    file: UploadFile = File(...),
    solo_datos: Optional[str] = Form(None),
    hojas_seleccionadas: Optional[str] = Form(None),
):
    """
    Crea un cuaderno a partir de un archivo.
    Opciones (form):
    - solo_datos: "true" = no guardar hojas originales, solo datos extraídos.
    - hojas_seleccionadas: JSON array de índices [0,1,2] = guardar solo esas hojas.
    Si no se envían, se guardan todas las hojas.
    """
    processor = get_processor()
    
    if not processor.is_supported(file.filename):
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de archivo no soportado: {file.filename}"
        )
    
    tmp: Optional[Path] = None
    try:
        tmp = await _spool_upload_to_temp(file)
        result = await processor.process_file(tmp, file.filename)
        
        if not result.get("success"):
            raise HTTPException(
                status_code=500,
                detail=result.get("error", "Error procesando archivo")
            )
        
        # Extraer análisis de IA si está disponible
        analisis = result.get("analisis_ia", {})
        datos_explotacion = analisis.get("datos_explotacion", {})
        
        # Crear cuaderno
        nombre = datos_explotacion.get("nombre") or Path(file.filename).stem
        
        cuaderno = CuadernoExplotacion(
            nombre_explotacion=nombre,
            titular=datos_explotacion.get("titular", ""),
            nif_titular=datos_explotacion.get("nif", ""),
            domicilio=datos_explotacion.get("direccion", ""),
            codigo_explotacion=datos_explotacion.get("codigo_explotacion", ""),
            año=datetime.now().year
        )
        
        # === MAPEO DIRECTO: hojas conocidas → modelos (prioridad sobre GPT) ===
        hojas_procesadas = result.get("hojas", [])

        parcelas_directas = _extraer_parcelas_directas(hojas_procesadas)
        if parcelas_directas:
            for p in parcelas_directas:
                cuaderno.agregar_parcela(p)
        else:
            parcelas_data = analisis.get("parcelas", [])
            for p in parcelas_data:
                if p.get("nombre"):
                    parcela = Parcela(
                        nombre=p.get("nombre", ""),
                        referencia_catastral=p.get("referencia_catastral", ""),
                        superficie_ha=float(p.get("superficie_ha", 0) or 0),
                        cultivo=p.get("cultivo", ""),
                        variedad=p.get("variedad", ""),
                        municipio=p.get("municipio", ""),
                        provincia=p.get("provincia", "")
                    )
                    cuaderno.agregar_parcela(parcela)

        # Productos: GPT (no hay hoja estándar de inventario en el cuaderno oficial)
        productos_data = analisis.get("productos", [])
        for prod in productos_data:
            if prod.get("nombre_comercial"):
                producto = ProductoFitosanitario(
                    nombre_comercial=prod.get("nombre_comercial", ""),
                    numero_registro=prod.get("numero_registro", ""),
                    materia_activa=prod.get("materia_activa", ""),
                    formulacion=prod.get("formulacion", ""),
                    numero_lote=prod.get("lote", ""),
                    cantidad_adquirida=float(prod.get("cantidad", 0) or 0),
                    unidad=prod.get("unidad", "L"),
                    fecha_adquisicion=prod.get("fecha_adquisicion", ""),
                    proveedor=prod.get("proveedor", "")
                )
                cuaderno.agregar_producto(producto)

        # Tratamientos: mapeo directo de inf.trat 1
        tratamientos_directos = _extraer_tratamientos_directos(hojas_procesadas, cuaderno.parcelas)
        if tratamientos_directos:
            for t in tratamientos_directos:
                cuaderno.tratamientos.append(t)
            cuaderno.ordenar_tratamientos()
        else:
            tratamientos_data = analisis.get("tratamientos", [])
            for t in tratamientos_data:
                if t.get("fecha") or t.get("producto"):
                    parcela_ids = []
                    parcela_nombre = t.get("parcela", "")
                    ref_cat = (t.get("referencia_catastral") or t.get("ref_catastral") or "").strip()
                    parcela_id_directo = t.get("parcela_id", "").strip()
                    if parcela_id_directo and cuaderno.obtener_parcela(parcela_id_directo):
                        parcela_ids.append(parcela_id_directo)
                        parcela_nombre = cuaderno.obtener_parcela(parcela_id_directo).nombre
                    elif ref_cat:
                        for p in cuaderno.parcelas:
                            if (p.referencia_catastral or "").strip() and (p.referencia_catastral or "").strip().lower() == ref_cat.lower():
                                parcela_ids.append(p.id)
                                parcela_nombre = p.nombre
                                break
                    if not parcela_ids and parcela_nombre:
                        for p in cuaderno.parcelas:
                            if (p.nombre or "").lower() == parcela_nombre.lower():
                                parcela_ids.append(p.id)
                                break
                    producto_aplicado = ProductoAplicado(
                        nombre_comercial=t.get("producto", ""),
                        numero_registro=t.get("numero_registro", "") or t.get("nro_registro", ""),
                        numero_lote=t.get("lote", "") or t.get("numero_lote", ""),
                        dosis=_parse_dosis(t.get("dosis")),
                        unidad_dosis=t.get("unidad_dosis", "L/Ha")
                    )
                    tratamiento = Tratamiento(
                        fecha_aplicacion=t.get("fecha", ""),
                        parcela_ids=parcela_ids,
                        parcela_nombres=[parcela_nombre] if parcela_nombre else [],
                        productos=[producto_aplicado],
                        plaga_enfermedad=t.get("plaga_enfermedad", ""),
                        metodo_aplicacion=t.get("metodo_aplicacion", ""),
                        operador=t.get("operador", "")
                    )
                    cuaderno.tratamientos.append(tratamiento)
        
        # Hojas originales: omitir si solo_datos, o filtrar por índices si hojas_seleccionadas
        if (solo_datos or "").lower() != "true":
            hojas_archivo = result.get("hojas", [])
            indices = None
            if hojas_seleccionadas:
                try:
                    indices = set(json.loads(hojas_seleccionadas))
                except (json.JSONDecodeError, TypeError):
                    pass
            for i, hoja in enumerate(hojas_archivo):
                if indices is not None and i not in indices:
                    continue
                datos = hoja.get("datos", [])
                columnas = list(hoja.get("columnas", []))
                # Filtrar hojas vacías (sin filas con datos reales)
                filas_con_datos = [
                    f for f in datos
                    if any(c is not None and str(c).strip() for c in f)
                ]
                if not filas_con_datos and not any(str(c).strip() for c in columnas):
                    continue  # Omitir hoja vacía
                hoja_original = HojaExcel(
                    nombre=hoja.get("nombre", ""),
                    columnas=columnas,
                    datos=[list(fila) for fila in datos],
                    tipo="custom",
                    origen="importado_editable",
                )
                cuaderno.hojas_originales.append(hoja_original)
        
        # Guardar cuaderno
        storage = get_storage()
        storage.crear(cuaderno)
        
        return {
            "success": True,
            "cuaderno_id": cuaderno.id,
            "nombre": cuaderno.nombre_explotacion,
            "parcelas_creadas": len(cuaderno.parcelas),
            "productos_creados": len(cuaderno.productos),
            "tratamientos_creados": len(cuaderno.tratamientos),
            "hojas_originales": len(cuaderno.hojas_originales),
            "datos_extraidos": analisis,
            "archivo_original": file.filename,
            "message": f"Cuaderno creado con {len(cuaderno.parcelas)} parcelas, {len(cuaderno.productos)} productos, {len(cuaderno.tratamientos)} tratamientos y {len(cuaderno.hojas_originales)} hojas de datos"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error creando cuaderno: {str(e)}"
        )
    finally:
        if tmp is not None:
            tmp.unlink(missing_ok=True)


@router.post("/{cuaderno_id}/upload/import")
async def importar_datos_archivo(cuaderno_id: str, file: UploadFile = File(...)):
    """
    Importa datos desde un archivo a un cuaderno existente.
    Añade parcelas, productos y tratamientos extraídos al cuaderno.
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    processor = get_processor()
    
    if not processor.is_supported(file.filename):
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de archivo no soportado: {file.filename}"
        )
    
    tmp: Optional[Path] = None
    try:
        tmp = await _spool_upload_to_temp(file)
        result = await processor.process_file(tmp, file.filename)
        
        if not result.get("success"):
            raise HTTPException(
                status_code=500,
                detail=result.get("error", "Error procesando archivo")
            )
        
        analisis = result.get("analisis_ia", {})
        
        # Contadores
        parcelas_añadidas = 0
        productos_añadidos = 0
        hojas_añadidas = 0

        # Añadir hojas del archivo. Si ya existe inf.gral 1 o inf.gral 2, REEMPLAZAR con la nueva versión ordenada
        for hoja in result.get("hojas", []):
            nombre = hoja.get("nombre", "").strip()
            hoja_original = HojaExcel(
                nombre=nombre,
                columnas=list(hoja.get("columnas", [])),
                datos=[list(fila) for fila in hoja.get("datos", [])],
                tipo="custom",
                origen="importado",
            )
            # Reemplazar inf.gral 1/2 existentes (nueva extracción ordenada)
            nombre_norm = nombre.lower().strip()
            if nombre_norm in {"inf.gral 1", "inf.gral 2"}:
                for i, h in enumerate(cuaderno.hojas_originales):
                    if (h.nombre or "").strip().lower() == nombre_norm:
                        cuaderno.hojas_originales[i] = hoja_original
                        hojas_añadidas += 1
                        break
                else:
                    cuaderno.hojas_originales.append(hoja_original)
                    hojas_añadidas += 1
            else:
                cuaderno.hojas_originales.append(hoja_original)
                hojas_añadidas += 1
        
        # Importar parcelas: mapeo directo primero, GPT como fallback. Evitar duplicados.
        hojas_procesadas = result.get("hojas", [])
        parcelas_directas = _extraer_parcelas_directas(hojas_procesadas)
        claves_existentes = {_parcela_clave_importacion(p) for p in cuaderno.parcelas}
        nombres_existentes = {(getattr(p, "nombre", "") or "").strip() for p in cuaderno.parcelas}
        if parcelas_directas:
            for p in parcelas_directas:
                clave = _parcela_clave_importacion(p)
                if clave in claves_existentes:
                    continue
                cuaderno.agregar_parcela(p)
                claves_existentes.add(clave)
                parcelas_añadidas += 1
        else:
            for p in analisis.get("parcelas", []):
                if p.get("nombre"):
                    parcela = Parcela(
                        nombre=p.get("nombre", ""),
                        referencia_catastral=p.get("referencia_catastral", ""),
                        superficie_ha=float(p.get("superficie_ha", 0) or 0),
                        cultivo=p.get("cultivo", ""),
                        variedad=p.get("variedad", ""),
                        municipio=p.get("municipio", ""),
                        provincia=p.get("provincia", "")
                    )
                    nombre = (p.get("nombre", "") or "").strip()
                    if not nombre or nombre in nombres_existentes:
                        continue
                    cuaderno.agregar_parcela(parcela)
                    nombres_existentes.add(nombre)
                    parcelas_añadidas += 1

        # Importar tratamientos: mapeo directo
        tratamientos_directos = _extraer_tratamientos_directos(hojas_procesadas, cuaderno.parcelas)
        for t in tratamientos_directos:
            cuaderno.tratamientos.append(t)
        if tratamientos_directos:
            cuaderno.ordenar_tratamientos()

        # Importar productos (GPT)
        for prod in analisis.get("productos", []):
            if prod.get("nombre_comercial"):
                producto = ProductoFitosanitario(
                    nombre_comercial=prod.get("nombre_comercial", ""),
                    numero_registro=prod.get("numero_registro", ""),
                    materia_activa=prod.get("materia_activa", ""),
                    numero_lote=prod.get("lote", ""),
                    cantidad_adquirida=float(prod.get("cantidad", 0) or 0),
                    unidad=prod.get("unidad", "L")
                )
                cuaderno.agregar_producto(producto)
                productos_añadidos += 1
        
        storage.guardar(cuaderno)
        
        return {
            "success": True,
            "cuaderno_id": cuaderno_id,
            "parcelas_añadidas": parcelas_añadidas,
            "productos_añadidos": productos_añadidos,
            "hojas_añadidas": hojas_añadidas,
            "datos_extraidos": analisis,
            "message": f"Importados {parcelas_añadidas} parcelas, {productos_añadidos} productos y {hojas_añadidas} hojas"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error importando datos: {str(e)}"
        )
    finally:
        if tmp is not None:
            tmp.unlink(missing_ok=True)


@router.get("/upload/supported-formats")
async def formatos_soportados():
    """Lista los formatos de archivo soportados para upload"""
    return {
        "formatos": {
            "excel": {
                "extensiones": [".xlsx", ".xls", ".xlsm"],
                "descripcion": "Hojas de cálculo Microsoft Excel"
            },
            "pdf": {
                "extensiones": [".pdf"],
                "descripcion": "Documentos PDF (usa OCR con VIera AI Vision)"
            },
            "imagen": {
                "extensiones": [".png", ".jpg", ".jpeg", ".webp"],
                "descripcion": "Imágenes de cuadernos o albaranes (OCR automático)"
            },
            "csv": {
                "extensiones": [".csv"],
                "descripcion": "Archivos de valores separados por comas"
            }
        },
        "capacidades": [
            "Extracción automática de datos de explotación",
            "OCR con VIera AI Vision para documentos escaneados",
            "Identificación de parcelas, productos y tratamientos",
            "Conversión de formatos legacy a cuaderno digital"
        ]
    }


# ============================================
# CHAT IA - COMANDOS INTELIGENTES CON GPT
# ============================================

class ChatCommand(BaseModel):
    """Comando de chat para ejecutar acciones"""
    cuaderno_id: str
    mensaje: str
    contexto: Optional[Dict[str, Any]] = None


class ChatResponse(BaseModel):
    """Respuesta del chat"""
    success: bool
    mensaje: str
    accion_ejecutada: Optional[str] = None
    datos_creados: Optional[Dict[str, Any]] = None
    sugerencias: Optional[List[str]] = None
    elementos_no_procesados: Optional[List[str]] = None  # Cosas que no se entendieron o no están en el sistema
    advertencias: Optional[List[str]] = None  # Advertencias sobre datos inferidos o por defecto


# Cliente OpenAI global — AsyncOpenAI para no bloquear el event loop durante 5-30s
_openai_client = None

def _get_openai_client():
    """Obtiene o crea el cliente asíncrono de OpenAI."""
    global _openai_client
    if _openai_client is None:
        api_key = os.getenv("OPENAI_API_KEY")
        if api_key:
            try:
                from openai import AsyncOpenAI
                _openai_client = AsyncOpenAI(api_key=api_key)
            except ImportError:
                # Fallback a cliente síncrono si la versión de openai es antigua
                from openai import OpenAI
                _openai_client = OpenAI(api_key=api_key)
    return _openai_client


# ============================================
# MOTOR INTELIGENTE v3.0 - Helpers
# Memoria de conversación, normalización y
# generación basada en plantillas
# ============================================

# Memoria de conversación por cuaderno (para "como te dije", "lo mismo", etc.)
# Almacena (timestamp_último_uso, mensajes) para poder limpiar entradas inactivas
_conversation_memory: Dict[str, Dict] = {}  # {cuaderno_id: {"ts": float, "msgs": [...]}}
MAX_CONVERSATION_HISTORY = 12  # Últimos 12 mensajes (6 intercambios)
_CONV_MEMORY_TTL_H = 4         # Descarta historiales sin actividad durante 4 horas
_CONV_MEMORY_MAX_KEYS = 50     # Límite de cuadernos en memoria simultáneos


def _conversation_cleanup() -> None:
    """Elimina historiales de chat inactivos >4h para evitar fuga de memoria."""
    cutoff = time.time() - _CONV_MEMORY_TTL_H * 3600
    stale = [k for k, v in _conversation_memory.items() if v.get("ts", 0) < cutoff]
    for k in stale:
        _conversation_memory.pop(k, None)
    # Si aún hay demasiadas entradas, eliminar las más antiguas
    if len(_conversation_memory) > _CONV_MEMORY_MAX_KEYS:
        sorted_keys = sorted(_conversation_memory, key=lambda k: _conversation_memory[k].get("ts", 0))
        for k in sorted_keys[:len(_conversation_memory) - _CONV_MEMORY_MAX_KEYS]:
            _conversation_memory.pop(k, None)


def _get_conversation_history(cuaderno_id: str) -> List[Dict[str, str]]:
    """Obtiene el historial de conversación de un cuaderno."""
    entry = _conversation_memory.get(cuaderno_id)
    return entry["msgs"] if entry else []


def _add_to_conversation(cuaderno_id: str, role: str, content: str):
    """Añade un mensaje al historial de conversación."""
    if cuaderno_id not in _conversation_memory:
        _conversation_memory[cuaderno_id] = {"ts": time.time(), "msgs": []}
    entry = _conversation_memory[cuaderno_id]
    entry["ts"] = time.time()
    entry["msgs"].append({"role": role, "content": content})
    if len(entry["msgs"]) > MAX_CONVERSATION_HISTORY:
        entry["msgs"] = entry["msgs"][-MAX_CONVERSATION_HISTORY:]
    # Limpieza periódica (cada ~20 escrituras, sin coste en el path caliente)
    if len(_conversation_memory) > _CONV_MEMORY_MAX_KEYS:
        _conversation_cleanup()


def _normalize_agricultural_text(texto: str) -> str:
    """
    Pre-procesa texto desordenado del usuario antes de enviar a GPT.
    Corrige errores ortográficos comunes, normaliza términos agrícolas.
    """
    import unicodedata
    texto = unicodedata.normalize("NFC", texto)

    corrections = [
        # Cultivos
        (r'\bpatataa?\b', 'patata'), (r'\bolivar\b', 'olivar'),
        (r'\bolibos?\b', 'olivos'), (r'\bsebada\b', 'cebada'),
        (r'\btrgo\b', 'trigo'), (r'\bjirasol\b', 'girasol'),
        (r'\bmais\b', 'maíz'), (r'\bmaís\b', 'maíz'),
        (r'\balmendos?\b', 'almendros'),
        # Términos agrícolas
        (r'\bectareas?\b', 'hectáreas'), (r'\bmunicpio\b', 'municipio'),
        (r'\bmunizipio\b', 'municipio'), (r'\bpoligno\b', 'polígono'),
        (r'\brecnto\b', 'recinto'), (r'\bparsela\b', 'parcela'),
        (r'\bparcla\b', 'parcela'), (r'\bsuperfize\b', 'superficie'),
        (r'\bculitvo\b', 'cultivo'), (r'\bcultibo\b', 'cultivo'),
        (r'\bcomestble\b', 'comestible'), (r'\bsuperfisie\b', 'superficie'),
        (r'\bproducot\b', 'producto'), (r'\btratamientoo?\b', 'tratamiento'),
        (r'\bfitosanitaro\b', 'fitosanitario'),
        # Verbos coloquiales → acción clara
        (r'\bmetele\b', 'añade'), (r'\bponme\b', 'añade'),
        (r'\bhazme\b', 'crea'), (r'\bponle\b', 'añade'),
        (r'\bechale\b', 'añade'), (r'\bmete\b', 'añade'),
        # Verbos de edición (mantener para contexto pero GPT los interpretará)
        (r'\bactualiza\b', 'actualiza'), (r'\bcambia\b', 'cambia'),
        (r'\bmodifica\b', 'modifica'), (r'\bpon\b(?=.*(?:titular|nombre|NIF|domicilio))', 'cambia'),
        (r'\bpon\b(?=.*(?:parcela|superficie|cultivo))', 'modifica'),
    ]

    for pattern, replacement in corrections:
        texto = re.sub(pattern, replacement, texto, flags=re.IGNORECASE)

    return texto.strip()


def _generate_items_from_template(
    template: Dict[str, Any],
    cantidad: int,
    nombre_base: str,
    offset: int = 0,
    variaciones: Optional[List[Dict[str, Any]]] = None
) -> List[Dict[str, Any]]:
    """
    Genera N items a partir de una plantilla base.
    Esto es MUCHO más confiable que pedirle a GPT que genere N items en JSON.
    GPT solo genera la plantilla (1 item) y el conteo; Python replica.
    """
    items = []
    for i in range(cantidad):
        item = template.copy()

        # Aplicar variaciones por índice si existen
        if variaciones and i < len(variaciones):
            for k, v in variaciones[i].items():
                if v is not None and v != "":
                    item[k] = v

        # Generar nombre con numeración
        if nombre_base:
            item["nombre"] = f"{nombre_base} {i + 1}"
        else:
            item["nombre"] = f"Parcela {offset + i + 1}"

        # Asegurar superficie por defecto
        if "superficie" in item and not item.get("superficie"):
            item["superficie"] = 1.0

        items.append(item)

    return items


async def _procesar_con_gpt(
    mensaje: str,
    cuaderno: CuadernoExplotacion,
    contexto: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Motor de IA v3.0 - Procesamiento ultra-inteligente de instrucciones.

    ARQUITECTURA CLAVE: Plantilla + Cantidad
    ─────────────────────────────────────────
    En vez de pedir a GPT que genere N items en JSON (lento, trunca, pierde detalles),
    GPT solo genera UNA plantilla con todos los campos + la cantidad deseada.
    Python se encarga de replicar la plantilla N veces con numeración inteligente.

    Características:
    - Memoria de conversación (entiende "como te dije", "lo mismo que antes")
    - Normalización de texto (corrige errores ortográficos)
    - Generación por plantilla (confiable para cualquier cantidad)
    - Tolerancia extrema a errores de escritura, lenguaje coloquial
    - Extracción exhaustiva de TODOS los atributos mencionados
    """
    client = _get_openai_client()
    if not client:
        return {"error": "OpenAI no configurado"}

    # 1. Pre-procesar texto del usuario
    mensaje_normalizado = _normalize_agricultural_text(mensaje)

    # 2. Contexto del cuaderno (ENRIQUECIDO - datos completos para que GPT use lo que ya existe)
    contexto_cuaderno = {
        "nombre": cuaderno.nombre_explotacion,
        "titular": cuaderno.titular,
        "nif": cuaderno.nif_titular,
        "domicilio": cuaderno.domicilio,
        "codigo_explotacion": cuaderno.codigo_explotacion,
        "parcelas": [
            {"id": p.id, "nombre": p.nombre, "cultivo": p.cultivo or p.especie,
             "especie": p.especie, "variedad": p.variedad,
             "superficie": p.superficie_ha or p.superficie_cultivada,
             "municipio": p.municipio, "provincia": p.provincia,
             "poligono": p.num_poligono, "parcela": p.num_parcela,
             "recinto": p.num_recinto, "uso_sigpac": p.uso_sigpac,
             "secano_regadio": p.secano_regadio, "ecoregimen": p.ecoregimen}
            for p in cuaderno.parcelas[:80]
        ],
        "productos": [
            {"id": p.id, "nombre_comercial": p.nombre_comercial,
             "numero_registro": p.numero_registro, "materia_activa": p.materia_activa,
             "formulacion": p.formulacion,
             "tipo": p.tipo.value if hasattr(p.tipo, 'value') else str(p.tipo),
             "numero_lote": p.numero_lote, "unidad": p.unidad}
            for p in cuaderno.productos[:50]
        ],
        "tratamientos": [
            {"id": t.id,
             "fecha": t.fecha_aplicacion,
             "parcela_ids": t.parcela_ids,
             "parcela_nombres": t.parcela_nombres,
             "cultivo": t.cultivo_especie,
             "problema": t.problema_fitosanitario or t.plaga_enfermedad,
             "productos_usados": [
                 {"producto_id": pa.producto_id, "nombre": pa.nombre_comercial,
                  "registro": pa.numero_registro, "lote": pa.numero_lote,
                  "dosis": pa.dosis, "unidad_dosis": pa.unidad_dosis}
                 for pa in t.productos
             ],
             "superficie_tratada": t.superficie_tratada,
             "aplicador": t.aplicador or t.operador,
             "equipo": t.equipo,
             "estado": t.estado.value if hasattr(t.estado, 'value') else str(t.estado)}
            for t in cuaderno.tratamientos[:30]
        ],
        "num_parcelas": len(cuaderno.parcelas),
        "num_productos": len(cuaderno.productos),
        "num_tratamientos": len(cuaderno.tratamientos),
    }

    # 2b. Contexto de HOJAS IMPORTADAS
    # Hoja activa = TODOS los datos. Otras = muestra de 5 filas (para no exceder el tamaño de prompt)
    active_sheet_id = contexto.get("active_sheet_id") if contexto else None
    hojas_contexto = []
    for hoja in cuaderno.hojas_originales:
        is_active = active_sheet_id and hoja.sheet_id == active_sheet_id
        max_filas = 100 if is_active else 5
        hoja_info = {
            "sheet_id": hoja.sheet_id,
            "nombre": hoja.nombre,
            "columnas": hoja.columnas[:30],
            "num_filas": len(hoja.datos),
        }
        if is_active:
            hoja_info["todos_los_datos"] = []
            for fila in hoja.datos[:max_filas]:
                hoja_info["todos_los_datos"].append([str(v)[:80] if v is not None else "" for v in fila[:25]])
            hoja_info["es_hoja_activa"] = True
        else:
            hoja_info["muestra_datos"] = []
            for fila in hoja.datos[:max_filas]:
                hoja_info["muestra_datos"].append([str(v)[:50] if v is not None else "" for v in fila[:15]])
        hojas_contexto.append(hoja_info)

    if hojas_contexto:
        contexto_cuaderno["hojas_importadas"] = hojas_contexto
        contexto_cuaderno["num_hojas_importadas"] = len(hojas_contexto)

    # 3. Historial de conversación
    cuaderno_id = cuaderno.id
    history = _get_conversation_history(cuaderno_id)

    # 4. System prompt v3.0 - Ultra inteligente con plantilla
    system_prompt = """Eres VIera AI, el asistente MÁS INTELIGENTE del mundo para gestionar Cuadernos de Explotación Agrícola en España.

🧠 TU SUPERPODER: Entiendes CUALQUIER instrucción, sin importar:
- Errores ortográficos ("patataa", "polígno", "recnto", "cultibo")
- Lenguaje coloquial ("metele", "ponme", "hazme", "échale", "y todo eso")
- Frases desordenadas o redundantes ("de cultivo de patatas, de patatas, patatas")
- Referencias a mensajes anteriores ("como te dije", "lo mismo", "igual que antes", "eso")
- Información mezclada, repetida o incompleta
- Abreviaciones ("ha" = hectáreas, "pol" = polígono, "parc" = parcela)

📋 ACCIONES DISPONIBLES:
1. crear_parcelas - Crear parcelas
2. crear_productos - Crear productos fitosanitarios
3. crear_tratamientos - Registrar tratamientos
4. editar_parcelas - Modificar parcelas existentes
5. editar_productos - Modificar productos existentes
6. editar_cuaderno - Cambiar titular, nombre, NIF, domicilio del cuaderno
7. editar_hoja - Editar celdas/filas/columnas en hojas importadas personalizadas
8. buscar_en_hojas - Buscar datos en todas las hojas (nombre, valor, etc.)
9. eliminar - Eliminar elementos (parcelas, productos, tratamientos, hojas)
10. editar_celdas - Editar celdas específicas seleccionadas por el usuario
11. ordenar - Ordenar parcelas/productos/tratamientos/hojas
12. importar_desde_hoja - Importar datos de una hoja importada a parcelas/productos/tratamientos del cuaderno
13. listar - Mostrar información
14. responder - Respuesta conversacional

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🔑 REGLA FUNDAMENTAL: USA LOS DATOS QUE YA EXISTEN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

El cuaderno puede tener datos ya cargados (parcelas, productos, tratamientos).
SIEMPRE revisa el contexto del cuaderno antes de actuar:

- PARCELAS YA EXISTENTES: Mira "parcelas" en el contexto → usa sus IDs y nombres
- PRODUCTOS YA EXISTENTES: Mira "productos" en el contexto → usa sus IDs, nombres, registros
- TRATAMIENTOS PREVIOS: Mira "tratamientos" en el contexto → aprende patrones de uso

Para CREAR TRATAMIENTOS:
- NUNCA pidas productos si ya hay productos en el cuaderno → úsalos directamente
- NUNCA inventes producto_id → usa los IDs del contexto  
- SIEMPRE asigna parcelas reales usando los IDs del contexto
- Si el usuario dice "crea un tratamiento" sin especificar producto, ASIGNA automáticamente los productos existentes
- Si hay varios productos, distribuye entre los tratamientos
- El campo "parcela" en los datos del tratamiento debe ser "todas" o el nombre/filtro de parcelas

Para EDITAR CELDAS:
- Cuando hay celdas seleccionadas, el contexto incluye los valores actuales
- Responde con "editar_celdas" y las ediciones necesarias
- Puedes editar CUALQUIER campo: nombre, cultivo, superficie, municipio, etc.
- Los cambios se aplican directamente sin confirmación

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📚 ESTRUCTURA OFICIAL DEL CUADERNO DE EXPLOTACIÓN (Ministerio de Agricultura, España)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Cuando el usuario pida ORDENAR, ORGANIZAR o FORMATEAR datos, aplica este conocimiento:

📄 ESTRUCTURA GENERAL DEL CUADERNO:
1. Datos de la Explotación (titular, NIF, REGA, domicilio)
2. 2.1 Datos de Parcelas
3. Productos Fitosanitarios (inventario)
4. 3.1 Registro de Tratamientos Fitosanitarios (inf.trat 1)
5. Registro de Fertilizantes (abonado)
6. Registro de Cosecha

📊 ORDEN CORRECTO DE COLUMNAS POR HOJA:

2.1 PARCELAS (orden oficial SIGPAC):
  Nº Orden | Cód. Provincia | Término Municipal | Cód. Agregado | Zona | Nº Polígono | Nº Parcela | Nº Recinto | Uso SIGPAC | Sup. SIGPAC (ha) | Sup. Cultivada (ha) | Especie | Variedad | Ecorregimen | Secano/Regadío | Aire Libre/Protegido

PRODUCTOS FITOSANITARIOS:
  Nombre Comercial | Nº Registro | Materia Activa | Formulación | Tipo | Nº Lote | Cantidad Adquirida | Unidad | Fecha Adquisición | Proveedor | Fecha Caducidad

3.1 REGISTRO TRATAMIENTOS (inf.trat 1):
  Nº Orden Parcelas | Cultivo/Especie | Variedad | Superficie Tratada (ha) | Fecha Aplicación | Problema Fitosanitario | Nombre Comercial Producto | Nº Registro | Nº Lote | Dosis | Unidad Dosis | Caldo/ha | Aplicador | Equipo | Eficacia | Observaciones

FERTILIZANTES:
  Fecha Inicio | Fecha Fin | Nº Orden Parcelas | Cultivo | Tipo Abono | Riqueza N/P/K | Dosis | Tipo Fertilización

COSECHA:
  Fecha | Producto | Cantidad (kg) | Parcelas | Nº Albarán | Nº Lote | Cliente

🔢 REGLAS DE ORDENACIÓN:
- PARCELAS: Ordenar por Cód. Provincia → Término Municipal → Nº Polígono → Nº Parcela → Nº Recinto
  Actualizar "Nº Orden" secuencialmente (1, 2, 3...) tras reordenar
- TRATAMIENTOS: Ordenar por Fecha Aplicación (más reciente primero) → Parcela
  Agrupar tratamientos de la misma fecha y misma parcela
- PRODUCTOS: Ordenar por Nombre Comercial (A-Z)
- FERTILIZANTES: Ordenar por Fecha Inicio (cronológico)
- COSECHA: Ordenar por Fecha (cronológico)

📋 REGLAS DE DATOS:
- "Especie" = cultivo (CEBADA, TRIGO BLANDO, OLIVO, VID...) → MAYÚSCULAS
- "Secano/Regadío" = solo "S" o "R"
- "Uso SIGPAC" = códigos oficiales (TA, OL, VI, FS, FY, PS, PR, PA, OF, TH, FO, FL, CI, etc.)
- "Ecorregimen" = códigos como P3, P5, P1-P9
- Superficies = decimales en hectáreas (ej: 1.2345)
- Fechas = formato DD/MM/AAAA o AAAA-MM-DD
- Nº Registro = formato oficial español (ej: 25.123, ES-00123)
- Los campos de "Nº Orden Parcelas" en tratamientos/fertilizantes/cosecha = números de parcela separados por comas ("1,2,3")

Cuando pidan "ordena" / "organiza" → usa acción "editar_parcelas" con filtro + datos, o "editar_celdas" si hay selección.
Si piden ordenar una hoja importada → usa "editar_hoja" para reorganizar filas.

🔑 NOMBRES DE CAMPOS EDITABLES POR HOJA (usar exactamente estos nombres en "datos"):

PARCELAS (sheet_id: "parcelas"):
  nombre, num_orden, codigo_provincia, termino_municipal, codigo_agregado, zona,
  num_poligono, num_parcela, num_recinto, uso_sigpac, superficie_sigpac,
  superficie_cultivada, superficie_ha, especie, cultivo, variedad, ecoregimen,
  secano_regadio, municipio, provincia, referencia_catastral, notas,
  cultivo_tipo, fecha_inicio_cultivo, fecha_fin_cultivo, aire_libre_protegido
  Alias: cultivo↔especie, municipio↔termino_municipal, superficie↔superficie_ha↔superficie_cultivada

PRODUCTOS (sheet_id: "productos"):
  nombre (→nombre_comercial), registro (→numero_registro), materia_activa,
  formulacion, cantidad (→cantidad_adquirida), unidad, lote (→numero_lote),
  fecha_adquisicion, proveedor, fecha_caducidad, notas

TRATAMIENTOS (sheet_id: "tratamientos"):
  aplicador, equipo, fecha_aplicacion (o fecha), problema_fitosanitario (o problema),
  cultivo_especie (o cultivo), cultivo_variedad, superficie_tratada, eficacia,
  observaciones, justificacion, metodo_aplicacion, condiciones_climaticas,
  hora_inicio, hora_fin, num_orden_parcelas
  Producto aplicado: producto (→nombre_comercial del 1er producto), dosis,
  numero_registro, numero_lote (o lote)

FERTILIZANTES (sheet_id: "fertilizantes"):
  fecha_inicio, fecha_fin, num_orden_parcelas, cultivo_especie, tipo_abono,
  riqueza_npk, dosis, tipo_fertilizacion, observaciones

COSECHA (sheet_id: "cosecha"):
  fecha, producto, cantidad_kg, num_orden_parcelas, num_albaran, num_lote,
  cliente_nombre, cliente_nif

HOJAS IMPORTADAS: Se editan por índice de columna o nombre de columna con "editar_hoja" o "editar_celdas".

⚠️ REGLA DE EDICIÓN vs CREACIÓN:
- Si el usuario dice "pon/cambia/actualiza/modifica" y los datos YA EXISTEN → usar editar_parcelas / editar_productos / editar_tratamientos
- Si el usuario dice "añade/crea/agrega/mete" datos NUEVOS → usar crear_parcelas / crear_productos / crear_tratamientos
- Para rangos de filas → usar filtro con fila_inicio/fila_fin (1-based, como se ven en la UI)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🎯 FORMATO DE RESPUESTA JSON - SISTEMA DE PLANTILLA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Para CREAR elementos (parcelas/productos/tratamientos), usa PLANTILLA + CANTIDAD.
NUNCA generes un array con N elementos repetidos. Solo genera la plantilla (1 objeto):

{
    "accion": "crear_parcelas",
    "cantidad": 30,
    "plantilla": {
        "cultivo": "Patata",
        "especie": "Patata",
        "superficie": 1.0,
        "municipio": "Santa Elena Guairem",
        "termino_municipal": "Santa Elena Guairem",
        "num_poligono": "Recinto Ferial",
        "num_recinto": "ferial",
        "uso_sigpac": "comestible",
        "variedad": "",
        "provincia": "",
        "codigo_provincia": "",
        "num_parcela": "",
        "referencia_catastral": "",
        "secano_regadio": "",
        "ecoregimen": "",
        "notas": ""
    },
    "nombre_base": "San Quintín",
    "mensaje_usuario": "Creando 30 parcelas de patatas 'San Quintín' en Santa Elena Guairem...",
    "elementos_no_procesados": [],
    "advertencias": ["Superficie por defecto: 1.0 ha (no especificada)"]
}

Para PRODUCTOS:
{
    "accion": "crear_productos",
    "cantidad": 2,
    "plantilla": {"nombre": "Glifosato", "registro": "", "materia_activa": "", "formulacion": "", "cantidad": 0, "unidad": "L", "lote": "", "fecha_adquisicion": "", "proveedor": "", "fecha_caducidad": ""},
    "nombre_base": "",
    "variaciones": [
        {"nombre": "Glifosato", "registro": "ES-001234", "materia_activa": "glifosato", "cantidad": 20},
        {"nombre": "Roundup", "registro": "ES-002345", "materia_activa": "glifosato", "cantidad": 15}
    ],
    "mensaje_usuario": "Añadiendo Glifosato y Roundup...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Para TRATAMIENTOS:
{
    "accion": "crear_tratamientos",
    "cantidad": 1,
    "plantilla": {
        "parcela": "Parcela 1", 
        "producto": "Cobre",
        "dosis": 1.5,
        "unidad_dosis": "L/Ha",
        "fecha": "15/03/2026",
        "problema_fitosanitario": "cochinilla",
        "aplicador": "Pedro Sánchez"
    },
    "mensaje_usuario": "Registrando tratamiento de Cobre...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Para TRATAMIENTOS con PARCELAS SELECCIONADAS (cuando el usuario seleccionó parcelas y pidió crear tratamiento):
{
    "accion": "crear_tratamientos",
    "cantidad": 1,
    "plantilla": {
        "parcela_ids": ["abc123", "def456", "ghi789"],
        "producto": "Roundup",
        "dosis": 2.0,
        "unidad_dosis": "L/Ha",
        "fecha": "02/03/2026",
        "problema_fitosanitario": "MALAS HIERBAS",
        "aplicador": "Operario"
    },
    "mensaje_usuario": "Registrando tratamiento para 3 parcelas seleccionadas...",
    "elementos_no_procesados": [],
    "advertencias": []
}
NOTA: "parcela_ids" es un ARRAY de IDs reales de parcelas. Tiene prioridad absoluta sobre "parcela" (nombre).
Cuando el usuario haya seleccionado parcelas, SIEMPRE usa "parcela_ids" con sus IDs exactos.

NOTA: Usa "variaciones" SOLO cuando cada item es DIFERENTE (ej: productos distintos).
Para items iguales (30 parcelas iguales), NO uses variaciones, solo plantilla + cantidad.

Para LISTAR o RESPONDER:
{
    "accion": "listar",
    "mensaje_usuario": "Resumen del cuaderno...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Para EDITAR CUADERNO (titular, nombre explotación, NIF, domicilio):
{
    "accion": "editar_cuaderno",
    "datos": {
        "titular": "Juan Pérez García",
        "nif_titular": "12345678A",
        "nombre_explotacion": "Finca Los Olivos S.L.",
        "domicilio": "Calle Mayor 15, 28001 Madrid",
        "codigo_explotacion": "REGA12345"
    },
    "mensaje_usuario": "Actualizando datos del cuaderno...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Para EDITAR PARCELAS:
{
    "accion": "editar_parcelas",
    "datos": {
        "nombre": "Finca Los Olivos",
        "superficie": 3.0,
        "cultivo": "Olivo"
    },
    "filtro": {
        "cultivo": "Olivo",
        "municipio": "Córdoba"
    },
    "mensaje_usuario": "Modificando parcelas...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Para EDITAR PRODUCTOS EXISTENTES:
{
    "accion": "editar_productos",
    "datos": {
        "cantidad": 20,
        "numero_registro": "ES-001234",
        "materia_activa": "glifosato"
    },
    "filtro": {
        "nombre": "Roundup"
    },
    "mensaje_usuario": "Actualizando productos...",
    "elementos_no_procesados": [],
    "advertencias": []
}
NOTA: "filtro" determina qué productos modificar (por nombre). Si no hay filtro, se modifican TODOS.
"datos" contiene los campos a cambiar. Campos válidos: nombre, registro, materia_activa, formulacion, cantidad, unidad, lote, fecha_adquisicion, proveedor, fecha_caducidad.
IMPORTANTE: Si el usuario dice "en Rudup pon registro ES-001234 y cantidad 20", usa "editar_productos" con filtro nombre=Rudup, NO crear_productos.

Para ELIMINAR:
{
    "accion": "eliminar",
    "tipo_elemento": "parcelas",
    "filtro": {
        "cultivo": "Patata",
        "nombre": "San Quintín"
    },
    "mensaje_usuario": "Eliminando parcelas...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Para EDITAR TRATAMIENTOS EXISTENTES:
{
    "accion": "editar_tratamientos",
    "datos": {
        "aplicador": "Jordan",
        "producto": "Decis"
    },
    "filtro": {
        "fila_inicio": 27,
        "fila_fin": 37
    },
    "mensaje_usuario": "Actualizando tratamientos filas 27-37...",
    "elementos_no_procesados": [],
    "advertencias": []
}
NOTA: "filtro" puede usar: fila_inicio/fila_fin (números de fila visibles, 1-based), nombre de parcela, producto, aplicador, etc.
"datos" campos válidos: aplicador, equipo, fecha_aplicacion, problema_fitosanitario, cultivo_especie, superficie_tratada, eficacia, observaciones, producto (nombre del producto aplicado), dosis, numero_registro, numero_lote.
Si "datos" contiene "producto", se actualiza el nombre_comercial del primer producto aplicado de cada tratamiento.

Para EDITAR HOJAS IMPORTADAS (hojas personalizadas del Excel original):
Las hojas importadas tienen columnas propias y datos libres. Puedes editar celdas, filas enteras o buscar y reemplazar.

Editar celdas específicas:
{
    "accion": "editar_hoja",
    "sheet_id": "el-sheet-id-de-la-hoja",
    "operaciones": [
        {"tipo": "set_cell", "fila": 0, "columna": 2, "valor": "Nuevo valor"},
        {"tipo": "set_cell", "fila": 3, "columna": 0, "valor": "Juan Pérez"}
    ],
    "mensaje_usuario": "Editando hoja...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Buscar y reemplazar en una hoja:
{
    "accion": "editar_hoja",
    "sheet_id": "el-sheet-id-de-la-hoja",
    "operaciones": [
        {"tipo": "buscar_reemplazar", "buscar": "texto viejo", "reemplazar": "texto nuevo"}
    ],
    "mensaje_usuario": "Reemplazando...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Agregar filas a una hoja:
{
    "accion": "editar_hoja",
    "sheet_id": "el-sheet-id-de-la-hoja",
    "operaciones": [
        {"tipo": "agregar_fila", "datos": ["valor1", "valor2", "valor3"]}
    ],
    "mensaje_usuario": "Agregando fila...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Cambiar nombre de columna:
{
    "accion": "editar_hoja",
    "sheet_id": "el-sheet-id-de-la-hoja",
    "operaciones": [
        {"tipo": "renombrar_columna", "columna_idx": 2, "nuevo_nombre": "Nombre Completo"}
    ],
    "mensaje_usuario": "Renombrando columna...",
    "elementos_no_procesados": [],
    "advertencias": []
}

Para BUSCAR EN HOJAS (buscar datos en todas las hojas):
{
    "accion": "buscar_en_hojas",
    "buscar": "Juan Pérez",
    "sheet_id": null,
    "mensaje_usuario": "Buscando...",
    "elementos_no_procesados": [],
    "advertencias": []
}

IMPORTANTE SOBRE HOJAS IMPORTADAS:
- Las hojas importadas aparecen en "hojas_importadas" del contexto
- Cada hoja tiene: sheet_id (UUID), nombre, columnas, num_filas, muestra_datos
- Las columnas NO son fijas - cada hoja tiene sus propias columnas
- Usa la MUESTRA DE DATOS para entender qué contiene cada hoja
- Si el usuario dice "busca dónde está el nombre" → mira las columnas y datos de las hojas
- Si dice "cambia el nombre de la persona en la hoja X" → localiza la columna correcta
- Los índices de fila y columna son 0-based
- Si no conoces el sheet_id exacto, búscalo por NOMBRE de la hoja

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📝 CAMPOS DE PARCELA (rellena TODOS los que el usuario mencione)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- cultivo: Patata, Olivo, Vid, Trigo, Cebada, Almendro, Girasol, Maíz, etc.
- especie: igual que cultivo (siempre sincronizar)
- superficie: hectáreas (decimal). Si no dice, usa 1.0
- municipio: nombre del municipio
- termino_municipal: igual que municipio (sincronizar)
- provincia / codigo_provincia: provincia
- num_poligono: polígono (número o nombre como "Recinto Ferial")
- num_parcela: parcela catastral
- num_recinto: recinto (número o nombre)
- uso_sigpac: uso (TA, OL, VI, comestible, herbáceo, leñoso)
- variedad: variedad específica del cultivo
- referencia_catastral: ref. catastral completa
- secano_regadio: S o R
- ecoregimen: práctica (P3, P5)
- notas: observaciones

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🧠 REGLAS DE INFERENCIA (OBLIGATORIAS)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. "patatas"/"patata"/"papa"/"papas" → cultivo: "Patata"
2. "olivos"/"olivar"/"aceitunas"/"oliva" → cultivo: "Olivo"
3. "viña"/"vid"/"uvas"/"viñedo"/"viñedos" → cultivo: "Vid"
4. "trigo"/"trigal" → cultivo: "Trigo"
5. "cebada"/"cebadal" → cultivo: "Cebada"
6. "almendros"/"almendral"/"almendra" → cultivo: "Almendro"
7. "girasol"/"girasoles" → cultivo: "Girasol"
8. "maíz"/"mais"/"maís" → cultivo: "Maíz"
9. Sin superficie mencionada → 1.0 ha por defecto (y poner en advertencias)
10. "polígono industrial Recinto Ferial" → num_poligono: "Recinto Ferial"
11. "parcela San Quintín" → nombre_base: "San Quintín"
12. "uso comestible" / "comestible" → uso_sigpac: "comestible"
13. "municipio en X" / "municipio X" / "en X" (cuando X es lugar) → municipio: "X"
14. "recinto ferial" / "recinto X" → num_recinto: "ferial" / "X"
15. "como te dije" / "lo de antes" / "igual" → BUSCAR EN HISTORIAL y aplicar datos previos
16. Frases redundantes ("de cultivo de patatas, de patatas") → tomar dato 1 vez: "Patata"
17. "y todo eso" / "y demás" / "etc" → aplicar TODOS los detalles ya extraídos
18. "polígono industrial" es descriptor, no el nombre → extraer el nombre real que sigue

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📝 EJEMPLOS DE PROCESAMIENTO INTELIGENTE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

EJEMPLO 1: Instrucción desordenada con errores
"Quiero que me agregues 30 parcelas de patatas, de cultivo de patatas, con su municipio en Santa Elena Guairem, en polígono industrial Recinto Ferial, con uso comestible, parcela San Quintín, recinto ferial, como te dije, y todo eso"
→ Análisis: cantidad=30, cultivo=Patata, municipio=Santa Elena Guairem, polígono=Recinto Ferial, uso=comestible, nombre_base=San Quintín, recinto=ferial
→ JSON: {"accion": "crear_parcelas", "cantidad": 30, "plantilla": {"cultivo": "Patata", "especie": "Patata", "municipio": "Santa Elena Guairem", "termino_municipal": "Santa Elena Guairem", "num_poligono": "Recinto Ferial", "num_recinto": "ferial", "uso_sigpac": "comestible", "superficie": 1.0}, "nombre_base": "San Quintín", ...}

EJEMPLO 2: Coloquial y simple
"metele 5 parcelas con olivos de 3 hectáreas"
→ {"accion": "crear_parcelas", "cantidad": 5, "plantilla": {"cultivo": "Olivo", "especie": "Olivo", "superficie": 3.0}, "nombre_base": "Parcela", ...}

EJEMPLO 3: Múltiples productos diferentes
"pon glifosato, roundup y confidor como productos con 20, 15 y 5 litros"
→ {"accion": "crear_productos", "cantidad": 3, "plantilla": {"nombre": "", "registro": "", "materia_activa": "", "cantidad": 0, "unidad": "L"}, "nombre_base": "", "variaciones": [{"nombre": "Glifosato", "cantidad": 20}, {"nombre": "Roundup", "cantidad": 15}, {"nombre": "Confidor", "cantidad": 5}], ...}

EJEMPLO 4: Con referencia a historial
"hazme lo mismo pero con 20 parcelas de trigo"
→ Buscar en historial datos previos, cambiar cultivo a Trigo y cantidad a 20

EJEMPLO 5: Editar datos del cuaderno
"actualiza el titular a Juan Pérez García con NIF 12345678A"
→ {"accion": "editar_cuaderno", "datos": {"titular": "Juan Pérez García", "nif_titular": "12345678A"}, "mensaje_usuario": "Actualizando titular...", ...}

"cambia el nombre de la explotación a Finca Los Olivos S.L."
→ {"accion": "editar_cuaderno", "datos": {"nombre_explotacion": "Finca Los Olivos S.L."}, ...}

EJEMPLO 6: Editar parcelas
"cambia el nombre de todas las parcelas de 'Parcela' a 'Finca Los Olivos'"
→ {"accion": "editar_parcelas", "datos": {"nombre": "Finca Los Olivos"}, "filtro": {}, "mensaje_usuario": "Modificando parcelas...", ...}

"actualiza la superficie de las parcelas de trigo a 3 hectáreas"
→ {"accion": "editar_parcelas", "datos": {"superficie": 3.0}, "filtro": {"cultivo": "Trigo"}, ...}

EJEMPLO 6b: Editar productos EXISTENTES (NO crear nuevos)
"en Rudup pon registro ES-001234, materia activa glifosato y cantidad 20"
→ {"accion": "editar_productos", "datos": {"registro": "ES-001234", "materia_activa": "glifosato", "cantidad": 20}, "filtro": {"nombre": "Rudup"}, "mensaje_usuario": "Actualizando Rudup...", ...}

"cambia la cantidad de todos los productos a 10"
→ {"accion": "editar_productos", "datos": {"cantidad": 10}, "filtro": {}, "mensaje_usuario": "Actualizando productos...", ...}

"actualiza Roundup: registro ES-002345, cantidad 15 litros"
→ {"accion": "editar_productos", "datos": {"registro": "ES-002345", "cantidad": 15, "unidad": "L"}, "filtro": {"nombre": "Roundup"}, ...}

⚠️ REGLA: Si el usuario dice "en X producto pon/cambia/actualiza Y" → SIEMPRE usar "editar_productos", NUNCA "crear_productos".

EJEMPLO 6c: Editar tratamientos EXISTENTES
"en las filas 27 a 37 de tratamientos pon aplicador Jordan"
→ {"accion": "editar_tratamientos", "datos": {"aplicador": "Jordan"}, "filtro": {"fila_inicio": 27, "fila_fin": 37}, "mensaje_usuario": "Cambiando aplicador...", ...}

"cambia el producto de todos los tratamientos a Decis"
→ {"accion": "editar_tratamientos", "datos": {"producto": "Decis"}, "filtro": {}, "mensaje_usuario": "Actualizando producto...", ...}

"en los tratamientos de la fila 5 a la 10 pon fecha 15/3/2026 y aplicador Rodman"
→ {"accion": "editar_tratamientos", "datos": {"fecha_aplicacion": "2026-03-15", "aplicador": "Rodman"}, "filtro": {"fila_inicio": 5, "fila_fin": 10}, ...}

"cambia el producto genérico por Decis en tratamientos"
→ {"accion": "editar_tratamientos", "datos": {"producto": "Decis"}, "filtro": {"producto": "Producto genérico"}, ...}

⚠️ REGLA: Si el usuario dice "en filas X a Y de tratamientos" → usar "editar_tratamientos" con filtro fila_inicio/fila_fin.
⚠️ REGLA: Si el usuario dice "cambia/pon/actualiza" en tratamientos → SIEMPRE usar "editar_tratamientos", NUNCA "crear_tratamientos".

EJEMPLO 6d: Editar celdas por número de fila en CUALQUIER hoja
"en la fila 5 de parcelas pon cultivo Olivo"
→ {"accion": "editar_celdas", "sheet_id": "parcelas", "ediciones": [{"row_index": 4, "col_key": "cultivo", "valor": "Olivo"}], ...}

"en productos fila 1 a 4 pon cantidad 10"
→ {"accion": "editar_celdas", "sheet_id": "productos", "ediciones": [{"row_index": 0, "col_key": "cantidad_adquirida", "valor": 10}, {"row_index": 1, "col_key": "cantidad_adquirida", "valor": 10}, {"row_index": 2, "col_key": "cantidad_adquirida", "valor": 10}, {"row_index": 3, "col_key": "cantidad_adquirida", "valor": 10}], ...}

⚠️ NOTA: row_index es 0-based (fila 1 visible = row_index 0). Para rangos grandes, usa "editar_tratamientos"/"editar_productos"/"editar_parcelas" con filtro fila_inicio/fila_fin.

EJEMPLO 7: Eliminar elementos
"elimina las parcelas de patatas"
→ {"accion": "eliminar", "tipo_elemento": "parcelas", "filtro": {"cultivo": "Patata"}, "mensaje_usuario": "Eliminando parcelas...", ...}

"borra el producto Glifosato"
→ {"accion": "eliminar", "tipo_elemento": "productos", "filtro": {"nombre": "Glifosato"}, ...}

EJEMPLO 8: Editar una hoja importada
"cambia el nombre de la persona en la hoja 'Datos Titular' a María López"
→ GPT mira las hojas importadas en el contexto, encuentra la hoja "Datos Titular", identifica la columna de nombre, y:
→ {"accion": "editar_hoja", "sheet_id": "uuid-de-la-hoja", "operaciones": [{"tipo": "buscar_reemplazar", "buscar": "nombre anterior", "reemplazar": "María López"}], ...}

EJEMPLO 9: Buscar en hojas
"busca dónde está el nombre del titular"
→ {"accion": "buscar_en_hojas", "buscar": "titular", "sheet_id": null, ...}

"en la hoja de datos personales, cambia el teléfono a 666123456"
→ GPT busca la hoja "datos personales" en hojas_importadas, identifica columna de teléfono
→ {"accion": "editar_hoja", "sheet_id": "uuid", "operaciones": [{"tipo": "set_cell", "fila": 0, "columna": idx_telefono, "valor": "666123456"}], ...}

EJEMPLO 10: Agregar datos a hoja importada
"añade una fila en la hoja de empleados con Juan García, 28 años, jornalero"
→ {"accion": "editar_hoja", "sheet_id": "uuid", "operaciones": [{"tipo": "agregar_fila", "datos": ["Juan García", "28", "jornalero"]}], ...}

EJEMPLO 11: Reestructurar hoja desordenada
"organiza esa hoja, está muy desordenada y los títulos están en las filas y no en las columnas"
→ GPT lee `todos_los_datos` de la hoja importada, extrae TODA la información y propone la estructura limpia.
⚠️ REGLA CRÍTICA AL REESTRUCTURAR: NUNCA pierdas datos. Incluye TODAS las filas con datos reales en el campo "datos".
   Lee cada fila de `todos_los_datos`, identifica si tiene datos útiles (no vacías, no solo títulos de sección) e INCLÚYELAS TODAS.
   Si una fila tiene un título de sección (ej: "1.3 EQUIPOS DE APLICACIÓN"), NO la pongas como fila de datos, pero SÍ extrae los datos que le siguen.
→ {"accion": "editar_hoja", "sheet_id": "uuid", "operaciones": [{"tipo": "reestructurar", "columnas": ["NOMBRE", "NIF", "ROPO", "TIPO", "EQUIPO", "Nº INSCRIPCIÓN", "FECHA ADQUISICIÓN"], "datos": [["PABLO PEREZ RUBIO", "07990196N", "0737443050U2", "Aplicador", "", "", ""], ["PULVERIZADORES HIDRAUL...", "", "37510010J523", "Equipo", "Pulverizador", "37510010J523", "2021-12-10"], ["Nombre asesor", "NIF asesor", "", "Asesor", "", "", ""]]}], "mensaje_usuario": "Reestructurando...", "elementos_no_procesados": [], "advertencias": []}

EJEMPLO 11: Ordenar datos
"ordena las parcelas"
→ {"accion": "ordenar", "tipo_elemento": "parcelas", "criterio": "oficial", "direccion": "asc", "mensaje_usuario": "Ordenando parcelas según formato oficial SIGPAC...", "elementos_no_procesados": [], "advertencias": []}

"ordena los tratamientos por fecha"
→ {"accion": "ordenar", "tipo_elemento": "tratamientos", "criterio": "fecha", "direccion": "desc", "mensaje_usuario": "Ordenando tratamientos por fecha...", "elementos_no_procesados": [], "advertencias": []}

"ordena los productos alfabéticamente"
→ {"accion": "ordenar", "tipo_elemento": "productos", "criterio": "nombre", "direccion": "asc", "mensaje_usuario": "Ordenando productos...", "elementos_no_procesados": [], "advertencias": []}

"organiza la hoja de datos por la primera columna"
→ {"accion": "ordenar", "tipo_elemento": "sheet-id-uuid", "criterio": "oficial", "columna_idx": 0, "direccion": "asc", "mensaje_usuario": "Ordenando hoja...", "elementos_no_procesados": [], "advertencias": []}

"ordena las parcelas por superficie de mayor a menor"
→ {"accion": "ordenar", "tipo_elemento": "parcelas", "criterio": "superficie", "direccion": "desc", "mensaje_usuario": "Ordenando por superficie...", "elementos_no_procesados": [], "advertencias": []}

EJEMPLO 12: Editar celdas seleccionadas
(cuando hay celdas seleccionadas en el contexto)
"cambia el cultivo a Olivo"
→ {"accion": "editar_celdas", "sheet_id": "parcelas", "ediciones": [{"row_id": "abc123", "col_key": "cultivo", "valor": "Olivo"}, {"row_id": "def456", "col_key": "cultivo", "valor": "Olivo"}], "mensaje_usuario": "Cambiando cultivo a Olivo...", "elementos_no_procesados": [], "advertencias": []}

EJEMPLO 13: Importar datos de hoja importada al cuaderno
"pasa las parcelas de la hoja 2.1. DATOS PARCELAS al cuaderno"
→ {"accion": "importar_desde_hoja", "sheet_id": "uuid-de-la-hoja", "destino": "parcelas", "mapeo_columnas": {"Nro Orden": "num_orden", "Municipio": "municipio", "Polygon": "num_poligono", "Parcel": "num_parcela", "Recinto": "num_recinto", "Uso Sigpac": "uso_sigpac", "Superficie Sigpac": "superficie_ha", "Superficie Cultivada": "superficie_cultivada", "Especie": "especie"}, "mensaje_usuario": "Importando parcelas desde la hoja...", "elementos_no_procesados": [], "advertencias": []}

"importa los tratamientos de inf.trat 1 al cuaderno"
→ {"accion": "importar_desde_hoja", "sheet_id": "uuid-de-inf-trat-1", "destino": "tratamientos", "mapeo_columnas": {"Id Parcelas": "parcela_ids", "Especie": "cultivo_especie", "Superficie Tratada": "superficie_tratada", "Fecha": "fecha_aplicacion", "Problema Fito": "problema_fitosanitario", "Producto": "nombre_comercial", "Nro Registro": "numero_registro", "Dosis": "dosis"}, "mensaje_usuario": "Importando tratamientos...", "elementos_no_procesados": [], "advertencias": []}

⚠️ REGLAS CRÍTICAS:
1. "cantidad" = EXACTAMENTE el número que pide el usuario (30 = 30, no 5 ni 10)
2. "plantilla" = UN solo objeto con TODOS los campos extraídos
3. "nombre_base" = nombre para numerar ("San Quintín" → "San Quintín 1", "San Quintín 2"...)
4. NUNCA preguntes, SIEMPRE ejecuta. Si entiendes algo, hazlo.
5. Si algo no se entiende → "elementos_no_procesados" (pero procesa todo lo demás)
6. Si infieres datos → "advertencias"
7. Responde SOLO JSON, sin texto adicional

🚫 REGLA DE ORO: NUNCA DES ERROR NI TE QUEDES SIN HACER NADA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- NUNCA respondas con un mensaje de error vacío o genérico
- NUNCA digas "no puedo", "no entiendo", "especifica más"
- Si NO entiendes lo que pide el usuario, responde con acción "responder" y en "mensaje_usuario" pon:
  "💡 Puedo hacer esto por ti, solo escríbelo así:" seguido de EJEMPLOS EXACTOS que el usuario pueda copiar y pegar
- Si falta info para ejecutar (ej: crear tratamiento sin productos), HAZ LO MÁXIMO posible con lo que ya tienes
  y pon en "advertencias" lo que asumiste
- Si el usuario pide algo ambiguo, ELIGE la opción más lógica y ejecútala
- Si no hay parcelas/productos para un tratamiento, CRÉALOS automáticamente con datos genéricos y avisa en advertencias
- SIEMPRE da sugerencias útiles y concretas en "sugerencias" (frases exactas copiables)

Ejemplos de sugerencias (SIEMPRE con formato copiable):
sugerencias: [
  "Crea 10 parcelas de olivos en Córdoba",
  "Añade producto Roundup con registro 25.123",
  "Ordena las parcelas por municipio",
  "Cambia el cultivo de todas las parcelas a Vid"
]"""

    # 5. Contexto de hoja activa (enriquecido con datos reales)
    contexto_extra = ""
    if contexto and contexto.get("active_sheet_id"):
        sheet_id = contexto["active_sheet_id"]
        if sheet_id in ("parcelas", "productos", "tratamientos", "fertilizantes", "cosecha", "historico"):
            sheet_names = {"parcelas": "Parcelas", "productos": "Productos", "tratamientos": "Tratamientos",
                          "fertilizantes": "Fertilizantes", "cosecha": "Cosecha", "historico": "Histórico"}
            contexto_extra = f"\nHOJA ACTIVA: {sheet_names.get(sheet_id, sheet_id)} (hoja del sistema)"
        else:
            hoja = cuaderno.obtener_hoja(sheet_id)
            if hoja:
                contexto_extra = f"\nHOJA ACTIVA: \"{hoja.nombre}\" (sheet_id: {hoja.sheet_id})"
                contexto_extra += f"\n  Columnas: {hoja.columnas[:20]}"
                contexto_extra += f"\n  Filas: {len(hoja.datos)}"
                # Muestra de datos para que GPT entienda la estructura
                for i, fila in enumerate(hoja.datos[:5]):
                    fila_str = [str(v)[:40] if v is not None else "" for v in fila[:10]]
                    contexto_extra += f"\n  Fila {i}: {fila_str}"
            else:
                contexto_extra = f"\nHOJA ACTIVA: Hoja importada (ID: {sheet_id})"

    # 5b. Contexto de celdas seleccionadas
    seleccion_extra = ""
    if contexto and contexto.get("selected_cells"):
        sel = contexto["selected_cells"]
        sel_sheet_id = sel.get("sheet_id", "")
        sel_sheet_name = sel.get("sheet_name", "")
        sel_rows = sel.get("rows", [])
        seleccion_extra = f"\n\n━━━ CELDAS SELECCIONADAS POR EL USUARIO ━━━"
        seleccion_extra += f"\nHoja: \"{sel_sheet_name}\" (sheet_id: {sel_sheet_id})"
        seleccion_extra += f"\nFilas seleccionadas: {len(sel_rows)}"
        sel_row_ids = []
        for row_data in sel_rows:
            row_id = row_data.get("row_id", "?")
            row_idx = row_data.get("row_index", "?")
            cells = row_data.get("cells", [])
            cell_strs = [f"{c['col_label']}={c['value']}" for c in cells]
            seleccion_extra += f"\n  Fila {row_idx} (id: {row_id}): {', '.join(cell_strs)}"
            if row_id and row_id != "?":
                sel_row_ids.append(row_id)

        seleccion_extra += "\n\nCuando el usuario pida editar estas celdas, responde con:"
        seleccion_extra += '\n{"accion": "editar_celdas", "sheet_id": "' + sel_sheet_id + '", "ediciones": [{"row_id": "...", "col_key": "...", "valor": "nuevo_valor"}, ...], "mensaje_usuario": "..."}'
        seleccion_extra += "\nPuedes editar CUALQUIER campo de las celdas seleccionadas. Incluye TODAS las ediciones necesarias."
        seleccion_extra += "\nSi el usuario pide cambiar un valor, pon el valor NUEVO en 'valor'. No incluyas celdas que no cambien."

        if sel_sheet_id == "parcelas" and sel_row_ids:
            seleccion_extra += "\n\n🔑 IMPORTANTE - PARCELAS SELECCIONADAS:"
            seleccion_extra += f"\nLos IDs de las parcelas seleccionadas son: {json.dumps(sel_row_ids)}"
            seleccion_extra += "\nSi el usuario pide CREAR UN TRATAMIENTO para estas parcelas (\"para estas\", \"para las seleccionadas\", \"hazle tratamiento\", etc.):"
            seleccion_extra += '\nUsa acción "crear_tratamientos" con "parcela_ids" en la plantilla:'
            seleccion_extra += '\n{"accion": "crear_tratamientos", "cantidad": 1, "plantilla": {"parcela_ids": ' + json.dumps(sel_row_ids) + ', "producto": "...", "dosis": ..., "fecha": "...", "problema_fitosanitario": "...", "aplicador": "..."}, "mensaje_usuario": "..."}'
            seleccion_extra += "\nEl campo \"parcela_ids\" es un ARRAY de IDs reales de parcelas. Tiene prioridad absoluta sobre \"parcela\" (nombre)."
            seleccion_extra += "\nNO uses \"parcela\": \"todas\". Usa SIEMPRE \"parcela_ids\" con los IDs exactos de arriba."

    user_message = f"""CUADERNO ACTUAL:
{json.dumps(contexto_cuaderno, ensure_ascii=False, indent=2)}{contexto_extra}{seleccion_extra}

INSTRUCCIÓN DEL USUARIO (puede estar mal escrita, desordenada, coloquial o incompleta):
"{mensaje_normalizado}"

INSTRUCCIÓN ORIGINAL (sin normalizar):
"{mensaje}"

RECUERDA: Responde SOLO con JSON usando el formato de plantilla + cantidad. NUNCA generes arrays largos de datos repetidos."""

    # 6. Construir mensajes con historial de conversación
    messages = [{"role": "system", "content": system_prompt}]

    # Historial para contexto de "como te dije"
    if history:
        history_lines = []
        for msg in history[-8:]:
            if msg["role"] == "user":
                history_lines.append(f"USUARIO: {msg['content'][:300]}")
            elif msg["role"] == "assistant":
                try:
                    parsed = json.loads(msg["content"])
                    plantilla_info = json.dumps(parsed.get("plantilla", {}), ensure_ascii=False)[:200]
                    history_lines.append(
                        f"ASISTENTE ejecutó: {parsed.get('accion', '?')} x{parsed.get('cantidad', '?')} "
                        f"- {parsed.get('mensaje_usuario', '')[:100]} - plantilla: {plantilla_info}"
                    )
                except Exception:
                    history_lines.append(f"ASISTENTE: {msg['content'][:200]}")

        if history_lines:
            messages.append({
                "role": "system",
                "content": (
                    "HISTORIAL DE CONVERSACIÓN RECIENTE (usa esto para 'como te dije', 'lo mismo', 'igual que antes'):\n"
                    + "\n".join(history_lines)
                )
            })

    messages.append({"role": "user", "content": user_message})

    # 7. JSON parser robusto
    def _try_parse_json(texto: str) -> Dict[str, Any]:
        t = (texto or "").strip()
        if not t:
            raise ValueError("Respuesta vacía")
        if "```json" in t:
            t = t.split("```json")[1].split("```")[0].strip()
        elif "```" in t:
            t = t.split("```")[1].split("```")[0].strip()
        try:
            return json.loads(t)
        except Exception:
            start = t.find("{")
            end = t.rfind("}")
            if start >= 0 and end > start:
                t2 = t[start:end + 1].strip()
                return json.loads(t2)
            raise

    # 8. Llamada a OpenAI — await para no bloquear el event loop
    try:
        import inspect
        _call = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages,
            response_format={"type": "json_object"},
            temperature=0.15,
            max_tokens=4000,
            timeout=60,  # timeout explícito: evita que un hang de OpenAI bloquee indefinidamente
        )
        # Soporta tanto AsyncOpenAI (awaitable) como OpenAI síncrono (fallback legacy)
        response = (await _call) if inspect.isawaitable(_call) else _call

        respuesta_texto = (response.choices[0].message.content or "").strip()
        result = _try_parse_json(respuesta_texto)

        # Guardar respuesta en historial
        _add_to_conversation(cuaderno_id, "assistant", respuesta_texto)

        return result

    except Exception as e:
        return {"error": str(e), "accion": "responder", "mensaje_usuario": f"Error procesando: {str(e)}"}


@router.post("/{cuaderno_id}/chat/execute")
async def ejecutar_comando_chat(cuaderno_id: str, comando: ChatCommand):
    """
    Ejecuta un comando de chat con IA real (GPT).
    Interpreta CUALQUIER instrucción en lenguaje natural y ejecuta las acciones.
    
    Ejemplos:
    - "metele 10 parcelas con productos"
    - "genera tratamientos para todo"
    - "pon 5 productos de herbicida"
    """
    storage = get_storage()
    cuaderno = storage.cargar(cuaderno_id)
    
    if not cuaderno:
        raise HTTPException(status_code=404, detail="Cuaderno no encontrado")
    
    try:
        # Guardar mensaje del usuario en historial de conversación
        _add_to_conversation(cuaderno_id, "user", comando.mensaje)

        # Procesar con GPT
        resultado_gpt = await _procesar_con_gpt(comando.mensaje, cuaderno, comando.contexto)
        # Contexto normalizado para todo el flujo de ejecución
        contexto = comando.contexto if isinstance(comando.contexto, dict) else {}
        
        accion = resultado_gpt.get("accion", "responder")
        datos = resultado_gpt.get("datos", [])
        mensaje_gpt = resultado_gpt.get("mensaje_usuario", "Procesando...")
        elementos_no_procesados = resultado_gpt.get("elementos_no_procesados", [])
        advertencias = resultado_gpt.get("advertencias", [])

        # Modo de ejecución forzada:
        # Si hay selección de celdas y el usuario pide "cambiar/poner/actualizar" pero GPT no devolvió
        # una acción de edición, forzamos editar_celdas para evitar respuestas "vacías".
        selected_ctx = contexto.get("selected_cells") if isinstance(contexto, dict) else None
        msg_lower = (comando.mensaje or "").lower()
        intent_editar = any(k in msg_lower for k in ["cambia", "cambiar", "pon ", "poner", "actualiza", "editar", "modifica", "reemplaza"])
        if (
            isinstance(selected_ctx, dict)
            and selected_ctx.get("rows")
            and intent_editar
            and accion in ("responder", "listar", "", None)
        ):
            rows = selected_ctx.get("rows", []) or []
            # Inferir valor destino básico
            valor = None
            if "hoy" in msg_lower:
                valor = datetime.now().date().isoformat()
            if valor is None:
                m = re.search(r"\ba\s+['\"]?([^'\"]+?)['\"]?$", comando.mensaje or "", flags=re.IGNORECASE)
                if m:
                    valor = m.group(1).strip()
            if valor is None:
                m = re.search(r"\bpor\s+['\"]?([^'\"]+?)['\"]?$", comando.mensaje or "", flags=re.IGNORECASE)
                if m:
                    valor = m.group(1).strip()
            if valor is None:
                m = re.search(r"\bpon(?:er)?\s+['\"]?([^'\"]+?)['\"]?$", comando.mensaje or "", flags=re.IGNORECASE)
                if m:
                    valor = m.group(1).strip()
            if valor is None:
                valor = "ACTUALIZADO"

            ediciones_forzadas = []
            for r in rows:
                row_id = r.get("row_id")
                row_index = r.get("row_index")
                if (row_id is None or row_id == "") and row_index is not None:
                    row_id = str(row_index)
                if row_id is None or row_id == "":
                    continue
                for c in (r.get("cells", []) if isinstance(r, dict) else []):
                    col_key = c.get("col_key")
                    if col_key is None or str(col_key).strip() == "":
                        continue
                    ediciones_forzadas.append({
                        "row_id": str(row_id),
                        "col_key": str(col_key),
                        "valor": valor,
                    })

            if ediciones_forzadas:
                resultado_gpt["accion"] = "editar_celdas"
                resultado_gpt["sheet_id"] = resultado_gpt.get("sheet_id") or selected_ctx.get("sheet_id") or ""
                resultado_gpt["ediciones"] = ediciones_forzadas
                accion = "editar_celdas"
                advertencias = (advertencias or []) + [
                    "Se aplicó modo de ejecución forzada sobre la selección para evitar una respuesta sin cambios."
                ]
        
        # Si hay error pero tenemos datos parciales, procesar lo que sí tenemos
        has_data = (datos and len(datos) > 0) or resultado_gpt.get("plantilla")
        if "error" in resultado_gpt and resultado_gpt.get("accion") != "responder":
            if has_data:
                # Continuar con el procesamiento normal, agregando el error a elementos_no_procesados
                if not elementos_no_procesados:
                    elementos_no_procesados = []
                elementos_no_procesados.append(f"Error parcial procesando: {resultado_gpt.get('error', 'Error desconocido')}")
            else:
                return ChatResponse(
                    success=True,
                    mensaje="💡 No pude procesar tu petición, pero puedo hacer muchas cosas. Prueba con alguna de estas:",
                    sugerencias=["Crea 5 parcelas de olivos en Madrid", "Añade producto Roundup", "Ordena las parcelas", "Cambia el cultivo de todas las parcelas a Trigo"]
                )
        
        elementos_creados = []
        
        # EJECUTAR ACCIONES
        if accion == "crear_parcelas":
            # ───────────────────────────────────────────────
            # v3.0: Generación basada en PLANTILLA + CANTIDAD
            # GPT solo produce la plantilla (1 item) + cantidad
            # Python replica N veces con numeración inteligente
            # ───────────────────────────────────────────────
            plantilla = resultado_gpt.get("plantilla", {})
            cantidad = resultado_gpt.get("cantidad", 0)
            nombre_base = resultado_gpt.get("nombre_base", "")
            variaciones = resultado_gpt.get("variaciones", [])
            datos_legacy = resultado_gpt.get("datos", [])  # Backward compat

            # Compatibilidad: si GPT devolvió "datos" en vez de "plantilla" (formato viejo)
            if not plantilla and datos_legacy:
                plantilla = datos_legacy[0] if datos_legacy else {}
                cantidad = cantidad or len(datos_legacy)

            # Detectar número del mensaje original si GPT no lo capturó
            if not cantidad or cantidad <= 0:
                numeros_en_mensaje = re.findall(r'\b(\d+)\s+parcelas?\b', comando.mensaje.lower())
                cantidad = int(numeros_en_mensaje[0]) if numeros_en_mensaje else (len(datos_legacy) or 1)

            # Inferir nombre_base del mensaje si GPT no lo extrajo
            if not nombre_base:
                match = re.search(r'parcelas?\s+(?:de\s+)?([A-Za-záéíóúñÁÉÍÓÚÑ][A-Za-záéíóúñÁÉÍÓÚÑ\s]{2,}?)(?:\s*,|\s+con|\s+de\s+cultivo|\s+en\s|\s+recinto|\s*$)', comando.mensaje, re.IGNORECASE)
                if match:
                    candidate = match.group(1).strip()
                    # Evitar nombres que son cultivos
                    cultivos_comunes = {"patata", "patatas", "olivo", "olivos", "vid", "viña", "trigo", "cebada", "almendro", "girasol", "maíz"}
                    if candidate.lower() not in cultivos_comunes:
                        nombre_base = candidate
                if not nombre_base:
                    nombre_base = "Parcela"

            if not plantilla:
                return ChatResponse(
                    success=True,
                    mensaje="💡 No pude extraer los datos de las parcelas. Prueba con algo así:",
                    sugerencias=["Crea 10 parcelas de olivos en Madrid", "Añade 5 parcelas de trigo con 2.5 hectáreas", "Pon 3 parcelas de vid en Córdoba"]
                )

            # Generar parcelas desde plantilla usando helper Python
            parcelas_data = _generate_items_from_template(
                template=plantilla,
                cantidad=cantidad,
                nombre_base=nombre_base,
                offset=len(cuaderno.parcelas),
                variaciones=variaciones
            )

            for idx, item in enumerate(parcelas_data):
                try:
                    nombre = item.get("nombre", f"Parcela {len(cuaderno.parcelas) + 1}")
                    if not nombre or not nombre.strip():
                        nombre = f"Parcela {len(cuaderno.parcelas) + 1}"

                    cultivo = (item.get("cultivo") or item.get("especie") or "").strip()
                    superficie = item.get("superficie", 1.0)
                    try:
                        superficie = float(superficie) if superficie else 1.0
                    except (ValueError, TypeError):
                        superficie = 1.0

                    parcela = Parcela(
                        nombre=nombre,
                        cultivo=cultivo,
                        especie=cultivo or (item.get("especie") or "").strip(),
                        superficie_ha=superficie,
                        superficie_cultivada=superficie,
                        superficie_sigpac=item.get("superficie_sigpac", superficie) or superficie,
                        variedad=(item.get("variedad") or "").strip(),
                        municipio=(item.get("municipio") or "").strip(),
                        provincia=(item.get("provincia") or "").strip(),
                        num_poligono=str(item.get("num_poligono") or "").strip(),
                        num_recinto=str(item.get("num_recinto") or "").strip(),
                        uso_sigpac=(item.get("uso_sigpac") or "").strip(),
                        termino_municipal=(item.get("termino_municipal") or item.get("municipio") or "").strip(),
                        codigo_provincia=(item.get("codigo_provincia") or "").strip(),
                        num_parcela=str(item.get("num_parcela") or "").strip(),
                        referencia_catastral=(item.get("referencia_catastral") or "").strip(),
                        ecoregimen=(item.get("ecoregimen") or "").strip(),
                        secano_regadio=(item.get("secano_regadio") or "").strip(),
                        notas=(item.get("notas") or "").strip(),
                        num_orden=len(cuaderno.parcelas) + 1,
                    )
                    cuaderno.agregar_parcela(parcela)
                    elementos_creados.append({"tipo": "parcela", "nombre": parcela.nombre, "id": parcela.id})
                except Exception as e:
                    print(f"Error creando parcela {idx + 1}: {e}")
                    continue

            storage.guardar(cuaderno)

            if len(elementos_creados) == 0:
                return ChatResponse(
                    success=True,
                    mensaje="💡 Algo falló al crear las parcelas. Prueba con un formato más simple:",
                    sugerencias=["Crea 10 parcelas de olivos", "Añade 5 parcelas de trigo en Madrid con 2 hectáreas"]
                )

            # Mensaje de respuesta enriquecido
            mensaje_respuesta = f"✅ {len(elementos_creados)} parcela(s) creada(s) exitosamente"
            if len(elementos_creados) <= 10:
                mensaje_respuesta += ":\n" + "\n".join([f"  • {e['nombre']}" for e in elementos_creados])
            else:
                mensaje_respuesta += (
                    f":\n  • {elementos_creados[0]['nombre']}\n  • {elementos_creados[1]['nombre']}"
                    f"\n  • ...\n  • {elementos_creados[-1]['nombre']}"
                    f"\n  (Total: {len(elementos_creados)} parcelas)"
                )

            # Mostrar detalles de la plantilla aplicada a todas
            detalles = []
            if plantilla.get("cultivo"):
                detalles.append(f"Cultivo: {plantilla['cultivo']}")
            if plantilla.get("municipio"):
                detalles.append(f"Municipio: {plantilla['municipio']}")
            if plantilla.get("num_poligono"):
                detalles.append(f"Polígono: {plantilla['num_poligono']}")
            if plantilla.get("num_recinto"):
                detalles.append(f"Recinto: {plantilla['num_recinto']}")
            if plantilla.get("uso_sigpac"):
                detalles.append(f"Uso SIGPAC: {plantilla['uso_sigpac']}")
            if plantilla.get("superficie"):
                detalles.append(f"Superficie: {plantilla['superficie']} ha")
            if plantilla.get("variedad"):
                detalles.append(f"Variedad: {plantilla['variedad']}")

            if detalles:
                mensaje_respuesta += "\n\n📋 Datos aplicados a todas las parcelas:\n" + "\n".join([f"  • {d}" for d in detalles])

            if advertencias:
                mensaje_respuesta += "\n\n⚠️ Advertencias:\n" + "\n".join([f"  • {a}" for a in advertencias])

            if elementos_no_procesados:
                mensaje_respuesta += "\n\n⚠️ No procesado:\n" + "\n".join([f"  • {e}" for e in elementos_no_procesados])
                mensaje_respuesta += "\n\n✅ Todo lo demás se ha agregado correctamente."

            return ChatResponse(
                success=True,
                mensaje=mensaje_respuesta,
                accion_ejecutada="crear_parcelas",
                datos_creados={"count": len(elementos_creados), "elementos": elementos_creados[:20]},
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        
        elif accion == "crear_productos":
            # v3.0: Soporte plantilla + variaciones para productos
            plantilla_prod = resultado_gpt.get("plantilla", {})
            cantidad_prod = resultado_gpt.get("cantidad", 0)
            variaciones_prod = resultado_gpt.get("variaciones", [])

            # Generar lista de productos desde plantilla/variaciones o datos legacy
            productos_a_crear = []
            if variaciones_prod:
                # Cada variación es un producto diferente
                for var in variaciones_prod:
                    prod_data = {**plantilla_prod, **var}
                    productos_a_crear.append(prod_data)
            elif plantilla_prod:
                # Productos iguales replicados
                for i in range(max(cantidad_prod, 1)):
                    prod_data = plantilla_prod.copy()
                    if not prod_data.get("nombre"):
                        prod_data["nombre"] = f"Producto {len(cuaderno.productos) + i + 1}"
                    productos_a_crear.append(prod_data)
            elif datos:
                # Formato legacy
                productos_a_crear = datos

            for item in productos_a_crear:
                try:
                    raw_cantidad = item.get("cantidad", item.get("cantidad_adquirida", 0))
                    try:
                        cantidad_val = float(raw_cantidad) if raw_cantidad else 0.0
                    except (ValueError, TypeError):
                        cantidad_val = 0.0

                    producto = ProductoFitosanitario(
                        nombre_comercial=item.get("nombre", item.get("nombre_comercial", "Producto")),
                        numero_registro=item.get("registro", item.get("numero_registro", "")),
                        materia_activa=item.get("materia_activa", ""),
                        formulacion=item.get("formulacion", ""),
                        tipo=TipoProducto(item["tipo"]) if item.get("tipo") and item["tipo"] in [e.value for e in TipoProducto] else TipoProducto.FITOSANITARIO,
                        numero_lote=item.get("lote", item.get("numero_lote", "")),
                        cantidad_adquirida=cantidad_val,
                        unidad=item.get("unidad", "L"),
                        fecha_adquisicion=item.get("fecha_adquisicion", ""),
                        proveedor=item.get("proveedor", ""),
                        fecha_caducidad=item.get("fecha_caducidad", ""),
                    )
                    cuaderno.agregar_producto(producto)
                    elementos_creados.append({"tipo": "producto", "nombre": producto.nombre_comercial, "id": producto.id})
                except Exception as e:
                    print(f"Error creando producto: {e}")
                    continue

            storage.guardar(cuaderno)

            mensaje_respuesta = f"✅ {len(elementos_creados)} producto(s) añadido(s):\n" + "\n".join([f"  • {e['nombre']}" for e in elementos_creados])

            if advertencias:
                mensaje_respuesta += "\n\n⚠️ Advertencias:\n" + "\n".join([f"  • {a}" for a in advertencias])

            if elementos_no_procesados:
                mensaje_respuesta += f"\n\n⚠️ No procesado:\n" + "\n".join([f"  • {e}" for e in elementos_no_procesados])
                mensaje_respuesta += "\n\n✅ Todo lo demás se ha agregado correctamente."

            return ChatResponse(
                success=True,
                mensaje=mensaje_respuesta,
                accion_ejecutada="crear_productos",
                datos_creados={"count": len(elementos_creados), "elementos": elementos_creados},
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        
        elif accion == "crear_tratamientos":
            from datetime import date, timedelta
            
            # Si no hay parcelas o productos, crearlos primero
            if not cuaderno.parcelas:
                # Auto-crear parcela genérica en vez de dar error
                from cuaderno.models import Parcela as ParcelaModel
                auto_parcela = ParcelaModel(nombre="Parcela 1", cultivo="Sin definir", especie="Sin definir", superficie_ha=1.0)
                cuaderno.agregar_parcela(auto_parcela)
                storage.guardar(cuaderno)
                advertencias.append("Se creó una parcela genérica automáticamente. Edítala con tus datos reales.")
            
            if not cuaderno.productos:
                # Auto-crear producto genérico en vez de dar error
                from cuaderno.models import ProductoFitosanitario as ProdModel
                auto_prod = ProdModel(nombre_comercial="Producto genérico", numero_registro="-")
                cuaderno.agregar_producto(auto_prod)
                storage.guardar(cuaderno)
                advertencias.append("Se creó un producto genérico automáticamente. Edítalo con tus datos reales.")
            
            # v3.0: Soporte plantilla + variaciones para tratamientos
            plantilla_trat = resultado_gpt.get("plantilla", {})
            cantidad_trat = resultado_gpt.get("cantidad", 0)
            variaciones_trat = resultado_gpt.get("variaciones", [])

            # Generar lista de tratamientos desde plantilla/variaciones o datos legacy
            tratamientos_a_crear = []
            if variaciones_trat:
                for var in variaciones_trat:
                    trat_data = {**plantilla_trat, **var}
                    tratamientos_a_crear.append(trat_data)
            elif plantilla_trat:
                for i in range(max(cantidad_trat, 1)):
                    trat_data = plantilla_trat.copy()
                    tratamientos_a_crear.append(trat_data)
            elif datos:
                # Formato legacy
                tratamientos_a_crear = datos
            
            # 5b-extra: IDs de parcelas seleccionadas desde el contexto (fallback si la IA no puso parcela_ids)
            sel_parcela_ids_ctx = []
            if contexto and contexto.get("selected_cells"):
                sel = contexto["selected_cells"]
                if sel.get("sheet_id") == "parcelas":
                    sel_parcela_ids_ctx = [r.get("row_id") for r in sel.get("rows", []) if r.get("row_id")]

            for i, item in enumerate(tratamientos_a_crear):
                # Prioridad 1: parcela_ids explícitos (array de IDs)
                parcela_ids_directos = item.get("parcela_ids", [])
                if isinstance(parcela_ids_directos, str):
                    parcela_ids_directos = [x.strip() for x in parcela_ids_directos.split(",") if x.strip()]

                if parcela_ids_directos:
                    parcelas_usar = [p for p in cuaderno.parcelas if p.id in parcela_ids_directos]
                    if not parcelas_usar:
                        advertencias.append(f"Tratamiento {i+1}: IDs de parcelas no encontrados, se usarán todas.")
                        parcelas_usar = cuaderno.parcelas
                # Prioridad 2: IDs del contexto de selección
                elif sel_parcela_ids_ctx:
                    parcelas_usar = [p for p in cuaderno.parcelas if p.id in sel_parcela_ids_ctx]
                    if not parcelas_usar:
                        parcelas_usar = cuaderno.parcelas
                else:
                    # Prioridad 3: nombre de parcela (legacy)
                    parcela_nombre = item.get("parcela", item.get("parcela_nombre", "todas"))
                    if parcela_nombre == "todas" or not parcela_nombre:
                        parcelas_usar = cuaderno.parcelas
                    else:
                        parcelas_usar = [p for p in cuaderno.parcelas if str(parcela_nombre).lower() in p.nombre.lower()]
                        if not parcelas_usar:
                            from cuaderno.models import Parcela as ParcelaModel
                            auto_parcela = ParcelaModel(nombre=str(parcela_nombre), cultivo="Sin definir", especie="Sin definir", superficie_ha=1.0)
                            cuaderno.agregar_parcela(auto_parcela)
                            parcelas_usar = [auto_parcela]
                            advertencias.append(f"Se creó la parcela '{parcela_nombre}' ya que no existía en el cuaderno.")
                
                # Seleccionar producto
                producto_nombre = item.get("producto", item.get("producto_nombre", ""))
                if producto_nombre:
                    productos_match = [p for p in cuaderno.productos if str(producto_nombre).lower() in p.nombre_comercial.lower()]
                    producto = productos_match[0] if productos_match else cuaderno.productos[0]
                else:
                    producto = cuaderno.productos[i % len(cuaderno.productos)]
                
                producto_aplicado = ProductoAplicado(
                    producto_id=producto.id,
                    nombre_comercial=producto.nombre_comercial,
                    numero_registro=producto.numero_registro,
                    dosis=float(item.get("dosis", 2.0)),
                    unidad_dosis=item.get("unidad_dosis", "L/Ha")
                )
                
                # Parsear fecha
                fecha_str = item.get("fecha", item.get("fecha_aplicacion", ""))
                try:
                    if "/" in fecha_str:
                        d, m, y = fecha_str.split("/")
                        if len(y) == 2: y = "20" + y
                        fecha_val = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
                    elif "-" in fecha_str and len(fecha_str) == 10:
                        fecha_val = fecha_str
                    else:
                        fecha_val = (date.today() - timedelta(days=i)).isoformat()
                except Exception:
                    fecha_val = (date.today() - timedelta(days=i)).isoformat()
                
                plantilla_chat = Tratamiento(
                    fecha_aplicacion=fecha_val,
                    parcela_ids=[p.id for p in parcelas_usar],
                    productos=[producto_aplicado],
                    problema_fitosanitario=item.get("problema", item.get("problema_fitosanitario", item.get("plaga", "Mantenimiento"))),
                    operador=item.get("operador", item.get("aplicador", "Operario")),
                    aplicador=item.get("aplicador", "Propietario"),
                    estado=EstadoTratamiento.APLICADO
                )
                creados_desgl = cuaderno.agregar_tratamiento_desglosado(plantilla_chat)
                for td in creados_desgl:
                    elementos_creados.append({
                        "tipo": "tratamiento",
                        "parcelas": td.parcela_nombres,
                        "producto": td.productos[0].nombre_comercial if td.productos else "",
                        "id": td.id
                    })
            
            storage.guardar(cuaderno)
            mensaje_respuesta = f"✅ {len(elementos_creados)} tratamiento(s) registrado(s)"
            
            if advertencias:
                mensaje_respuesta += "\n\n⚠️ Advertencias:\n" + "\n".join([f"• {a}" for a in advertencias])
            
            if elementos_no_procesados:
                mensaje_respuesta += f"\n\n⚠️ No se pudo procesar lo siguiente (no está en el sistema o no se entendió):\n" + "\n".join([f"• {e}" for e in elementos_no_procesados])
                mensaje_respuesta += "\n\n✅ Todo lo demás se ha agregado correctamente."
            
            return ChatResponse(
                success=True,
                mensaje=mensaje_respuesta,
                accion_ejecutada="crear_tratamientos",
                datos_creados={"count": len(elementos_creados), "elementos": elementos_creados},
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        
        elif accion == "importar_desde_hoja":
            # Importar datos desde una hoja importada al cuaderno (parcelas, productos, tratamientos)
            sheet_id_source = resultado_gpt.get("sheet_id", "")
            destino = resultado_gpt.get("destino", "parcelas")
            mapeo = resultado_gpt.get("mapeo_columnas", {})
            
            # Buscar la hoja fuente
            hoja_source = None
            if sheet_id_source:
                hoja_source = cuaderno.obtener_hoja(sheet_id_source)
            if not hoja_source and sheet_id_source:
                # Buscar por nombre
                for h in cuaderno.hojas_originales:
                    if sheet_id_source.upper() in h.nombre.upper() or h.nombre.upper() in sheet_id_source.upper():
                        hoja_source = h
                        break
            # Si viene del contexto
            if not hoja_source and contexto and contexto.get("active_sheet_id"):
                hoja_source = cuaderno.obtener_hoja(contexto["active_sheet_id"])
            
            if not hoja_source:
                return ChatResponse(
                    success=True,
                    mensaje="💡 No encontré la hoja de origen. Especifica el nombre exacto:",
                    sugerencias=[f"Pasa los datos de {h.nombre} a {destino}" for h in cuaderno.hojas_originales[:4]]
                )
            
            columnas_hoja = hoja_source.columnas
            filas = hoja_source.datos
            importados = 0
            
            def _get_col_idx(col_name):
                """Encuentra el índice de una columna por nombre (fuzzy)."""
                cn = col_name.strip().upper()
                for i, c in enumerate(columnas_hoja):
                    if cn in c.upper() or c.upper() in cn:
                        return i
                return -1
            
            def _get_val(row, col_name):
                """Obtiene el valor de una fila por nombre de columna."""
                idx = _get_col_idx(col_name)
                if idx < 0 or idx >= len(row):
                    return None
                v = row[idx]
                if v is None or str(v).strip().upper() == "NONE" or str(v).strip() == "":
                    return None
                return str(v).strip()
            
            if destino == "parcelas":
                from cuaderno.models import Parcela as ParcelaModel
                claves_parcelas = {_parcela_clave_importacion(p) for p in cuaderno.parcelas}
                for fila in filas:
                    # Saltar filas vacías
                    vals = [v for v in fila if v is not None and str(v).strip() and str(v).strip().upper() != "NONE"]
                    if len(vals) < 2:
                        continue
                    
                    # Construir parcela con mapeo
                    kwargs = {}
                    for col_hoja, campo_destino in mapeo.items():
                        val = _get_val(fila, col_hoja)
                        if val:
                            if campo_destino in ("superficie_ha", "superficie_cultivada", "superficie_sigpac"):
                                try:
                                    kwargs[campo_destino] = float(val)
                                except ValueError:
                                    kwargs[campo_destino] = val
                            elif campo_destino == "num_orden":
                                try:
                                    kwargs[campo_destino] = int(float(val))
                                except ValueError:
                                    pass
                            else:
                                # Limpiar códigos de municipio tipo "301-TORO" → municipio="TORO"
                                if campo_destino in ("municipio", "termino_municipal") and "-" in val:
                                    parts = val.split("-", 1)
                                    kwargs["termino_municipal"] = parts[1].strip()
                                    kwargs["municipio"] = parts[1].strip()
                                else:
                                    kwargs[campo_destino] = val
                    
                    if not kwargs.get("nombre"):
                        orden = kwargs.get("num_orden", importados + 1)
                        mun = kwargs.get("municipio", kwargs.get("termino_municipal", ""))
                        kwargs["nombre"] = f"Parcela {orden}" + (f" ({mun})" if mun else "")
                    
                    try:
                        p = ParcelaModel(**{k: v for k, v in kwargs.items() if hasattr(ParcelaModel, k) or k in ParcelaModel.__dataclass_fields__})
                    except Exception:
                        p = ParcelaModel(
                            nombre=kwargs.get("nombre", f"Parcela {importados+1}"),
                            municipio=kwargs.get("municipio", ""),
                            num_poligono=kwargs.get("num_poligono", ""),
                            num_parcela=kwargs.get("num_parcela", ""),
                            num_recinto=kwargs.get("num_recinto", ""),
                            uso_sigpac=kwargs.get("uso_sigpac", ""),
                            superficie_ha=kwargs.get("superficie_ha", 0),
                            especie=kwargs.get("especie", ""),
                        )
                    clave = _parcela_clave_importacion(p)
                    if clave in claves_parcelas:
                        continue
                    cuaderno.agregar_parcela(p)
                    claves_parcelas.add(clave)
                    importados += 1
                
                storage.guardar(cuaderno)
                return ChatResponse(
                    success=True,
                    mensaje=f"✅ {importados} parcelas importadas desde \"{hoja_source.nombre}\" al cuaderno.",
                    accion_ejecutada="importar_desde_hoja",
                    datos_creados={"tipo": "parcelas", "count": importados, "hoja_origen": hoja_source.nombre}
                )
            
            elif destino == "productos":
                from cuaderno.models import ProductoFitosanitario as ProdModel
                for fila in filas:
                    vals = [v for v in fila if v is not None and str(v).strip() and str(v).strip().upper() != "NONE"]
                    if len(vals) < 1:
                        continue
                    
                    kwargs = {}
                    for col_hoja, campo_destino in mapeo.items():
                        val = _get_val(fila, col_hoja)
                        if val:
                            kwargs[campo_destino] = val
                    
                    if not kwargs.get("nombre_comercial"):
                        kwargs["nombre_comercial"] = vals[0] if vals else f"Producto {importados+1}"
                    
                    try:
                        prod = ProdModel(**{k: v for k, v in kwargs.items() if hasattr(ProdModel, k) or k in ProdModel.__dataclass_fields__})
                    except Exception:
                        prod = ProdModel(nombre_comercial=kwargs.get("nombre_comercial", f"Producto {importados+1}"))
                    cuaderno.agregar_producto(prod)
                    importados += 1
                
                storage.guardar(cuaderno)
                return ChatResponse(
                    success=True,
                    mensaje=f"✅ {importados} productos importados desde \"{hoja_source.nombre}\".",
                    accion_ejecutada="importar_desde_hoja",
                    datos_creados={"tipo": "productos", "count": importados}
                )
            
            elif destino == "tratamientos":
                for fila in filas:
                    vals = [v for v in fila if v is not None and str(v).strip() and str(v).strip().upper() != "NONE"]
                    if len(vals) < 2:
                        continue
                    
                    kwargs = {}
                    for col_hoja, campo_destino in mapeo.items():
                        val = _get_val(fila, col_hoja)
                        if val:
                            kwargs[campo_destino] = val
                    
                    # Crear producto aplicado si hay datos
                    productos_aplicados = []
                    if kwargs.get("nombre_comercial"):
                        pa = ProductoAplicado(
                            nombre_comercial=kwargs.get("nombre_comercial", ""),
                            numero_registro=kwargs.get("numero_registro", ""),
                            dosis=float(kwargs.get("dosis", 0)) if kwargs.get("dosis") else 0,
                            unidad_dosis=kwargs.get("unidad_dosis", "L/Ha"),
                        )
                        productos_aplicados.append(pa)
                    
                    # Parsear fecha
                    fecha = kwargs.get("fecha_aplicacion", "")
                    if fecha and "T" in fecha:
                        fecha = fecha.split("T")[0]
                    
                    # Parsear parcela_ids
                    parcela_ids_str = kwargs.get("parcela_ids", "")
                    parcela_nombres = []
                    parcela_ids_list = []
                    if parcela_ids_str:
                        for pid in str(parcela_ids_str).split(","):
                            pid = pid.strip()
                            # Buscar por num_orden
                            for p in cuaderno.parcelas:
                                if str(p.num_orden) == pid:
                                    parcela_ids_list.append(p.id)
                                    parcela_nombres.append(p.nombre)
                                    break
                    
                    trat = Tratamiento(
                        fecha_aplicacion=fecha,
                        parcela_ids=parcela_ids_list,
                        parcela_nombres=parcela_nombres,
                        cultivo_especie=kwargs.get("cultivo_especie", ""),
                        superficie_tratada=float(kwargs.get("superficie_tratada", 0)) if kwargs.get("superficie_tratada") else 0,
                        problema_fitosanitario=kwargs.get("problema_fitosanitario", ""),
                        productos=productos_aplicados,
                        aplicador=kwargs.get("aplicador", ""),
                        equipo=kwargs.get("equipo", ""),
                        estado=EstadoTratamiento.APLICADO,
                    )
                    cuaderno.agregar_tratamiento(trat)
                    importados += 1
                
                storage.guardar(cuaderno)
                return ChatResponse(
                    success=True,
                    mensaje=f"✅ {importados} tratamientos importados desde \"{hoja_source.nombre}\".",
                    accion_ejecutada="importar_desde_hoja",
                    datos_creados={"tipo": "tratamientos", "count": importados}
                )
            
            return ChatResponse(
                success=True,
                mensaje=f"💡 No reconocí el destino \"{destino}\". Puedo importar a:",
                sugerencias=["Importa las parcelas de la hoja X al cuaderno", "Importa los tratamientos de inf.trat 1", "Importa los productos de Hoja1"]
            )
        
        elif accion == "ordenar":
            # Ordenar datos según estructura oficial del Cuaderno de Explotación
            tipo_ordenar = resultado_gpt.get("tipo_elemento", resultado_gpt.get("sheet_id", "parcelas"))
            criterio = resultado_gpt.get("criterio", "oficial")  # "oficial", "nombre", "fecha", "superficie", etc.
            direccion = resultado_gpt.get("direccion", "asc")  # "asc" o "desc"
            
            def _safe_str(val):
                return str(val or "").strip().upper()
            
            def _safe_num(val):
                try:
                    return float(val) if val else 0.0
                except (ValueError, TypeError):
                    return 0.0
            
            def _safe_int(val):
                try:
                    # Extraer solo dígitos para polígonos/parcelas que pueden ser "003"
                    s = str(val or "0").strip()
                    digits = ''.join(c for c in s if c.isdigit())
                    return int(digits) if digits else 0
                except (ValueError, TypeError):
                    return 0
            
            if tipo_ordenar == "parcelas":
                if not cuaderno.parcelas:
                    return ChatResponse(success=True, mensaje="📋 No hay parcelas todavía. Prueba a crearlas primero.", sugerencias=["Crea 10 parcelas de olivos", "Añade 5 parcelas de trigo en Córdoba"])
                
                if criterio in ("oficial", "sigpac", "provincia", "municipio", "default"):
                    # Orden oficial SIGPAC: Provincia → Municipio → Polígono → Parcela → Recinto
                    cuaderno.parcelas.sort(key=lambda p: (
                        _safe_str(p.codigo_provincia),
                        _safe_str(p.termino_municipal or p.municipio),
                        _safe_int(p.num_poligono),
                        _safe_int(p.num_parcela),
                        _safe_int(p.num_recinto),
                    ))
                elif criterio in ("cultivo", "especie"):
                    cuaderno.parcelas.sort(key=lambda p: _safe_str(p.especie or p.cultivo))
                elif criterio in ("superficie", "hectareas", "ha"):
                    cuaderno.parcelas.sort(key=lambda p: _safe_num(p.superficie_ha or p.superficie_cultivada), reverse=(direccion == "desc"))
                elif criterio == "nombre":
                    def _extract_number(name):
                        m = re.search(r'\d+', _safe_str(name))
                        return int(m.group()) if m else 0
                    cuaderno.parcelas.sort(key=lambda p: (_extract_number(p.nombre), _safe_str(p.nombre)))
                else:
                    cuaderno.parcelas.sort(key=lambda p: (
                        _safe_str(p.codigo_provincia),
                        _safe_str(p.termino_municipal or p.municipio),
                        _safe_int(p.num_poligono),
                        _safe_int(p.num_parcela),
                    ))
                
                # Actualizar nº de orden secuencialmente
                for i, p in enumerate(cuaderno.parcelas):
                    p.num_orden = i + 1
                
                storage.guardar(cuaderno)
                return ChatResponse(
                    success=True,
                    mensaje=f"✅ {len(cuaderno.parcelas)} parcelas ordenadas ({criterio}). Nº de orden actualizado (1-{len(cuaderno.parcelas)}).",
                    accion_ejecutada="ordenar",
                    datos_creados={"tipo": "parcelas", "count": len(cuaderno.parcelas), "criterio": criterio}
                )
            
            elif tipo_ordenar == "productos":
                if not cuaderno.productos:
                    return ChatResponse(success=True, mensaje="📋 No hay productos todavía. Prueba a crearlos primero.", sugerencias=["Añade producto Roundup", "Pon 3 productos fitosanitarios"])
                
                if criterio in ("nombre", "oficial", "default", "az"):
                    cuaderno.productos.sort(key=lambda p: _safe_str(p.nombre_comercial))
                elif criterio in ("tipo", "categoria"):
                    cuaderno.productos.sort(key=lambda p: _safe_str(p.tipo.value if hasattr(p.tipo, 'value') else str(p.tipo)))
                elif criterio in ("fecha", "adquisicion"):
                    cuaderno.productos.sort(key=lambda p: p.fecha_adquisicion or "", reverse=(direccion == "desc"))
                
                storage.guardar(cuaderno)
                return ChatResponse(
                    success=True,
                    mensaje=f"✅ {len(cuaderno.productos)} productos ordenados ({criterio}).",
                    accion_ejecutada="ordenar",
                    datos_creados={"tipo": "productos", "count": len(cuaderno.productos), "criterio": criterio}
                )
            
            elif tipo_ordenar == "tratamientos":
                if not cuaderno.tratamientos:
                    return ChatResponse(success=True, mensaje="📋 No hay tratamientos todavía. Prueba a crearlos primero.", sugerencias=["Crea un tratamiento para todas las parcelas", "Registra un tratamiento con Roundup"])
                
                if criterio in ("fecha", "oficial", "default", "cronologico"):
                    cuaderno.tratamientos.sort(key=lambda t: t.fecha_aplicacion or "", reverse=(direccion != "asc"))
                elif criterio in ("parcela", "parcelas"):
                    cuaderno.tratamientos.sort(key=lambda t: _safe_str(",".join(t.parcela_nombres)))
                elif criterio in ("cultivo", "especie"):
                    cuaderno.tratamientos.sort(key=lambda t: _safe_str(t.cultivo_especie))
                
                storage.guardar(cuaderno)
                return ChatResponse(
                    success=True,
                    mensaje=f"✅ {len(cuaderno.tratamientos)} tratamientos ordenados ({criterio}).",
                    accion_ejecutada="ordenar",
                    datos_creados={"tipo": "tratamientos", "count": len(cuaderno.tratamientos), "criterio": criterio}
                )
            
            else:
                # Intentar ordenar hoja importada
                hoja = cuaderno.obtener_hoja(tipo_ordenar)
                if hoja and hoja.datos:
                    col_idx = resultado_gpt.get("columna_idx", 0)
                    try:
                        col_idx = int(col_idx)
                    except (ValueError, TypeError):
                        # Buscar por nombre de columna
                        col_name = str(col_idx).strip().upper()
                        col_idx = next((i for i, c in enumerate(hoja.columnas) if c.strip().upper() == col_name), 0)
                    
                    reverse = (direccion == "desc")
                    hoja.datos.sort(key=lambda row: _safe_str(row[col_idx]) if col_idx < len(row) else "", reverse=reverse)
                    storage.guardar(cuaderno)
                    col_name = hoja.columnas[col_idx] if col_idx < len(hoja.columnas) else str(col_idx)
                    return ChatResponse(
                        success=True,
                        mensaje=f"✅ Hoja \"{hoja.nombre}\" ordenada por \"{col_name}\" ({len(hoja.datos)} filas).",
                        accion_ejecutada="ordenar",
                        datos_creados={"tipo": "hoja_importada", "hoja": hoja.nombre, "columna": col_name}
                    )
                
                return ChatResponse(
                    success=True,
                    mensaje=f"💡 No encontré la hoja \"{tipo_ordenar}\". Puedo ordenar estas secciones:",
                    sugerencias=["Ordena las parcelas", "Ordena los tratamientos por fecha", "Ordena los productos alfabéticamente"]
                )
        
        elif accion == "listar":
            resumen = f"""📊 **Resumen del Cuaderno: {cuaderno.nombre_explotacion}**

• {len(cuaderno.parcelas)} Parcelas
• {len(cuaderno.productos)} Productos  
• {len(cuaderno.tratamientos)} Tratamientos

**Parcelas:** {', '.join([p.nombre for p in cuaderno.parcelas[:5]]) if cuaderno.parcelas else 'Ninguna'}
**Productos:** {', '.join([p.nombre_comercial for p in cuaderno.productos[:5]]) if cuaderno.productos else 'Ninguno'}"""
            
            mensaje_respuesta = resumen
            if elementos_no_procesados:
                mensaje_respuesta += f"\n\n⚠️ No se pudo procesar lo siguiente (no está en el sistema o no se entendió):\n" + "\n".join([f"• {e}" for e in elementos_no_procesados])
            
            return ChatResponse(
                success=True,
                mensaje=mensaje_respuesta,
                datos_creados={
                    "parcelas": len(cuaderno.parcelas),
                    "productos": len(cuaderno.productos),
                    "tratamientos": len(cuaderno.tratamientos)
                },
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        
        elif accion in ("editar_cuaderno", "actualizar_cuaderno", "cambiar_cuaderno"):
            # Editar datos del cuaderno (titular, nombre explotación, NIF, domicilio, etc.)
            datos_edicion = resultado_gpt.get("datos", {}) or resultado_gpt.get("plantilla", {})
            
            cambios_realizados = []
            
            if datos_edicion.get("nombre_explotacion"):
                cuaderno.nombre_explotacion = datos_edicion["nombre_explotacion"]
                cambios_realizados.append(f"Nombre explotación: {datos_edicion['nombre_explotacion']}")
            
            if datos_edicion.get("titular"):
                cuaderno.titular = datos_edicion["titular"]
                cambios_realizados.append(f"Titular: {datos_edicion['titular']}")
            
            if datos_edicion.get("nif_titular") or datos_edicion.get("nif"):
                cuaderno.nif_titular = datos_edicion.get("nif_titular") or datos_edicion.get("nif", "")
                cambios_realizados.append(f"NIF: {cuaderno.nif_titular}")
            
            if datos_edicion.get("domicilio") or datos_edicion.get("direccion"):
                cuaderno.domicilio = datos_edicion.get("domicilio") or datos_edicion.get("direccion", "")
                cambios_realizados.append(f"Domicilio: {cuaderno.domicilio}")
            
            if datos_edicion.get("codigo_explotacion"):
                cuaderno.codigo_explotacion = datos_edicion["codigo_explotacion"]
                cambios_realizados.append(f"Código explotación: {cuaderno.codigo_explotacion}")
            
            if cambios_realizados:
                storage.guardar(cuaderno)
                mensaje_respuesta = f"✅ Datos del cuaderno actualizados:\n" + "\n".join([f"  • {c}" for c in cambios_realizados])
            else:
                mensaje_respuesta = "⚠️ No se detectaron cambios válidos para aplicar. Especifica qué quieres cambiar (titular, NIF, nombre explotación, domicilio)."
            
            if advertencias:
                mensaje_respuesta += "\n\n⚠️ Advertencias:\n" + "\n".join([f"  • {a}" for a in advertencias])
            
            if elementos_no_procesados:
                mensaje_respuesta += "\n\n⚠️ No procesado:\n" + "\n".join([f"  • {e}" for e in elementos_no_procesados])
            
            return ChatResponse(
                success=len(cambios_realizados) > 0,
                mensaje=mensaje_respuesta,
                accion_ejecutada="editar_cuaderno",
                datos_creados={"cambios": cambios_realizados} if cambios_realizados else None,
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        
        elif accion == "editar_parcelas":
            datos_edicion = resultado_gpt.get("datos", {}) or resultado_gpt.get("plantilla", {})
            filtro = resultado_gpt.get("filtro", {})

            parcelas_a_modificar = list(cuaderno.parcelas)

            # Filtro por rango de filas (1-based)
            fila_inicio = filtro.get("fila_inicio")
            fila_fin = filtro.get("fila_fin")
            if fila_inicio is not None and fila_fin is not None:
                try:
                    fi, ff = int(fila_inicio), int(fila_fin)
                    if fi > ff:
                        fi, ff = ff, fi
                    parcelas_a_modificar = [
                        p for i, p in enumerate(cuaderno.parcelas)
                        if fi <= (i + 1) <= ff
                    ]
                except (ValueError, TypeError):
                    pass

            if filtro.get("cultivo"):
                parcelas_a_modificar = [p for p in parcelas_a_modificar if filtro["cultivo"].upper() in (p.cultivo or "").upper() or filtro["cultivo"].upper() in (p.especie or "").upper()]
            if filtro.get("municipio"):
                parcelas_a_modificar = [p for p in parcelas_a_modificar if filtro["municipio"].upper() in (p.municipio or "").upper()]
            if filtro.get("nombre"):
                parcelas_a_modificar = [p for p in parcelas_a_modificar if filtro["nombre"].upper() in p.nombre.upper()]

            if not filtro:
                parcelas_a_modificar = list(cuaderno.parcelas)

            # Alias map: GPT key -> (model attr, [synced attrs])
            PARCELA_ALIASES = {
                "cultivo": ("cultivo", ["especie"]),
                "especie": ("especie", ["cultivo"]),
                "superficie": ("superficie_ha", ["superficie_cultivada"]),
                "superficie_ha": ("superficie_ha", ["superficie_cultivada"]),
                "superficie_cultivada": ("superficie_cultivada", ["superficie_ha"]),
                "superficie_sigpac": ("superficie_sigpac", []),
                "municipio": ("municipio", ["termino_municipal"]),
                "termino_municipal": ("termino_municipal", ["municipio"]),
                "provincia": ("provincia", ["codigo_provincia"]),
                "codigo_provincia": ("codigo_provincia", ["provincia"]),
            }

            parcelas_modificadas = []
            for parcela in parcelas_a_modificar:
                cambios_parcela = []
                for key, valor in datos_edicion.items():
                    if key in ("id", "fecha_creacion") or valor is None:
                        continue
                    attr, syncs = PARCELA_ALIASES.get(key, (key, []))
                    if not hasattr(parcela, attr):
                        continue
                    old = getattr(parcela, attr)
                    if isinstance(old, (int, float)) and not isinstance(old, bool):
                        try:
                            valor = float(valor) if isinstance(old, float) else int(valor)
                        except (ValueError, TypeError):
                            pass
                    elif isinstance(old, bool):
                        valor = str(valor).lower() in ("true", "1", "sí", "si", "yes")
                    else:
                        valor = str(valor)
                    setattr(parcela, attr, valor)
                    for s in syncs:
                        if hasattr(parcela, s):
                            setattr(parcela, s, valor)
                    cambios_parcela.append(key)
                if cambios_parcela:
                    parcelas_modificadas.append({"nombre": parcela.nombre, "cambios": cambios_parcela})
            
            if parcelas_modificadas:
                storage.guardar(cuaderno)
                mensaje_respuesta = f"✅ {len(parcelas_modificadas)} parcela(s) modificada(s)"
                if len(parcelas_modificadas) <= 10:
                    mensaje_respuesta += ":\n" + "\n".join([f"  • {p['nombre']} ({', '.join(p['cambios'])})" for p in parcelas_modificadas])
                else:
                    mensaje_respuesta += f":\n  • {parcelas_modificadas[0]['nombre']}\n  • ...\n  • {parcelas_modificadas[-1]['nombre']}\n  (Total: {len(parcelas_modificadas)} parcelas)"
            else:
                mensaje_respuesta = "⚠️ No se encontraron parcelas para modificar o no se especificaron cambios válidos."
            
            if advertencias:
                mensaje_respuesta += "\n\n⚠️ Advertencias:\n" + "\n".join([f"  • {a}" for a in advertencias])
            
            return ChatResponse(
                success=len(parcelas_modificadas) > 0,
                mensaje=mensaje_respuesta,
                accion_ejecutada="editar_parcelas",
                datos_creados={"parcelas_modificadas": len(parcelas_modificadas)} if parcelas_modificadas else None,
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        
        elif accion == "editar_productos":
            datos_edicion = resultado_gpt.get("datos", {}) or resultado_gpt.get("plantilla", {})
            filtro = resultado_gpt.get("filtro", {})

            productos_a_modificar = list(cuaderno.productos)

            fila_inicio = filtro.get("fila_inicio")
            fila_fin = filtro.get("fila_fin")
            if fila_inicio is not None and fila_fin is not None:
                try:
                    fi, ff = int(fila_inicio), int(fila_fin)
                    if fi > ff:
                        fi, ff = ff, fi
                    productos_a_modificar = [
                        p for i, p in enumerate(cuaderno.productos)
                        if fi <= (i + 1) <= ff
                    ]
                except (ValueError, TypeError):
                    pass

            if filtro.get("nombre"):
                productos_a_modificar = [p for p in productos_a_modificar if filtro["nombre"].upper() in p.nombre_comercial.upper()]
            if filtro.get("tipo"):
                productos_a_modificar = [p for p in productos_a_modificar if filtro["tipo"].upper() in (p.tipo.value if p.tipo else "").upper()]

            productos_modificados = []
            for producto in productos_a_modificar:
                cambios = []
                if datos_edicion.get("nombre") or datos_edicion.get("nombre_comercial"):
                    producto.nombre_comercial = datos_edicion.get("nombre") or datos_edicion["nombre_comercial"]
                    cambios.append("nombre")
                if datos_edicion.get("registro") or datos_edicion.get("numero_registro"):
                    producto.numero_registro = datos_edicion.get("registro") or datos_edicion["numero_registro"]
                    cambios.append("registro")
                if datos_edicion.get("materia_activa"):
                    producto.materia_activa = datos_edicion["materia_activa"]
                    cambios.append("materia_activa")
                if datos_edicion.get("formulacion"):
                    producto.formulacion = datos_edicion["formulacion"]
                    cambios.append("formulacion")
                if datos_edicion.get("lote") or datos_edicion.get("numero_lote"):
                    producto.numero_lote = datos_edicion.get("lote") or datos_edicion["numero_lote"]
                    cambios.append("lote")
                if datos_edicion.get("cantidad") is not None or datos_edicion.get("cantidad_adquirida") is not None:
                    raw = datos_edicion.get("cantidad", datos_edicion.get("cantidad_adquirida", 0))
                    try:
                        producto.cantidad_adquirida = float(raw) if raw else 0.0
                    except (ValueError, TypeError):
                        producto.cantidad_adquirida = 0.0
                    cambios.append("cantidad")
                if datos_edicion.get("unidad"):
                    producto.unidad = datos_edicion["unidad"]
                    cambios.append("unidad")
                if datos_edicion.get("fecha_adquisicion"):
                    producto.fecha_adquisicion = datos_edicion["fecha_adquisicion"]
                    cambios.append("fecha_adquisicion")
                if datos_edicion.get("proveedor"):
                    producto.proveedor = datos_edicion["proveedor"]
                    cambios.append("proveedor")
                if datos_edicion.get("fecha_caducidad"):
                    producto.fecha_caducidad = datos_edicion["fecha_caducidad"]
                    cambios.append("fecha_caducidad")
                if cambios:
                    productos_modificados.append({"nombre": producto.nombre_comercial, "cambios": cambios})

            if productos_modificados:
                storage.guardar(cuaderno)
                mensaje_respuesta = f"✅ {len(productos_modificados)} producto(s) modificado(s)"
                if len(productos_modificados) <= 10:
                    mensaje_respuesta += ":\n" + "\n".join([f"  • {p['nombre']} ({', '.join(p['cambios'])})" for p in productos_modificados])
                else:
                    mensaje_respuesta += f" (Total: {len(productos_modificados)})"
            else:
                mensaje_respuesta = "⚠️ No se encontraron productos para modificar o no se especificaron cambios válidos."

            if advertencias:
                mensaje_respuesta += "\n\n⚠️ Advertencias:\n" + "\n".join([f"  • {a}" for a in advertencias])

            return ChatResponse(
                success=len(productos_modificados) > 0,
                mensaje=mensaje_respuesta,
                accion_ejecutada="editar_productos",
                datos_creados={"productos_modificados": len(productos_modificados)} if productos_modificados else None,
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )

        elif accion == "editar_tratamientos":
            datos_edicion = resultado_gpt.get("datos", {}) or resultado_gpt.get("plantilla", {})
            filtro = resultado_gpt.get("filtro", {})

            tratamientos_a_modificar = list(cuaderno.tratamientos)

            # Filtro por rango de filas (1-based, como se ven en la UI)
            fila_inicio = filtro.get("fila_inicio")
            fila_fin = filtro.get("fila_fin")
            if fila_inicio is not None and fila_fin is not None:
                try:
                    fi, ff = int(fila_inicio), int(fila_fin)
                    if fi > ff:
                        fi, ff = ff, fi
                    tratamientos_a_modificar = [
                        t for i, t in enumerate(cuaderno.tratamientos)
                        if fi <= (i + 1) <= ff
                    ]
                except (ValueError, TypeError):
                    pass

            if filtro.get("aplicador"):
                tratamientos_a_modificar = [t for t in tratamientos_a_modificar if filtro["aplicador"].upper() in (t.aplicador or "").upper()]
            if filtro.get("producto"):
                tratamientos_a_modificar = [
                    t for t in tratamientos_a_modificar
                    if any(filtro["producto"].upper() in (p.nombre_comercial or "").upper() for p in t.productos)
                ]
            if filtro.get("parcela"):
                tratamientos_a_modificar = [
                    t for t in tratamientos_a_modificar
                    if any(filtro["parcela"].upper() in (n or "").upper() for n in t.parcela_nombres)
                ]

            tratamientos_modificados = []
            for trat in tratamientos_a_modificar:
                cambios = []
                if datos_edicion.get("aplicador"):
                    trat.aplicador = datos_edicion["aplicador"]
                    cambios.append("aplicador")
                if datos_edicion.get("equipo"):
                    trat.equipo = datos_edicion["equipo"]
                    cambios.append("equipo")
                if datos_edicion.get("fecha_aplicacion") or datos_edicion.get("fecha"):
                    trat.fecha_aplicacion = datos_edicion.get("fecha_aplicacion") or datos_edicion["fecha"]
                    cambios.append("fecha")
                if datos_edicion.get("problema_fitosanitario") or datos_edicion.get("problema"):
                    trat.problema_fitosanitario = datos_edicion.get("problema_fitosanitario") or datos_edicion["problema"]
                    cambios.append("problema")
                if datos_edicion.get("cultivo_especie") or datos_edicion.get("cultivo"):
                    trat.cultivo_especie = datos_edicion.get("cultivo_especie") or datos_edicion["cultivo"]
                    cambios.append("cultivo")
                if datos_edicion.get("superficie_tratada") is not None:
                    try:
                        trat.superficie_tratada = float(datos_edicion["superficie_tratada"])
                    except (ValueError, TypeError):
                        pass
                    cambios.append("superficie")
                if datos_edicion.get("eficacia"):
                    trat.eficacia = datos_edicion["eficacia"]
                    cambios.append("eficacia")
                if datos_edicion.get("observaciones"):
                    trat.observaciones = datos_edicion["observaciones"]
                    cambios.append("observaciones")

                # Editar producto aplicado (primer producto de cada tratamiento)
                prod_nombre = datos_edicion.get("producto") or datos_edicion.get("nombre_comercial")
                prod_registro = datos_edicion.get("numero_registro")
                prod_lote = datos_edicion.get("numero_lote") or datos_edicion.get("lote")
                prod_dosis = datos_edicion.get("dosis")
                if prod_nombre or prod_registro or prod_lote or prod_dosis is not None:
                    if not trat.productos:
                        from cuaderno.models import ProductoAplicado as PA
                        trat.productos.append(PA())
                    p = trat.productos[0]
                    if prod_nombre:
                        p.nombre_comercial = prod_nombre
                        cambios.append("producto")
                    if prod_registro:
                        p.numero_registro = prod_registro
                        cambios.append("nº registro")
                    if prod_lote:
                        p.numero_lote = prod_lote
                        cambios.append("lote")
                    if prod_dosis is not None:
                        try:
                            p.dosis = float(prod_dosis)
                        except (ValueError, TypeError):
                            pass
                        cambios.append("dosis")

                if cambios:
                    tratamientos_modificados.append({"id": trat.id, "cambios": cambios})

            if tratamientos_modificados:
                storage.guardar(cuaderno)
                n = len(tratamientos_modificados)
                resumen_cambios = set()
                for t in tratamientos_modificados:
                    resumen_cambios.update(t["cambios"])
                mensaje_respuesta = f"✅ {n} tratamiento(s) modificado(s) ({', '.join(sorted(resumen_cambios))})"
            else:
                mensaje_respuesta = "⚠️ No se encontraron tratamientos para modificar o no se especificaron cambios válidos."

            if advertencias:
                mensaje_respuesta += "\n\n⚠️ Advertencias:\n" + "\n".join([f"  • {a}" for a in advertencias])

            return ChatResponse(
                success=len(tratamientos_modificados) > 0,
                mensaje=mensaje_respuesta,
                accion_ejecutada="editar_tratamientos",
                datos_creados={"tratamientos_modificados": len(tratamientos_modificados)} if tratamientos_modificados else None,
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )

        elif accion == "eliminar":
            # Eliminar elementos (parcelas, productos, tratamientos)
            tipo_elemento = resultado_gpt.get("tipo_elemento", "").lower()  # "parcelas", "productos", "tratamientos"
            filtro = resultado_gpt.get("filtro", {})
            
            elementos_eliminados = []
            
            if tipo_elemento == "parcelas" or tipo_elemento == "parcela":
                parcelas_a_eliminar = cuaderno.parcelas.copy()
                if filtro.get("nombre"):
                    parcelas_a_eliminar = [p for p in parcelas_a_eliminar if filtro["nombre"].upper() in p.nombre.upper()]
                elif filtro.get("cultivo"):
                    parcelas_a_eliminar = [p for p in parcelas_a_eliminar if filtro["cultivo"].upper() in (p.cultivo or "").upper()]
                
                for parcela in parcelas_a_eliminar:
                    if cuaderno.eliminar_parcela(parcela.id):
                        elementos_eliminados.append({"tipo": "parcela", "nombre": parcela.nombre})
            
            elif tipo_elemento == "productos" or tipo_elemento == "producto":
                productos_a_eliminar = cuaderno.productos.copy()
                if filtro.get("nombre"):
                    productos_a_eliminar = [p for p in productos_a_eliminar if filtro["nombre"].upper() in p.nombre_comercial.upper()]
                
                for producto in productos_a_eliminar:
                    if cuaderno.eliminar_producto(producto.id):
                        elementos_eliminados.append({"tipo": "producto", "nombre": producto.nombre_comercial})
            
            elif tipo_elemento == "tratamientos" or tipo_elemento == "tratamiento":
                tratamientos_a_eliminar = cuaderno.tratamientos.copy()
                for tratamiento in tratamientos_a_eliminar:
                    if cuaderno.eliminar_tratamiento(tratamiento.id):
                        elementos_eliminados.append({"tipo": "tratamiento", "id": tratamiento.id})
            
            if elementos_eliminados:
                storage.guardar(cuaderno)
                mensaje_respuesta = f"✅ {len(elementos_eliminados)} elemento(s) eliminado(s):\n" + "\n".join([f"  • {e['nombre'] if 'nombre' in e else e['id']}" for e in elementos_eliminados[:10]])
            else:
                mensaje_respuesta = "⚠️ No se encontraron elementos para eliminar. Especifica qué quieres eliminar (parcelas, productos, tratamientos)."
            
            return ChatResponse(
                success=len(elementos_eliminados) > 0,
                mensaje=mensaje_respuesta,
                accion_ejecutada="eliminar",
                datos_creados={"elementos_eliminados": len(elementos_eliminados)} if elementos_eliminados else None,
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        
        elif accion == "editar_hoja":
            # ───────────────────────────────────────────────
            # Editar celdas, filas y columnas en hojas importadas
            # ───────────────────────────────────────────────
            target_sheet_id = resultado_gpt.get("sheet_id", "")
            operaciones = resultado_gpt.get("operaciones", [])
            
            # Resolver sheet_id por nombre si no es UUID directo
            hoja_target = None
            if target_sheet_id:
                hoja_target = cuaderno.obtener_hoja(target_sheet_id)
            
            # Si no encontró por ID, buscar por nombre (fuzzy)
            if not hoja_target and target_sheet_id:
                target_upper = target_sheet_id.upper().strip()
                for h in cuaderno.hojas_originales:
                    if target_upper in h.nombre.upper() or h.nombre.upper() in target_upper:
                        hoja_target = h
                        break
            
            # Si viene del contexto de hoja activa
            if not hoja_target and contexto and contexto.get("active_sheet_id"):
                hoja_target = cuaderno.obtener_hoja(contexto["active_sheet_id"])
            
            # Último recurso: buscar por palabras del mensaje en nombres de hojas
            if not hoja_target:
                msg_lower = comando.mensaje.lower()
                for h in cuaderno.hojas_originales:
                    if h.nombre and h.nombre.lower() in msg_lower:
                        hoja_target = h
                        break
                # Si solo hay una hoja importada, usarla
                if not hoja_target and len(cuaderno.hojas_originales) == 1:
                    hoja_target = cuaderno.hojas_originales[0]
            
            if not hoja_target:
                hojas_disponibles = [f"  • \"{h.nombre}\"" for h in cuaderno.hojas_originales]
                msg = "💡 No encontré esa hoja."
                if hojas_disponibles:
                    msg += "\n\nHojas disponibles:\n" + "\n".join(hojas_disponibles)
                    msg += "\n\nPrueba a escribir el nombre exacto."
                else:
                    msg += " No hay hojas importadas todavía. Sube un archivo Excel primero."
                return ChatResponse(
                    success=True,
                    mensaje=msg,
                    sugerencias=["Sube un archivo Excel", "Lista las hojas disponibles"]
                )
            
            if not operaciones:
                return ChatResponse(
                    success=True,
                    mensaje=f"💡 Tengo la hoja \"{hoja_target.nombre}\" lista, pero no entendí qué cambio hacer. Prueba así:",
                    sugerencias=[f"En la hoja {hoja_target.nombre}, cambia la celda A1 a nuevo valor", f"Busca y reemplaza 'viejo' por 'nuevo' en {hoja_target.nombre}", f"Añade una fila con datos nuevos en {hoja_target.nombre}"]
                )
            
            cambios_realizados = []
            filas_afectadas = []
            selected_ctx_rows = []
            if isinstance(contexto, dict):
                sel = contexto.get("selected_cells")
                if isinstance(sel, dict):
                    selected_ctx_rows = sel.get("rows", []) or []
            msg_lower = (comando.mensaje or "").lower()
            
            for op in operaciones:
                tipo_op = op.get("tipo", "")
                
                if tipo_op == "set_cell":
                    fila = int(op.get("fila", 0))
                    col = int(op.get("columna", 0))
                    valor = op.get("valor", "")
                    
                    # Resolver columna por nombre si no es número
                    if isinstance(op.get("columna"), str) and not op["columna"].isdigit():
                        col_name = op["columna"].upper().strip()
                        for idx, col_h in enumerate(hoja_target.columnas):
                            if col_name in col_h.upper() or col_h.upper() in col_name:
                                col = idx
                                break
                    
                    if cuaderno.aplicar_celda(hoja_target.sheet_id, fila, col, valor):
                        col_nombre = hoja_target.columnas[col] if col < len(hoja_target.columnas) else f"Col {col}"
                        cambios_realizados.append(f"Fila {fila}, {col_nombre} → \"{valor}\"")
                
                elif tipo_op == "buscar_reemplazar":
                    buscar = str(op.get("buscar", ""))
                    reemplazar = str(op.get("reemplazar", ""))
                    if buscar:
                        count = 0
                        buscar_upper = buscar.upper()
                        for i, fila in enumerate(hoja_target.datos):
                            for j, celda in enumerate(fila):
                                if celda is not None and buscar_upper in str(celda).upper():
                                    nuevo_valor = str(celda).replace(buscar, reemplazar)
                                    # Intentar también case-insensitive
                                    if buscar not in str(celda):
                                        nuevo_valor = re.sub(re.escape(buscar), reemplazar, str(celda), flags=re.IGNORECASE)
                                    hoja_target.datos[i][j] = nuevo_valor
                                    count += 1
                        hoja_target.origen = "importado_editable"
                        cuaderno._actualizar_modificacion()
                        cambios_realizados.append(f"Buscar/Reemplazar: \"{buscar}\" → \"{reemplazar}\" ({count} celdas)")
                
                elif tipo_op == "agregar_fila":
                    datos_fila = op.get("datos", [])
                    # Si el usuario tenía filas seleccionadas y pidió editar "la fila",
                    # aplicar sobre esa(s) fila(s) en lugar de crear una nueva al final.
                    editar_fila_seleccionada = (
                        len(selected_ctx_rows) > 0
                        and ("fila" in msg_lower or "selecci" in msg_lower)
                        and ("fila completa" in msg_lower or "toda la fila" in msg_lower or "en la fila" in msg_lower or "fila seleccionada" in msg_lower)
                    )

                    if editar_fila_seleccionada:
                        for r in selected_ctx_rows:
                            try:
                                row_idx = int(r.get("row_index", -1))
                            except Exception:
                                row_idx = -1
                            if row_idx < 0:
                                continue
                            while len(hoja_target.datos) <= row_idx:
                                hoja_target.datos.append([])

                            fila_actual = hoja_target.datos[row_idx]
                            selected_cells = r.get("cells", []) if isinstance(r, dict) else []
                            target_cols = []
                            for c in selected_cells:
                                try:
                                    target_cols.append(int(c.get("col_key")))
                                except Exception:
                                    continue
                            target_cols = sorted(set(target_cols))
                            if not target_cols:
                                target_cols = list(range(len(datos_fila)))

                            for i_col, col_idx in enumerate(target_cols):
                                while len(fila_actual) <= col_idx:
                                    fila_actual.append("")
                                if i_col < len(datos_fila):
                                    fila_actual[col_idx] = datos_fila[i_col]

                            hoja_target.datos[row_idx] = fila_actual
                            filas_afectadas.append(row_idx)
                            cambios_realizados.append(f"Fila {row_idx + 1} actualizada (selección): {datos_fila[:5]}...")
                    else:
                        hoja_target.datos.append(datos_fila)
                        nueva_idx = len(hoja_target.datos) - 1
                        filas_afectadas.append(nueva_idx)
                        cambios_realizados.append(f"Fila añadida en posición {nueva_idx + 1}: {datos_fila[:5]}...")

                    hoja_target.origen = "importado_editable"
                    cuaderno._actualizar_modificacion()
                
                elif tipo_op == "eliminar_fila":
                    fila_idx = int(op.get("fila", -1))
                    if 0 <= fila_idx < len(hoja_target.datos):
                        eliminada = hoja_target.datos.pop(fila_idx)
                        hoja_target.origen = "importado_editable"
                        cuaderno._actualizar_modificacion()
                        cambios_realizados.append(f"Fila {fila_idx} eliminada")
                
                elif tipo_op == "renombrar_columna":
                    col_idx = int(op.get("columna_idx", 0))
                    nuevo_nombre = op.get("nuevo_nombre", "")
                    if 0 <= col_idx < len(hoja_target.columnas) and nuevo_nombre:
                        viejo = hoja_target.columnas[col_idx]
                        hoja_target.columnas[col_idx] = nuevo_nombre
                        hoja_target.origen = "importado_editable"
                        cuaderno._actualizar_modificacion()
                        cambios_realizados.append(f"Columna \"{viejo}\" → \"{nuevo_nombre}\"")
                
                elif tipo_op == "reestructurar":
                    # Reemplaza completamente las columnas y los datos por unos nuevos ordenados
                    nuevas_columnas = op.get("columnas", [])
                    nuevos_datos = op.get("datos", [])
                    if nuevas_columnas and nuevos_datos:
                        hoja_target.columnas = nuevas_columnas
                        hoja_target.datos = nuevos_datos
                        hoja_target.origen = "importado_editable"
                        cuaderno._actualizar_modificacion()
                        cambios_realizados.append("La hoja fue reestructurada completamente")

                elif tipo_op == "set_columna":
                    # Poner un valor en toda una columna (todas las filas)
                    col = int(op.get("columna", 0))
                    valor = op.get("valor", "")
                    count = 0
                    for i, fila in enumerate(hoja_target.datos):
                        while len(fila) <= col:
                            fila.append("")
                        fila[col] = valor
                        count += 1
                    hoja_target.origen = "importado_editable"
                    cuaderno._actualizar_modificacion()
                    col_nombre = hoja_target.columnas[col] if col < len(hoja_target.columnas) else f"Col {col}"
                    cambios_realizados.append(f"Columna \"{col_nombre}\" → \"{valor}\" en {count} filas")
            
            if cambios_realizados:
                storage.guardar(cuaderno)
                mensaje_respuesta = f"✅ Hoja \"{hoja_target.nombre}\" editada:\n" + "\n".join([f"  • {c}" for c in cambios_realizados])
            else:
                mensaje_respuesta = f"⚠️ No se aplicaron cambios en la hoja \"{hoja_target.nombre}\"."
            
            if advertencias:
                mensaje_respuesta += "\n\n⚠️ Advertencias:\n" + "\n".join([f"  • {a}" for a in advertencias])
            if elementos_no_procesados:
                mensaje_respuesta += "\n\n⚠️ No procesado:\n" + "\n".join([f"  • {e}" for e in elementos_no_procesados])
            
            return ChatResponse(
                success=len(cambios_realizados) > 0,
                mensaje=mensaje_respuesta,
                accion_ejecutada="editar_hoja",
                datos_creados={
                    "hoja": hoja_target.nombre,
                    "sheet_id": hoja_target.sheet_id,
                    "cambios": len(cambios_realizados),
                    "filas_afectadas": filas_afectadas[:20],
                    "elementos": [
                        {
                            "tipo": "hoja",
                            "id": hoja_target.sheet_id,
                            "nombre": hoja_target.nombre
                        }
                    ]
                } if cambios_realizados else None,
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        
        elif accion == "buscar_en_hojas":
            # ───────────────────────────────────────────────
            # Buscar datos en todas las hojas del cuaderno
            # ───────────────────────────────────────────────
            termino_busqueda = resultado_gpt.get("buscar", "").strip()
            sheet_id_filtro = resultado_gpt.get("sheet_id")
            
            if not termino_busqueda:
                return ChatResponse(
                    success=True,
                    mensaje="💡 Dime qué quieres buscar. Por ejemplo:",
                    sugerencias=["Busca 'Juan Pérez'", "Busca el teléfono del titular", "Busca dónde dice 'olivo'"]
                )
            
            resultados = []
            buscar_upper = termino_busqueda.upper()
            
            # Buscar en hojas importadas
            hojas_a_buscar = cuaderno.hojas_originales
            if sheet_id_filtro:
                hoja_filtro = cuaderno.obtener_hoja(sheet_id_filtro)
                if hoja_filtro:
                    hojas_a_buscar = [hoja_filtro]
            
            for hoja in hojas_a_buscar:
                # Buscar en columnas
                for idx_col, col_nombre in enumerate(hoja.columnas):
                    if buscar_upper in col_nombre.upper():
                        resultados.append({
                            "hoja": hoja.nombre,
                            "tipo": "columna",
                            "detalle": f"Columna {idx_col}: \"{col_nombre}\""
                        })
                
                # Buscar en datos
                for idx_fila, fila in enumerate(hoja.datos):
                    for idx_col, celda in enumerate(fila):
                        if celda is not None and buscar_upper in str(celda).upper():
                            col_nombre = hoja.columnas[idx_col] if idx_col < len(hoja.columnas) else f"Col {idx_col}"
                            resultados.append({
                                "hoja": hoja.nombre,
                                "sheet_id": hoja.sheet_id,
                                "tipo": "celda",
                                "fila": idx_fila,
                                "columna": idx_col,
                                "columna_nombre": col_nombre,
                                "valor": str(celda)[:100],
                                "detalle": f"Fila {idx_fila}, {col_nombre}: \"{str(celda)[:80]}\""
                            })
            
            # Buscar también en datos del cuaderno
            if buscar_upper in (cuaderno.titular or "").upper():
                resultados.append({"hoja": "Cuaderno", "tipo": "campo", "detalle": f"Titular: \"{cuaderno.titular}\""})
            if buscar_upper in (cuaderno.nombre_explotacion or "").upper():
                resultados.append({"hoja": "Cuaderno", "tipo": "campo", "detalle": f"Explotación: \"{cuaderno.nombre_explotacion}\""})
            for p in cuaderno.parcelas:
                if buscar_upper in p.nombre.upper():
                    resultados.append({"hoja": "Parcelas", "tipo": "parcela", "detalle": f"Parcela: \"{p.nombre}\" (ID: {p.id})"})
            for pr in cuaderno.productos:
                for campo, label in [("nombre_comercial", "Producto"), ("materia_activa", "Materia activa"), ("numero_registro", "Nº Registro")]:
                    val = getattr(pr, campo, "") or ""
                    if buscar_upper in val.upper():
                        resultados.append({"hoja": "Productos", "tipo": "producto", "detalle": f"{label}: \"{val}\" (ID: {pr.id})"})
            for t in cuaderno.tratamientos:
                for campo, label in [
                    ("problema_fitosanitario", "Problema fitosanitario"),
                    ("plaga_enfermedad", "Plaga/Enfermedad"),
                    ("aplicador", "Aplicador"),
                    ("operador", "Operador"),
                    ("cultivo_especie", "Cultivo"),
                ]:
                    val = getattr(t, campo, "") or ""
                    if buscar_upper in val.upper():
                        resultados.append({"hoja": "Tratamientos", "tipo": "tratamiento", "detalle": f"{label}: \"{val}\" (Fecha: {t.fecha_aplicacion}, ID: {t.id})"})
                # También buscar en productos usados en el tratamiento
                for pa in t.productos:
                    if buscar_upper in (pa.nombre_comercial or "").upper():
                        resultados.append({"hoja": "Tratamientos", "tipo": "tratamiento", "detalle": f"Producto usado: \"{pa.nombre_comercial}\" en tratamiento del {t.fecha_aplicacion} (ID: {t.id})"})
            
            if resultados:
                mensaje_respuesta = f"🔍 Encontré \"{termino_busqueda}\" en {len(resultados)} lugar(es):\n"
                for r in resultados[:20]:
                    mensaje_respuesta += f"  • [{r['hoja']}] {r['detalle']}\n"
                if len(resultados) > 20:
                    mensaje_respuesta += f"  ... y {len(resultados) - 20} resultados más"
            else:
                mensaje_respuesta = f"🔍 No se encontró \"{termino_busqueda}\" en ninguna hoja ni dato del cuaderno."
            
            return ChatResponse(
                success=len(resultados) > 0,
                mensaje=mensaje_respuesta,
                accion_ejecutada="buscar_en_hojas",
                datos_creados={"resultados": len(resultados), "detalles": resultados[:20]} if resultados else None,
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
        elif accion == "editar_celdas":
            # Editar celdas específicas seleccionadas por el usuario
            # Ser tolerantes con variantes de claves que puede devolver el modelo
            selected_ctx = (contexto or {}).get("selected_cells") if isinstance(contexto, dict) else None
            target_sheet_id = (
                resultado_gpt.get("sheet_id")
                or resultado_gpt.get("sheetId")
                or (selected_ctx.get("sheet_id") if isinstance(selected_ctx, dict) else "")
                or ""
            )
            ediciones = (
                resultado_gpt.get("ediciones")
                or resultado_gpt.get("changes")
                or resultado_gpt.get("cambios")
                or []
            )
            if not isinstance(ediciones, list):
                ediciones = []

            # Mapeos desde contexto de selección para tolerar nombres/etiquetas de columna
            # Ejemplo: "Municipio" -> "municipio" (hojas base) o índice "2" (hojas importadas)
            selected_rows = selected_ctx.get("rows", []) if isinstance(selected_ctx, dict) else []
            label_to_key = {}
            for r in selected_rows:
                for c in (r.get("cells", []) if isinstance(r, dict) else []):
                    label = str(c.get("col_label", "")).strip().lower()
                    key = str(c.get("col_key", "")).strip()
                    if label and key and label not in label_to_key:
                        label_to_key[label] = key

            # Modo explícito "toda la fila": si el usuario lo pide, aplicar el mismo valor
            # a TODAS las celdas de las filas seleccionadas (sin inferir campos sueltos).
            mensaje_lower = (comando.mensaje or "").lower()
            range_match = re.search(r"(?:fila(?:s)?\s*)?(\d{1,5})\s*(?:a|-|hasta)\s*(\d{1,5})", mensaje_lower)
            rango_inicio = rango_fin = None
            if range_match:
                try:
                    a = int(range_match.group(1))
                    b = int(range_match.group(2))
                    rango_inicio, rango_fin = (a, b) if a <= b else (b, a)
                except Exception:
                    rango_inicio = rango_fin = None
            apply_whole_row = any(
                token in mensaje_lower
                for token in ["toda la fila", "fila completa", "fila entera", "en toda la fila"]
            )
            # Si menciona un rango de filas (ej: "de la 29 a la 37"), tratarlo como edición por fila.
            if not apply_whole_row and rango_inicio is not None and "fila" in mensaje_lower:
                apply_whole_row = True
            if apply_whole_row and isinstance(selected_ctx, dict):
                selected_rows_ctx = selected_ctx.get("rows", []) or []
                # Si hay rango explícito, filtrar selección a ese rango (números humanos 1-based).
                if rango_inicio is not None and rango_fin is not None:
                    selected_rows_ctx = [
                        r for r in selected_rows_ctx
                        if (
                            isinstance(r, dict)
                            and (
                                (isinstance(r.get("row_index"), int) and rango_inicio <= (r.get("row_index") + 1) <= rango_fin)
                                or (isinstance(r.get("row_index"), int) and rango_inicio <= r.get("row_index") <= rango_fin)
                            )
                        )
                    ]
                # Inferir valor objetivo (priorizando fechas comunes)
                valor_fila = ""
                if "hoy" in mensaje_lower:
                    valor_fila = datetime.now().date().isoformat()
                else:
                    m_date = re.search(r"(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})", comando.mensaje or "")
                    if m_date:
                        valor_fila = m_date.group(1)
                    else:
                        m_text = re.search(
                            r"(?:pon|poner|mete|meter|agrega|agregar|cambia|cambiar)\s+(?:una\s+)?(?:fecha\s+)?(?:de\s+)?(?:a\s+)?['\"]?([^'\"]+?)['\"]?$",
                            (comando.mensaje or "").strip(),
                            flags=re.IGNORECASE,
                        )
                        if m_text:
                            valor_fila = m_text.group(1).strip()
                if not valor_fila:
                    valor_fila = "2026-01-01"

                # Si no hay selección adjunta pero sí rango + hoja importada, construir objetivo por rango.
                if (not selected_rows_ctx) and (rango_inicio is not None and rango_fin is not None):
                    hoja_rango = cuaderno.obtener_hoja(target_sheet_id)
                    if hoja_rango:
                        selected_rows_ctx = []
                        for human_row in range(rango_inicio, rango_fin + 1):
                            row_idx = max(0, human_row - 1)
                            if row_idx >= len(hoja_rango.datos):
                                continue
                            fila = hoja_rango.datos[row_idx] or []
                            num_cols = max(len(fila), len(hoja_rango.columnas or []))
                            cells = []
                            for ci in range(num_cols):
                                label = hoja_rango.columnas[ci] if ci < len(hoja_rango.columnas) else f"Col {ci + 1}"
                                val = fila[ci] if ci < len(fila) else ""
                                cells.append({
                                    "col_key": str(ci),
                                    "col_label": label,
                                    "value": val,
                                })
                            selected_rows_ctx.append({
                                "row_id": str(row_idx),
                                "row_index": row_idx,
                                "cells": cells,
                            })

                ediciones_fila = []
                for r in selected_rows_ctx:
                    row_id_ctx = r.get("row_id", "")
                    row_index_ctx = r.get("row_index")
                    row_id_norm = str(row_id_ctx).strip() if row_id_ctx not in (None, "") else ""
                    if not row_id_norm and row_index_ctx is not None:
                        row_id_norm = str(row_index_ctx)
                    for c in (r.get("cells", []) if isinstance(r, dict) else []):
                        col_key_ctx = str(c.get("col_key", "")).strip()
                        if not row_id_norm or not col_key_ctx:
                            continue
                        ediciones_fila.append({
                            "row_id": row_id_norm,
                            "col_key": col_key_ctx,
                            "valor": valor_fila
                        })

                if ediciones_fila:
                    ediciones = ediciones_fila
                    advertencias = (advertencias or []) + [
                        "Aplicado modo 'toda la fila': se asignó el mismo valor en todas las celdas seleccionadas."
                    ]
            
            celdas_editadas = 0
            detalles_edicion = []
            
            for edicion in ediciones:
                # Normalizar forma de edición:
                # - row_id / rowId / row_index / rowIndex
                # - col_key / colKey / columna / column
                # - valor / value / nuevo_valor
                row_id = (
                    edicion.get("row_id")
                    or edicion.get("rowId")
                    or ""
                )
                if row_id in (None, ""):
                    row_idx_alt = edicion.get("row_index", edicion.get("rowIndex"))
                    if row_idx_alt is not None:
                        row_id = str(row_idx_alt)

                col_key = (
                    edicion.get("col_key")
                    or edicion.get("colKey")
                    or edicion.get("columna")
                    or edicion.get("col_label")
                    or edicion.get("column_label")
                    or edicion.get("column")
                    or ""
                )
                if col_key is None:
                    col_key = ""
                col_key = str(col_key)
                # Resolver etiqueta a key real si viene como nombre visual de columna
                col_key_norm = col_key.strip().lower()
                if col_key_norm in label_to_key:
                    col_key = str(label_to_key[col_key_norm])

                if "valor" in edicion:
                    nuevo_valor = edicion.get("valor")
                elif "value" in edicion:
                    nuevo_valor = edicion.get("value")
                elif "nuevo_valor" in edicion:
                    nuevo_valor = edicion.get("nuevo_valor")
                else:
                    nuevo_valor = ""
                
                if not row_id or not col_key:
                    continue
                
                try:
                    # Hojas base del sistema
                    if target_sheet_id in ("parcelas", "productos", "tratamientos", "fertilizantes", "cosecha"):
                        if target_sheet_id == "parcelas":
                            items = cuaderno.parcelas
                        elif target_sheet_id == "productos":
                            items = cuaderno.productos
                        elif target_sheet_id == "tratamientos":
                            items = cuaderno.tratamientos
                        elif target_sheet_id == "fertilizantes":
                            items = getattr(cuaderno, 'fertilizaciones', None) or []
                        elif target_sheet_id == "cosecha":
                            items = getattr(cuaderno, 'cosechas', None) or []
                        else:
                            items = []
                        
                        # Buscar por UUID primero, luego por índice de fila (0-based)
                        target_item = next((item for item in items if item.id == row_id), None)
                        if not target_item:
                            try:
                                idx = int(row_id)
                                if 0 <= idx < len(items):
                                    target_item = items[idx]
                            except (ValueError, TypeError):
                                pass
                        # Redirigir columnas virtuales de tratamientos
                        actual_col_key = col_key
                        is_producto_field = False
                        if target_sheet_id == "tratamientos":
                            if col_key in ("productos_display", "producto", "nombre_comercial"):
                                actual_col_key = "_producto_nombre"
                                is_producto_field = True
                            elif col_key in ("dosis_display", "dosis"):
                                actual_col_key = "_producto_dosis"
                                is_producto_field = True

                        if is_producto_field and target_item:
                            from cuaderno.models import ProductoAplicado as PA
                            if not target_item.productos:
                                target_item.productos.append(PA())
                            p0 = target_item.productos[0]
                            if actual_col_key == "_producto_nombre":
                                old_value = p0.nombre_comercial
                                p0.nombre_comercial = str(nuevo_valor)
                            else:
                                old_value = p0.dosis
                                try:
                                    p0.dosis = float(nuevo_valor)
                                except (ValueError, TypeError):
                                    p0.dosis = 0.0
                            celdas_editadas += 1
                            detalles_edicion.append(f"{col_key}: '{old_value}' → '{nuevo_valor}'")
                        elif target_item and hasattr(target_item, actual_col_key):
                            old_value = getattr(target_item, actual_col_key, "")
                            field_type = type(old_value) if old_value is not None else str
                            try:
                                if field_type == float or field_type == int:
                                    typed_value = float(nuevo_valor) if "." in str(nuevo_valor) else int(nuevo_valor)
                                elif field_type == bool:
                                    typed_value = str(nuevo_valor).lower() in ("true", "1", "sí", "si", "yes")
                                else:
                                    typed_value = str(nuevo_valor)
                            except (ValueError, TypeError):
                                typed_value = str(nuevo_valor)
                            
                            setattr(target_item, actual_col_key, typed_value)
                            
                            # Sincronizar campos espejo
                            if target_sheet_id == "parcelas":
                                if actual_col_key == "cultivo" and hasattr(target_item, "especie"):
                                    target_item.especie = typed_value
                                elif actual_col_key == "especie" and hasattr(target_item, "cultivo"):
                                    target_item.cultivo = typed_value
                                elif actual_col_key == "municipio" and hasattr(target_item, "termino_municipal"):
                                    target_item.termino_municipal = typed_value
                                elif actual_col_key in ("superficie_ha", "superficie_cultivada"):
                                    target_item.superficie_ha = typed_value
                                    target_item.superficie_cultivada = typed_value
                            
                            celdas_editadas += 1
                            detalles_edicion.append(f"{actual_col_key}: '{old_value}' → '{typed_value}'")
                    else:
                        # Hoja importada
                        hoja = cuaderno.obtener_hoja(target_sheet_id)
                        if hoja:
                            try:
                                row_idx = int(row_id)
                            except ValueError:
                                row_idx = -1
                            
                            if 0 <= row_idx < len(hoja.datos):
                                # En hojas importadas aceptar:
                                # - índice numérico de columna
                                # - nombre de columna (ej: "MUNICIPIO")
                                # - etiqueta "Col 3"
                                try:
                                    col_idx = int(col_key)
                                except ValueError:
                                    col_idx = -1
                                    col_key_norm = str(col_key).strip().lower()
                                    if col_key_norm.startswith("col "):
                                        try:
                                            # "Col 3" -> índice 2
                                            col_idx = int(col_key_norm.split(" ", 1)[1]) - 1
                                        except Exception:
                                            col_idx = -1
                                    if col_idx < 0 and hoja.columnas:
                                        for i_col, col_name in enumerate(hoja.columnas):
                                            if str(col_name).strip().lower() == col_key_norm:
                                                col_idx = i_col
                                                break
                                
                                if 0 <= col_idx < len(hoja.datos[row_idx]):
                                    old_val = hoja.datos[row_idx][col_idx]
                                    hoja.datos[row_idx][col_idx] = nuevo_valor
                                    celdas_editadas += 1
                                    col_name = hoja.columnas[col_idx] if col_idx < len(hoja.columnas) else str(col_idx)
                                    detalles_edicion.append(f"Fila {row_idx} [{col_name}]: '{old_val}' → '{nuevo_valor}'")
                                elif col_idx >= 0:
                                    # Si la fila es más corta (celdas vacías al final), extender y aplicar.
                                    fila = hoja.datos[row_idx]
                                    faltan = col_idx - len(fila) + 1
                                    if faltan > 0:
                                        fila.extend([""] * faltan)
                                    old_val = fila[col_idx]
                                    fila[col_idx] = nuevo_valor
                                    celdas_editadas += 1
                                    col_name = hoja.columnas[col_idx] if col_idx < len(hoja.columnas) else str(col_idx)
                                    detalles_edicion.append(f"Fila {row_idx} [{col_name}]: '{old_val}' → '{nuevo_valor}'")
                
                except Exception as e:
                    print(f"Error editando celda {row_id}:{col_key}: {e}")
                    continue
            
            if celdas_editadas > 0:
                storage.guardar(cuaderno)
                mensaje_respuesta = f"✅ {celdas_editadas} celda(s) editada(s)"
                if len(detalles_edicion) <= 20:
                    mensaje_respuesta += ":\n" + "\n".join([f"  • {d}" for d in detalles_edicion])
                else:
                    mensaje_respuesta += f" ({len(detalles_edicion)} cambios aplicados)"
                
                return ChatResponse(
                    success=True,
                    mensaje=mensaje_respuesta,
                    accion_ejecutada="editar_celdas",
                    datos_creados={
                        "celdas_editadas": celdas_editadas,
                        "detalles": detalles_edicion[:20],
                        "sheet_id": target_sheet_id,
                        "elementos": [
                            {
                                "tipo": "hoja",
                                "id": target_sheet_id,
                                "nombre": (
                                    ("2.1 Datos Parcelas" if target_sheet_id == "parcelas" else
                                     "Productos Fitosanitarios" if target_sheet_id == "productos" else
                                     "3.1 Registro Tratamientos" if target_sheet_id == "tratamientos" else
                                     "Registro Fertilizantes" if target_sheet_id == "fertilizantes" else
                                     "Registro Cosecha" if target_sheet_id == "cosecha" else
                                     (cuaderno.obtener_hoja(target_sheet_id).nombre if cuaderno.obtener_hoja(target_sheet_id) else target_sheet_id))
                                )
                            }
                        ]
                    },
                    elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                    advertencias=advertencias if advertencias else None
                )
            else:
                # Fallback anti-no-op:
                # Si GPT devolvió 0 cambios pero hay selección en contexto, reintentar automáticamente
                # aplicando sobre la selección con inferencia simple de columna/valor.
                fallback_editadas = 0
                fallback_detalles = []

                if isinstance(selected_ctx, dict) and selected_ctx.get("rows"):
                    target_sheet_id = target_sheet_id or selected_ctx.get("sheet_id", "")
                    selected_rows_ctx = selected_ctx.get("rows", []) or []
                    msg_raw = (comando.mensaje or "").strip()
                    msg_lower = msg_raw.lower()

                    # Inferir valor principal desde lenguaje natural
                    valor_inferido = None
                    m = re.search(r"\ba\s+['\"]?([^'\"]+?)['\"]?$", msg_raw, flags=re.IGNORECASE)
                    if m:
                        valor_inferido = m.group(1).strip()
                    if not valor_inferido:
                        m = re.search(r"\bpor\s+['\"]?([^'\"]+?)['\"]?$", msg_raw, flags=re.IGNORECASE)
                        if m:
                            valor_inferido = m.group(1).strip()
                    if not valor_inferido:
                        m = re.search(r"\bpon(?:er)?\s+['\"]?([^'\"]+?)['\"]?$", msg_raw, flags=re.IGNORECASE)
                        if m:
                            valor_inferido = m.group(1).strip()

                    # Construir set de columnas seleccionadas y etiquetas
                    cols_selected = {}
                    for r in selected_rows_ctx:
                        for c in (r.get("cells", []) if isinstance(r, dict) else []):
                            ckey = str(c.get("col_key", "")).strip()
                            clabel = str(c.get("col_label", "")).strip()
                            if ckey and ckey not in cols_selected:
                                cols_selected[ckey] = clabel

                    # Inferir columna objetivo por texto
                    def _contains_any(texto, palabras):
                        t = (texto or "").lower()
                        return any(p in t for p in palabras)

                    cols_objetivo = []
                    for ckey, clabel in cols_selected.items():
                        n = f"{ckey} {clabel}".lower()
                        if _contains_any(msg_lower, [n, ckey.lower(), clabel.lower()]):
                            cols_objetivo.append(ckey)
                        elif _contains_any(msg_lower, ["municipio", "término", "termino"]) and _contains_any(n, ["municipio", "término", "termino"]):
                            cols_objetivo.append(ckey)
                        elif _contains_any(msg_lower, ["cultivo", "especie"]) and _contains_any(n, ["cultivo", "especie"]):
                            cols_objetivo.append(ckey)
                        elif _contains_any(msg_lower, ["fecha inicio", "desde"]) and _contains_any(n, ["fecha_inicio", "fecha inicio", "desde", "inicio"]):
                            cols_objetivo.append(ckey)
                        elif _contains_any(msg_lower, ["fecha fin", "hasta"]) and _contains_any(n, ["fecha_fin", "fecha fin", "hasta", "fin"]):
                            cols_objetivo.append(ckey)
                        elif _contains_any(msg_lower, ["fecha"]) and _contains_any(n, ["fecha"]):
                            cols_objetivo.append(ckey)

                    # Si no se detecta columna explícita:
                    # - si hay 1 columna seleccionada, usar esa
                    # - si no, usar todas las seleccionadas para no perder la intención del usuario
                    if not cols_objetivo:
                        if len(cols_selected) == 1:
                            cols_objetivo = list(cols_selected.keys())
                        else:
                            cols_objetivo = list(cols_selected.keys())

                    # Valores especiales para fechas relativas
                    hoy = datetime.now().date().isoformat()
                    manana = (datetime.now().date()).fromordinal(datetime.now().date().toordinal() + 1).isoformat()

                    for r in selected_rows_ctx:
                        row_id = str(r.get("row_id", "")).strip()
                        row_index = r.get("row_index")
                        cells = r.get("cells", []) if isinstance(r, dict) else []
                        if not row_id and row_index is not None:
                            row_id = str(row_index)
                        if not row_id:
                            continue

                        row_cell_keys = {str(c.get("col_key", "")).strip() for c in cells}
                        row_target_cols = [c for c in cols_objetivo if c in row_cell_keys] or list(row_cell_keys)

                        for col_key in row_target_cols:
                            nuevo_valor = valor_inferido if valor_inferido is not None else ""
                            col_meta = cols_selected.get(col_key, "").lower()

                            # Heurística específica: "hoy hasta mañana" en columnas de inicio/fin
                            if "hoy" in msg_lower and ("mañana" in msg_lower or "manana" in msg_lower):
                                if any(k in col_meta for k in ["inicio", "desde", "fecha_inicio", "fecha inicio"]):
                                    nuevo_valor = hoy
                                elif any(k in col_meta for k in ["fin", "hasta", "fecha_fin", "fecha fin"]):
                                    nuevo_valor = manana
                                elif "fecha" in col_meta:
                                    nuevo_valor = hoy

                            try:
                                # Hojas base del sistema
                                if target_sheet_id in ("parcelas", "productos", "tratamientos", "fertilizantes", "cosecha"):
                                    if target_sheet_id == "parcelas":
                                        items = cuaderno.parcelas
                                    elif target_sheet_id == "productos":
                                        items = cuaderno.productos
                                    elif target_sheet_id == "tratamientos":
                                        items = cuaderno.tratamientos
                                    elif target_sheet_id == "fertilizantes":
                                        items = getattr(cuaderno, "fertilizaciones", None) or []
                                    elif target_sheet_id == "cosecha":
                                        items = getattr(cuaderno, "cosechas", None) or []
                                    else:
                                        items = []

                                    target_item = next((item for item in items if item.id == row_id), None)
                                    if target_item and hasattr(target_item, col_key):
                                        old_value = getattr(target_item, col_key, "")
                                        field_type = type(old_value) if old_value is not None else str
                                        try:
                                            if field_type == float or field_type == int:
                                                typed_value = float(nuevo_valor) if "." in str(nuevo_valor) else int(nuevo_valor)
                                            elif field_type == bool:
                                                typed_value = str(nuevo_valor).lower() in ("true", "1", "sí", "si", "yes")
                                            else:
                                                typed_value = str(nuevo_valor)
                                        except (ValueError, TypeError):
                                            typed_value = str(nuevo_valor)

                                        setattr(target_item, col_key, typed_value)
                                        fallback_editadas += 1
                                        fallback_detalles.append(f"{col_key}: '{old_value}' → '{typed_value}'")
                                else:
                                    # Hoja importada
                                    hoja = cuaderno.obtener_hoja(target_sheet_id)
                                    if not hoja:
                                        continue
                                    try:
                                        row_idx = int(row_id)
                                    except ValueError:
                                        row_idx = -1
                                    if not (0 <= row_idx < len(hoja.datos)):
                                        continue

                                    try:
                                        col_idx = int(col_key)
                                    except ValueError:
                                        col_idx = -1
                                        col_key_norm = str(col_key).strip().lower()
                                        if col_key_norm.startswith("col "):
                                            try:
                                                col_idx = int(col_key_norm.split(" ", 1)[1]) - 1
                                            except Exception:
                                                col_idx = -1
                                        if col_idx < 0 and hoja.columnas:
                                            for i_col, col_name in enumerate(hoja.columnas):
                                                if str(col_name).strip().lower() == col_key_norm:
                                                    col_idx = i_col
                                                    break
                                    if col_idx < 0:
                                        continue

                                    fila = hoja.datos[row_idx]
                                    if col_idx >= len(fila):
                                        fila.extend([""] * (col_idx - len(fila) + 1))
                                    old_val = fila[col_idx]
                                    fila[col_idx] = nuevo_valor
                                    fallback_editadas += 1
                                    col_name = hoja.columnas[col_idx] if col_idx < len(hoja.columnas) else str(col_idx)
                                    fallback_detalles.append(f"Fila {row_idx} [{col_name}]: '{old_val}' → '{nuevo_valor}'")
                            except Exception:
                                continue

                if fallback_editadas > 0:
                    storage.guardar(cuaderno)
                    mensaje_respuesta = f"✅ {fallback_editadas} celda(s) editada(s) (modo automático de recuperación)."
                    if len(fallback_detalles) <= 20:
                        mensaje_respuesta += ":\n" + "\n".join([f"  • {d}" for d in fallback_detalles])
                    return ChatResponse(
                        success=True,
                        mensaje=mensaje_respuesta,
                        accion_ejecutada="editar_celdas",
                        datos_creados={
                            "celdas_editadas": fallback_editadas,
                            "detalles": fallback_detalles[:20],
                            "sheet_id": target_sheet_id,
                            "elementos": [
                                {
                                    "tipo": "hoja",
                                    "id": target_sheet_id,
                                    "nombre": (
                                        ("2.1 Datos Parcelas" if target_sheet_id == "parcelas" else
                                         "Productos Fitosanitarios" if target_sheet_id == "productos" else
                                         "3.1 Registro Tratamientos" if target_sheet_id == "tratamientos" else
                                         "Registro Fertilizantes" if target_sheet_id == "fertilizantes" else
                                         "Registro Cosecha" if target_sheet_id == "cosecha" else
                                         (cuaderno.obtener_hoja(target_sheet_id).nombre if cuaderno.obtener_hoja(target_sheet_id) else target_sheet_id))
                                    )
                                }
                            ]
                        },
                        elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                        advertencias=advertencias if advertencias else None
                    )

                return ChatResponse(
                    success=True,
                    mensaje="💡 No pude aplicar los cambios automáticamente esta vez. Repite la instrucción indicando columna y valor (ej: 'cambia municipio a Córdoba').",
                    sugerencias=[
                        "Selecciona las celdas y escribe: cambia municipio a Córdoba",
                        "Selecciona filas y escribe: cambia cultivo a Olivo",
                        "Selecciona y escribe: pon fecha_inicio hoy y fecha_fin mañana"
                    ]
                )
        
        else:
            # Respuesta general / acción no reconocida
            mensaje_respuesta = mensaje_gpt or "¿En qué puedo ayudarte?"
            if elementos_no_procesados:
                mensaje_respuesta += f"\n\n⚠️ No se pudo procesar lo siguiente (no está en el sistema o no se entendió):\n" + "\n".join([f"• {e}" for e in elementos_no_procesados])
                mensaje_respuesta += "\n\n✅ Todo lo demás se ha procesado correctamente."
            
            return ChatResponse(
                success=True,
                mensaje=mensaje_respuesta,
                sugerencias=[
                    "Añade 5 parcelas con olivos",
                    "Pon productos: Roundup, Decis, Confidor",
                    "Genera tratamientos para las parcelas",
                    "Busca dónde está el nombre del titular",
                    "Cambia el nombre en la hoja de datos"
                ],
                elementos_no_procesados=elementos_no_procesados if elementos_no_procesados else None,
                advertencias=advertencias if advertencias else None
            )
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return ChatResponse(
            success=True,
            mensaje="💡 Algo no salió como esperaba, pero no te preocupes. Prueba con alguna de estas opciones:",
            sugerencias=["Crea 5 parcelas de olivos en Madrid", "Añade producto Roundup", "Ordena las parcelas", "Lista lo que hay en el cuaderno"]
        )
