"""
PRUEBA 7 — Enterprise End-to-End
Valida: Lectura → Preview → Commit → Excel final auditable
"""
import sys
import json
import hashlib
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from src.ai_processor import AIProcessor, resolve_date
from src.parcel_manager import ParcelManager

# ============================================================
# CONFIGURACIÓN
# ============================================================
INPUT_FILE = "/Volumes/Uniclick4TB/organizadorhndezbueno/PABLO PEREZ RUBIO 2025.xlsx"
OUTPUT_ORGANIZED = "/tmp/OUT_organizado.xlsx"
OUTPUT_TREATED = "/tmp/OUT_organizado_tratado.xlsx"

# Comando complejo a probar
TEST_COMMAND = "Aplica Glifosato 3 l/ha hoy en todas las CEBADAS de RASUEROS, excepto recinto 2. Notas: preventivo."


def print_section(title):
    print(f"\n{'='*70}")
    print(f"  {title}")
    print(f"{'='*70}")


def calculate_sheet_fingerprint(filepath):
    """Calcula fingerprint de sheets (nombre, rows, cols)"""
    wb = load_workbook(filepath, data_only=True)
    fingerprint = {}
    for name in wb.sheetnames:
        sheet = wb[name]
        fingerprint[name] = {
            "rows": sheet.max_row,
            "cols": sheet.max_column
        }
    wb.close()
    return fingerprint


def validate_excel_structure(filepath):
    """Valida que la estructura del Excel esté correcta"""
    errors = []
    wb = load_workbook(filepath, data_only=True)
    
    # Sheets críticas que deben existir
    critical_sheets = ["2.1. DATOS PARCELAS"]
    for sheet in critical_sheets:
        if sheet not in wb.sheetnames:
            errors.append(f"Falta hoja crítica: {sheet}")
    
    wb.close()
    return errors


def validate_treatment_rows(filepath, expected_count, exclude_recintos=None):
    """Valida las filas de tratamiento generadas"""
    exclude_recintos = exclude_recintos or []
    results = {
        "rows_found": 0,
        "duplicates": 0,
        "invalid_dates": 0,
        "surface_mismatches": 0,
        "excluded_recintos_present": 0,
        "errors": []
    }
    
    wb = load_workbook(filepath, data_only=True)
    
    if "Tratamientos_IA" not in wb.sheetnames:
        results["errors"].append("No se encontró hoja Tratamientos_IA")
        wb.close()
        return results
    
    sheet = wb["Tratamientos_IA"]
    
    # Leer filas (saltando header)
    seen_keys = set()
    for row in range(2, sheet.max_row + 1):
        results["rows_found"] += 1
        
        # Columnas esperadas: Nº, FechaIni, FechaFin, Producto, ..., Polígono, Parcela, Recinto
        fecha = sheet.cell(row, 2).value
        producto = sheet.cell(row, 4).value
        poligono = sheet.cell(row, 11).value
        parcela = sheet.cell(row, 12).value
        recinto = sheet.cell(row, 13).value
        
        # Validar fecha
        if fecha:
            try:
                if isinstance(fecha, str) and "/" in fecha:
                    datetime.strptime(fecha, "%d/%m/%Y")
            except ValueError:
                results["invalid_dates"] += 1
        
        # Validar recinto excluido
        try:
            recinto_int = int(recinto) if recinto else 0
            if recinto_int in exclude_recintos:
                results["excluded_recintos_present"] += 1
        except:
            pass
        
        # Detectar duplicados
        key = (str(poligono), str(parcela), str(recinto), str(fecha), str(producto))
        if key in seen_keys:
            results["duplicates"] += 1
        seen_keys.add(key)
    
    wb.close()
    return results


def test_7_enterprise():
    """Prueba Enterprise End-to-End"""
    
    checklist = {
        "input_exists": False,
        "structure_ok": False,
        "parcels_loaded": False,
        "ai_intent_ok": False,
        "preview_ok": False,
        "commit_ok": False,
        "rows_match": False,
        "no_duplicates": False,
        "no_invalid_dates": False,
        "no_excluded_recintos": False,
    }
    
    # ============================================================
    # PASO 0: Verificar archivo de entrada
    # ============================================================
    print_section("PASO 0: Verificar Input")
    
    if not Path(INPUT_FILE).exists():
        print(f"❌ Archivo no encontrado: {INPUT_FILE}")
        return checklist
    
    checklist["input_exists"] = True
    print(f"✅ Archivo existe: {INPUT_FILE}")
    
    # Fingerprint original
    fp_original = calculate_sheet_fingerprint(INPUT_FILE)
    print(f"📊 Sheets en original: {list(fp_original.keys())}")
    
    # Validar estructura
    errors = validate_excel_structure(INPUT_FILE)
    if errors:
        print(f"❌ Errores de estructura: {errors}")
    else:
        checklist["structure_ok"] = True
        print("✅ Estructura OK")
    
    # ============================================================
    # PASO 1: Cargar Parcelas
    # ============================================================
    print_section("PASO 1: Cargar Parcelas")
    
    pm = ParcelManager(INPUT_FILE)
    parcels = pm.get_parcels()
    context = pm.get_context_for_ai()
    
    if parcels:
        checklist["parcels_loaded"] = True
        print(f"✅ Parcelas cargadas: {len(parcels)}")
        print(f"   Cultivos: {context['crops']}")
        print(f"   Municipios: {context['municipios']}")
    else:
        print("❌ No se cargaron parcelas")
        return checklist
    
    # Verificar que tenemos CEBADAS en RASUEROS
    cebadas_rasueros = [p for p in parcels if "CEBADA" in p.crop and "RASUEROS" in p.municipio]
    print(f"📊 CEBADAS en RASUEROS: {len(cebadas_rasueros)}")
    for p in cebadas_rasueros[:5]:
        print(f"   - Pol {p.polygon}, Parc {p.parcel}, Rec {p.recinto}: {p.surface_ha} ha")
    
    # ============================================================
    # PASO 2: Interpretar con IA (Preview)
    # ============================================================
    print_section("PASO 2: Interpretar Comando con IA")
    
    print(f"💬 Comando: \"{TEST_COMMAND}\"")
    
    ai = AIProcessor()
    intent = ai.interpret_command(TEST_COMMAND, context)
    
    if "error" in intent:
        print(f"❌ Error IA: {intent['error']}")
        return checklist
    
    checklist["ai_intent_ok"] = True
    print(f"✅ Intent recibido:")
    print(json.dumps(intent, indent=2, ensure_ascii=False))
    
    # Extraer datos del intent
    filters = intent.get("filters", {})
    targets = filters.get("targets", {})
    logic = filters.get("logic", "AND")
    scope = targets.get("scope", "")
    treatment_data = intent.get("data", {})
    exclude_recinto = targets.get("exclude_recinto", [])
    
    print(f"\n📊 Verificaciones del Intent:")
    print(f"   • Scope: {scope}")
    print(f"   • Crop: {targets.get('crop')}")
    print(f"   • Municipio: {targets.get('municipio')}")
    print(f"   • exclude_recinto: {exclude_recinto}")
    print(f"   • Notas: {treatment_data.get('notes')}")
    
    # ============================================================
    # PASO 3: Filtrar y Generar Preview
    # ============================================================
    print_section("PASO 3: Generar Preview")
    
    matched = pm.filter_parcels(scope, targets, logic)
    
    print(f"📊 Parcelas coincidentes: {len(matched)}")
    requires_confirmation = len(matched) > 20
    print(f"   requires_confirmation: {requires_confirmation}")
    
    if not matched:
        print("❌ No hay parcelas coincidentes")
        return checklist
    
    # Verificar que todas son CEBADA y RASUEROS
    all_cebada = all("CEBADA" in p.crop for p in matched)
    all_rasueros = all("RASUEROS" in p.municipio for p in matched)
    no_recinto_2 = all(p.recinto not in exclude_recinto for p in matched)
    
    print(f"   • Todas CEBADA: {'✅' if all_cebada else '❌'}")
    print(f"   • Todas RASUEROS: {'✅' if all_rasueros else '❌'}")
    print(f"   • Recinto 2 excluido: {'✅' if no_recinto_2 else '❌'}")
    
    if all_cebada and all_rasueros and no_recinto_2:
        checklist["preview_ok"] = True
        print("✅ Preview OK")
    
    for p in matched[:5]:
        print(f"   - Pol {p.polygon}, Parc {p.parcel}, Rec {p.recinto}: {p.crop} @ {p.municipio}")
    
    # ============================================================
    # PASO 4: Commit (Escribir tratamientos)
    # ============================================================
    print_section("PASO 4: Commit")
    
    # Resolver fecha
    date_obj = treatment_data.get("date", {})
    resolved_date = resolve_date(date_obj)
    print(f"📅 Fecha resuelta: {resolved_date}")
    
    # Copiar archivo original a output
    import shutil
    shutil.copy(INPUT_FILE, OUTPUT_TREATED)
    
    # Crear ParcelManager sobre la copia
    pm_out = ParcelManager(OUTPUT_TREATED)
    
    # Generar filas
    rows = pm_out.generate_treatment_rows(matched, treatment_data, resolved_date, start_order=1)
    print(f"📝 Filas generadas: {len(rows)}")
    
    # Escribir
    output_path, count = pm_out.write_treatments(rows, output_path=OUTPUT_TREATED)
    print(f"💾 Escrito: {count} filas en {output_path}")
    
    if count == len(matched):
        checklist["commit_ok"] = True
        print("✅ Commit OK")
    
    # ============================================================
    # PASO 5: Validación Final
    # ============================================================
    print_section("PASO 5: Validación Final")
    
    validation = validate_treatment_rows(OUTPUT_TREATED, len(matched), exclude_recinto)
    
    print(f"📊 Resultados de validación:")
    print(f"   • Filas encontradas: {validation['rows_found']} (esperadas: {len(matched)})")
    print(f"   • Duplicados: {validation['duplicates']}")
    print(f"   • Fechas inválidas: {validation['invalid_dates']}")
    print(f"   • Recintos excluidos presentes: {validation['excluded_recintos_present']}")
    
    checklist["rows_match"] = validation["rows_found"] == len(matched)
    checklist["no_duplicates"] = validation["duplicates"] == 0
    checklist["no_invalid_dates"] = validation["invalid_dates"] == 0
    checklist["no_excluded_recintos"] = validation["excluded_recintos_present"] == 0
    
    # ============================================================
    # RESUMEN FINAL
    # ============================================================
    print_section("RESUMEN CHECKLIST")
    
    for key, passed in checklist.items():
        status = "✅" if passed else "❌"
        print(f"  {status} {key}")
    
    all_passed = all(checklist.values())
    print(f"\n{'🎉 PRUEBA ENTERPRISE PASADA!' if all_passed else '⚠️ Algunos checks fallaron'}")
    print(f"\n📁 Archivo generado: {OUTPUT_TREATED}")
    print(f"   → Abre en Excel y verifica manualmente la hoja 'Tratamientos_IA'")
    
    return checklist


if __name__ == "__main__":
    test_7_enterprise()
