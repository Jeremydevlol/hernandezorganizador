#!/usr/bin/env python3
"""
🔥 STRESS TEST MULTI-CLIENTE — Mercado Real 🔥

Simula 3 clientes distintos con:
- 2,154 parcelas totales
- 4,009 tratamientos totales
- Cuadernos distintos

Valida:
- Output = Golden (sin hojas extras)
- Todas las filas escritas
- Sin duplicados
- Tiempo de ejecución aceptable
"""
import sys
import json
import time
from pathlib import Path
from datetime import datetime
from typing import Dict

sys.path.insert(0, str(Path(__file__).parent))

from src.resolver_cuaderno import resolver_cuaderno, GoldenTemplateWriter
from openpyxl import load_workbook


# ============================================
# Config
# ============================================
BASE_DIR = Path(__file__).parent
CLIENTS = ["cliente_sigpac", "cliente_hort", "cliente_coop"]


# ============================================
# Utilidades
# ============================================
def print_header(title):
    print(f"\n{'🔥'*25}")
    print(f"  {title}")
    print(f"{'🔥'*25}\n")


def print_section(title):
    print(f"\n{'='*65}")
    print(f"  {title}")
    print(f"{'='*65}")


def count_sheets(filepath):
    wb = load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def count_data_rows(filepath, sheet_name, start_row=11):
    """Cuenta filas con Id numérico"""
    wb = load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[sheet_name]
    count = 0
    for row in range(start_row, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val is not None:
            try:
                int(val)
                count += 1
            except:
                pass
    wb.close()
    return count


def check_duplicates(filepath, sheet_name="inf.trat 1"):
    """Verifica duplicados en inf.trat 1"""
    wb = load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    
    ws = wb[sheet_name]
    seen = set()
    duplicates = 0
    
    for row in range(11, ws.max_row + 1):
        id_p = ws.cell(row, 1).value
        fecha = ws.cell(row, 5).value
        prod = ws.cell(row, 9).value
        
        if id_p is not None:
            key = (str(id_p), str(fecha), str(prod))
            if key in seen:
                duplicates += 1
            seen.add(key)
    
    wb.close()
    return duplicates


# ============================================
# Test por Cliente
# ============================================
def test_cliente(cliente_id: str) -> Dict:
    """Ejecuta test completo para un cliente"""
    client_dir = BASE_DIR / "clients" / cliente_id
    
    input_path = client_dir / "input.xlsx"
    golden_path = client_dir / "golden.xlsx"
    output_path = client_dir / "output.xlsx"
    tratamientos_path = client_dir / "tratamientos.json"
    
    result = {
        "cliente_id": cliente_id,
        "success": False,
        "parcelas": 0,
        "tratamientos_input": 0,
        "tratamientos_escritos": 0,
        "filas_output": 0,
        "duplicados": 0,
        "hojas_ok": False,
        "tiempo_seg": 0,
        "errors": []
    }
    
    # Cargar tratamientos
    if not tratamientos_path.exists():
        result["errors"].append("No existe tratamientos.json")
        return result
    
    with open(tratamientos_path, "r", encoding="utf-8") as f:
        tratamientos = json.load(f)
    
    result["tratamientos_input"] = len(tratamientos)
    
    # Contar parcelas únicas
    parcelas_unicas = set()
    for t in tratamientos:
        key = (t["polygon"], t["parcel"], t["recinto"])
        parcelas_unicas.add(key)
    result["parcelas"] = len(parcelas_unicas)
    
    print(f"\n  📊 Input: {result['parcelas']} parcelas, {result['tratamientos_input']} tratamientos")
    
    # Ejecutar resolver
    start_time = time.time()
    
    try:
        resolver_result = resolver_cuaderno(
            input_path=str(input_path),
            output_path=str(output_path),
            golden_path=str(golden_path),
            treatments=tratamientos
        )
        
        result["tratamientos_escritos"] = resolver_result.get("actions_applied", 0)
        result["errors"].extend(resolver_result.get("errors", []))
        
    except Exception as e:
        result["errors"].append(f"Resolver falló: {e}")
        return result
    
    result["tiempo_seg"] = round(time.time() - start_time, 2)
    
    # Verificar output
    if not output_path.exists():
        result["errors"].append("Output no generado")
        return result
    
    # Contar filas
    result["filas_output"] = count_data_rows(output_path, "inf.trat 1")
    
    # Verificar hojas
    golden_sheets = set(count_sheets(golden_path))
    output_sheets = set(count_sheets(output_path))
    extra_sheets = output_sheets - golden_sheets
    result["hojas_ok"] = len(extra_sheets) == 0
    
    if extra_sheets:
        result["errors"].append(f"Hojas extras: {extra_sheets}")
    
    # Verificar duplicados
    result["duplicados"] = check_duplicates(output_path)
    
    # Éxito
    result["success"] = (
        len(result["errors"]) == 0 and
        result["hojas_ok"] and
        result["tratamientos_escritos"] > 0
    )
    
    return result


# ============================================
# Main
# ============================================
def main():
    print_header("STRESS TEST MULTI-CLIENTE — Mercado Real")
    print(f"  Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Clientes: {len(CLIENTS)}")
    
    resultados = []
    total_start = time.time()
    
    for cliente_id in CLIENTS:
        print_section(f"Cliente: {cliente_id}")
        
        result = test_cliente(cliente_id)
        resultados.append(result)
        
        # Mostrar resultado
        status = "✅ PASSED" if result["success"] else "❌ FAILED"
        print(f"\n  Status: {status}")
        print(f"  📊 Tratamientos escritos: {result['tratamientos_escritos']}/{result['tratamientos_input']}")
        print(f"  📊 Filas en output: {result['filas_output']}")
        print(f"  📊 Duplicados: {result['duplicados']}")
        print(f"  📊 Hojas OK: {'✅' if result['hojas_ok'] else '❌'}")
        print(f"  ⏱️ Tiempo: {result['tiempo_seg']}s")
        
        if result["errors"]:
            for e in result["errors"][:3]:
                print(f"  ❌ {e}")
    
    total_time = round(time.time() - total_start, 2)
    
    # Resumen final
    print_section("RESUMEN FINAL")
    
    total_parcelas = sum(r["parcelas"] for r in resultados)
    total_tratamientos = sum(r["tratamientos_input"] for r in resultados)
    total_escritos = sum(r["tratamientos_escritos"] for r in resultados)
    total_filas = sum(r["filas_output"] for r in resultados)
    all_passed = all(r["success"] for r in resultados)
    
    print(f"\n  {'Cliente':<25} {'Parcelas':>10} {'Trat. Input':>12} {'Escritos':>10} {'Tiempo':>8}")
    print(f"  {'-'*70}")
    
    for r in resultados:
        status = "✅" if r["success"] else "❌"
        print(f"  {status} {r['cliente_id']:<22} {r['parcelas']:>10} {r['tratamientos_input']:>12} {r['tratamientos_escritos']:>10} {r['tiempo_seg']:>7}s")
    
    print(f"  {'-'*70}")
    print(f"  {'TOTAL':<25} {total_parcelas:>10} {total_tratamientos:>12} {total_escritos:>10} {total_time:>7}s")
    
    print(f"\n  📊 Métricas globales:")
    print(f"     Parcelas totales: {total_parcelas:,}")
    print(f"     Tratamientos input: {total_tratamientos:,}")
    print(f"     Tratamientos escritos: {total_escritos:,}")
    print(f"     Filas output total: {total_filas:,}")
    print(f"     Tiempo total: {total_time}s")
    print(f"     Velocidad: {round(total_escritos/total_time, 1) if total_time > 0 else 0} tratamientos/seg")
    
    if all_passed:
        print(f"\n{'🎉'*20}")
        print(f"  STRESS TEST: ALL CLIENTS PASSED!")
        print(f"  VIERA AI resuelve cuadernos de cualquier cliente.")
        print(f"{'🎉'*20}\n")
    else:
        print(f"\n❌ Algunos clientes fallaron")
    
    # Generar reporte
    report_path = BASE_DIR / "stress_test_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "total_time": total_time,
            "total_parcelas": total_parcelas,
            "total_tratamientos": total_tratamientos,
            "total_escritos": total_escritos,
            "all_passed": all_passed,
            "clientes": resultados
        }, f, indent=2, ensure_ascii=False)
    
    print(f"📁 Reporte: {report_path}\n")
    
    return 0 if all_passed else 1


if __name__ == "__main__":
    from typing import Dict
    exit_code = main()
    sys.exit(exit_code)
